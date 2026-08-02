"""LangChain agent that analyzes irrigation data via @tool functions.

The agent is built on demand from AISettings (configured in admin). Each tool
queries the ORM directly — same process, no HTTP hop — so the LLM gets real data
to ground its analysis on.

Tool-calling follows the LangChain v1 pattern: runtime context (the per-session
workspace thread id) is declared via ``context_schema`` and read inside tools
through ``ToolRuntime`` rather than a module-level ContextVar.
"""
import csv
import json
import logging
import os
import subprocess
import sys
from datetime import date, timedelta
from functools import lru_cache
from typing import TypedDict

from django.conf import settings
from django.utils import timezone
from langchain.agents import create_agent
from langchain.tools import tool, ToolRuntime
from langchain_openai import ChatOpenAI
from langgraph.checkpoint.sqlite import SqliteSaver

from core.models import (
    AISettings, WorkReport, WorkReportEntry, WorkItem, Project,
    Zone, Patch, WeatherData, WaterRequest,
)

logger = logging.getLogger(__name__)

# Per-run context passed to agent.stream(context=...). Tools access it through
# ToolRuntime instead of a global ContextVar — the LangChain v1 recommendation
# (replaces the older InjectedState / get_runtime / ContextVar patterns).
class WorkorderContext(TypedDict):
    thread_id: str


# ── Code-execution workspace management ───────────────────────────────────

WORKSPACE_ROOT = os.path.join(settings.MEDIA_ROOT, 'ai_workspaces')


def ensure_workspace(thread_id):
    """Return the per-session workspace dir, creating it and preloading data on first use."""
    ws = os.path.join(WORKSPACE_ROOT, thread_id)
    marker = os.path.join(ws, '.populated')
    if not os.path.isdir(ws):
        os.makedirs(ws, exist_ok=True)
    if not os.path.exists(marker):
        _populate_workspace_data(ws)
        with open(marker, 'w') as f:
            f.write(timezone.now().isoformat())
    return ws


# Section labels are reused across the CSV export and the report tools.
_SECTION_LABELS = dict(WorkItem.SECTION_CHOICES)


def _report_section_summary(report):
    """Comma-joined section-label summary for a report, derived from its tree entries.

    The legacy WorkReport.work_category FK was removed; the new architecture stores
    work content as WorkReportEntry rows under the WorkItem template tree. This
    collapses a report's entries to the distinct top-level sections (章节) they
    fall under, e.g. "常规维护, 报修应急".
    """
    sections = []
    seen = set()
    for e in report.entries.all():
        sec = getattr(e.work_item, 'section', '') if e.work_item_id else ''
        if sec and sec not in seen:
            seen.add(sec)
            sections.append(_SECTION_LABELS.get(sec, sec))
    return ', '.join(sections)


def _populate_workspace_data(ws):
    """Export recent business data to CSVs the generated code can pd.read_csv."""
    today = date.today()
    since = today - timedelta(days=90)
    # WorkReports (last 90 days). The old 工作类别 column is gone (work_category
    # was removed); we export the tree-derived 章节 summary instead, plus the
    # free-text 工作内容 column already on the model.
    wr_qs = (WorkReport.objects.filter(date__gte=since)
             .select_related('worker', 'location')
             .prefetch_related('entries__work_item')
             .order_by('-date'))
    with open(os.path.join(ws, 'work_reports.csv'), 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['日期', '处理人', '位置编号', '位置名称', '班次', '章节',
                    '灌溉组工时', '第三方工时', '疑难', '待修', '工作内容', '备注'])
        for r in wr_qs:
            w.writerow([
                r.date.isoformat(), str(r.worker) if r.worker else '',
                getattr(r.location, 'code', '') if r.location else '',
                str(r.location) if r.location else '',
                r.shift or '',
                _report_section_summary(r),
                r.team_hours or 0, r.third_party_hours or 0,
                '是' if r.is_difficult else '否',
                '是' if r.is_pending_repair else '否',
                (r.work_content or '')[:200],
                (r.remark or '')[:200],
            ])
    # WorkReportEntries (structured 现场作业记录 tree content, last 90 days)
    we_qs = WorkReportEntry.objects.filter(
        work_report__date__gte=since
    ).select_related('work_report', 'work_item', 'project').order_by('-work_report__date')
    with open(os.path.join(ws, 'work_entries.csv'), 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['日期', '章节', '节点编码', '节点名称', '值类型', '项目', '数量', '状态', '文本'])
        for e in we_qs:
            wi = e.work_item
            w.writerow([
                e.work_report.date.isoformat(),
                _SECTION_LABELS.get(wi.section, wi.section),
                wi.code, wi.name_zh, wi.value_type,
                str(e.project) if e.project else '',
                e.count, e.status, (e.text_value or '')[:200],
            ])
    # Zones (all)
    with open(os.path.join(ws, 'zones.csv'), 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['编号', '通用名称', '片区', '面积㎡', '优先级', '灌水器类型', '当前状态'])
        for z in Zone.objects.select_related('patch'):
            w.writerow([
                z.code, z.name, str(z.patch) if z.patch else '',
                z.area_sqm or '', z.get_priority_display() if z.priority else '',
                z.sprinkler_type or '', z.current_status or '',
            ])

    # ── 跨表关联分析用的扩展表 ──────────────────────────────────────────
    # 这些 CSV 让 run_python_code 能回答"项目预算消耗""待修×停水关联""天气×工单"
    # 等跨表问题，而不只是单表查询。数据范围与上面一致（近90天），项目表为全量。
    from core.models import (
        Project, WaterRequest, MaxicomWeatherLog, PMWorkOrder,
        InventoryTransaction,
    )

    # Projects (all) — 项目主数据 + 预算/费率/日期，配合 work_entries.csv 的项目列
    # 可算"某项目已用工时 vs 人工预算""哪些项目超预算/快完工"。
    with open(os.path.join(ws, 'projects.csv'), 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['项目名称', '类别', '子类别', '代号', 'Code', '是否完工',
                    '开工日期', '计划完工', '材料预算', '人工预算',
                    '灌溉组费率', '第三方费率', '备注'])
        for p in Project.objects.all().order_by('category', 'subcategory', 'name'):
            w.writerow([
                p.name, p.get_category_display(),
                p.get_subcategory_display() if p.subcategory else '',
                p.symbol or '', p.code or '',
                '是' if p.is_completed else '否',
                p.start_date.isoformat() if p.start_date else '',
                p.planned_end_date.isoformat() if p.planned_end_date else '',
                float(p.material_budget_amount) if p.material_budget_amount else '',
                float(p.labor_budget_amount) if p.labor_budget_amount else '',
                float(p.team_rate) if p.team_rate else '',
                float(p.third_party_rate) if p.third_party_rate else '',
                (p.notes or '')[:200],
            ])

    # Water requests (近90天) — 浇水协调需求，可关联 work_reports 的待修/区域。
    with open(os.path.join(ws, 'water_requests.csv'), 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['提交时间', '提交人', '用户类型', '需求类型', '状态',
                    '起始时间', '结束时间', '涉及区域'])
        wr_qs = WaterRequest.objects.filter(created_at__date__gte=since).select_related('submitter')
        status_map = dict(WaterRequest.STATUS_CHOICES)
        for req in wr_qs.order_by('-created_at'):
            zones = ', '.join(z.code for z in req.zones.all()[:8]) or '—'
            w.writerow([
                timezone.localtime(req.created_at).strftime('%Y-%m-%d %H:%M'),
                str(req.submitter) if req.submitter else '',
                req.user_type, req.get_request_type_display(),
                status_map.get(req.status, req.status),
                timezone.localtime(req.start_datetime).strftime('%Y-%m-%d %H:%M') if req.start_datetime else '',
                timezone.localtime(req.end_datetime).strftime('%Y-%m-%d %H:%M') if req.end_datetime else '',
                zones,
            ])

    # Weather (近90天) — Maxicom 气象日志，可关联 work_reports 的日期做"天气×工单量"。
    with open(os.path.join(ws, 'weather.csv'), 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['日期', '站点', '温度', '最高温', '最低温', '降雨mm', '湿度', '风速', 'ET'])
        # timestamp is a compact string (YYYYMMDDHHMM); take the date prefix.
        wlog_qs = MaxicomWeatherLog.objects.filter(timestamp__gte=since.strftime('%Y%m%d'))
        for log in wlog_qs.order_by('-timestamp')[:2000]:
            ts = log.timestamp or ''
            day = f'{ts[:4]}-{ts[4:6]}-{ts[6:8]}' if len(ts) >= 8 else ''
            w.writerow([
                day, str(log.weather_station) if log.weather_station_id else '',
                log.temperature, log.max_temp, log.min_temp,
                log.rainfall, log.humidity, log.wind_run, log.et,
            ])

    # Project budget / consumption — 出库-项目的实际消耗（已确认的），按项目汇总。
    # 材料余额 = 材料预算 − Σ(数量×单价)。配合 projects.csv 算余额。
    # 单价取自 InventoryPriceRecord 的"当前价格"（与库存管理页 current_price_for 同口径）。
    with open(os.path.join(ws, 'project_consumption.csv'), 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['日期', '项目', '操作', '去向', '消耗模式', '物料编码', '物料名称',
                    '单位', '数量', '单价', '金额'])
        from core.inventory_tree_views import current_price_for
        txn_qs = (InventoryTransaction.objects.filter(date__gte=since)
                  .select_related('related_project').prefetch_related('lines__category'))
        for txn in txn_qs.order_by('-date'):
            proj = str(txn.related_project) if txn.related_project_id else ''
            for line in txn.lines.all():
                cat = line.category
                # 当前价格：优先 is_current，否则最新一条采购价。无价格记录→空。
                price_rec = current_price_for(cat) if cat else None
                unit_price = float(price_rec.unit_price) if price_rec and price_rec.unit_price else ''
                qty = line.quantity or 0
                amount = round(unit_price * qty, 2) if unit_price != '' else ''
                w.writerow([
                    txn.date.isoformat(), proj, txn.operation,
                    txn.entry_subtype or '', txn.consumption_mode or '',
                    getattr(cat, 'code', '') if cat else '',
                    getattr(cat, 'name_zh', '') if cat else '',
                    getattr(cat, 'unit', '') if cat else '',
                    qty, unit_price, amount,
                ])

    # PM work orders (近90天) — PM 完成工单，含工时/班次/区域。PM 工单与普通工单
    # 独立序列，不在 work_reports.csv 里，这里单独导出便于分析计划性维修执行情况。
    with open(os.path.join(ws, 'pm_workorders.csv'), 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['日期', '工单号', '处理人', '位置', '班次', '灌溉组工时', '第三方工时', '区域', '备注'])
        pm_qs = PMWorkOrder.objects.filter(date__gte=since).select_related('worker', 'location')
        for pm in pm_qs.order_by('-date', '-id'):
            w.writerow([
                pm.date.isoformat(), pm.display_number,
                str(pm.worker) if pm.worker_id else '',
                str(pm.location) if pm.location_id else '',
                pm.shift or '', pm.team_hours or 0, pm.third_party_hours or 0,
                pm.zone_names or '', (pm.remark or '')[:200],
            ])



# ── Tool implementations (call ORM directly) ──────────────────────────────


@tool
def query_work_reports(
    start_date: str = "",
    end_date: str = "",
    limit: int = 20,
) -> str:
    """查询维修工作日报。可按日期范围过滤。返回工单列表（日期、处理人、位置、班次、章节、工时、内容摘要）。

    Args:
        start_date: 起始日期 YYYY-MM-DD，留空默认最近7天
        end_date: 结束日期 YYYY-MM-DD，留空默认今天
        limit: 最多返回条数，默认20
    """
    today = date.today()
    end = _parse_date(end_date) or today
    start = _parse_date(start_date) or (end - timedelta(days=7))
    qs = (WorkReport.objects.filter(date__gte=start, date__lte=end)
          .select_related('worker', 'location')
          .prefetch_related('entries__work_item')
          .order_by('-date', '-id')[:limit])
    rows = []
    for r in qs:
        rows.append({
            '日期': r.date.isoformat(),
            '处理人': str(r.worker) if r.worker else '',
            '位置': str(r.location) if r.location else '',
            '班次': r.shift or '',
            '章节': _report_section_summary(r),
            '灌溉组工时': r.team_hours,
            '第三方工时': r.third_party_hours,
            '疑难': r.is_difficult,
            '待修': r.is_pending_repair,
            '工作内容': (r.work_content or r.remark or '')[:120],
        })
    return json.dumps({
        'count': WorkReport.objects.filter(date__gte=start, date__lte=end).count(),
        'returned': len(rows),
        'date_range': [start.isoformat(), end.isoformat()],
        'records': rows,
    }, ensure_ascii=False)


@tool
def query_work_report_stats(
    start_date: str = "",
    end_date: str = "",
) -> str:
    """统计维修工作日报的汇总数据：总工单数、总工时、按班次/章节/处理人的分布。用于趋势分析和报告制作。

    Args:
        start_date: 起始日期 YYYY-MM-DD，留空默认最近30天
        end_date: 结束日期 YYYY-MM-DD，留空默认今天
    """
    from django.db.models import Sum, Count, Q

    today = date.today()
    end = _parse_date(end_date) or today
    start = _parse_date(start_date) or (end - timedelta(days=30))
    base = WorkReport.objects.filter(date__gte=start, date__lte=end)
    total = base.count()
    agg = base.aggregate(
        team_hours=Sum('team_hours'),
        third_hours=Sum('third_party_hours'),
        difficult=Count('id', filter=Q(is_difficult=True)),
        pending=Count('id', filter=Q(is_pending_repair=True)),
    )

    # by shift
    shift_dist = {}
    for row in base.values('shift').annotate(c=Count('id')):
        shift_dist[row['shift'] or '未指定'] = row['c']
    # by top-level section (章节). The legacy 工作类别 FK was removed; a workorder's
    # section(s) now come from its WorkReportEntry rows under the WorkItem tree, so
    # the count here is "report × section" (a report touching two sections counts once
    # in each), which matches how the work-content is actually recorded.
    section_dist = {}
    entry_qs = (WorkReportEntry.objects
                .filter(work_report__in=base)
                .values('work_item__section')
                .annotate(c=Count('work_report', distinct=True))
                .order_by('-c'))
    for row in entry_qs:
        sec = row['work_item__section'] or '未指定'
        section_dist[_SECTION_LABELS.get(sec, sec)] = row['c']
    # by worker (top 10)
    worker_dist = []
    for row in base.values('worker__full_name').annotate(
        c=Count('id'), h=Sum('team_hours')
    ).order_by('-c')[:10]:
        worker_dist.append({
            '处理人': row['worker__full_name'] or '未知',
            '工单数': row['c'],
            '工时': row['h'] or 0,
        })
    return json.dumps({
        'date_range': [start.isoformat(), end.isoformat()],
        '总工单数': total,
        '总灌溉组工时': agg['team_hours'] or 0,
        '总第三方工时': agg['third_hours'] or 0,
        '疑难工单数': agg['difficult'] or 0,
        '待修工单数': agg['pending'] or 0,
        '按班次分布': shift_dist,
        '按章节分布': section_dist,
        '处理人工单Top10': worker_dist,
    }, ensure_ascii=False)


@tool
def query_work_entries_stats(
    start_date: str = "",
    end_date: str = "",
    section: str = "",
    project: str = "",
    top: int = 15,
) -> str:
    """统计工单「工作内容」明细(树形结构数据)：按章节/节点/灌溉项目汇总填报次数与数量。

    覆盖新现场作业记录表单的结构化明细，区别于 query_work_report_stats 的工单级汇总。
    可分析：哪些工作节点填报最频繁、各章节工作量分布、各灌溉项目投入等。

    Args:
        start_date: 起始日期 YYYY-MM-DD，留空默认最近30天
        end_date: 结束日期 YYYY-MM-DD，留空默认今天
        section: 章节过滤(可选)，值见返回中的「章节选项」，如 routine_maint/irrigation_project/repair_emergency
        project: 灌溉项目名称或代号模糊匹配(可选)
        top: 节点维度最多返回条数，默认15
    """
    from django.db.models import Sum, Count, Q

    today = date.today()
    end = _parse_date(end_date) or today
    start = _parse_date(start_date) or (end - timedelta(days=30))

    base = WorkReportEntry.objects.filter(
        work_report__date__gte=start, work_report__date__lte=end,
        work_item__active=True,
    )
    if section:
        base = base.filter(work_item__section=section)
    if project:
        base = base.filter(
            Q(project__name__icontains=project) | Q(project__code__icontains=project)
        )

    section_labels = dict(WorkItem.SECTION_CHOICES)
    total_entries = base.count()
    count_sum = base.filter(work_item__value_type='count').aggregate(s=Sum('count'))['s'] or 0

    # by section
    section_dist = {}
    for row in base.values('work_item__section').annotate(
        entries=Count('id'), counts=Sum('count')
    ).order_by('-entries'):
        key = row['work_item__section']
        section_dist[section_labels.get(key, key)] = {
            '填报次数': row['entries'],
            '合计数量': row['counts'] or 0,
        }

    # top nodes actually filled
    node_dist = []
    for row in base.values(
        'work_item__code', 'work_item__name_zh', 'work_item__value_type'
    ).annotate(entries=Count('id'), counts=Sum('count')).order_by('-entries')[:top]:
        node_dist.append({
            '编码': row['work_item__code'],
            '节点': row['work_item__name_zh'],
            '值类型': row['work_item__value_type'],
            '填报次数': row['entries'],
            '合计数量': row['counts'] or 0,
        })

    # by irrigation project
    project_dist = []
    for row in base.exclude(project__isnull=True).values(
        'project__name', 'project__category'
    ).annotate(entries=Count('id'), counts=Sum('count')).order_by('-entries'):
        project_dist.append({
            '项目': row['project__name'],
            '类别': row['project__category'],
            '填报次数': row['entries'],
            '合计数量': row['counts'] or 0,
        })

    return json.dumps({
        'date_range': [start.isoformat(), end.isoformat()],
        '明细总数': total_entries,
        '计数型合计': count_sum,
        '按章节分布': section_dist,
        '填报最频节点Top': node_dist,
        '按灌溉项目分布': project_dist,
        '章节选项': section_labels,
    }, ensure_ascii=False)


@tool
def query_pending_repairs(start_date: str = "", limit: int = 30) -> str:
    """查询待修工单（is_pending_repair=True）的明细：哪些工单标记了待修、待修的具体内容、负责人、位置、日期。
    用于跟进未完成的维修任务。

    Args:
        start_date: 起始日期 YYYY-MM-DD，留空默认最近90天
        limit: 最多返回条数，默认30
    """
    today = date.today()
    start = _parse_date(start_date) or (today - timedelta(days=90))
    qs = (WorkReport.objects.filter(date__gte=start, is_pending_repair=True)
          .select_related('worker', 'location')
          .order_by('-date', '-id')[:limit])
    rows = []
    for r in qs:
        rows.append({
            '日期': r.date.isoformat(),
            '处理人': str(r.worker) if r.worker else '',
            '位置': str(r.location) if r.location else '',
            '工作内容': (r.work_content or r.remark or '')[:200],
            '是否疑难': r.is_difficult,
            '疑难已处理': r.is_difficult_resolved,
        })
    return json.dumps({
        '待修工单总数': WorkReport.objects.filter(date__gte=start, is_pending_repair=True).count(),
        'returned': len(rows),
        'date_range_start': start.isoformat(),
        'records': rows,
    }, ensure_ascii=False)


@tool
def query_difficult_workorders(start_date: str = "", resolved_only: bool = False, limit: int = 30) -> str:
    """查询疑难问题工单（is_difficult=True）的明细：疑难问题描述、是否已处理、处理人。
    可筛选未处理的疑难，便于跟进。

    Args:
        start_date: 起始日期 YYYY-MM-DD，留空默认最近90天
        resolved_only: True 只看已处理的，False(默认)看全部疑难
        limit: 最多返回条数，默认30
    """
    today = date.today()
    start = _parse_date(start_date) or (today - timedelta(days=90))
    qs = WorkReport.objects.filter(date__gte=start, is_difficult=True)
    if resolved_only:
        qs = qs.filter(is_difficult_resolved=True)
    qs = (qs.select_related('worker', 'location')
          .order_by('-date', '-id')[:limit])
    rows = []
    for r in qs:
        rows.append({
            '日期': r.date.isoformat(),
            '处理人': str(r.worker) if r.worker else '',
            '位置': str(r.location) if r.location else '',
            '工作内容': (r.work_content or r.remark or '')[:200],
            '已处理': r.is_difficult_resolved,
            '是否待修': r.is_pending_repair,
        })
    total = WorkReport.objects.filter(date__gte=start, is_difficult=True).count()
    resolved = WorkReport.objects.filter(date__gte=start, is_difficult=True, is_difficult_resolved=True).count()
    return json.dumps({
        '疑难工单总数': total,
        '已处理数': resolved,
        '未处理数': total - resolved,
        'returned': len(rows),
        'date_range_start': start.isoformat(),
        'records': rows,
    }, ensure_ascii=False)


@tool
def query_water_requests(start_date: str = "", status: str = "", limit: int = 30) -> str:
    """查询浇水协调需求(WaterRequest)：停水/新苗程序/调水量等需求，按状态(已提交/已批准/已拒绝)筛选，
    含提交人、涉及区域、需求时段。用于了解各区域的浇水协调情况。

    Args:
        start_date: 起始日期 YYYY-MM-DD，留空默认最近30天
        status: 状态过滤 submitted(已提交)/approved(已批准)/rejected(已拒绝)/info_needed(需补充)，留空看全部
        limit: 最多返回条数，默认30
    """
    today = date.today()
    start = _parse_date(start_date) or (today - timedelta(days=30))
    qs = WaterRequest.objects.filter(created_at__date__gte=start)
    if status:
        qs = qs.filter(status=status)
    qs = qs.select_related('submitter').prefetch_related('zones').order_by('-created_at')[:limit]
    status_map = dict(WaterRequest.STATUS_CHOICES)
    rows = []
    for req in qs:
        rows.append({
            '提交时间': timezone.localtime(req.created_at).strftime('%Y-%m-%d %H:%M'),
            '提交人': str(req.submitter) if req.submitter else '',
            '用户类型': req.user_type,
            '需求类型': req.request_type,
            '状态': status_map.get(req.status, req.status),
            '起始时间': timezone.localtime(req.start_datetime).strftime('%Y-%m-%d %H:%M') if req.start_datetime else '',
            '结束时间': timezone.localtime(req.end_datetime).strftime('%Y-%m-%d %H:%M') if req.end_datetime else '',
            '涉及区域': ', '.join(z.code for z in req.zones.all()[:8]) or '—',
        })
    return json.dumps({
        '需求总数': WaterRequest.objects.filter(created_at__date__gte=start).count(),
        'returned': len(rows),
        'date_range_start': start.isoformat(),
        'records': rows,
    }, ensure_ascii=False)


@tool
def query_projects(
    q: str = "",
    category: str = "",
    include_completed: bool = True,
    limit: int = 30,
) -> str:
    """查询灌溉项目(Project)：项目名称/类别/子类别/代号、是否完工、开工/计划完工日期、
    材料/人工预算，以及每个项目的已用工时（从关联工单自动汇总）。

    项目是灌溉工程的管理单元（FAM/FES/WDI/绿化等），工单填报时可选关联项目；
    本工具汇总每个项目的关联工单工时，便于了解项目进度与投入。

    Args:
        q: 项目名称/代号/Code 模糊匹配（部分匹配即可），留空返回全部
        category: 类别过滤 IRRIGATION(灌溉)/DRAINAGE(排水)/OTHER(其他)，留空看全部
        include_completed: True=含已完工，False=只看进行中（未完工）
        limit: 最多返回条数，默认30
    """
    from django.db.models import Q as _Q, Sum

    qs = Project.objects.all()
    if q:
        qs = qs.filter(_Q(name__icontains=q) | _Q(symbol__icontains=q)
                       | _Q(code__icontains=q) | _Q(notes__icontains=q))
    if category in {c for c, _ in Project.CATEGORY_CHOICES}:
        qs = qs.filter(category=category)
    if not include_completed:
        qs = qs.filter(is_completed=False)
    projects = list(qs.order_by('category', 'subcategory', 'name')[:limit])

    # 汇总每个项目的关联工单工时：WorkReportEntry.project → WorkReport 的 team_hours/third_party_hours。
    # 一张工单可能关联多个项目，工时按项目计入（与 _project_summaries 口径一致）。
    entry_pids = [e['project_id'] for e in WorkReportEntry.objects
                  .filter(project__in=[p.id for p in projects])
                  .values('project_id', 'work_report_id')]
    rep_ids_by_pid = {}
    for pid in {p.id for p in projects}:
        rep_ids_by_pid[pid] = set()
    for e in WorkReportEntry.objects.filter(project__in=[p.id for p in projects]).values('project_id', 'work_report_id'):
        rep_ids_by_pid[e['project_id']].add(e['work_report_id'])
    # 每张关联工单的工时
    all_rep_ids = set().union(*rep_ids_by_pid.values()) if rep_ids_by_pid else set()
    rep_hours = {r['id']: (r['team_hours'] or 0, r['third_party_hours'] or 0)
                 for r in WorkReport.objects.filter(id__in=all_rep_ids).values('id', 'team_hours', 'third_party_hours')}
    # 关联工单数
    rep_counts = {pid: len(ids) for pid, ids in rep_ids_by_pid.items()}

    rows = []
    for p in projects:
        team_h = third_h = 0.0
        for rid in rep_ids_by_pid.get(p.id, ()):
            th, tph = rep_hours.get(rid, (0, 0))
            team_h += th; third_h += tph
        rows.append({
            '项目名称': p.name,
            '类别': p.get_category_display(),
            '子类别': p.get_subcategory_display() if p.subcategory else '',
            '代号': p.symbol or '',
            '是否完工': p.is_completed,
            '开工日期': p.start_date.isoformat() if p.start_date else '',
            '计划完工': p.planned_end_date.isoformat() if p.planned_end_date else '',
            '材料预算': float(p.material_budget_amount) if p.material_budget_amount else None,
            '人工预算': float(p.labor_budget_amount) if p.labor_budget_amount else None,
            '关联工单数': rep_counts.get(p.id, 0),
            '已用灌溉组工时': round(team_h, 1),
            '已用第三方工时': round(third_h, 1),
        })
    return json.dumps({
        '项目总数': qs.count(),
        'returned': len(rows),
        'records': rows,
        '类别选项': dict(Project.CATEGORY_CHOICES),
    }, ensure_ascii=False)


@tool
def query_zones(zone_code: str = "", limit: int = 50) -> str:
    """查询区域(Zone)信息：编号、通用名称、所属片区、面积、优先级、灌水器类型等。可按编号模糊查找。

    Args:
        zone_code: 区域编号或通用名称，部分匹配即可（如 "1-1" 或 "BOH"）；留空返回汇总
        limit: 最多返回条数
    """
    if zone_code:
        from django.db.models import Q
        qs = Zone.objects.filter(
            Q(code__icontains=zone_code) | Q(name__icontains=zone_code)
        ).select_related('patch')[:limit]
        rows = []
        for z in qs:
            rows.append({
                '编号': z.code,
                '通用名称': z.name,
                '片区': str(z.patch) if z.patch else '',
                '面积㎡': z.area_sqm,
                '优先级': z.get_priority_display() if z.priority else '',
                '灌水器类型': z.sprinkler_type,
                '当前状态': z.current_status,
            })
        return json.dumps({'returned': len(rows), 'zones': rows}, ensure_ascii=False)
    # summary
    total = Zone.objects.count()
    by_priority = {}
    for z in Zone.objects.values('priority'):
        p = z['priority'] or '未指定'
        by_priority[p] = by_priority.get(p, 0) + 1
    return json.dumps({
        '总区域数': total,
        '按优先级分布': by_priority,
        '提示': '请用 zone_code 参数查具体区域',
    }, ensure_ascii=False)


@tool
def query_weather(days: int = 7) -> str:
    """查询最近若干天的天气数据记录（逐时数据汇总：最高/最低温度、降水总量、主要天气描述）。

    Args:
        days: 查询最近多少天，默认7
    """
    end = date.today()
    start = end - timedelta(days=days)
    qs = WeatherData.objects.filter(date__gte=start).order_by('-date')[:days]
    rows = []
    for w in qs:
        # hourly_data is a list of {hour, temp, humidity, precip, wind, code}.
        hourly = w.hourly_data or []
        temps = [h.get('temp') for h in hourly if h.get('temp') is not None]
        precip = sum((h.get('precip') or 0) for h in hourly)
        # Dominant weather code = the one covering the most daytime hours (6-20).
        from collections import Counter
        day_codes = [h.get('code') for h in hourly
                     if h.get('code') is not None and 6 <= (h.get('hour') or 0) <= 20]
        code_counts = Counter(day_codes)
        dom_code = code_counts.most_common(1)[0][0] if code_counts else None
        rows.append({
            '日期': w.date.isoformat() if w.date else '',
            '最高温度℃': round(max(temps), 1) if temps else None,
            '最低温度℃': round(min(temps), 1) if temps else None,
            '降水总量mm': round(precip, 1),
            '主要天气': w.get_weather_description(dom_code) if dom_code is not None else '',
            '逐时记录数': len(hourly),
        })
    return json.dumps({'returned': len(rows), 'weather': rows}, ensure_ascii=False)


@tool
def query_irrigation_overview() -> str:
    """查询灌溉系统总览统计：片区/站点/控制器/流量区域/气象站/事件数量、锁定站点数。用于系统健康度概览。"""
    from core.models import (
        MaxicomController, MaxicomSchedule, MaxicomFlowZone,
        MaxicomWeatherStation, MaxicomWeatherLog, MaxicomEvent,
)
    return json.dumps({
        '片区数': Patch.objects.count(),
        '控制器数': MaxicomController.objects.count(),
        '站点数': Patch.objects.filter(parent__isnull=False).count(),
        '计划数': MaxicomSchedule.objects.count(),
        '流量区域数': MaxicomFlowZone.objects.count(),
        '气象站数': MaxicomWeatherStation.objects.count(),
        '气象日志数': MaxicomWeatherLog.objects.count(),
        '事件数': MaxicomEvent.objects.count(),
        '锁定站点数': Patch.objects.filter(parent__isnull=False, lockout=True).count(),
    }, ensure_ascii=False)


# ── Helpers ───────────────────────────────────────────────────────────────


def _parse_date(s: str):
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except (ValueError, TypeError):
        return None


# ── Code execution tool ───────────────────────────────────────────────────


# Generated code runs in a stripped-down environment. No Django/DB credentials
# leak through — the subprocess only sees a minimal PATH and the preloaded CSVs.
_SANDBOX_ENV = {
    'PATH': '/usr/bin:/bin:/usr/local/bin',
    'LANG': 'en_US.UTF-8',
    'LC_ALL': 'en_US.UTF-8',
    'HOME': '/tmp',
    'MPLBACKEND': 'Agg',  # matplotlib headless, in case it's ever installed
}

# Cap output preview to keep the LLM context small.
_MAX_STDOUT = 2000
_MAX_STDERR = 1000
# Allowed output extensions + size guard.
_ALLOWED_EXTS = {'.xlsx', '.xls', '.csv', '.json', '.txt'}
_MAX_FILE_BYTES = 20 * 1024 * 1024  # 20 MB


def _fault_matrix_columns():
    """Fixed fault-matrix columns for the Excel export: every count leaf node in
    the routine_maint section, as [(work_item_id, path_segments), ...], sorted by
    path then id so columns stay identical across exports. ``path_segments`` is
    the group hierarchy with the section root dropped, e.g.
    ``['计划性维修', '喷头', '喷嘴丢/坏']`` — so the caller can render a merged
    multi-row header. Mirrors work_reports_excel's _work_report_count_columns.
    """
    items = {it.id: it for it in WorkItem.objects.filter(section='routine_maint')}

    def segs_of(it):
        parts, cur, seen = [], it, set()
        while cur and cur.id not in seen:
            seen.add(cur.id)
            parts.append(cur.name_zh)
            cur = items.get(cur.parent_id) if cur.parent_id else None
        parts.reverse()
        return parts[1:] or [it.name_zh]   # drop the section-root name itself

    cols = [(it.id, segs_of(it)) for it in items.values() if it.value_type == 'count']
    cols.sort(key=lambda c: (c[1], c[0]))
    return cols


def _dedup_zone_names(raw):
    """Collapse repeated zone names in a stored zone_names string, preserving
    first-seen order. Older reports list the same name N times when several
    same-named zones were selected (e.g. 'BOH, BOH, BOH'). Mirrors the helper
    in views.py to avoid a cross-module import.
    """
    if not raw:
        return ''
    out, seen = [], set()
    for p in (part.strip() for part in raw.split(',')):
        if p and p not in seen:
            seen.add(p)
            out.append(p)
    return ', '.join(out)


@tool
def export_work_reports_excel(start_date: str = "", end_date: str = "", runtime: ToolRuntime = None) -> str:
    """导出维修工单为 Excel 报表（reporttemplate 格式），生成 .xlsx 文件并返回下载链接。

    报表结构：一行一个工单，固定列头 = 常规维护章节下所有计数型节点（故障类型矩阵），
    无值留空。基础列含日期/处理人/位置/工作分类/故障位置/备注/疑难/待修。
    不同时间范围导出格式一致。文件可在对话中直接下载。

    Args:
        start_date: 起始日期 YYYY-MM-DD，留空默认最近30天
        end_date: 结束日期 YYYY-MM-DD，留空默认今天
        runtime: LangChain 运行时（自动注入，提供会话 context.thread_id）
    """
    import io
    from django.db.models import Prefetch
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ctx = (runtime.context if runtime is not None else None) or {}
    thread_id = ctx.get('thread_id')
    if not thread_id:
        return json.dumps({'error': '无法确定会话工作区'}, ensure_ascii=False)
    ws = ensure_workspace(thread_id)

    today = date.today()
    end = _parse_date(end_date) or today
    start = _parse_date(start_date) or (end - timedelta(days=30))

    # Fixed fault-matrix columns (deterministic across exports).
    count_nodes = _fault_matrix_columns()

    qs = (WorkReport.objects.filter(date__gte=start, date__lte=end)
          .select_related('worker', 'location')
          .prefetch_related(Prefetch(
              'entries',
              queryset=WorkReportEntry.objects.select_related('work_item').filter(work_item__value_type='count'),
              to_attr='_count_entries'))
          .order_by('date', 'id'))

    wb = Workbook()
    sh = wb.active
    sh.title = '维修记录'
    base_header = ['序号', '日期', '处理人', '位置', '工作分类', '故障/事件位置',
                   '备注', '信息来源', '疑难问题', '疑难已处理']
    n_base = len(base_header)
    n_cols = n_base + len(count_nodes)
    hdr_rows = max((len(segs) for _, segs in count_nodes), default=1)

    # Style every header cell up front so merged ranges keep their borders.
    hfill = PatternFill('solid', fgColor='1B4332')
    hfont = Font(color='FFFFFF', bold=True, size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='FFFFFF')
    hborder = Border(left=thin, right=thin, top=thin, bottom=thin)
    for rr in range(1, hdr_rows + 1):
        for cc in range(1, n_cols + 1):
            cell = sh.cell(row=rr, column=cc)
            cell.fill = hfill; cell.font = hfont; cell.alignment = center; cell.border = hborder

    # Base columns span every header row (vertical merge).
    for ci, label in enumerate(base_header, 1):
        sh.cell(row=1, column=ci, value=label)
        if hdr_rows > 1:
            sh.merge_cells(start_row=1, start_column=ci, end_row=hdr_rows, end_column=ci)

    # Count columns: merge consecutive siblings sharing the same path prefix.
    for depth in range(hdr_rows):
        row = depth + 1
        ci = 0
        while ci < len(count_nodes):
            segs = count_nodes[ci][1]
            if depth >= len(segs):
                ci += 1
                continue
            prefix = segs[:depth + 1]
            cj = ci + 1
            while cj < len(count_nodes) and count_nodes[cj][1][:depth + 1] == prefix:
                cj += 1
            col_start = n_base + ci + 1
            col_end = n_base + cj
            sh.cell(row=row, column=col_start, value=segs[depth])
            need_h = col_end > col_start
            need_v = depth == len(segs) - 1 and row < hdr_rows
            if need_h and need_v:
                sh.merge_cells(start_row=row, start_column=col_start, end_row=hdr_rows, end_column=col_end)
            elif need_h:
                sh.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
            elif need_v:
                sh.merge_cells(start_row=row, start_column=col_start, end_row=hdr_rows, end_column=col_start)
            ci = cj

    id_to_col = {wid: n_base + i + 1 for i, (wid, _) in enumerate(count_nodes)}

    for idx, r in enumerate(qs, 1):
        # 工作分类 = 该工单明细的 section（优先非常规维护）
        secs = [e.work_item.section for e in getattr(r, '_count_entries', [])]
        cat = '常规维护'
        for s in secs:
            if s and s != 'routine_maint':
                cat = _SECTION_LABELS.get(s, s); break
        row = [idx, r.date.isoformat() if r.date else '',
               r.worker.full_name if r.worker_id and r.worker else '',
               r.location.code if r.location_id and r.location else '',
               cat, _dedup_zone_names(r.zone_names),
               (r.work_content or r.remark or ''), '',
               '是' if r.is_difficult else '',
               '是' if r.is_difficult_resolved else '']
        row += [None] * len(count_nodes)
        for e in getattr(r, '_count_entries', []):
            ci = id_to_col.get(e.work_item_id)
            if ci is not None and e.count:
                row[ci - 1] = (row[ci - 1] or 0) + e.count
        sh.append(row)

    # Data cell styling: thin grey borders, centre numerics, wrap 备注.
    gside = Side(style='thin', color='D0D0D0')
    gborder = Border(left=gside, right=gside, top=gside, bottom=gside)
    dcenter = Alignment(horizontal='center', vertical='center')
    dwrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    last_row = hdr_rows + qs.count()
    for rr in range(hdr_rows + 1, last_row + 1):
        for cc in range(1, n_cols + 1):
            cell = sh.cell(row=rr, column=cc)
            cell.border = gborder
            cell.alignment = dwrap if cc == 7 else dcenter

    # Column widths + freeze the header rows and first four ID columns.
    from openpyxl.utils import get_column_letter
    for ci in range(1, n_cols + 1):
        if ci <= n_base:
            width = 28 if ci == 7 else 12
        else:
            leaf = count_nodes[ci - n_base - 1][1][-1]
            width = max(8, min(20, len(leaf) * 1.7 + 2))
        sh.column_dimensions[get_column_letter(ci)].width = width
    sh.freeze_panes = sh.cell(row=hdr_rows + 1, column=5)

    fname = f'workreports_{start}_{end}.xlsx'
    out_path = os.path.join(ws, fname)
    wb.save(out_path)
    size = os.path.getsize(out_path)

    return json.dumps({
        'file': {
            'name': fname,
            'size': size,
            'url': f'{settings.MEDIA_URL}ai_workspaces/{thread_id}/{fname}',
        },
        'report_count': qs.count(),
        'matrix_columns': len(count_nodes),
        'date_range': [start.isoformat(), end.isoformat()],
    }, ensure_ascii=False)


@tool
def send_report_email(report_type: str, start_date: str, end_date: str,
                      recipients: str, runtime: ToolRuntime = None) -> str:
    """生成指定类型的 Excel 报表并通过 SMTP 邮件发送给收件人。

    覆盖全部 7 种报表类型，与 export_work_reports_excel（仅维修工单、仅生成文件）
    互补：本工具覆盖全部类型并直接发送邮件。

    支持的 report_type：
      - work_reports      维修工单记录
      - irrigation        灌溉运行报表
      - projects          项目管理
      - inventory_catalog 库存目录
      - inventory_txn     库存出入库流水
      - purchase_orders   采购订单
      - zones             区域/地块目录

    zones / inventory_catalog / purchase_orders 为目录类报表，与日期范围无关
    （start_date/end_date 仍需传入占位值，会被忽略）。

    Args:
        report_type: 报表类型（见上）
        start_date: 起始日期 YYYY-MM-DD
        end_date: 结束日期 YYYY-MM-DD
        recipients: 收件人邮箱，多个用逗号或分号分隔
        runtime: LangChain 运行时（自动注入，保留参数）
    """
    import re
    from core.report_export import REPORT_TYPES, REPORT_TYPE_LABELS, build_report
    from core.smtp_utils import (
        is_smtp_configured, get_email_connection, resolve_from_email,
    )
    from django.core.mail import EmailMessage

    if report_type not in REPORT_TYPES:
        return json.dumps({
            'error': f'未知报表类型: {report_type}',
            'valid': REPORT_TYPES,
            'labels': REPORT_TYPE_LABELS,
        }, ensure_ascii=False)

    start = _parse_date(start_date)
    end = _parse_date(end_date)
    if start is None or end is None:
        return json.dumps({
            'error': 'start_date/end_date 格式无效，需 YYYY-MM-DD',
        }, ensure_ascii=False)

    rcpt_list = [r.strip() for r in re.split(r'[,\n;]', recipients or '') if r.strip()]
    if not rcpt_list:
        return json.dumps({'error': '未提供收件人邮箱'}, ensure_ascii=False)

    if not is_smtp_configured():
        return json.dumps({
            'error': 'SMTP 未配置，无法发送邮件。请先在「用户管理 → 邮件报表」配置 SMTP 服务器。',
        }, ensure_ascii=False)

    label = REPORT_TYPE_LABELS.get(report_type, report_type)
    try:
        fname, buf = build_report(report_type, start=start, end=end)
    except Exception as e:
        return json.dumps({'error': f'生成报表失败: {e}'}, ensure_ascii=False)

    period = (start.strftime('%Y-%m-%d') if start == end
              else f'{start.strftime("%Y-%m-%d")} 至 {end.strftime("%Y-%m-%d")}')
    subject = f'【{label}】{period} 报表'
    body = (f'您好，\n\n这是由 AI 助手生成的报表邮件。\n'
            f'报表类型：{label}\n数据范围：{period}\n\n请见附件。')
    try:
        msg = EmailMessage(
            subject=subject, body=body,
            from_email=resolve_from_email(), to=rcpt_list,
            connection=get_email_connection(),
        )
        msg.attach(fname, buf.getvalue(),
                   'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        msg.send(fail_silently=False)
    except Exception as e:
        return json.dumps({'error': f'邮件发送失败: {e}'}, ensure_ascii=False)

    return json.dumps({
        'sent': True,
        'report_type': report_type,
        'label': label,
        'filename': fname,
        'date_range': [start.isoformat(), end.isoformat()],
        'recipients': rcpt_list,
        'recipient_count': len(rcpt_list),
    }, ensure_ascii=False)


@tool
def run_python_code(code: str, description: str = "", runtime: ToolRuntime = None) -> str:
    """运行 Python 代码进行数据分析、计算、并生成报表文件。

    工作目录已预置以下 CSV 数据文件，用 pandas 读取：
    - work_reports.csv（最近90天维修工单：日期/处理人/位置/班次/章节/工时/待修/内容）
    - work_entries.csv（最近90天工单明细：日期/章节/节点编码/节点名称/值类型/项目/数量/状态/文本）
    - zones.csv（全部区域：编号/通用名称/片区/面积/优先级/灌水器类型）

    用法规则：
    - 用 `import pandas as pd` 然后 `pd.read_csv('work_reports.csv')` 加载数据
    - 生成文件时用相对路径保存到当前目录，如 df.to_excel('工时统计.xlsx', index=False) 或 df.to_csv('result.csv')
    - 仅允许扩展名：.xlsx .csv .json .txt
    - 可用库：pandas, openpyxl, xlsxwriter, json, datetime, math, statistics, collections
    - 执行超时30秒；禁止访问网络/数据库/其他目录
    - 用 print() 输出中间结果和摘要

    Args:
        code: 完整的 Python 代码（可多行）
        description: 对这段代码目的的简短描述
        runtime: LangChain 运行时（保留参数名，自动注入；提供会话 context.thread_id）
    """
    # `runtime` is a reserved param (auto-injected, hidden from the LLM). The
    # default None keeps the function callable for direct unit tests.
    ctx = (runtime.context if runtime is not None else None) or {}
    thread_id = ctx.get('thread_id')
    if not thread_id:
        return json.dumps({'error': '无法确定会话工作区'}, ensure_ascii=False)
    ws = ensure_workspace(thread_id)

    # Snapshot existing files so we can detect newly-created outputs.
    before = set(os.listdir(ws))

    try:
        result = subprocess.run(
            [sys.executable, '-c', code],
            cwd=ws,
            env=_SANDBOX_ENV,
            capture_output=True,
            text=True,
            timeout=30,
        )
    except subprocess.TimeoutExpired:
        return json.dumps({'error': '代码执行超时（30秒）'}, ensure_ascii=False)
    except Exception as e:  # noqa: BLE001
        return json.dumps({'error': f'执行失败: {e}'}, ensure_ascii=False)

    # Collect newly-created files matching allowed extensions.
    files = []
    for name in os.listdir(ws):
        if name in before or name.startswith('.'):
            continue
        ext = os.path.splitext(name)[1].lower()
        if ext not in _ALLOWED_EXTS:
            continue
        path = os.path.join(ws, name)
        if not os.path.isfile(path):
            continue
        size = os.path.getsize(path)
        if size > _MAX_FILE_BYTES:
            continue
        files.append({
            'name': name,
            'size': size,
            'url': f'{settings.MEDIA_URL}ai_workspaces/{thread_id}/{name}',
        })

    return json.dumps({
        'exit_code': result.returncode,
        'stdout': (result.stdout or '')[:_MAX_STDOUT],
        'stderr': (result.stderr or '')[:_MAX_STDERR],
        'files': files,
        'generated_count': len(files),
    }, ensure_ascii=False)


ALL_TOOLS = [
    query_work_reports,
    query_work_report_stats,
    query_work_entries_stats,
    query_pending_repairs,
    query_difficult_workorders,
    query_water_requests,
    query_projects,
    query_zones,
    query_weather,
    query_irrigation_overview,
    export_work_reports_excel,
    send_report_email,
    run_python_code,
]


@lru_cache(maxsize=1)
def _checkpoint_saver():
    """A persistent SqliteSaver shared across agent builds (conversation memory).

    Unlike the previous InMemorySaver (RAM only), this writes checkpoints to a
    dedicated sqlite file so conversation history survives:
      - server / worker restarts (production deploys)
      - multiple WSGI workers (each worker reads the same file)
      - the Django dev server's autoreloader restarting the process on code edits

    The file lives under MEDIA_ROOT (kept out of the app db to avoid write-lock
    contention with Django's own sqlite). WAL mode lets multiple workers read
    concurrently while one writes — AI chat is low-concurrency (a handful of
    managers), so sqlite is plenty here; for higher scale use PostgresSaver.
    """
    import sqlite3
    db_path = os.path.join(settings.MEDIA_ROOT, 'ai_checkpoints.sqlite3')
    conn = sqlite3.connect(db_path, check_same_thread=False)
    # WAL: readers don't block the writer, so concurrent workers stay responsive.
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA busy_timeout=5000')  # wait up to 5s on a locked db
    saver = SqliteSaver(conn)
    saver.setup()  # create the checkpoint tables (idempotent)
    return saver


def build_agent():
    """Build a fresh LangChain agent from current AISettings.

    Rebuilt per request so config changes (api_key/model) in admin take effect
    immediately. The sqlite checkpointer is reused (file-backed, shared across
    processes/restarts) so a thread_id keeps conversation history across
    requests even when the server has multiple workers or restarts.
    """
    cfg = AISettings.get_settings()
    if not (cfg.enabled and cfg.api_base_url and cfg.api_key and cfg.model_name):
        raise RuntimeError('AI 助手未启用或配置不完整，请在管理后台填写 base_url / api_key / model')

    # 当前时间注入到系统提示词（get_today_date 工具已移除，时间直接给到模型）。
    # 每个请求重新 build，所以时间总是新的。
    now = timezone.localtime(timezone.now())
    _weekday = ['周一', '周二', '周三', '周四', '周五', '周六', '周日'][now.weekday()]
    time_block = (
        f'\n## 当前时间（系统注入，无需调用工具查询）\n'
        f'- 今天：{now.date().isoformat()}（{_weekday}）\n'
        f'- 当前时刻：{now.strftime("%Y-%m-%d %H:%M")}（北京时间）\n'
        f'用户说"今天/昨天/本周/本月/最近N天"时，直接用上面的日期推算。\n'
    )

    model = ChatOpenAI(
        base_url=cfg.api_base_url,
        api_key=cfg.api_key,
        model=cfg.model_name,
        temperature=cfg.temperature,
        # Many third-party OpenAI-compatible proxies do not support stream_options;
        # disable streaming usage metadata to avoid 400 errors.
        stream_usage=False,
    )
    base_prompt = cfg.get_system_prompt()
    # 时间块插在系统提示词最前面，保证模型一开始就知道当前日期。
    agent = create_agent(
        model=model,
        tools=ALL_TOOLS,
        system_prompt=time_block + base_prompt,
        checkpointer=_checkpoint_saver(),
        # Per-run context (the workspace thread id) declared as a schema so tools
        # read it via ToolRuntime.context instead of a global — LangChain v1 pattern.
        context_schema=WorkorderContext,
    )
    return agent


def is_configured() -> bool:
    """Quick check whether the agent can be built (used by the view to show a hint)."""
    cfg = AISettings.get_settings()
    return bool(cfg.enabled and cfg.api_base_url and cfg.api_key and cfg.model_name)
