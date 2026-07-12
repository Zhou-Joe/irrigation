"""Import PM plans from the Maximo PM list Excel.

Reads ``PM list 260711-架构.xlsx`` and creates JobPlanTemplate +
MaintenancePlan records. Each row maps to one PM (one asset group +
one frequency).

Three asset-resolution modes are supported (driven by SHEET_CONFIG):
  - ``zone_col``  : zone codes parsed from column 3 (顿号/逗号/分号/斜杠)
  - ``desc_sat``  : SAT code extracted from the description (e.g. "1-1 西门")
  - ``asset_ccu`` : Maximo asset number → CCU (asset[1:3] → Patch code)

DRAFT sheets (per the md architecture doc "暂停启用") are imported with
``active=False`` so they don't participate in dispatch.

Usage:
    python manage.py import_pm_plans
    python manage.py import_pm_plans --file /path/to/PM.xlsx
"""

import os
import re

from django.core.management.base import BaseCommand, CommandError
from django.conf import settings

from core.models import (
    JobPlanTemplate, MaintenancePlan, Zone, Satellite, Patch,
)


# Each sheet → config. ``resolve`` picks the asset-resolution mode.
#   zone_col : parse zone codes from the Zone/资产 column (col 3)
#   desc_sat : extract SAT code from the description (col 1)
#   asset_ccu: map Maximo asset number (col 3) → CCU Patch
SHEET_CONFIG = {
    # ── A. Zone组（超级重点）— 已有，保留 ──
    '超级重点区目检3-30': {
        'job_plan_name': '超级重点区目检',
        'asset_level': 'zone_group',
        'description': '每天目检，对应3处通用位置28个Zone',
        'default_unit': 'days',
        'resolve': 'zone_col',
    },
    '超级重点区水检3-30': {
        'job_plan_name': '超级重点区水检',
        'asset_level': 'zone_group',
        'description': '每周水检，对应3处通用位置28个Zone',
        'default_unit': 'weeks',
        'resolve': 'zone_col',
    },
    # ── B. Zone组（重点）— 14处通用位置172个Zone ──
    '重点区目检14-172': {
        'job_plan_name': '喷头目检重点区域',
        'asset_level': 'zone_group',
        'description': '每周目检，对应14处通用位置172个Zone',
        'default_unit': 'weeks',
        'resolve': 'zone_col',
    },
    '重点区水检14-172': {
        'job_plan_name': '喷头水检重点区域',
        'asset_level': 'zone_group',
        'description': '每2周水检，对应14处通用位置172个Zone',
        'default_unit': 'weeks',
        'resolve': 'zone_col',
    },
    # ── C. SAT级（一般/次要）— 每 SAT 一个 PM ──
    '一般区水检': {
        'job_plan_name': '喷头水检一般位置',
        'asset_level': 'sat',
        'description': '每3周水检，对应137处SAT',
        'default_unit': 'weeks',
        'resolve': 'desc_sat',
    },
    '次要区水检': {
        'job_plan_name': '喷头水检次要位置',
        'asset_level': 'sat',
        'description': '每5周水检，对应138处SAT',
        'default_unit': 'weeks',
        'resolve': 'desc_sat',
    },
    # ── D. 设备级（POC阀类）— 每 POC 一个 PM ──
    '主阀29': {
        'job_plan_name': '电动主阀检查',
        'asset_level': 'ccu',
        'description': '每2-3周检查，对应29处POC',
        'default_unit': 'weeks',
        'resolve': 'asset_ccu',
    },
    '隔离阀29': {
        'job_plan_name': '隔离阀开关检查',
        'asset_level': 'ccu',
        'description': '每24周开关检查，对应29处POC',
        'default_unit': 'months',
        'resolve': 'asset_ccu',
    },
    '冲洗阀29': {
        'job_plan_name': '冲洗阀检查',
        'asset_level': 'ccu',
        'description': '每2周检查，对应29处POC',
        'default_unit': 'weeks',
        'resolve': 'asset_ccu',
    },
    '洗滤网29': {
        'job_plan_name': 'POC滤网清洗',
        'asset_level': 'ccu',
        'description': '每9周清洗，对应29处POC',
        'default_unit': 'weeks',
        'resolve': 'asset_ccu',
    },
    # ── E. 设备级（CCU/SAT设备）— 暂停启用(DRAFT) ──
    'CCU22': {
        'job_plan_name': 'CCU现场检查',
        'asset_level': 'ccu',
        'description': '每24周检查，对应22处CCU（暂停启用）',
        'default_unit': 'months',
        'resolve': 'asset_ccu',
        'draft': True,
    },
    '控制器138': {
        'job_plan_name': '控制器保养',
        'asset_level': 'sat',
        'description': '每24周保养，对应138处SAT（暂停启用）',
        'default_unit': 'months',
        'resolve': 'desc_sat',
        'draft': True,
    },
    '电磁阀138': {
        'job_plan_name': '电磁阀球阀调压器检查',
        'asset_level': 'sat',
        'description': '每48周检查，对应138处SAT',
        'default_unit': 'months',
        'resolve': 'desc_sat',
    },
}


class Command(BaseCommand):
    help = 'Import PM plans from the Maximo PM list Excel (all 13 sheets).'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            default=str(settings.BASE_DIR.parent / 'PM list 260711-架构.xlsx'),
            help='Path to PM list Excel (default: <repo>/PM list 260711-架构.xlsx)',
        )

    def handle(self, *args, **options):
        import openpyxl

        path = options['file']
        if not os.path.exists(path):
            raise CommandError(f'PM list file not found: {path}')

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        total_created = 0
        total_skipped = 0

        for sheet_name, config in SHEET_CONFIG.items():
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f'Sheet "{sheet_name}" not found, skipping.'))
                continue

            # Ensure the JobPlanTemplate exists.
            job_plan, jp_created = JobPlanTemplate.objects.get_or_create(
                name=config['job_plan_name'],
                defaults={
                    'description': config['description'],
                    'asset_level': config['asset_level'],
                },
            )
            if jp_created:
                self.stdout.write(self.style.SUCCESS(
                    f'Created JobPlanTemplate: {job_plan.name} ({job_plan.asset_level})'))

            ws = wb[sheet_name]
            created, skipped = self._import_sheet(ws, job_plan, config)
            total_created += created
            total_skipped += skipped
            self.stdout.write(f'  {sheet_name}: {created} created, {skipped} skipped.')

        self.stdout.write(self.style.SUCCESS(
            f'\nDone: {total_created} PM plans created, {total_skipped} skipped.'))

    def _import_sheet(self, ws, job_plan, config):
        """Import one sheet's rows into MaintenancePlan records."""
        from datetime import date, datetime

        created = 0
        skipped = 0
        is_draft = config.get('draft', False)

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Columns: PM | 描述 | 位置 | 资产/Zone | ... | 频率(14) | Units(15)
            pm_number = str(row[0]).strip() if row[0] else ''
            description = str(row[1]).strip() if row[1] else ''
            asset_col = str(row[3]).strip() if row[3] else ''
            freq_value = row[14] if len(row) > 14 and row[14] else 1
            freq_unit_raw = str(row[15]).strip().upper() if len(row) > 15 and row[15] else ''

            if not description or description == 'None':
                skipped += 1
                continue

            # Map unit + frequency.
            unit = self._map_unit(freq_unit_raw, config['default_unit'])
            try:
                freq_value = int(freq_value)
            except (ValueError, TypeError):
                freq_value = 1

            # Resolve the asset link based on the sheet's mode (needed before
            # generating a PM number, so the number can encode the asset).
            resolve = config['resolve']
            zones = []
            satellite = None
            patch = None

            if resolve == 'zone_col':
                zone_codes = self._parse_zone_codes(asset_col)
                if not zone_codes:
                    skipped += 1
                    continue
                zones = list(Zone.objects.filter(code__in=zone_codes))
                if not zones:
                    skipped += 1
                    continue
            elif resolve == 'desc_sat':
                sat_code = self._extract_sat_from_desc(description)
                if sat_code:
                    satellite = Satellite.objects.filter(code=sat_code).first()
                if not satellite:
                    skipped += 1
                    continue
            elif resolve == 'asset_ccu':
                patch = self._extract_ccu_from_asset(asset_col)
                if not patch:
                    skipped += 1
                    continue

            # Generate a readable PM number when the Excel doesn't provide one.
            # Uses the resolved asset so the number is meaningful, e.g.
            # "喷头水检一般位置-SAT1-1" or "喷头目检重点区域-3-5-1".
            if not pm_number or pm_number == 'None':
                pm_number = self._gen_pm_number(description, job_plan, zones, satellite, patch)

            # Skip if already exists (idempotent).
            if MaintenancePlan.objects.filter(pm_number=pm_number).exists():
                skipped += 1
                continue

            # Parse start date from "上次开始日期" (col 10) if available.
            from django.utils import timezone as _tz
            start_date = _tz.localdate()
            last_start = row[10] if len(row) > 10 else None
            if isinstance(last_start, datetime):
                start_date = last_start.date()
            elif isinstance(last_start, date):
                start_date = last_start

            plan = MaintenancePlan.objects.create(
                pm_number=pm_number,
                job_plan=job_plan,
                frequency_value=freq_value,
                frequency_unit=unit,
                start_date=start_date,
                lead_days=1,
                active=not is_draft,   # DRAFT sheets → inactive
                remark_template=description,
                satellite=satellite,
                patch=patch,
            )
            if zones:
                plan.zones.set(zones)

            created += 1
            asset_desc = (f'{len(zones)} zones' if zones
                          else f'SAT {satellite.code}' if satellite
                          else f'CCU {patch.code}' if patch else '?')
            self.stdout.write(f'    {pm_number}: {asset_desc}, freq={freq_value} {unit}'
                              + (' [DRAFT]' if is_draft else ''))

        return created, skipped

    def _parse_zone_codes(self, text):
        """Parse zone-code lists from the PM Excel's Zone column.

        Handles mixed separators: `、`(顿号) `；`(中文分号) `;` `，`(中文逗号) `,` newline.
        Formats:
        - ``3-5-11,`` → ['3-5-11']
        - ``3-5-1、3-12-3/6、3-13-1`` → ['3-5-1','3-12-3','3-12-6','3-13-1']
          (slashes = discrete segment numbers under the same prefix)
        - ``7-8-6/7/10/12/16`` → ['7-8-6','7-8-7','7-8-10','7-8-12','7-8-16']
        """
        if not text or text == 'None':
            return []
        codes = []
        # Normalize ALL separators to semicolons, then split.
        text = (text.replace('、', ';').replace('；', ';').replace(';', ';')
                    .replace('，', ';').replace(',', ';').replace('\n', ';')
                    .strip().rstrip(';'))
        for group in text.split(';'):
            group = group.strip()
            if not group:
                continue
            # Match "prefix nums" where prefix = "CCU-SAT-" and nums may have slashes.
            m = re.match(r'^((\d+-)+)(.+)$', group)
            if not m:
                continue
            prefix = m.group(1)       # "7-8-"
            rest = m.group(3)          # "6/7/10/12/16"
            for seg in rest.split('/'):
                seg = seg.strip()
                if seg.isdigit():
                    codes.append(f'{prefix}{seg}')
        return codes

    def _extract_sat_from_desc(self, description):
        """Extract a SAT code (e.g. '1-1') from a description like
        '每3周喷头水检一般位置 1-1  西门'."""
        m = re.search(r'(\d+-\d+)', description)
        return m.group(1) if m else None

    def _extract_ccu_from_asset(self, asset_code):
        """Map a Maximo asset number (e.g. '713101') to the CCU Patch.

        Asset numbers encode the CCU in digits [1:3]: 713101 → 13 → Patch 'CCU13'.
        Returns the Patch or None.
        """
        if not asset_code or asset_code == 'None':
            return None
        asset_code = asset_code.strip()
        if len(asset_code) < 3 or not asset_code[1:3].isdigit():
            return None
        ccu_num = str(int(asset_code[1:3]))   # '013' → '13', '02' → '2'
        return Patch.objects.filter(code='CCU' + ccu_num).first()

    def _map_unit(self, raw, default):
        """Map Excel 'WEEKS'/'DAYS'/'MONTHS' → model 'weeks'/'days'/'months'."""
        raw = (raw or '').upper().strip()
        if 'DAY' in raw:
            return 'days'
        if 'WEEK' in raw:
            return 'weeks'
        if 'MONTH' in raw:
            return 'months'
        return default

    def _gen_pm_number(self, description, job_plan, zones=None, satellite=None, patch=None):
        """Generate a readable PM number when the Excel doesn't provide one.

        Encodes the asset so the number is meaningful:
          SAT-level  → "喷头水检一般位置-SAT1-1"
          Zone-group → "喷头目检重点区域-3-5-1" (first zone code + count)
          CCU-level  → "冲洗阀检查-CCU13"
        Falls back to a short hash suffix only if the asset is ambiguous.
        Includes a zone count suffix for multi-zone groups to avoid collisions
        between two distinct zone sets that share the same first zone.
        """
        # Short job-plan label: drop common prefixes for brevity.
        label = job_plan.name
        if satellite:
            return f'{label}-SAT{satellite.code}'
        if patch:
            return f'{label}-{patch.code}'
        if zones:
            base = f'{label}-{zones[0].code}'
            if len(zones) > 1:
                base += f'-{len(zones)}'
            return base
        # Last resort: hash of description.
        import hashlib
        h = hashlib.md5(description.encode()).hexdigest()[:6]
        return f'{label}-{h}'
