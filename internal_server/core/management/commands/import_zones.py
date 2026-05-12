"""Import zones from 'Zone list V0.xlsx' into Django database.

Synchronizes CCU patch names and imports/updates all zones.
"""

import os
from django.core.management.base import BaseCommand
from core.models import Patch, Region, Zone

try:
    import openpyxl
except ImportError:
    openpyxl = None


# CCU编号 → primary 灌溉分区 name (most frequent in Excel)
CCU_NAMES = {
    '1': '西门',
    '2': '东门',
    '3': 'TL',
    '5': 'FL',
    '6': 'FL',
    '7': '探险岛',
    '8': '宝藏湾',
    '9': '奇想花园',
    '10': '玩具总动员酒店',
    '11': '乐园酒店',
    '12': '小镇',
    '13': '蓝天大道',
    '14': '西游客停车场',
    '35': '北泵站',
    '36': '申迪北路',
    '37': '西泵站',
    '38': '申迪西路',
    '39': '南泵站',
    '40': '奇妙路',
    '43': '东大湖',
    '44': '南大湖',
    '45': '西大湖',
}

# Excel 位置重要程度 → Zone priority choice
PRIORITY_MAP = {
    '超级重点位置': Zone.PRIORITY_CRITICAL,
    '重点位置': Zone.PRIORITY_HIGH,
    '一般位置': Zone.PRIORITY_MEDIUM,
    '次要位置': Zone.PRIORITY_LOW,
    '废除': Zone.PRIORITY_ABOLISHED,
}


class Command(BaseCommand):
    help = 'Import zones from Zone list V0.xlsx'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str,
                            default=os.path.abspath(os.path.join(
                                os.path.dirname(__file__),
                                '../../../../',
                                'Zone list V0.xlsx')),
                            help='Path to the Excel file')
        parser.add_argument('--dry-run', action='store_true',
                            help='Preview changes without writing to DB')

    def handle(self, *args, **options):
        if openpyxl is None:
            self.stderr.write(self.style.ERROR(
                'openpyxl is required: uv pip install openpyxl'))
            return

        file_path = os.path.normpath(options['file'])
        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f'File not found: {file_path}'))
            return

        dry_run = options['dry_run']
        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN — no changes written'))

        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb['Zone info 仅色块']

        # --- Step 1: Sync patches ---
        self.stdout.write('\n=== Step 1: Sync CCU patches ===')
        region = Region.objects.first()
        if not region:
            region = Region.objects.create(name='OC1', active=True)
            self.stdout.write(f'Created Region: {region.name}')

        for ccu_num, ccu_name in CCU_NAMES.items():
            code = f'CCU{ccu_num}'
            patch = Patch.objects.filter(code=code).first()
            if patch:
                old_name = patch.name
                patch.name = ccu_name
                patch.region = region
                patch.active = True
                if not dry_run:
                    patch.save()
                status = 'updated' if old_name != ccu_name else 'ok'
                self.stdout.write(f'  {code}: "{old_name}" → "{ccu_name}" [{status}]')
            else:
                if not dry_run:
                    Patch.objects.create(
                        code=code, name=ccu_name, region=region,
                        active=True)
                self.stdout.write(f'  {code}: created "{ccu_name}"')

        # --- Step 2: Import zones ---
        self.stdout.write('\n=== Step 2: Import zones ===')
        created = 0
        updated = 0
        skipped = 0

        # Preload patch lookup
        patch_map = {}
        for p in Patch.objects.filter(code__startswith='CCU'):
            num = p.code.replace('CCU', '')
            patch_map[num] = p

        # Excel column indices (0-based)
        COL_CODE = 0        # A: 编号
        COL_PRIORITY = 1    # B: 位置重要程度
        COL_NAME = 2        # C: 通用名称
        COL_DESC = 3        # D: 灌溉管理用的位置
        COL_STATUS = 4      # E: 当前状态
        COL_SPRINKLER = 5   # F: 灌水器类型
        COL_INTENSITY = 6   # G: 灌溉强度 mm/h
        COL_AREA = 7        # H: 区域面积
        COL_IRR_ZONE = 8    # I: 灌溉分区
        COL_CCU_NUM = 9     # J: CCU编号
        COL_VALVE_SIZE = 10 # K: 电磁阀尺寸
        COL_LC = 11         # L: 景观系数
        COL_PLANT_TYPE = 12 # M: 植物类型
        COL_IRR_FOREMAN = 13 # N: 灌溉领班
        COL_GREEN_ZONE = 14 # O: 绿化分区
        COL_GREEN_FOREMAN = 15 # P: 绿化领班
        COL_PEST_ZONE = 16  # Q: 植保分区
        COL_PEST_FOREMAN = 17 # R: 植保领班
        COL_TERRAIN = 18    # S: 地形特点
        COL_PLANT_FEATURE = 19 # T: 植物特点
        COL_SOIL = 20       # U: 土壤湿度
        COL_EQUIP_MAINT = 21 # V: 灌溉设备维护记录
        COL_IRR_MGMT = 22   # W: 灌溉管理以往记录

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=23, values_only=True):
            zone_code = row[COL_CODE]
            if not zone_code or not isinstance(zone_code, str):
                skipped += 1
                continue

            zone_code = zone_code.strip()
            parts = zone_code.split('-')
            if len(parts) < 2:
                skipped += 1
                continue

            ccu_prefix = parts[0]
            patch = patch_map.get(ccu_prefix)
            if not patch:
                skipped += 1
                self.stdout.write(f'  SKIP {zone_code}: no CCU{ccu_prefix} patch')
                continue

            name = str(row[COL_NAME] or '').strip() or zone_code
            description = str(row[COL_DESC] or '').strip()
            priority_raw = str(row[COL_PRIORITY] or '').strip()
            priority = PRIORITY_MAP.get(priority_raw, Zone.PRIORITY_MEDIUM)

            def _float(val):
                if val is None or val == '':
                    return None
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return None

            zone = Zone.objects.filter(code=zone_code).first()
            zone_data = dict(
                name=name,
                description=description,
                patch=patch,
                priority=priority,
                current_status=str(row[COL_STATUS] or '').strip(),
                sprinkler_type=str(row[COL_SPRINKLER] or '').strip(),
                irrigation_intensity=_float(row[COL_INTENSITY]),
                solenoid_valve_size=_float(row[COL_VALVE_SIZE]),
                landscape_coefficient=_float(row[COL_LC]),
                plant_type=str(row[COL_PLANT_TYPE] or '').strip(),
                irrigation_foreman=str(row[COL_IRR_FOREMAN] or '').strip(),
                greenery_zone=str(row[COL_GREEN_ZONE] or '').strip(),
                greenery_foreman=str(row[COL_GREEN_FOREMAN] or '').strip(),
                pest_control_zone=str(row[COL_PEST_ZONE] or '').strip(),
                pest_control_foreman=str(row[COL_PEST_FOREMAN] or '').strip(),
                terrain_feature=str(row[COL_TERRAIN] or '').strip(),
                plant_feature=str(row[COL_PLANT_FEATURE] or '').strip(),
                soil_moisture=str(row[COL_SOIL] or '').strip(),
                equipment_maintenance_notes=str(row[COL_EQUIP_MAINT] or '').strip(),
                irrigation_management_notes=str(row[COL_IRR_MGMT] or '').strip(),
            )

            if zone:
                for k, v in zone_data.items():
                    setattr(zone, k, v)
                if not dry_run:
                    zone.save(update_fields=['updated_at'] + list(zone_data.keys()))
                updated += 1
            else:
                if not dry_run:
                    Zone.objects.create(code=zone_code, **zone_data)
                created += 1

        wb.close()

        self.stdout.write(self.style.SUCCESS(
            f'\nDone: {created} created, {updated} updated, {skipped} skipped'))
        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN — no changes were written'))

        # --- Summary ---
        self.stdout.write('\n=== Zone count per CCU ===')
        for ccu_num in sorted(patch_map.keys(), key=lambda x: int(x)):
            patch = patch_map[ccu_num]
            count = Zone.objects.filter(patch=patch).count()
            self.stdout.write(f'  CCU{ccu_num} ({patch.name}): {count} zones')
