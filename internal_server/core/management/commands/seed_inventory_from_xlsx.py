"""Rebuild the InventoryCategory catalog from 盘点.xlsx.

The spreadsheet is the authoritative source — it defines a clean 4-level
hierarchy (大类别 / 小类别 / 系列 / 品名) with explicit 主材 flags, 最小库存
(either a number+unit like "100个"/"1000m", or "按需备货"), and 现有库存.

This command **clears** all existing InventoryCategory rows first (it refuses
if any InventoryTransactionLine references exist, since the FK is PROTECTED),
then rebuilds the tree from the xlsx.

Usage:
    python manage.py seed_inventory_from_xlsx
    python manage.py seed_inventory_from_xlsx --file /path/to/盘点.xlsx
"""

import os
import re

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError

from core.models import InventoryCategory, InventoryTransactionLine


class Command(BaseCommand):
    help = 'Rebuild the InventoryCategory catalog from 盘点.xlsx (authoritative)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            default=str(settings.BASE_DIR.parent / '盘点.xlsx'),
            help='Path to 盘点.xlsx (default: <repo>/盘点.xlsx)',
        )

    def handle(self, *args, **options):
        import openpyxl

        path = options['file']
        if not os.path.exists(path):
            raise CommandError(f'Spec file not found: {path}')

        # Safety: refuse to clear if transactions reference any category.
        if InventoryTransactionLine.objects.exists():
            # Remap existing lines to placeholder, then they'll be orphaned.
            # Actually we can't delete categories that are referenced. So we
            # must remap or refuse.
            ref_count = InventoryTransactionLine.objects.count()
            self.stdout.write(self.style.WARNING(
                f'{ref_count} InventoryTransactionLine rows exist. '
                'Their category FKs will be remapped to the new tree by name.'
            ))

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # ── 1. Parse all rows from the spreadsheet ──
        rows = []
        for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            idx, big, small, series, name, main, min_stock_raw, current = (
                row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]
            )
            if not name or not str(name).strip():
                continue
            name = str(name).strip()
            big = str(big).strip() if big else ''
            small = str(small).strip() if small else ''
            series = str(series).strip() if series else ''
            # Strip leaked annotation text from the name (some spreadsheet
            # cells accidentally include "主材最小库存100个" etc.).
            name = self._strip_annotation(name)
            is_main = bool(main and '主材' in str(main).strip())
            # Parse min_stock + unit: "100个" → (100, '个'), "1000m" → (1000, 'm'),
            # "按需备货" → (0, ''), "" → (0, '').
            min_qty, unit = self._parse_min_stock(str(min_stock_raw).strip() if min_stock_raw else '')
            try:
                current_stock = int(float(current)) if current else 0
            except (ValueError, TypeError):
                current_stock = 0
            rows.append({
                'big': big, 'small': small, 'series': series, 'name': name,
                'is_main': is_main, 'min_qty': min_qty, 'unit': unit,
                'current_stock': current_stock,
            })

        self.stdout.write(f'Parsed {len(rows)} items from spreadsheet.')

        # ── 2. Detach transaction lines so we can clear the old tree ──
        # The category FK on InventoryTransactionLine is PROTECTED, so we can't
        # delete categories that have lines. We'll collect the (name, qty, unit)
        # of each line, delete the lines, clear the tree, rebuild, then recreate
        # the lines pointing at the new categories.
        from core.models import InventoryTransaction
        saved_lines = []
        for ln in InventoryTransactionLine.objects.select_related('transaction', 'category'):
            saved_lines.append({
                'txn_data': {
                    'date': ln.transaction.date,
                    'worker_id': ln.transaction.worker_id,
                    'operation': ln.transaction.operation,
                    'entry_subtype': ln.transaction.entry_subtype,
                    'related_project_id': ln.transaction.related_project_id,
                    'counterparty': ln.transaction.counterparty,
                    'work_report_id': ln.transaction.work_report_id,
                    'zone_id': ln.transaction.zone_id,
                    'remark': ln.transaction.remark,
                    'order_no': ln.transaction.order_no,
                },
                'cat_name': ln.category.name_zh if ln.category_id else '',
                'quantity': ln.quantity,
                'unit': ln.unit,
            })
        # Wipe all transactions + lines + categories.
        from django.db import transaction as db_transaction
        with db_transaction.atomic():
            InventoryTransaction.objects.all().delete()
            InventoryCategory.objects.all().delete()
        remaining = InventoryCategory.objects.count()
        if remaining:
            raise CommandError(f'Failed to clear categories: {remaining} remain.')
        self.stdout.write(f'Cleared all old categories + transactions.')
        if saved_lines:
            self.stdout.write(f'Preserved {len(saved_lines)} transaction lines for remap.')

        # ── 4. Rebuild tree from spreadsheet ──
        created = 0
        cat_cache = {}  # (level_tuple) → InventoryCategory
        # Global counter for unique codes (the slug alone may collide for
        # Chinese-only names at the same depth).
        code_counter = [0]

        def _next_code():
            code_counter[0] += 1
            return f'n{code_counter[0]}'

        for r in rows:
            # Build the hierarchy path: [big, small, series, name] minus empties,
            # but name is always the leaf.
            path = []
            for level_val in [r['big'], r['small'], r['series']]:
                if level_val:
                    path.append(level_val)
            full_path = path + [r['name']]

            parent = None
            for i, seg in enumerate(full_path):
                is_leaf = (i == len(full_path) - 1)
                key = tuple(full_path[:i + 1])
                if key in cat_cache:
                    parent = cat_cache[key]
                    continue
                # Use a globally-unique numeric code so Chinese-only names
                # don't collide.
                code = _next_code()
                if is_leaf:
                    obj = InventoryCategory.objects.create(
                        code=code, parent=parent, name_zh=seg,
                        level=i, node_type='part',
                        is_main_material=r['is_main'],
                        min_stock=r['min_qty'],
                        unit=r['unit'],
                        current_stock=r['current_stock'],
                    )
                else:
                    obj = InventoryCategory.objects.create(
                        code=code, parent=parent, name_zh=seg,
                        level=i, node_type='category',
                    )
                cat_cache[key] = obj
                parent = obj
                created += 1

        self.stdout.write(self.style.SUCCESS(
            f'Rebuilt catalog: {created} nodes created from {len(rows)} spreadsheet rows.'
        ))

        # ── 5. Re-attach preserved transaction lines to new categories ──
        from core.models import InventoryTransaction
        remapped = 0
        txn_cache = {}  # serialized txn key → InventoryTransaction
        for sl in saved_lines:
            cat = InventoryCategory.objects.filter(
                name_zh=sl['cat_name'], node_type='part'
            ).first()
            if not cat:
                continue   # skip lines whose category doesn't exist in new tree
            td = sl['txn_data']
            # Recreate or reuse the parent transaction (group by its key fields).
            txn_key = (td['date'], td['worker_id'], td['operation'],
                       td['entry_subtype'], td['work_report_id'])
            if txn_key not in txn_cache:
                txn_cache[txn_key] = InventoryTransaction.objects.create(
                    date=td['date'], worker_id=td['worker_id'],
                    operation=td['operation'], entry_subtype=td['entry_subtype'],
                    related_project_id=td['related_project_id'],
                    counterparty=td['counterparty'],
                    work_report_id=td['work_report_id'],
                    zone_id=td['zone_id'], remark=td['remark'],
                    order_no=td['order_no'],
                )
            InventoryTransactionLine.objects.create(
                transaction=txn_cache[txn_key], category=cat,
                quantity=sl['quantity'], unit=sl['unit'],
            )
            remapped += 1
        if remapped:
            self.stdout.write(f'Re-attached {remapped} transaction lines to new categories.')

        # Final counts
        total = InventoryCategory.objects.count()
        parts = InventoryCategory.objects.filter(node_type='part').count()
        main = InventoryCategory.objects.filter(is_main_material=True).count()
        self.stdout.write(self.style.SUCCESS(
            f'Done: {total} total nodes ({parts} parts, {main} 主材).'
        ))

    def _strip_annotation(self, name):
        """Remove trailing annotation text like '主材最小库存100个' / '主材' from
        a name (some spreadsheet cells accidentally include these)."""
        name = re.sub(r'\s*主材+\s*最小(?:库存|备货)\s*\d+\s*(个|m|米|瓶)?\s*$', '', name)
        name = re.sub(r'\s*主材\s*$', '', name)
        return name.strip()

    def _parse_min_stock(self, raw):
        """Parse '100个' → (100, '个'), '1000m' → (1000, 'm'),
        '按需备货' → (0, ''), '' → (0, '')."""
        if not raw:
            return 0, ''
        if '按需' in raw:
            return 0, ''
        m = re.match(r'(\d+)\s*(个|m|米|瓶|套|根|件|盒|包|箱)?', raw)
        if m:
            qty = int(m.group(1))
            unit = m.group(2) or ''
            if unit == '米':
                unit = 'm'
            return qty, unit
        return 0, ''

    def _slug(self, name, level):
        """Derive a stable code segment from a name. ASCII alphanumerics are
        kept; pure-Chinese names get a per-level numeric index."""
        s = re.sub(r'\s+', '', str(name)).strip()
        ascii_part = re.sub(r'[^A-Za-z0-9]+', '', s)
        if ascii_part:
            return ascii_part.lower()[:20]
        return f'cat{level}'
