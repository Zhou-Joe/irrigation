"""Views for the refactored 现场作业记录 (work order) form.

Server-rendered, mobile-first single-page form driven by the WorkItem
template tree (seeded from 工单记录格式.md). One view handles create + edit:
header fields are written to WorkReport, every filled node becomes a
WorkReportEntry row. See
docs/superpowers/specs/2026-06-16-workorder-form-refactor-design.md.
"""

import json
import os
from datetime import date, datetime, timedelta
from functools import lru_cache

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST
from django.core.files.storage import default_storage
from django.db import transaction
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render

from core.models import (
    Patch, Project, Worker, WorkItem,
    WorkReport, WorkReportEntry, Zone,
)


# ── PM (preventive maintenance) helpers ─────────────────────────────────────

@lru_cache(maxsize=1)
def _pm_leaf_id():
    """Cached id of the WorkItem with code '1.2.pm' (PM作业计划 leaf).

    This is static reference data (seeded by migration 0094), so we cache it
    to avoid a DB query on every PM work-order edit-modal render.
    """
    leaf = WorkItem.objects.filter(code='1.2.pm').only('id').first()
    return leaf.id if leaf else None


def mark_pm_completed(report):
    """Mark a PM-dispatched GeneratedWorkOrder as completed.

    Called from BOTH submit paths (mobile `workorder_mobile_v2` and desktop
    `_handle_save`) so that completing a PM task via either UI records the
    completion. Safe to call on non-PM reports (no-op when the report has no
    `pm_generation`).

    Re-fetches the GWO with select_for_update so concurrent submits by two
    crew members (the whole crew sees each task) serialize at the DB level —
    the second caller sees status='completed' and becomes a no-op.
    """
    from core.models import GeneratedWorkOrder
    try:
        gwo = (GeneratedWorkOrder.objects
               .select_for_update()
               .get(work_report=report))
    except GeneratedWorkOrder.DoesNotExist:
        return
    if gwo.status not in ('dispatched', 'overdue'):
        return
    from django.utils import timezone as _tz
    gwo.status = 'completed'
    gwo.completed_at = _tz.now()
    gwo.save(update_fields=['status', 'completed_at'])


# ── serialization ──────────────────────────────────────────────────────────

def serialize_workitem_tree():
    """Emit the 12-section WorkItem tree as nested JSON (depth-first)."""
    qs = (WorkItem.objects.filter(active=True)
          .order_by('section', 'order', 'code')
          .values('id', 'code', 'name_zh', 'parent_id', 'section', 'value_type',
                  'unit', 'is_project_scoped', 'status_options'))
    nodes = {n['id']: {**n, 'name': n['name_zh'], 'children': []} for n in qs}
    roots = []
    for n in qs:
        node = nodes[n['id']]
        pid = n['parent_id']
        if pid in nodes:
            nodes[pid]['children'].append(node)
        else:
            roots.append(node)
    return roots


def serialize_projects():
    return [
        {'id': p.id, 'name': p.name, 'category': p.category,
         'category_display': p.get_category_display(),
         'subcategory': p.subcategory, 'subcategory_display': p.get_subcategory_display(),
         'symbol': p.symbol, 'code': p.code}
        for p in Project.objects.filter(active=True).order_by('category', 'subcategory', 'name')
    ]


def IRRIGATION_SUBCATEGORIES():
    return [{'code': c, 'label': label} for c, label in Project.SUBCATEGORY_CHOICES]


@login_required(login_url='core:login')
def project_create_api(request):
    """Create a Project instance (manager / admin only).

    POST {name, category, subcategory, symbol, code} → Project (get_or_create).
    Used by the mobile workorder form when no existing project matches.
    """
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        return JsonResponse({'error': '无权限创建项目（仅经理/管理员）'}, status=403)
    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        data = {}
    name = (data.get('name') or '').strip()
    category = (data.get('category') or 'IRRIGATION').strip()
    if category not in {c for c, _ in Project.CATEGORY_CHOICES}:
        category = 'IRRIGATION'
    subcategory = (data.get('subcategory') or '').strip()
    if subcategory and subcategory not in {c for c, _ in Project.SUBCATEGORY_CHOICES}:
        subcategory = ''
    if category != 'IRRIGATION':
        subcategory = ''  # subcategory only for irrigation
    if not name:
        return JsonResponse({'error': '项目名称不能为空'}, status=400)
    proj, created = Project.objects.get_or_create(
        category=category, subcategory=subcategory, name=name,
        defaults={'active': True,
                  'symbol': (data.get('symbol') or '').strip(),
                  'code': (data.get('code') or '').strip()})
    return JsonResponse({
        'id': proj.id, 'name': proj.name, 'category': proj.category,
        'category_display': proj.get_category_display(),
        'subcategory': proj.subcategory, 'subcategory_display': proj.get_subcategory_display(),
        'symbol': proj.symbol, 'code': proj.code, 'created': created,
    })


# ── 项目管理 page ──────────────────────────────────────────────────────────

_PROJECT_SECTIONS = ('irrigation_project', 'drainage_project', 'other_project')


def _phase_map():
    """Map work_item_id → its phase (the level-1 child of the project section root:
    设计 / 费用评估 / 材料准备 / 施工) for the three project sections."""
    items = WorkItem.objects.filter(section__in=_PROJECT_SECTIONS).values('id', 'name_zh', 'parent_id')
    by_id = {it['id']: it for it in items}
    phase = {}
    for it in items:
        cur = it
        # walk up until cur's parent is the section root (root has parent_id None)
        while cur and cur['parent_id'] and by_id.get(cur['parent_id'], {}).get('parent_id'):
            cur = by_id.get(cur['parent_id'])
        phase[it['id']] = cur['name_zh'] if cur else ''
    return phase


def _project_summaries(projects):
    """Per-project work summary: {project_id: {reports, hours, third_hours, phases}}."""
    pmap = _phase_map()
    summ = {p.id: {'reports': set(), 'hours': 0, 'third_hours': 0, 'phases': {}}
            for p in projects}
    entries = (WorkReportEntry.objects.filter(project__in=[p.id for p in projects])
               .select_related('work_item'))
    rep_to_projects = {}
    for e in entries:
        s = summ[e.project_id]
        s['reports'].add(e.work_report_id)
        ph = pmap.get(e.work_item_id, '其他')
        s['phases'][ph] = s['phases'].get(ph, 0) + (e.count or 0)
        rep_to_projects.setdefault(e.work_report_id, set()).add(e.project_id)
    # Fetch both labor fields in one query.
    rep_hours = {r['id']: (r['team_hours'] or 0, r['third_party_hours'] or 0)
                 for r in WorkReport.objects.filter(id__in=rep_to_projects.keys())
                 .values('id', 'team_hours', 'third_party_hours')}
    for rid, pids in rep_to_projects.items():
        team_h, third_h = rep_hours.get(rid, (0, 0))
        for pid in pids:
            summ[pid]['hours'] += team_h
            summ[pid]['third_hours'] += third_h
    return summ


def _project_budget_data(projects):
    """Per-project consumed material + linked POs + budget amounts, for the panel.

    Returns {project_id: {material_consumed, consumption_records,
    material_budget_amount, labor_budget_amount, linked_pos, work_reports}}.
    - material_consumed: aggregated from 出库-项目 transactions' lines, by category
      (used to show what materials were actually pulled for this project).
    - material_budget_amount / labor_budget_amount: the project's amount fields.
    - linked_pos: [{id, order_number, po_number, po_amount_untaxed}].
    材料余额 = material_budget_amount − Σ(linked_pos.po_amount_untaxed) is computed
    here as material_balance and exposed to the table + panel.
    """
    from core.models import (ProjectPurchaseOrder, InventoryTransaction)
    from core.views import inventory_category_paths, _po_received_parts_batch
    pids = [p.id for p in projects]

    # Precompute every InventoryCategory's full display path once (1 query).
    cat_paths = inventory_category_paths()

    # Material consumed: aggregate 出库-项目 lines by category, AND keep the
    # raw per-transaction records (date/worker/source/lines) for the 出库记录 list.
    # Only 'actual' consumption counts toward the balance — estimated (预估消耗)
    # transactions are tracked separately below and shown as a reminder, not
    # deducted from the budget until confirmed.
    consumed = {}
    records = {}
    txns = (InventoryTransaction.objects
            .filter(related_project_id__in=pids, operation='出库', entry_subtype='项目',
                    consumption_mode='actual')
            .select_related('worker', 'work_report', 'zone')
            .prefetch_related('lines__category')
            .order_by('-date', '-id'))
    for t in txns:
        # Aggregate by category (consumption summary table).
        for ln in t.lines.all():
            cat = ln.category
            entry = consumed.setdefault(t.related_project_id, {}).get(cat.id)
            if entry is None:
                entry = {'category_id': cat.id, 'name': cat.name_zh,
                         'path': cat_paths.get(cat.id, cat.name_zh),
                         'unit': ln.unit or cat.unit, 'qty': 0}
                consumed.setdefault(t.related_project_id, {})[cat.id] = entry
            entry['qty'] += ln.quantity
        # Per-transaction record (for the 出库记录 list).
        rec = {
            'id': t.id, 'date': t.date.isoformat() if t.date else '',
            'worker': t.worker.full_name if t.worker else '—',
            'source': ('工单 #' + str(t.work_report_id)) if t.work_report_id else '出库登记',
            'work_report_id': t.work_report_id,
            'zone': t.zone.name if t.zone else '',
            'remark': t.remark or '',
            'lines': [{'name': ln.category.name_zh,
                       'path': cat_paths.get(ln.category_id, ln.category.name_zh),
                       'qty': ln.quantity, 'unit': ln.unit or ln.category.unit}
                      for ln in t.lines.all()],
        }
        records.setdefault(t.related_project_id, []).append(rec)
    consumed_list = {pid: list(d.values()) for pid, d in consumed.items()}

    # Estimated (预估消耗) consumption — pending, not yet deducted from stock or
    # the budget. Surfaced as a per-project reminder so managers can see expected
    # usage before it is confirmed. Aggregated by category per project.
    estimated = {}
    est_txns = (InventoryTransaction.objects
                .filter(related_project_id__in=pids, operation='出库', entry_subtype='项目',
                        consumption_mode='estimated')
                .prefetch_related('lines__category')
                .order_by('-date', '-id'))
    for t in est_txns:
        for ln in t.lines.all():
            cat = ln.category
            entry = estimated.setdefault(t.related_project_id, {}).get(cat.id)
            if entry is None:
                entry = {'category_id': cat.id, 'name': cat.name_zh,
                         'path': cat_paths.get(cat.id, cat.name_zh),
                         'unit': ln.unit or cat.unit, 'qty': 0}
                estimated.setdefault(t.related_project_id, {})[cat.id] = entry
            entry['qty'] += ln.quantity
    estimated_list = {pid: list(d.values()) for pid, d in estimated.items()}

    # Linked POs — with their received parts (so the panel can show, for materials
    # the project also consumes, how many that order purchased in total). Batched:
    # one query for all linked POs' received parts, not one per PO.
    links = {}
    links_po_total = {}  # {project_id: Σ po_amount_untaxed}
    link_rows = list((ProjectPurchaseOrder.objects
                      .filter(project_id__in=pids)
                      .select_related('purchase_order')
                      .order_by('-created_at')))
    linked_po_ids = [lk.purchase_order_id for lk in link_rows]
    po_parts_map = _po_received_parts_batch(linked_po_ids, cat_paths)
    for lk in link_rows:
        po = lk.purchase_order
        amt = po.po_amount_untaxed
        po_parts, _ = po_parts_map.get(po.id, ([], 0))
        links.setdefault(lk.project_id, []).append({
            'id': po.id, 'order_number': po.order_number,
            'po_number': po.po_number,
            'po_amount_untaxed': ('' if amt is None else f'{amt:.2f}'),
            'parts': po_parts,
        })
        if amt is not None:
            links_po_total[lk.project_id] = links_po_total.get(lk.project_id, 0) + amt

    # Related work reports (工单) — via WorkReportEntry.project, distinct + with hours.
    from core.models import WorkReport, WorkReportEntry
    reports = {}
    rep_ids_by_proj = {}
    for e in (WorkReportEntry.objects.filter(project_id__in=pids)
              .values('project_id', 'work_report_id').distinct()):
        rep_ids_by_proj.setdefault(e['project_id'], set()).add(e['work_report_id'])
    all_rep_ids = set().union(*rep_ids_by_proj.values()) if rep_ids_by_proj else set()
    rep_map = {r['id']: r for r in (WorkReport.objects.filter(id__in=all_rep_ids)
                .values('id', 'date', 'team_hours', 'third_party_hours'))}
    for pid, rids in rep_ids_by_proj.items():
        for rid in sorted(rids, reverse=True):
            r = rep_map.get(rid)
            if not r:
                continue
            reports.setdefault(pid, []).append({
                'id': r['id'],
                'date': r['date'].isoformat() if r['date'] else '',
                'team_hours': r['team_hours'] or 0,
                'third_hours': r['third_party_hours'] or 0,
            })

    # Split consumption records: those from a work order are grouped under that work
    # report (shown in 关联工单 with its material lines); the rest (standalone 出库
    # form) stay in consumption_records as 「其他出库记录」.
    standalone_records = {}
    for pid, recs in records.items():
        standalone_records[pid] = [r for r in recs if not r['work_report_id']]
    # Attach material-out lines to each work report.
    wr_by_pid = {pid: {wr['id']: wr for wr in wrs} for pid, wrs in reports.items()}
    for pid, recs in records.items():
        for r in recs:
            wid = r['work_report_id']
            if wid and pid in wr_by_pid and wid in wr_by_pid[pid]:
                wr_by_pid[pid][wid].setdefault('materials', []).append({
                    'date': r['date'], 'worker': r['worker'],
                    'zone': r['zone'], 'remark': r['remark'],
                    'lines': r['lines'],
                })

    # 材料余额 = 材料预算金额 − Σ关联采购订单金额（无预算时为 None；预算为0则按0算）。
    out = {}
    for p in projects:
        mba = p.material_budget_amount
        po_sum = links_po_total.get(p.id)
        if mba is None:
            balance = None
        else:
            balance = mba - (po_sum or 0)
        out[p.id] = {
            'material_consumed': consumed_list.get(p.id, []),
            'estimated_consumed': estimated_list.get(p.id, []),
            'consumption_records': standalone_records.get(p.id, []),
            'material_budget_amount': ('' if mba is None else f'{mba:.2f}'),
            'labor_budget_amount': ('' if p.labor_budget_amount is None
                                    else f'{p.labor_budget_amount:.2f}'),
            'team_rate': ('' if p.team_rate is None else f'{p.team_rate:.2f}'),
            'third_party_rate': ('' if p.third_party_rate is None
                                 else f'{p.third_party_rate:.2f}'),
            'material_balance': ('' if balance is None else f'{balance:.2f}'),
            'po_amount_total': ('' if po_sum is None else f'{po_sum:.2f}'),
            'linked_pos': links.get(p.id, []),
            'work_reports': reports.get(p.id, []),
        }
    return out


@never_cache
@login_required(login_url='core:login')
def project_management(request):
    """List/create/edit projects grouped by category, with per-project work summary."""
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    from collections import OrderedDict
    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        messages.error(request, '无权限访问项目管理')
        return redirect('core:dashboard')

    # Filters: q (name/symbol/code/notes), category (ALL/...), completed (all/0/1).
    from django.db.models import Q
    fq = (request.GET.get('q') or '').strip()
    fcat = (request.GET.get('category') or '').strip()
    fcompleted = (request.GET.get('completed') or '').strip()
    ffrom = _parse_date(request.GET.get('from'))
    fto = _parse_date(request.GET.get('to'))

    qs = Project.objects.all()
    if fq:
        qs = qs.filter(Q(name__icontains=fq) | Q(symbol__icontains=fq)
                       | Q(code__icontains=fq) | Q(notes__icontains=fq))
    if fcat in {c for c, _ in Project.CATEGORY_CHOICES}:
        qs = qs.filter(category=fcat)
    if fcompleted in ('0', '1'):
        qs = qs.filter(is_completed=(fcompleted == '1'))
    # Date range filters against 开工日期. A project with no start_date is only
    # shown when no date range is applied (can't place it on the timeline).
    if ffrom or fto:
        qs = qs.exclude(start_date__isnull=True)
        if ffrom:
            qs = qs.filter(start_date__gte=ffrom)
        if fto:
            qs = qs.filter(start_date__lte=fto)
    projects = list(qs.order_by('category', 'subcategory', 'name'))

    summ = _project_summaries(projects)
    budget_data = _project_budget_data(projects)
    groups = OrderedDict((c, {'label': lbl, 'items': []}) for c, lbl in Project.CATEGORY_CHOICES)
    for p in projects:
        s = summ[p.id]
        b = budget_data.get(p.id, {})
        # 人工余额 = 人工预算金额 − (灌溉组已用工时 × 灌溉组rate + 第三方已用工时 × 第三方rate)。
        # 已用工时来自 _project_summaries（关联工单汇总）。预算或费率任一缺失则余额为空。
        from decimal import Decimal
        tr = p.team_rate
        tpr = p.third_party_rate
        labor_cost = (Decimal(str(s['hours'])) * tr if tr is not None else Decimal('0')) + \
                     (Decimal(str(s['third_hours'])) * tpr if tpr is not None else Decimal('0'))
        if p.labor_budget_amount is not None:
            labor_balance = p.labor_budget_amount - labor_cost
            b['labor_balance'] = f'{labor_balance:.2f}'
            b['labor_cost'] = f'{labor_cost:.2f}'
            b['team_hours'] = f'{s["hours"]:.1f}'
            b['third_hours'] = f'{s["third_hours"]:.1f}'
            labor_balance_str = f'{labor_balance:.2f}'
        else:
            b['labor_balance'] = ''
            b['labor_cost'] = f'{labor_cost:.2f}' if labor_cost else ''
            b['team_hours'] = f'{s["hours"]:.1f}'
            b['third_hours'] = f'{s["third_hours"]:.1f}'
            labor_balance_str = ''
        groups[p.category]['items'].append({
            'project': p,
            'report_count': len(s['reports']),
            'team_hours': s['hours'],
            'third_hours': s['third_hours'],
            'phases': s['phases'],
            'material_budget_amount': b.get('material_budget_amount', ''),
            'labor_budget_amount': b.get('labor_budget_amount', ''),
            'material_balance': b.get('material_balance', ''),
            'labor_balance': labor_balance_str,
            'budget_json': json.dumps(b, ensure_ascii=False),
            'edit_json': json.dumps({
                'id': p.id, 'name': p.name, 'category': p.category,
                'subcategory': p.subcategory, 'symbol': p.symbol, 'code': p.code,
                'active': p.active, 'is_completed': p.is_completed, 'notes': p.notes,
                'start_date': p.start_date.isoformat() if p.start_date else '',
                'planned_end_date': p.planned_end_date.isoformat() if p.planned_end_date else '',
            }, ensure_ascii=False),
        })
    edit_id = request.GET.get('edit')
    edit_obj = Project.objects.filter(pk=edit_id).first() if edit_id else None

    # All POs available for linking (the 关联订单 dropdown).
    from core.models import PurchaseOrder
    all_pos = [{'id': po.id, 'order_number': po.order_number, 'po_number': po.po_number}
               for po in PurchaseOrder.objects.all().order_by('-created_at')]

    ctx = {
        'groups': groups,
        'subcategories': Project.SUBCATEGORY_CHOICES,
        'categories': Project.CATEGORY_CHOICES,
        'edit_obj': edit_obj,
        'all_pos_json': json.dumps(all_pos, ensure_ascii=False),
        # filter values (to re-populate the filter bar)
        'fq': fq, 'fcat': fcat, 'fcompleted': fcompleted,
        'ffrom': ffrom.isoformat() if ffrom else '',
        'fto': fto.isoformat() if fto else '',
        'total_projects': len(projects),
    }
    return render(request, 'core/project_management.html', ctx)


@login_required(login_url='core:login')
def project_export_excel(request):
    """Export all projects (with current filters) to an Excel workbook.

    Columns mirror the management table plus two material columns: 材料预算 and
    材料消耗. Within each cell, multiple materials are newline-separated, and the
    rows in the two columns correspond by material (budget-only / consumed-only /
    both), so the line-up stays consistent.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        return redirect('core:dashboard')

    # Apply the same filters as the management page.
    from django.db.models import Q
    fq = (request.GET.get('q') or '').strip()
    fcat = (request.GET.get('category') or '').strip()
    fcompleted = (request.GET.get('completed') or '').strip()
    ffrom = _parse_date(request.GET.get('from'))
    fto = _parse_date(request.GET.get('to'))
    qs = Project.objects.all()
    if fq:
        qs = qs.filter(Q(name__icontains=fq) | Q(symbol__icontains=fq)
                       | Q(code__icontains=fq) | Q(notes__icontains=fq))
    if fcat in {c for c, _ in Project.CATEGORY_CHOICES}:
        qs = qs.filter(category=fcat)
    if fcompleted in ('0', '1'):
        qs = qs.filter(is_completed=(fcompleted == '1'))
    if ffrom or fto:
        qs = qs.exclude(start_date__isnull=True)
        if ffrom:
            qs = qs.filter(start_date__gte=ffrom)
        if fto:
            qs = qs.filter(start_date__lte=fto)
    projects = list(qs.order_by('category', 'subcategory', 'name'))
    summ = _project_summaries(projects)
    budget_data = _project_budget_data(projects)

    cat_labels = dict(Project.CATEGORY_CHOICES)
    sub_labels = dict(Project.SUBCATEGORY_CHOICES)

    wb = Workbook()
    ws = wb.active
    ws.title = '项目管理'
    headers = ['项目', '类别', '子类别', '开工日期', '计划完工日期',
               '材料预算金额', '人工预算金额', '材料余额', '人工余额',
               '灌溉组工时', '第三方工时', '工单数', '状态', '备注']
    ws.append(headers)

    # Styles.
    header_fill = PatternFill('solid', fgColor='1B4332')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    thin = Side(style='thin', color='D9D0C0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical='top')
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = header_fill; c.font = header_font; c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')

    for p in projects:
        s = summ[p.id]
        b = budget_data.get(p.id, {})
        # 人工余额 = 人工预算 − 已用工时×费率。Decimal 与 float 不能直接运算，统一转 float。
        tr = p.team_rate
        tpr = p.third_party_rate
        labor_cost = (s['hours'] * float(tr) if tr is not None else 0) + \
                     (s['third_hours'] * float(tpr) if tpr is not None else 0)
        labor_bal = (float(p.labor_budget_amount) - labor_cost) \
            if p.labor_budget_amount is not None else ''
        row = [
            p.name,
            cat_labels.get(p.category, p.category),
            sub_labels.get(p.subcategory, p.subcategory or ''),
            p.start_date.isoformat() if p.start_date else '',
            p.planned_end_date.isoformat() if p.planned_end_date else '',
            float(p.material_budget_amount) if p.material_budget_amount is not None else '',
            float(p.labor_budget_amount) if p.labor_budget_amount is not None else '',
            float(b.get('material_balance') or 0) if b.get('material_balance') != '' else '',
            labor_bal,
            round(s['hours'], 1),
            round(s['third_hours'], 1),
            len(s['reports']),
            '已完成' if p.is_completed else '进行中',
            p.notes,
        ]
        ws.append(row)
        r = ws.max_row
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).border = border
            ws.cell(row=r, column=ci).alignment = wrap_top

    # Column widths.
    widths = [20, 10, 10, 12, 12, 14, 14, 14, 14, 10, 10, 8, 8, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="projects.xlsx"'
    return resp


def _clean_project_fields(post):
    name = (post.get('name') or '').strip()
    category = (post.get('category') or 'IRRIGATION').strip()
    if category not in {c for c, _ in Project.CATEGORY_CHOICES}:
        category = 'IRRIGATION'
    subcategory = (post.get('subcategory') or '').strip()
    if subcategory and subcategory not in {c for c, _ in Project.SUBCATEGORY_CHOICES}:
        subcategory = ''
    if category != 'IRRIGATION':
        subcategory = ''
    return name, category, subcategory


def _parse_date(raw):
    """Parse a YYYY-MM-DD string into a date, or None if blank/invalid."""
    from datetime import date as _date
    s = (str(raw) if raw else '').strip()
    if not s:
        return None
    try:
        return _date.fromisoformat(s)
    except ValueError:
        return None


@login_required(login_url='core:login')
def project_save(request):
    """Create or update a Project (manager / admin only)."""
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        messages.error(request, '无权限')
        return redirect('core:dashboard')
    if request.method != 'POST':
        return redirect('core:project_management')
    name, category, subcategory = _clean_project_fields(request.POST)
    if not name:
        messages.error(request, '项目名称不能为空')
        return redirect('core:project_management')
    pid = request.POST.get('id')
    symbol = (request.POST.get('symbol') or '').strip()
    code = (request.POST.get('code') or '').strip()
    notes = (request.POST.get('notes') or '').strip()
    active = bool(request.POST.get('active'))
    is_completed = bool(request.POST.get('is_completed'))
    start_date = _parse_date(request.POST.get('start_date'))
    planned_end_date = _parse_date(request.POST.get('planned_end_date'))
    if pid:
        proj = get_object_or_404(Project, pk=pid)
        proj.name = name; proj.category = category; proj.subcategory = subcategory
        proj.symbol = symbol; proj.code = code; proj.notes = notes
        proj.active = active; proj.is_completed = is_completed
        proj.start_date = start_date; proj.planned_end_date = planned_end_date
        proj.save()
        messages.success(request, f'项目已更新：{proj.name}')
    else:
        proj, created = Project.objects.get_or_create(
            category=category, subcategory=subcategory, name=name,
            defaults={'symbol': symbol, 'code': code, 'notes': notes,
                      'active': active, 'is_completed': is_completed,
                      'start_date': start_date, 'planned_end_date': planned_end_date})
        messages.success(request, ('项目已创建：' if created else '项目已存在：') + proj.name)
    return redirect('core:project_management')


@login_required(login_url='core:login')
def project_delete(request, pk):
    """Delete a Project (manager / admin only). Entries' project FK is SET_NULL."""
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        messages.error(request, '无权限')
        return redirect('core:dashboard')
    if request.method == 'POST':
        proj = get_object_or_404(Project, pk=pk)
        name = proj.name
        proj.delete()
        messages.success(request, f'项目已删除：{name}')
    return redirect('core:project_management')


@login_required(login_url='core:login')
@require_POST
def project_budget_save(request, pk):
    """Save a project's material budget amount + labor budget amount.

    材料余额 = 材料预算金额 − Σ关联采购订单金额，在前端/明细面板实时展示，
    此处只持久化两个金额字段。
    """
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    from decimal import Decimal, InvalidOperation
    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)
    proj = get_object_or_404(Project, pk=pk)

    def _parse_amount(key, label):
        raw = (request.POST.get(key) or '').strip()
        if not raw:
            return None, None
        try:
            val = Decimal(raw)
        except (InvalidOperation, ValueError):
            return None, f'{label}格式无效'
        if val < 0:
            return None, f'{label}不能为负'
        return val, None

    mba, err = _parse_amount('material_budget_amount', '材料预算金额')
    if err:
        return JsonResponse({'success': False, 'message': err}, status=400)
    lba, err = _parse_amount('labor_budget_amount', '人工预算金额')
    if err:
        return JsonResponse({'success': False, 'message': err}, status=400)
    tr, err = _parse_amount('team_rate', '灌溉组工时费率')
    if err:
        return JsonResponse({'success': False, 'message': err}, status=400)
    tpr, err = _parse_amount('third_party_rate', '第三方工时费率')
    if err:
        return JsonResponse({'success': False, 'message': err}, status=400)
    proj.material_budget_amount = mba
    proj.labor_budget_amount = lba
    proj.team_rate = tr
    proj.third_party_rate = tpr
    proj.save(update_fields=['material_budget_amount', 'labor_budget_amount',
                             'team_rate', 'third_party_rate'])
    return JsonResponse({'success': True, 'message': '预算已保存'})


@login_required(login_url='core:login')
@require_POST
def project_pos_save(request, pk):
    """Save a project's linked purchase orders (JSON [po_id, ...], replace)."""
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    from core.models import ProjectPurchaseOrder, PurchaseOrder
    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)
    proj = get_object_or_404(Project, pk=pk)

    try:
        po_ids = json.loads(request.POST.get('po_ids', '[]'))
    except (json.JSONDecodeError, TypeError):
        po_ids = []
    if not isinstance(po_ids, list):
        po_ids = []
    po_ids = [int(x) for x in po_ids if str(x).isdigit()]

    # Replace: drop links not in the payload, add the rest.
    ProjectPurchaseOrder.objects.filter(project=proj).exclude(purchase_order_id__in=po_ids).delete()
    existing = set(ProjectPurchaseOrder.objects.filter(project=proj)
                   .values_list('purchase_order_id', flat=True))
    valid = set(PurchaseOrder.objects.filter(pk__in=po_ids).values_list('id', flat=True))
    for pid in valid:
        if pid not in existing:
            ProjectPurchaseOrder.objects.create(project=proj, purchase_order_id=pid)
    return JsonResponse({'success': True, 'message': '关联订单已更新'})


@login_required(login_url='core:login')
def planned_maintenance_pending(request):
    """Past WorkReports (in the given zones, or all) flagged 待修 — the PM backlog.

    GET ?zones=CODE,CODE → {reports: [{id, date, worker, items: [path,...]}]}.
    Used by the 计划性维修 work-content view.
    """
    zone_codes = [z for z in (request.GET.get('zones') or '').split(',') if z]
    qs = WorkReport.objects.filter(is_pending_repair=True).distinct()
    if zone_codes:
        qs = qs.filter(zones__code__in=zone_codes)
    qs = qs.select_related('worker').order_by('-date', '-id')[:60]
    out = []
    for r in qs:
        items = []
        content = (r.work_content or '').strip()
        if content:
            items.append(content)
        remark = (r.remark or '').strip()
        if remark:
            items.append(remark)
        if not items:
            items = ['待修']
        out.append({
            'id': r.id,
            'date': r.date.isoformat() if r.date else '',
            'worker': r.worker.full_name if r.worker else '',
            'items': items,
        })
    return JsonResponse({'reports': out})


def serialize_existing_entries(report):
    """Pre-fill list for edit mode (JS indexes by work_item id)."""
    return [
        {'work_item': e.work_item_id, 'project': e.project_id,
         'count': e.count, 'status': e.status, 'text_value': e.text_value,
         'photos': e.photos or []}
        for e in report.entries.all()
    ]


# ── listing helpers: build readable entry paths + grouped summaries ─────────
# Used by both the desktop 维修日志 table and the mobile 工作记录 card list so
# they can render the tree-based WorkReportEntry content (full ancestor path
# per node, grouped by section) instead of the legacy 故障详情 rows.

_section_labels = dict(WorkItem.SECTION_CHOICES)
_section_order = [code for code, _ in WorkItem.SECTION_CHOICES]


def workitem_path_map():
    """Return ``{work_item_id: '常规维护 › 喷头 › 待修'}`` for every WorkItem.

    Loads the whole tree once and walks parent pointers in memory (memoized on
    the class, so a single request pays the cost once regardless of how many
    reports are listed). The path includes the section-root name so it reads
    naturally when an entry is shown outside a grouped context.
    """
    cache = getattr(workitem_path_map, '_cache', None)
    if cache is not None:
        return cache
    rows = WorkItem.objects.values('id', 'name_zh', 'parent_id')
    by_id = {r['id']: {'name': r['name_zh'], 'parent_id': r['parent_id']} for r in rows}
    path_of = {}

    def build(wid):
        if wid in path_of:
            return path_of[wid]
        node = by_id.get(wid)
        if not node or not node['parent_id']:
            path_of[wid] = node['name'] if node else ''
            return path_of[wid]
        up = build(node['parent_id'])
        path_of[wid] = (up + ' › ' + node['name']) if up else node['name']
        return path_of[wid]

    for wid in by_id:
        build(wid)
    workitem_path_map._cache = path_of
    return path_of


def entry_value_label(entry):
    """Human-readable value for one WorkReportEntry (mirrors mobile woValueLabel)."""
    wi = entry.work_item
    vtype = getattr(wi, 'value_type', 'count')
    if vtype == 'count':
        if entry.count:
            return f"{entry.count}{wi.unit}" if getattr(wi, 'unit', '') else str(entry.count)
        return ''
    if vtype == 'status':
        return entry.status or ''
    if vtype in ('text', 'text_photo'):
        return entry.text_value or ''
    # toggle — no numeric value; presence is the value.
    return '✓'


def enrich_reports(reports, path_map=None):
    """Attach grouped/summarized entry data to each report in ``reports``.

    Adds, per report:
      * ``entry_groups`` — [{section, section_label, items:[{path, value, project, photos}]}, …]
        ordered by SECTION_CHOICES; only sections with entries appear.
      * ``entry_count`` — int
      * ``section_labels`` — ordered list of section display labels present

    Expects ``entries__work_item`` and ``entries__project`` to be prefetched on the queryset.
    """
    path_map = path_map if path_map is not None else workitem_path_map()
    order_index = {code: i for i, code in enumerate(_section_order)}
    for report in reports:
        groups = {}
        for e in report.entries.all():
            wi = e.work_item
            sec = getattr(wi, 'section', '') or ''
            g = groups.setdefault(sec, {'section': sec,
                                        'section_label': _section_labels.get(sec, sec),
                                        'items': []})
            g['items'].append({
                'path': path_map.get(wi.id, wi.name_zh),
                'value': entry_value_label(e),
                'project': e.project.name if e.project_id and e.project else '',
                'photos': e.photos or [],
            })
        ordered = sorted(groups.values(), key=lambda g: order_index.get(g['section'], 999))
        report.entry_groups = ordered
        report.entry_count = sum(len(g['items']) for g in ordered)
        report.section_labels = [g['section_label'] for g in ordered]


def attach_zone_hierarchy(reports):
    """Attach a deduplicated Land → 通用名称 → [codes] hierarchy to each report.

    A report often references many zones that share a land/name; this collapses them
    so list/detail pages don't list the same name dozens of times. Expects
    ``zones__land`` to be prefetched on each report. Sets ``report.zone_hierarchy``
    and ``report.zone_summary`` (a flat "Land·name/name" string for compact display).
    """
    for report in reports:
        lands = {}      # land_name -> { name -> [codes] }
        order = []
        for z in report.zones.all():
            ln = (z.land.name if z.land_id and z.land else '其它') or '其它'
            nm = z.name or z.code
            if ln not in lands:
                lands[ln] = {}
                order.append(ln)
            lands[ln].setdefault(nm, []).append(z.code)
        report.zone_hierarchy = [
            {'land': ln,
             'names': [{'name': nm, 'codes': codes, 'count': len(codes)}
                        for nm, codes in sorted(lands[ln].items())],
             'zone_count': sum(len(v) for v in lands[ln].values())}
            for ln in order
        ]
        parts = [ln + '·' + '/'.join(sorted(lands[ln].keys())) for ln in order]
        report.zone_summary = '、'.join(parts) if parts else ''


# ── photo + hours helpers ───────────────────────────────────────────────────

def _save_photo(report, uploaded):
    """Persist one uploaded file under media/workorder_photos/<report_id>/ and
    generate a small thumbnail alongside it.

    Returns the original's relative path (unchanged contract). The thumbnail is
    saved at the same path with a ``_thumb`` suffix + ``.jpg`` extension, and can
    be derived via :func:`thumb_path`. The list page loads thumbnails (a few KB)
    instead of multi-MB originals so it doesn't choke on a cloud tunnel.

    Raises ValueError for files failing the allow-list / content check so the
    caller can reject the whole submission rather than silently dropping a file.
    """
    from core.upload_security import validate_upload
    ok, err = validate_upload(uploaded)
    if not ok:
        raise ValueError(f'{getattr(uploaded, "name", "")}: {err}')

    subdir = f'workorder_photos/{report.id}'
    name = default_storage.get_available_name(os.path.join(subdir, uploaded.name))
    saved = default_storage.save(name, uploaded)

    # Best-effort thumbnail — never let a thumbnail failure break the upload.
    try:
        _make_thumbnail(saved, uploaded)
    except Exception:
        pass
    return saved


def thumb_path(original_path):
    """Derive the thumbnail path for an original media path.

    ``workorder_photos/12/IMG_1234.jpg`` → ``workorder_photos/12/IMG_1234_thumb.jpg``
    Works for both photos and videos (video poster is a jpg).
    """
    base, ext = os.path.splitext(original_path)
    return base + '_thumb.jpg'


def _delete_media(relative_path):
    """Best-effort delete of a media file and its thumbnail from default_storage.

    Used when a user removes an existing report/entry photo on edit. Errors are
    swallowed — a stale DB row should never block a save.
    """
    try:
        if relative_path and default_storage.exists(relative_path):
            default_storage.delete(relative_path)
        thumb = thumb_path(relative_path)
        if thumb and default_storage.exists(thumb):
            default_storage.delete(thumb)
    except Exception:  # noqa: BLE001
        pass


def _make_thumbnail(original_path, uploaded):
    """Create a ~300px-wide JPEG thumbnail.

    Photos are resized with Pillow. Videos get their first frame extracted via
    ffmpeg (if available) into a poster image. The thumbnail is written next to
    the original in default_storage.
    """
    thumb = thumb_path(original_path)
    # Determine if it's a video by extension (uploaded.content_type may be absent).
    ext = os.path.splitext(original_path)[1].lower()
    video_exts = {'.mp4', '.mov', '.avi', '.m4v', '.webm', '.mkv'}

    if ext in video_exts:
        _make_video_thumbnail(original_path, thumb)
        return

    # Photo: resize with Pillow.
    from PIL import Image
    from io import BytesIO
    uploaded.seek(0)
    with Image.open(uploaded) as img:
        img = img.convert('RGB')
        # 300px wide, preserve aspect, cap height.
        max_w, max_h = 300, 300
        img.thumbnail((max_w, max_h), Image.LANCZOS)
        buf = BytesIO()
        img.save(buf, format='JPEG', quality=80)
        buf.seek(0)
        if default_storage.exists(thumb):
            default_storage.delete(thumb)
        default_storage.save(thumb, buf)


def _make_video_thumbnail(original_path, thumb):
    """Extract the first frame of a video as a thumbnail via ffmpeg."""
    import subprocess
    import tempfile
    abs_in = os.path.join(settings.MEDIA_ROOT, original_path)
    if not os.path.exists(abs_in):
        return
    abs_thumb = os.path.join(settings.MEDIA_ROOT, thumb)
    os.makedirs(os.path.dirname(abs_thumb), exist_ok=True)
    # Seek ~1s in (or 10% of duration) to skip a black lead-in, grab one frame.
    try:
        subprocess.run(
            ['ffmpeg', '-y', '-ss', '1', '-i', abs_in,
             '-frames:v', '1', '-vf', 'scale=300:-1', '-q:v', '4', abs_thumb],
            capture_output=True, timeout=20, check=False,
        )
    except (FileNotFoundError, subprocess.TimeoutExpired):
        # ffmpeg not installed or hung — skip thumbnail (poster-less <video> still works).
        pass


def _collect_entry_photos(request):
    """Map work_item_id → [uploaded files] from field names like ``ep_<id>``."""
    photos = {}
    for field, files in request.FILES.lists():
        if not field.startswith('ep_'):
            continue
        try:
            wid = int(field[3:])
        except ValueError:
            continue
        photos.setdefault(wid, []).extend(files)
    return photos


def _calc_hours(start, end, headcount):
    """Work-hours = headcount × duration, rounded to nearest 0.5h."""
    if not start or not end or headcount <= 0:
        return 0.0
    today = date.today()
    base = datetime.combine(today, start)
    end_dt = datetime.combine(today, end)
    if end_dt < base:                      # overnight shift wraps to next day
        end_dt = datetime.combine(today + timedelta(days=1), end)
    hours = (end_dt - base).total_seconds() / 3600
    return round(hours * headcount * 2) / 2


# ── view ────────────────────────────────────────────────────────────────────

@login_required(login_url='core:login')
def workorder_tree_form(request, report_id=None):
    report = None
    if report_id:
        report = get_object_or_404(WorkReport, pk=report_id)

    if request.method == 'POST':
        return _handle_save(request, report)
    return _handle_render(request, report)


def _handle_render(request, report):
    tree = serialize_workitem_tree()
    # Inventory catalog tree + the report's current material cart, so the form
    # can render the material-consumption section (lazy import: views.py imports
    # from this module, so the reverse import must stay inside the function).
    from core.views import serialize_inventory_tree
    ctx = {
        'tree_json': json.dumps(tree, ensure_ascii=False),
        'projects_json': json.dumps(serialize_projects(), ensure_ascii=False),
        'locations': Patch.objects.filter(active=True).order_by('order'),
        'zones': Zone.objects.order_by('code'),
        'grouped_zones': _build_grouped_zones(Zone.objects.select_related('land', 'patch').order_by('code')),
        'today': date.today().isoformat(),
        'report': report,
        'existing_json': json.dumps(serialize_existing_entries(report), default=str)
                         if report else '{}',
        'report_photos_json': json.dumps(report.photos or []) if report else '[]',
        'header_json': json.dumps(_report_header_dict(report), ensure_ascii=False)
                        if report else '{}',
        'inventory_tree_json': json.dumps(serialize_inventory_tree(), ensure_ascii=False),
        'existing_materials_json': json.dumps(_serialize_workorder_materials(report))
                                   if report else '[]',
        'material_dest_json': json.dumps(_serialize_workorder_material_dest(report))
                              if report else 'null',
    }
    return render(request, 'core/workorder_tree_form.html', ctx)


def _report_header_dict(report):
    """Header field values to pre-fill the form in edit mode."""
    zones = list(report.zones.values_list('code', flat=True))
    header = {
        'h_date': report.date.isoformat(),
        'h_weather': report.weather,
        'h_shift': report.shift,
        'h_location': report.location_id,
        'h_zone_names': report.zone_names,
        'h_zones': zones,
        'h_remark': report.remark,
        'h_is_pending_repair': report.is_pending_repair,
        'h_is_difficult': report.is_difficult,
        'h_is_difficult_resolved': report.is_difficult_resolved,
        'h_team_size': report.team_size,
        'h_third_party_count': report.third_party_count,
        'h_work_start_time': report.work_start_time.strftime('%H:%M') if report.work_start_time else '',
        'h_work_end_time': report.work_end_time.strftime('%H:%M') if report.work_end_time else '',
    }
    # PM auto-dispatched work orders carry the JobPlan name so the edit modal
    # can pre-fill the 工作类别 (常规维护 › 维保定期检查) + a work-content entry.
    gwo = getattr(report, 'pm_generation', None)
    if gwo and gwo.plan_id:
        header['h_pm_job_plan'] = gwo.plan.job_plan.name if gwo.plan.job_plan_id else ''
        header['h_pm_work_item'] = _pm_leaf_id()
        header['h_pm_gwo_id'] = gwo.id
    return header


def _record_edit(report, user, note=''):
    """编辑保存成功后写入一条编辑记录。仅在 edit（report 已存在）时调用。

    编辑人通过 resolve_or_create_worker 解析，任意账号类型都能正确归因。
    调用方应将其置于保存事务内，使编辑记录与工单同生共死。
    """
    from core.models import WorkReportEditLog
    from core.role_utils import resolve_or_create_worker
    editor, _created = resolve_or_create_worker(user)
    WorkReportEditLog.objects.create(
        work_report=report, editor=editor, note=(note or '')[:200],
    )


def _handle_save(request, report):
    # Resolve the real submitter. Field workers have a linked Worker row;
    # managers/admins do not, so provision one from their profile (idempotent).
    from core.role_utils import resolve_or_create_worker
    worker, _created = resolve_or_create_worker(request.user)
    if not worker:
        messages.error(request, '当前用户未关联处理人账号')
        return redirect('core:work_reports')

    # Required header fields — browser enforces too; this is the safety net.
    # 位置/CCU is NOT required here: it auto-derives from the selected zones'
    # 所属位置 (patch) below, so reports created under a zone group (e.g. 酒店3)
    # whose Patch isn't directly selectable still save.
    missing = [label for label, key in (
        ('日期', 'date'),)
        if not request.POST.get(key)]
    if missing:
        messages.error(request, '请填写必填项：' + '、'.join(missing))
        return _handle_render(request, report)

    zone_codes = request.POST.getlist('zones')
    zones = list(Zone.objects.filter(code__in=zone_codes))

    start = _parse_time(request.POST.get('work_start_time'))
    end = _parse_time(request.POST.get('work_end_time'))
    team_size = int(request.POST.get('team_size') or 1)
    third_party_count = int(request.POST.get('third_party_count') or 0)

    with transaction.atomic():
        is_new = report is None
        if is_new:
            report = WorkReport(worker=worker)
        report.date = request.POST.get('date') or date.today()
        report.weather = request.POST.get('weather', '')
        report.shift = request.POST.get('shift', '')
        # 位置/CCU: explicit selection wins; otherwise auto-derive from the first
        # selected zone's 所属位置 (patch). Leaves None when neither is available,
        # which is now valid (the column is nullable).
        loc_id = request.POST.get('location') or None
        if not loc_id and zones:
            loc_id = next((z.patch_id for z in zones if z.patch_id), None)
        report.location_id = loc_id
        report.zone_names = request.POST.get('zone_names', '')
        report.remark = request.POST.get('remark', '')
        report.is_pending_repair = bool(request.POST.get('is_pending_repair'))
        report.is_difficult = bool(request.POST.get('is_difficult'))
        report.is_difficult_resolved = bool(request.POST.get('is_difficult_resolved'))
        report.team_size = team_size
        report.third_party_count = third_party_count
        report.work_start_time = start
        report.work_end_time = end
        report.team_hours = _calc_hours(start, end, team_size)
        report.third_party_hours = _calc_hours(start, end, third_party_count)
        # Save first so report.id exists for photo paths.
        report.save()
        report.zones.set(zones)
        # 编辑历史：仅 edit（非新建）记录一次，新建工单不记。
        if not is_new:
            _record_edit(report, request.user)

        # Report-level photos (1.1.12). Merge instead of replace: drop any photos
        # the user marked for removal (deleting their files + thumbnails), then
        # append newly uploaded ones. Untouched photos are preserved — previously
        # any upload silently wiped the whole list on edit.
        existing = list(report.photos or [])
        # Removals arrive as a comma-joined list of relative media paths.
        remove_set = {p for p in (request.POST.get('report_photos_remove') or '').split(',') if p}
        if remove_set:
            kept = []
            for p in existing:
                if p in remove_set:
                    _delete_media(p)            # best-effort; ignore missing
                else:
                    kept.append(p)
            existing = kept
        # New uploads append to the surviving list.
        # Validate uploads BEFORE the save loop: _save_photo raises ValueError on
        # a disallowed extension/size/content, which would otherwise abort the
        # whole atomic block as a 500. Pre-checking gives a clean user-facing error.
        from core.upload_security import validate_upload
        for f in request.FILES.getlist('report_photos'):
            ok, err = validate_upload(f)
            if not ok:
                messages.error(request, f'{getattr(f, "name", "")}: {err}')
                return _handle_render(request, report)
        new_photos = [_save_photo(report, f)
                      for f in request.FILES.getlist('report_photos')]
        if new_photos:
            existing.extend(new_photos)
        if remove_set or new_photos:
            report.photos = existing
            report.save(update_fields=['photos'])

        # Work-content entries.
        entries = json.loads(request.POST.get('entries', '[]') or '[]')
        entry_photos = _collect_entry_photos(request)
        # Validate entry-level uploads too (same pre-check pattern as report photos).
        from core.upload_security import validate_upload
        for _wid, files in entry_photos.items():
            for f in files:
                ok, err = validate_upload(f)
                if not ok:
                    messages.error(request, f'{getattr(f, "name", "")}: {err}')
                    return _handle_render(request, report)
        _save_entries(report, entries, entry_photos)

        # Material consumption (材料消耗): rebuild the report's outbound
        # transaction from the cart. _save_workorder_materials rolls back the
        # prior linked transaction first, so edits keep stock correct. The
        # destination is auto-derived from the work category (routine→日常维护,
        # project sections→项目) unless the user picked one for 'other' categories.
        materials = json.loads(request.POST.get('materials', '[]') or '[]')
        m_dest, m_proj, m_cp = _resolve_material_dest(request, entries)
        _save_workorder_materials(report, materials, entry_subtype=m_dest,
                                  related_project_id=m_proj, counterparty=m_cp)

        # PM completion: mirror the mobile path so a PM task submitted via the
        # desktop tree form also marks its GeneratedWorkOrder completed.
        mark_pm_completed(report)

    messages.success(request, f'现场作业记录已保存 (ID: {report.id})')
    if request.POST.get('save_and_new'):
        return redirect('core:workorder_tree_form')
    return redirect('core:work_report_detail', report_id=report.id)


def _save_entries(report, entries, entry_photos):
    """Replace all entries; create one row per node with a real value.

    A row is kept when it carries a value (count/status/text/photos). A row that is
    empty *but* references a group/category node (value_type='group') is also kept —
    it represents "this category was worked under" and prevents the report from being
    mislabeled as 旧版记录 when the user picked a category without drilling into
    specific leaf content.
    """
    report.entries.all().delete()
    # Pre-fetch the WorkItems referenced, to tell group/category nodes from leaves.
    wids = [e.get('work_item') for e in entries if e.get('work_item')]
    group_wids = set()
    if wids:
        group_wids = set(
            WorkItem.objects.filter(id__in=wids, value_type='group').values_list('id', flat=True)
        )
    for e in entries:
        wid = e.get('work_item')
        if not wid:
            continue
        count = int(e.get('count') or 0)
        status = (e.get('status') or '').strip()
        text_value = (e.get('text_value') or '').strip()
        project_id = e.get('project') or None
        photos = entry_photos.get(int(wid), [])
        has_value = bool(count or status or text_value or photos)
        is_category_marker = wid in group_wids
        if not has_value and not is_category_marker:
            continue
        saved_paths = [_save_photo(report, f) for f in photos]
        WorkReportEntry.objects.update_or_create(
            work_report=report, work_item_id=wid, project_id=project_id,
            defaults={'count': count, 'status': status,
                      'text_value': text_value, 'photos': saved_paths},
        )

    # Safety net: 待修 flag ⇒ 疑难 / 疑难未解决 (mirrors the mobile UI behavior).
    if report.is_pending_repair:
        report.is_difficult = True
        report.is_difficult_resolved = False
        report.save(update_fields=['is_difficult', 'is_difficult_resolved'])


def _resolve_material_dest(request, entries):
    """Resolve the outbound destination for a workorder's material consumption.

    Priority:
    1. Explicit POST fields (mat_dest / mat_project_id / mat_counterparty) — set
       by the client when the user picks a destination chip for 'other' categories.
    2. Auto-derived from the filled entries' WorkItem sections:
       - project section (irrigation/drainage/other_project) → '项目' + the entry's project
       - routine_maint → '日常维护'
       - anything else → '日常维护' (safe default; the client overrides for 'other')
    Returns (entry_subtype, project_id, counterparty).
    """
    from core.models import WorkItem
    PROJECT_SECTIONS = {'irrigation_project', 'drainage_project', 'other_project'}
    entry_subtype = (request.POST.get('mat_dest') or '').strip()
    project_id = request.POST.get('mat_project_id') or None
    counterparty = (request.POST.get('mat_counterparty') or '').strip()

    if not entry_subtype:
        # Auto-derive from the submitted entries' sections.
        wid_ids = [e.get('work_item') for e in entries if e.get('work_item')]
        sections = set()
        if wid_ids:
            sections = set(WorkItem.objects.filter(id__in=wid_ids).values_list('section', flat=True))
        proj_sections = sections & PROJECT_SECTIONS
        if proj_sections:
            entry_subtype = '项目'
            # Use the first project-scoped entry's project if not explicitly set.
            if not project_id:
                for e in entries:
                    node = WorkItem.objects.filter(id=e.get('work_item')).first()
                    if node and node.section in PROJECT_SECTIONS and e.get('project'):
                        project_id = e.get('project')
                        break
        else:
            entry_subtype = '日常维护'
    return entry_subtype, project_id, counterparty


def _save_workorder_materials(report, lines, entry_subtype='日常维护',
                              related_project_id=None, counterparty=''):
    """工单材料消耗：先回滚该工单旧的关联出库单，再按新的 lines 建出库单并扣减库存。

    Must run inside the caller's ``transaction.atomic()`` block. Idempotent —
    every save rebuilds the report's material-consumption transaction so the
    stock reflects exactly the current cart (editing then re-saving corrects
    both the ledger and the stock).

    ``lines = [{category: <id>, quantity: <num>, unit: <str>}, ...]``

    The outbound destination mirrors the standalone inventory form:
    ``entry_subtype`` ∈ 日常维护/项目/借用/其他, with ``related_project_id`` for
    项目 and ``counterparty`` for 借用. Callers derive the default from the work
    category (routine→日常维护, project sections→项目) and let the user override
    for other categories.
    """
    from django.db.models import F
    from core.models import (
        InventoryTransaction, InventoryTransactionLine, InventoryCategory, Project,
    )

    # 1) Roll back any prior outbound transaction linked to this report: refund
    #    each line's quantity back to current_stock, then drop txn + lines.
    old_txns = InventoryTransaction.objects.filter(work_report=report)
    for txn in old_txns:
        for ln in txn.lines.all():
            InventoryCategory.objects.filter(pk=ln.category_id).update(
                current_stock=F('current_stock') + ln.quantity,
            )
    old_txns.delete()

    if not lines:
        return

    # 2) Build one outbound transaction for the whole cart.
    first_zone = report.zones.first()
    project = None
    if related_project_id:
        project = Project.objects.filter(pk=related_project_id).first()
    txn = InventoryTransaction.objects.create(
        date=report.date if isinstance(report.date, date) else (report.date or date.today()),
        worker=report.worker,
        operation='出库',
        entry_subtype=entry_subtype or '日常维护',
        work_report=report,
        related_project=project,
        counterparty=(counterparty or '').strip(),
        zone=first_zone,
        remark=f'工单 #{report.id} 材料消耗',
    )

    # 3) One line per material; subtract stock atomically (F avoids races).
    created = 0
    for ln in lines:
        cat_id = ln.get('category')
        qty = ln.get('quantity') or 0
        try:
            qty = abs(float(qty))
        except (ValueError, TypeError):
            continue
        cat = InventoryCategory.objects.filter(pk=cat_id).first()
        if not cat or qty <= 0:
            continue
        InventoryTransactionLine.objects.create(
            transaction=txn, category=cat,
            quantity=qty, unit=(ln.get('unit') or '').strip(),
        )
        InventoryCategory.objects.filter(pk=cat_id).update(
            current_stock=F('current_stock') - qty,
        )
        created += 1

    # If nothing survived validation, drop the empty txn (no stock moved).
    if created == 0:
        txn.delete()


def _serialize_workorder_materials(report):
    """Return the report's current material cart as a list for edit prefill.

    Shape: ``[{category, quantity, unit, name}, ...]`` — ``name`` is the leaf's
    display name so the cart can render without re-fetching the inventory tree.
    """
    if not report:
        return []
    out = []
    for txn in report.material_consumptions.all():
        for ln in txn.lines.all():
            out.append({
                'category': ln.category_id,
                'quantity': ln.quantity,
                'unit': ln.unit,
                'name': ln.category.name_zh,
            })
    return out


def _serialize_workorder_material_dest(report):
    """Return the destination of the report's last material-consumption txn, for
    edit prefill of the destination selector. ``{entry_subtype, project_id, counterparty}``
    or ``None`` when no material txn exists yet."""
    if not report:
        return None
    txn = report.material_consumptions.order_by('-id').first()
    if not txn:
        return None
    return {
        'entry_subtype': txn.entry_subtype,
        'project_id': txn.related_project_id,
        'counterparty': txn.counterparty,
    }


def _resolve_pending_repairs(pm_report, report_ids):
    """Mark past 待修 reports as resolved by this 计划性维修 report.

    For each id: clear its 待修 flag (so it drops out of the planned-maintenance
    backlog, which filters on is_pending_repair=True) and append the
    计划性维修 work-order number to its remark.
    """
    note = f'已由计划性维修工单 #{pm_report.id} 处理'
    for rid in report_ids:
        past = WorkReport.objects.filter(pk=rid).first()
        if not past:
            continue
        past.is_pending_repair = False
        past.resolved_by_pm = pm_report  # FK link for the manager inbox
        remark = past.remark or ''
        if note not in remark:
            past.remark = (remark + '\n' + note).strip() if remark else note
            past.save(update_fields=['is_pending_repair', 'resolved_by_pm', 'remark'])
        else:
            past.save(update_fields=['is_pending_repair', 'resolved_by_pm'])


def _parse_time(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%H:%M').time()
    except ValueError:
        return None


def _build_grouped_zones(zones):
    """Group zones for the <select optgroup> layout by 所属Land (the
    land-name-zone hierarchy). Falls back to the Patch name — then '其它' — when a
    zone has no Land, matching build_zone_hierarchy's grouping on the list page.
    """
    groups = {}
    order = []
    for z in zones:
        if z.land_id:
            key = ('land', z.land_id)
            label = z.land.name
        elif z.patch_id:
            key = ('patch', z.patch_id)
            label = z.patch.name
        else:
            key = ('none', 0)
            label = '其它'
        if key not in groups:
            groups[key] = {'name': label, 'zones': []}
            order.append(key)
        groups[key]['zones'].append({'name': z.name, 'code': z.code})
    return [groups[k] for k in order]
