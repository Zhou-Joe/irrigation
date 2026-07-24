import json
from datetime import timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from django.db.models import Count, Q, Avg, Sum, F
from django.db.models.functions import Coalesce
from django.views.decorators.http import require_POST
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from core.models import Zone, Patch, WorkItem, WorkReportEntry

def auto_close_boundary_points(boundary_data):
    """
    Auto-close incomplete polygons in boundary data.

    If a user defines points but doesn't click "完成当前区域" before saving,
    this function will automatically close the polygon by adding the first point
    at the end if it has at least 3 points but isn't explicitly closed.

    Args:
        boundary_data: List of polygons, each polygon is a list of {lat, lng} points.
                       Format: [[{lat, lng}, ...], [{lat, lng}, ...]]

    Returns:
        List of properly closed polygons with at least 3 points each.
    """
    if not boundary_data or not isinstance(boundary_data, list):
        return []

    result = []
    for polygon in boundary_data:
        if not polygon or not isinstance(polygon, list):
            continue

        # Convert to list of dicts if points are in [lat, lng] array format
        points = []
        for p in polygon:
            if isinstance(p, dict) and 'lat' in p and 'lng' in p:
                points.append({'lat': float(p['lat']), 'lng': float(p['lng'])})
            elif isinstance(p, (list, tuple)) and len(p) >= 2:
                points.append({'lat': float(p[0]), 'lng': float(p[1])})

        # Skip if less than 3 points - cannot form a polygon
        if len(points) < 3:
            continue

        # Check if polygon is already closed (first point == last point)
        first_point = points[0]
        last_point = points[-1]

        is_closed = (
            abs(first_point['lat'] - last_point['lat']) < 0.000001 and
            abs(first_point['lng'] - last_point['lng']) < 0.000001
        )

        if not is_closed:
            # Auto-close by appending the first point
            points.append({'lat': first_point['lat'], 'lng': first_point['lng']})

        result.append(points)

    return result


def _wants_json(request):
    """Return JSON for async form saves while preserving normal HTML fallback."""
    return (
        request.headers.get('x-requested-with') == 'XMLHttpRequest'
        or 'application/json' in request.headers.get('accept', '')
    )


def _parse_float(val):
    if not val:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _get_zone_dropdown_options():
    """Get distinct values from DB for zone dropdown fields."""
    from .models import Zone
    opts = {}
    for field in ['plant_type', 'sprinkler_type', 'soil_moisture', 'terrain_feature']:
        vals = Zone.objects.exclude(**{field: ''}).exclude(**{field: None}) \
                   .values_list(field, flat=True).distinct().order_by(field)
        opts[field] = list(vals)
    # Foreman candidates: distinct names already used + manager/dept user names
    from .models import ManagerProfile, DepartmentUserProfile
    from django.contrib.auth.models import User
    foreman_names = set()
    for field in ['irrigation_foreman', 'greenery_foreman', 'pest_control_foreman']:
        foreman_names.update(
            Zone.objects.exclude(**{field: ''}).exclude(**{field: None})
            .values_list(field, flat=True)
        )
    foreman_names.update(
        ManagerProfile.objects.filter(active=True).values_list('full_name', flat=True)
    )
    foreman_names.update(
        DepartmentUserProfile.objects.filter(active=True).values_list('full_name', flat=True)
    )
    # Also add user first_names (non-superuser, non-staff)
    for u in User.objects.filter(is_superuser=False):
        name = u.first_name or u.username
        if name:
            foreman_names.add(name)
    opts['foreman_names'] = sorted(foreman_names)
    return opts


def _zone_save_response(request, zone, message, created=False):
    if _wants_json(request):
        resp = {
            'success': True,
            'message': message,
            'created': created,
            'zone_id': zone.id,
            'zone_name': zone.name,
            'edit_url': reverse('core:zone_edit', args=[zone.id]),
        }
        # Return parsed notes for editor reload
        for field in ('equipment_maintenance_notes', 'irrigation_management_notes'):
            raw = getattr(zone, field, '') or ''
            try:
                parsed = json.loads(raw) if raw else []
                resp[field] = parsed
                resp[field + '_count'] = len(parsed)
            except (json.JSONDecodeError, TypeError):
                resp[field] = []
                resp[field + '_count'] = 0
        resp['boundary_count'] = len(zone.boundary_points) if zone.boundary_points else 0
        resp['area_display'] = zone.area_display
        resp['plant_count'] = zone.plants.count()
        resp['equipment_count'] = zone.equipments.count()
        return JsonResponse(resp)

    messages.success(request, message)
    return redirect('core:zone_edit', zone_id=zone.id)


def _build_grouped_zones(zones_qs=None, group_by='patch'):
    """Build grouped structure for template rendering.

    group_by='patch': Groups zones by their Patch FK.
    group_by='priority': Groups zones by priority level.
    Returns a list of dicts: {id, name, code, type, zone_count, zones: [{...}]}
    """
    from core.models import Patch

    if zones_qs is None:
        zones_qs = Zone.objects.all().order_by('code')

    zones_data = []
    for z in zones_qs:
        zones_data.append({
            'id': z.id,
            'name': z.name,
            'code': z.code,
            'description': z.description or '',
            'patch_id': z.patch_id,
            'patch_name': z.patch.name if z.patch else None,
            'patch_code': z.patch.code if z.patch else None,
            'boundary_points': z.active_boundary_points,
            'boundary_color': z.boundary_color,
            'priority': z.priority,
            'priority_display': z.get_priority_display(),
            'current_status': z.current_status,
            'sprinkler_type': z.sprinkler_type,
            'irrigation_intensity': z.irrigation_intensity,
            'solenoid_valve_size': z.solenoid_valve_size,
            'landscape_coefficient': z.landscape_coefficient,
            'plant_type': z.plant_type,
            'irrigation_foreman': z.irrigation_foreman,
            'greenery_zone': z.greenery_zone,
            'greenery_foreman': z.greenery_foreman,
            'pest_control_zone': z.pest_control_zone,
            'pest_control_foreman': z.pest_control_foreman,
            'terrain_feature': z.terrain_feature,
            'plant_feature': z.plant_feature,
            'soil_moisture': z.soil_moisture,
            'equipment_maintenance_notes': z.equipment_maintenance_notes,
            'irrigation_management_notes': z.irrigation_management_notes,
            'has_remarks': bool(z.remarks and json.loads(z.remarks)) if z.remarks else False,
            'has_confirmed_remarks': bool(z.confirmed_remarks and json.loads(z.confirmed_remarks)) if z.confirmed_remarks else False,
            'plant_names': list(z.plants.values_list('name', flat=True)),
        })

    if group_by == 'priority':
        priority_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3, 'abolished': 4}
        groups = {}
        for z in zones_data:
            p = z['priority']
            groups.setdefault(p, []).append(z)
        grouped = []
        for pkey in sorted(groups.keys(), key=lambda k: priority_order.get(k, 99)):
            grouped.append({
                'type': 'priority',
                'id': pkey,
                'name': groups[pkey][0]['priority_display'],
                'code': pkey,
                'type_display': '优先级',
                'zones': groups[pkey],
                'zone_count': len(groups[pkey]),
            })
        return grouped

    # Default: group by patch_id
    groups = {}
    orphans = []
    for z in zones_data:
        if z['patch_id'] is not None:
            groups.setdefault(z['patch_id'], []).append(z)
        else:
            orphans.append(z)

    grouped = []
    for patch_id, zones in sorted(groups.items()):
        grouped.append({
            'id': patch_id,
            'name': zones[0]['patch_name'],
            'code': zones[0]['patch_code'],
            'zones': zones,
            'zone_count': len(zones),
        })

    if orphans:
        grouped.append({
            'name': '未分配片区',
            'zones': orphans,
            'zone_count': len(orphans),
        })

    return grouped


def _get_reference_map_data(exclude_zone_id=None, exclude_pipeline_id=None):
    """Build JSON data for rendering existing zones and pipelines as reference layers on edit maps."""
    from core.models import Pipeline

    ref_zones = []
    for z in Zone.objects.all():
        if exclude_zone_id and z.id == exclude_zone_id:
            continue
        if z.active_boundary_points:
            ref_zones.append({
                'id': z.id, 'name': z.name, 'code': z.code,
                'boundary_points': z.active_boundary_points,
                'boundary_color': z.boundary_color or '#52B788',
                'smooth_override': z.smooth_override,
                'ring_display_modes': z.ring_display_modes or {},
            })

    ref_pipelines = []
    for p in Pipeline.objects.all():
        if exclude_pipeline_id and p.id == exclude_pipeline_id:
            continue
        if p.line_points and len(p.line_points) >= 2:
            ref_pipelines.append({
                'id': p.id, 'name': p.name, 'code': p.code,
                'pipeline_type': p.pipeline_type,
                'line_points': p.line_points,
                'line_color': p.line_color,
                'line_weight': p.line_weight,
            })

    return json.dumps(ref_zones), json.dumps(ref_pipelines)


def _auto_pipeline_name_code(zone_ids, pipeline_type):
    """Generate unique pipeline name and code from zone IDs and type."""
    from .models import Pipeline

    type_label = '灌溉水管' if pipeline_type == 'irrigation' else '冲洗水管'
    type_prefix = 'IRR' if pipeline_type == 'irrigation' else 'FLU'

    zones = Zone.objects.filter(id__in=zone_ids).order_by('code')
    zone_names = list(zones.values_list('name', flat=True))
    zone_codes = list(zones.values_list('code', flat=True))

    if not zone_names:
        return '', ''

    base_name = '、'.join(zone_names) + ' - ' + type_label
    base_code = type_prefix + '-' + '-'.join(zone_codes)

    # Ensure uniqueness
    name, code = base_name, base_code
    suffix = 2
    while Pipeline.objects.filter(code=code).exists():
        name = f"{base_name} ({suffix})"
        code = f"{base_code}-{suffix}"
        suffix += 1

    return name, code


from django.views.decorators.csrf import ensure_csrf_cookie


@ensure_csrf_cookie
def user_login(request):
    """Login page for frontend with role-based redirect."""
    from .models import DepartmentUserProfile

    if request.user.is_authenticated:
        return redirect('core:dashboard')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)

            # Determine redirect based on role
            redirect_url = 'core:dashboard'

            # Check if dept user - redirect to requests page (water requests focus)
            try:
                DepartmentUserProfile.objects.get(user=user, active=True)
                redirect_url = 'core:water_requests'
            except DepartmentUserProfile.DoesNotExist:
                pass

            next_url = request.GET.get('next', redirect_url)
            # Only honor relative (same-origin) redirects — a bare host/path like
            # "/dashboard/" passes; an absolute URL to another site is rejected,
            # preventing open-redirect phishing after login.
            if not url_has_allowed_host_and_scheme(
                    url=next_url,
                    allowed_hosts={request.get_host()},
                    require_https=request.is_secure()):
                next_url = redirect_url
            return redirect(next_url)
        else:
            messages.error(request, '用户名或密码错误')

    return render(request, 'core/login.html')


def user_logout(request):
    """Logout view."""
    logout(request)
    return redirect('core:login')




@login_required(login_url='core:login')
def profile_page(request):
    """User profile page - view and edit personal information."""
    from core.models import Worker, ManagerProfile, DepartmentUserProfile

    user = request.user
    profile_data = {}
    profile_type = None

    # Priority: ManagerProfile > DepartmentUserProfile > Worker > System User
    # This ensures admin users show as managers even if they also have Worker records
    try:
        manager = ManagerProfile.objects.get(user=user, active=True)
        profile_type = 'manager'
        profile_data = {
            'type': '管理员',
            'employee_id': manager.employee_id,
            'full_name': manager.full_name,
            'phone': manager.phone,
            'api_token': str(manager.api_token),
            'is_super_admin': manager.is_super_admin,
            'can_approve_registrations': manager.can_approve_registrations,
            'can_approve_work_orders': manager.can_approve_work_orders,
        }
    except ManagerProfile.DoesNotExist:
        pass

    if not profile_type:
        try:
            dept_user = DepartmentUserProfile.objects.get(user=user, active=True)
            profile_type = 'dept_user'
            profile_data = {
                'type': '部门用户',
                'employee_id': dept_user.employee_id,
                'full_name': dept_user.full_name,
                'phone': dept_user.phone,
                'api_token': str(dept_user.api_token),
                'department': dept_user.get_department_display_name(),
            }
        except DepartmentUserProfile.DoesNotExist:
            pass

    if not profile_type:
        try:
            worker = Worker.objects.get(user=user, active=True)
            profile_type = 'worker'
            profile_data = {
                'type': '灌溉一线',
                'employee_id': worker.employee_id,
                'full_name': worker.full_name,
                'phone': worker.phone,
                'department': worker.get_department_display_name(),
                'api_token': str(worker.api_token),
            }
        except Worker.DoesNotExist:
            pass

    if not profile_type:
        profile_data = {
            'type': '系统用户',
            'username': user.username,
            'email': user.email,
        }

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'regenerate_token' and profile_type in ['worker', 'manager', 'dept_user']:
            try:
                if profile_type == 'worker':
                    profile = Worker.objects.get(user=user, active=True)
                elif profile_type == 'manager':
                    profile = ManagerProfile.objects.get(user=user, active=True)
                else:
                    profile = DepartmentUserProfile.objects.get(user=user, active=True)
                profile.regenerate_token()
                messages.success(request, 'API令牌已重新生成')
                return redirect('core:profile')
            except Exception:
                messages.error(request, '未找到用户档案')
        elif action == 'update_profile':
            phone = request.POST.get('phone', '').strip()
            full_name = request.POST.get('full_name', '').strip()
            employee_id = request.POST.get('employee_id', '').strip()
            email = request.POST.get('email', '').strip()

            # Update profile if exists
            if profile_type in ['worker', 'manager', 'dept_user']:
                try:
                    if profile_type == 'worker':
                        profile = Worker.objects.get(user=user, active=True)
                    elif profile_type == 'manager':
                        profile = ManagerProfile.objects.get(user=user, active=True)
                    else:
                        profile = DepartmentUserProfile.objects.get(user=user, active=True)

                    if full_name:
                        profile.full_name = full_name
                    if phone:
                        profile.phone = phone
                    if employee_id:
                        # Check uniqueness for employee_id
                        model_class = Worker if profile_type == 'worker' else (ManagerProfile if profile_type == 'manager' else DepartmentUserProfile)
                        if model_class.objects.filter(employee_id=employee_id).exclude(pk=profile.pk).exists():
                            messages.error(request, f'工号 {employee_id} 已被使用')
                        else:
                            profile.employee_id = employee_id
                    profile.save()
                except Exception as e:
                    messages.error(request, f'Profile更新失败: {str(e)}')

            # Update user email
            if email:
                user.email = email
                user.save()

            messages.success(request, '个人信息已更新')
            return redirect('core:profile')
        elif action == 'change_password':
            from django.contrib.auth import update_session_auth_hash
            from django.contrib.auth import password_validation
            from django.core.exceptions import ValidationError

            old_password = request.POST.get('old_password', '')
            new_password = request.POST.get('new_password', '')
            new_password2 = request.POST.get('new_password2', '')

            if not user.check_password(old_password):
                messages.error(request, '当前密码不正确')
            elif new_password != new_password2:
                messages.error(request, '两次输入的新密码不一致')
            elif not new_password:
                messages.error(request, '新密码不能为空')
            else:
                # Validate against the configured AUTH_PASSWORD_VALIDATORS (min
                # length, common/numeric/attribute-similarity rules).
                try:
                    password_validation.validate_password(new_password, user=user)
                    user.set_password(new_password)
                    user.save()
                    # Keep the user logged in after the password hash changes.
                    update_session_auth_hash(request, user)
                    messages.success(request, '密码已修改')
                    return redirect('core:profile')
                except ValidationError as e:
                    messages.error(request, '；'.join(e.messages))

    return render(request, 'core/profile.html', {
        'profile_data': profile_data,
        'profile_type': profile_type,
        'user': user,
    })


def get_zone_center(boundary_points):
    """Calculate the center point of a zone from its boundary points. Supports multi-polygon format."""
    if not boundary_points or len(boundary_points) == 0:
        return None

    lats = []
    lngs = []

    def _extract_coords(points):
        for point in points:
            if isinstance(point, list) and len(point) >= 2:
                lats.append(point[0])
                lngs.append(point[1])
            elif isinstance(point, dict) and 'lat' in point and 'lng' in point:
                lats.append(point['lat'])
                lngs.append(point['lng'])

    # Detect multi-polygon format: [[{lat,lng},...], [{lat,lng},...]]
    first = boundary_points[0]
    is_multi = isinstance(first, list) and len(first) > 0 and (
        isinstance(first[0], list) or (isinstance(first[0], dict) and 'lat' in first[0])
    )

    # Detect nested multi-group: [[[{lat,lng},...], ...], ...]
    is_nested = False
    if is_multi and isinstance(first, list) and len(first) > 0:
        inner = first[0]
        if isinstance(inner, list) and len(inner) > 0:
            innermost = inner[0]
            if isinstance(innermost, (list, dict)):
                is_nested = True

    if is_nested:
        # Flatten: [[ring1, ring2], [ring3]] → [ring1, ring2, ring3]
        for group in boundary_points:
            if isinstance(group, list):
                for ring in group:
                    _extract_coords(ring)
    elif is_multi:
        for ring in boundary_points:
            _extract_coords(ring)
    else:
        _extract_coords(boundary_points)

    if lats and lngs:
        return {
            'lat': round(sum(lats) / len(lats), 6),
            'lng': round(sum(lngs) / len(lngs), 6)
        }
    return None


def _safe_remark_items(raw):
    """Parse a zone remarks JSON string into a compact list for the dashboard.

    Returns a list of {date, content, author} (confirmed entries also keep their
    confirm_* fields). Empty/malformed → []. The dashboard hover tooltip renders
    these so users can read pending remarks without navigating to the zone page.
    """
    if not raw:
        return []
    try:
        items = json.loads(raw)
    except (ValueError, TypeError):
        return []
    if not isinstance(items, list):
        return []
    out = []
    for it in items:
        if not isinstance(it, dict):
            continue
        out.append({
            'date': it.get('date', ''),
            'content': it.get('content', '') or it.get('confirm_reply', ''),
            'author': it.get('author', '') or it.get('confirm_author', ''),
        })
    return out


def _cached(key, ttl, builder):
    """TTL cache wrapper with a safe fallback.

    Uses the framework cache when a real backend is configured; if the backend
    is DummyCache (or unset) the builder just runs every call — still correct,
    only slower. ``builder`` may return None legitimately; we cache a sentinel.
    """
    from django.core.cache import caches
    try:
        backend = caches['default']
        from django.core.cache.backends.dummy import DummyCache
        if isinstance(backend, DummyCache):
            return builder()
    except Exception:
        return builder()
    from django.core.cache import cache
    sentinel = object()
    val = cache.get(key, sentinel)
    if val is sentinel:
        val = builder()
        try:
            cache.set(key, val, ttl)
        except Exception:
            pass
    return val


def _build_zones_payload(today, week_ago):
    """Build the per-zone JSON payload used by the dashboard map + zone-detail cards.

    Extracted from `dashboard` so the same payload can be served:
      - inline in the dashboard HTML (legacy), and
      - as its own cacheable/gzip-friendly JSON endpoint (api/zones-payload/), so the
        ~5MB of zone data is no longer inlined in the page HTML and can be cached by
        the browser across navigations. All heavy group-by queries live here so the
        two paths can never drift apart.
    """
    from datetime import timedelta
    from collections import defaultdict
    from django.db.models import Count, Sum
    from django.db.models.functions import Coalesce
    from core.models import (
        Zone, Plant, ZoneEquipment, WorkReport, WaterRequest,
    )

    thirty_days_ago = today - timedelta(days=30)

    # Base zone queryset: select_related for patch/region, .only() the fields we
    # actually serialize to the client. Counts are computed via separate group-by
    # queries below to avoid the expensive Count(distinct=True) cross-product that
    # a single mega-annotated query produced on 2500+ zones.
    zones = (Zone.objects.select_related('patch', 'patch__region', 'land')
             .only('id', 'code', 'name', 'description', 'boundary_points',
                   'dxf_boundary_points', 'boundary_source',
                   'current_status', 'priority', 'remarks', 'confirmed_remarks',
                   'sprinkler_type', 'irrigation_intensity', 'solenoid_valve_size',
                   'landscape_coefficient', 'plant_type', 'irrigation_foreman',
                   'greenery_zone', 'greenery_foreman', 'pest_control_zone',
                   'pest_control_foreman', 'terrain_feature', 'plant_feature',
                   'soil_moisture', 'equipment_maintenance_notes',
                   'irrigation_management_notes', 'label_lat', 'label_lng',
                   'label_scale', 'label_angle', 'smooth_override',
                   'ring_display_modes', 'area_sqm',
                   'patch__id', 'patch__name', 'patch__code',
                   'patch__region__id', 'patch__region__name',
                   'land__id', 'land__name'))

    zone_ids = list(zones.values_list('id', flat=True))

    # Group-by counts (one query each, no distinct=True cross-product).
    def _count_map(model, **filters):
        out = {}
        qs = model.objects.filter(zone_id__in=zone_ids, **filters).values('zone_id')
        for row in qs.annotate(c=Count('id')):
            out[row['zone_id']] = row['c']
        return out

    plant_count_map = _count_map(Plant)
    equipment_count_map = _count_map(ZoneEquipment)
    # WaterRequest uses a multi-zone M2M; count per zone across all its zones.
    water_count_map = {}
    for row in (WaterRequest.objects.filter(zones__in=zone_ids)
                .values('zones').annotate(c=Count('id'))):
        zid = int(row['zones'])
        water_count_map[zid] = water_count_map.get(zid, 0) + int(row['c'])

    # ── Bulk: zones with an UNRESOLVED 待修 work report (is_pending_repair AND not yet
    # closed by a 计划性维修). Drives the orange "needs attention" boundary color on the
    # map. One group-by query instead of per-zone lookups.
    pending_repair_zone_ids = set(
        WorkReport.objects.filter(is_pending_repair=True, resolved_by_pm__isnull=True, zones__in=zone_ids)
        .values_list('zones', flat=True).distinct()
    )

    # ── Bulk: pending water requests for today ──
    # A request spans multiple zones (M2M), so attribute it to every zone it covers.
    pending_water_map = {}  # zone_id -> list of {id, type, type_display}
    for wr in WaterRequest.objects.filter(
        status='submitted',
        start_datetime__date__lte=today,
        end_datetime__date__gte=today
    ).prefetch_related('zones'):
        for z in wr.zones.all():
            if z.id in zone_ids:
                pending_water_map.setdefault(z.id, []).append({
                    'id': wr.id, 'type': 'water', 'type_display': '浇水协调',
                })

    # ── Bulk: recent water requests (top 3 per zone) — a request spans multiple zones ──
    recent_water_map = defaultdict(list)
    for w in WaterRequest.objects.prefetch_related('zones').order_by('-created_at'):
        item = {
            'id': w.id,
            'type': w.get_request_type_display(),
            'status': w.status,
            'status_display': w.get_status_display(),
            'start': w.start_datetime.strftime('%m-%d %H:%M'),
            'end': w.end_datetime.strftime('%m-%d %H:%M'),
        }
        for z in w.zones.all():
            if z.id in zone_ids:
                lst = recent_water_map[z.id]
                if len(lst) < 3:
                    lst.append(item)

    # ── Bulk: plant names per zone ──
    plant_names_map = defaultdict(list)
    for zone_id, name in Plant.objects.filter(
        zone_id__in=zone_ids
    ).values_list('zone_id', 'name'):
        plant_names_map[zone_id].append(name)

    # ── Bulk: recent workorders per zone (top N, last 90 days) ──
    # For each zone, attach up to N recent workorders with: id, date, sections
    # (work-type labels), first_entry_path (work-content prefix), detail_url.
    # Bounded window keeps the M2M join small; client-side the list shows under
    # the zone-popup card next to 设备维护记录.
    from core.workorder_tree_views import workitem_path_map as _wim_path_map
    from core.models import WorkItem, WorkReportEntry
    from django.urls import reverse as _reverse
    RECENT_WO_WINDOW_DAYS = 90
    RECENT_WO_PER_ZONE = 5
    _section_label_lookup = dict(WorkItem.SECTION_CHOICES)
    wo_window_start = today - timedelta(days=RECENT_WO_WINDOW_DAYS)
    # Step 1: zone→workreport pairs in window (M2M), keep most recent per zone.
    recent_wo_map = defaultdict(list)  # zone_id -> list of workreport ids
    pairs = (WorkReport.objects
             .filter(date__gte=wo_window_start, zones__in=zone_ids)
             .values_list('zones', 'id', 'date')
             .order_by('-date', '-id'))
    seen_per_zone = defaultdict(int)
    wr_ids_set = set()
    for zid, wrid, d in pairs:
        if seen_per_zone[zid] >= RECENT_WO_PER_ZONE:
            continue
        recent_wo_map[zid].append(wrid)
        wr_ids_set.add(wrid)
        seen_per_zone[zid] += 1
    # Step 2: prefetch the workreports + first entry's WorkItem for path lookup.
    workorder_info_map = {}  # wr_id -> dict
    if wr_ids_set:
        path_lookup = _wim_path_map()
        # Pull one entry per report (the first by id) — that's the "前级" prefix.
        first_entries = (WorkReportEntry.objects
                         .filter(work_report_id__in=wr_ids_set)
                         .select_related('work_item', 'project')
                         .order_by('work_report_id', 'id'))
        first_entry_per_report = {}
        sections_per_report = defaultdict(set)
        for fe in first_entries:
            rid = fe.work_report_id
            if rid not in first_entry_per_report:
                wi = fe.work_item
                first_entry_per_report[rid] = {
                    'path': path_lookup.get(wi.id, wi.name_zh) if wi else '',
                    'value': fe.count if hasattr(fe, 'count') else '',
                }
            sec = getattr(fe.work_item, 'section', '') if fe.work_item else ''
            if sec:
                sections_per_report[rid].add(_section_label_lookup.get(sec, sec))
        # Pull the workreport date (pairs above already gave us this but we want
        # a single source of truth).
        for wr in WorkReport.objects.filter(id__in=wr_ids_set).values('id', 'date'):
            wid = wr['id']
            workorder_info_map[wid] = {
                'id': wid,
                'date': wr['date'].isoformat() if wr['date'] else '',
                'sections': sorted(sections_per_report[wid]),
                'first_entry_path': first_entry_per_report.get(wid, {}).get('path', ''),
                'detail_url': _reverse('core:work_report_detail', args=[wid]),
            }
    # Build per-zone list (preserves the date-desc order from pairs query).
    recent_workorders_map = {}  # zone_id -> list of workorder dicts
    for zid, wr_ids in recent_wo_map.items():
        recent_workorders_map[zid] = [workorder_info_map[wid] for wid in wr_ids if wid in workorder_info_map]


    # ── Build zones_list (no per-zone DB queries) ──
    zones_list = []
    for zone in zones:
        center = get_zone_center(zone.active_boundary_points)
        # Parse remarks TextFields once each (was 4× per zone: bool(json.loads) + _safe_remark_items).
        zone_remarks = _safe_remark_items(zone.remarks)
        zone_confirmed = _safe_remark_items(zone.confirmed_remarks)
        # "Needs attention" = has unconfirmed remarks OR an unresolved 待修 report.
        # Drives the orange boundary color on the map (default is green; selected is red).
        needs_attention = bool(zone_remarks) or (zone.id in pending_repair_zone_ids)
        zones_list.append({
            'id': zone.id,
            'code': zone.code,
            'name': zone.name,
            'description': zone.description,
            'boundary_points': zone.active_boundary_points,
            'needs_attention': needs_attention,
            'status': zone.get_today_status(),
            'statusDisplay': zone.get_status_display(),
            'plant_count': plant_count_map.get(zone.id, 0),
            'plant_names': plant_names_map.get(zone.id, []),
            'equipment_count': equipment_count_map.get(zone.id, 0),
            'center': center,
            'pending_requests': pending_water_map.get(zone.id, []),
            # Patch info
            'patch_id': zone.patch.id if zone.patch else None,
            'patch_name': zone.patch.name if zone.patch else None,
            'patch_code': zone.patch.code if zone.patch else None,
            # Region info
            'region_id': zone.patch.region_id if zone.patch and zone.patch.region else None,
            'region_name': zone.patch.region.name if zone.patch and zone.patch.region else None,
            # Land info (top-level grouping for sidebar)
            'land_id': zone.land.id if zone.land else None,
            'land_name': zone.land.name if zone.land else None,
            # Priority
            'priority': zone.priority,
            'priority_display': zone.get_priority_display(),
            # Zone attributes from Excel
            'current_status': zone.current_status,
            'sprinkler_type': zone.sprinkler_type,
            'irrigation_intensity': zone.irrigation_intensity,
            'solenoid_valve_size': zone.solenoid_valve_size,
            'landscape_coefficient': zone.landscape_coefficient,
            'plant_type': zone.plant_type,
            'irrigation_foreman': zone.irrigation_foreman,
            'greenery_zone': zone.greenery_zone,
            'greenery_foreman': zone.greenery_foreman,
            'pest_control_zone': zone.pest_control_zone,
            'pest_control_foreman': zone.pest_control_foreman,
            'terrain_feature': zone.terrain_feature,
            'plant_feature': zone.plant_feature,
            'soil_moisture': zone.soil_moisture,
            'equipment_maintenance_notes': zone.equipment_maintenance_notes,
            'irrigation_management_notes': zone.irrigation_management_notes,
            'has_remarks': bool(zone_remarks),
            'has_confirmed_remarks': bool(zone_confirmed),
            'remarks': zone_remarks,
            'confirmed_remarks': zone_confirmed,
            # Label settings
            'label_lat': zone.label_lat,
            'label_lng': zone.label_lng,
            'label_scale': zone.label_scale,
            'label_angle': zone.label_angle,
            'smooth_override': zone.smooth_override,
            'ring_display_modes': zone.ring_display_modes or {},
            'area_sqm': zone.area_sqm,
            'area_display': zone.area_display,
            # Counts (from group-by maps, not annotations)
            'water_count': water_count_map.get(zone.id, 0),
            # Recent items (from bulk queries)
            'recent_water': recent_water_map.get(zone.id, []),
            'recent_workorders': recent_workorders_map.get(zone.id, []),
        })
    return zones_list


@login_required(login_url='core:login')
def zones_payload_api(request):
    """Serve the dashboard zone payload as a standalone, cacheable JSON document.

    Same data as the inline `zones_json` in the dashboard HTML, but separated so:
      - the HTML page renders instantly without blocking on ~5MB of inline JSON, and
      - the browser caches this response (1 min) so re-navigating to the dashboard
        reuses it instead of re-downloading/re-parsing. GZipMiddleware compresses it
        to ~0.3MB on the wire.
    """
    from datetime import date, timedelta
    today = date.today()
    week_ago = today - timedelta(days=7)
    payload = _build_zones_payload(today, week_ago)
    resp = JsonResponse(payload, safe=False, json_dumps_params={'ensure_ascii': False})
    # Zones' status/counts are day-scoped; cache briefly on the browser to make
    # back-to-back dashboard loads near-instant, but not so long that stale status lingers.
    resp['Cache-Control'] = 'private, max-age=60'
    resp['Vary'] = 'Accept-Encoding'
    return resp


@login_required(login_url='core:login')
def today_weather_api(request):
    """Return a short weather string for the CURRENT hour (for workorder auto-fill).

    The workorder form pre-fills 天气 with this single-hour snapshot taken at the
    moment of submission, e.g. "阴天, 22.5°C". Falls back to the latest record.
    """
    from datetime import date
    from django.utils import timezone
    from core.models import WeatherData
    now = timezone.localtime()
    today = now.date()
    wd = WeatherData.objects.filter(date=today).first()
    if not wd:
        wd = WeatherData.objects.order_by('-date').first()
    if not wd or not wd.hourly_data:
        return JsonResponse({'summary': '', 'has_data': False})
    # Current hour; if the record is from another day, use hour 0.
    cur_hour = now.hour if wd.date == today else 0
    # Closest available hour record (exact match, else the nearest).
    hours = wd.hourly_data
    exact = next((h for h in hours if h.get('hour') == cur_hour), None)
    if exact is None:
        exact = min(hours, key=lambda h: abs((h.get('hour') or 0) - cur_hour))
    desc = wd.get_weather_description(exact.get('code'))
    temp = exact.get('temp')
    summary = desc
    if temp is not None:
        summary = f"{desc}, {temp:g}°C" if desc else f"{temp:g}°C"
    return JsonResponse({'summary': summary, 'has_data': True})


def _pm_tasks_for_field_worker(user):
    """Today's pending PM tasks for a field worker.

    Moved to core.notifications.pm_tasks_for_field_worker so the context
    processor can call it without importing views. Kept here as a thin
    re-export for backward compatibility.
    """
    from .notifications import pm_tasks_for_field_worker
    return pm_tasks_for_field_worker(user)


@login_required(login_url='core:login')
def dashboard(request):
    """
    Main dashboard view with interactive map showing irrigation zones.
    """
    from datetime import date, timedelta
    from django.db.models.functions import TruncDate
    from core.models import (
        WaterRequest,
        ManagerProfile, DepartmentUserProfile, RegistrationRequest, Worker,
        Pipeline, Plant, MapStyleSettings, ZoneEquipment, WorkReport,
    )

    user = request.user
    today = date.today()
    week_ago = today - timedelta(days=7)

    # Determine user role
    is_admin = user.is_superuser or user.is_staff
    is_manager = False
    is_dept_user = False
    is_field_worker = False

    if not is_admin:
        try:
            ManagerProfile.objects.get(user=user, active=True)
            is_manager = True
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        try:
            DepartmentUserProfile.objects.get(user=user, active=True)
            is_dept_user = True
        except DepartmentUserProfile.DoesNotExist:
            pass

    if not is_admin and not is_dept_user:
        try:
            Worker.objects.get(user=user, active=True)
            is_field_worker = True
        except Worker.DoesNotExist:
            pass

    # Zone payload (queries + serialization) is shared with the cacheable JSON endpoint.
    zones_list = _build_zones_payload(today, week_ago)

    # Group zones by Land → common-name (zone.name) for sidebar display
    from .models import Land
    lands = Land.objects.filter(active=True).order_by('order', 'name')

    # Bucket zones by land_id
    land_buckets = {}
    orphan_zones = []
    for z in zones_list:
        lid = z.get('land_id')
        if lid is not None:
            land_buckets.setdefault(lid, []).append(z)
        else:
            orphan_zones.append(z)

    def _build_name_groups(zones):
        """Group zones by their common name (zone.name), preserving first-seen order."""
        order = []
        grouped = {}
        for z in zones:
            nm = z.get('name') or '(未命名)'
            if nm not in grouped:
                grouped[nm] = []
                order.append(nm)
            grouped[nm].append(z)
        result = []
        for nm in order:
            result.append({
                'name': nm,
                'zones': grouped[nm],
                'zone_count': len(grouped[nm]),
            })
        return result

    grouped_zones = []
    for land in lands:
        lz = land_buckets.pop(land.id, None)
        if lz:
            name_groups = _build_name_groups(lz)
            grouped_zones.append({
                'type': 'land',
                'id': land.id,
                'name': land.name,
                'name_groups': name_groups,
                'zone_count': len(lz),
            })

    # Zones whose Land is inactive / missing (land_id set but no active Land row)
    extra_orphan_zones = []
    for lid, lz in land_buckets.items():
        extra_orphan_zones.extend(lz)
    all_orphan = extra_orphan_zones + orphan_zones
    if all_orphan:
        grouped_zones.append({
            'type': 'orphan',
            'name': '未分配Land',
            'name_groups': _build_name_groups(all_orphan),
            'zone_count': len(all_orphan),
        })

    # Only admins see pending counts
    pending_counts = None
    if is_admin:
        pending_counts = {
            'registrations': RegistrationRequest.objects.filter(status='pending').count(),
            'water': WaterRequest.objects.filter(status='submitted').count(),
        }

    # Dashboard statistics
    status_distribution = {
        'unarranged': sum(1 for z in zones_list if z['status'] == 'unarranged'),
        'in_progress': sum(1 for z in zones_list if z['status'] == 'in_progress'),
        'completed': sum(1 for z in zones_list if z['status'] == 'completed'),
        'canceled': sum(1 for z in zones_list if z['status'] == 'canceled'),
        'delayed': sum(1 for z in zones_list if z['status'] == 'delayed'),
    }

    # Recent activity (last 7 days)
    recent_activity = []
    for req in WaterRequest.objects.prefetch_related('zones').filter(created_at__date__gte=week_ago).order_by('-created_at')[:5]:
        recent_activity.append({
            'type': 'water',
            'type_display': '浇水协调',
            'zone': ', '.join(z.name for z in req.all_zones),
            'date': req.created_at.strftime('%m-%d %H:%M'),
            'status': req.get_status_display(),
        })

    # Sort by date
    recent_activity.sort(key=lambda x: x['date'], reverse=True)
    recent_activity = recent_activity[:10]

    # Prepare pipelines data for map (prefetch zones to avoid N+1 per pipeline).
    # Pipelines rarely change — cache the built list for 5 min.
    pipelines_list = _cached('dashboard:pipelines', 300, lambda: [
        {
            'id': p.id, 'code': p.code, 'name': p.name,
            'pipeline_type': p.pipeline_type,
            'pipeline_type_display': p.get_pipeline_type_display(),
            'line_points': p.line_points, 'line_color': p.line_color,
            'line_weight': p.line_weight,
            'zone_names': [z.name for z in p.zones.all()],
        } for p in Pipeline.objects.prefetch_related('zones')
    ])

    all_plant_names = _cached('dashboard:plant_names', 300,
                              lambda: list(Plant.objects.values_list('name', flat=True).distinct().order_by('name')))

    # Landmark data for dashboard map and filter (rarely changes → cached 5 min).
    from .models import Landmark, ZoneLandmarkAssignment

    def _build_landmarks():
        lms = [{
            'id': lm.id, 'name': lm.name, 'boundary_points': lm.boundary_points,
            'boundary_color': lm.boundary_color, 'center': lm.center,
        } for lm in Landmark.objects.order_by('order', 'name')]
        zlm = {}
        for assignment in ZoneLandmarkAssignment.objects.select_related('landmark').all():
            zlm.setdefault(assignment.zone_id, []).append({
                'id': assignment.landmark_id, 'name': assignment.landmark.name,
            })
        return {'landmarks': lms, 'zone_landmark_map': zlm}

    _lm = _cached('dashboard:landmarks', 300, _build_landmarks)
    landmarks_data = _lm['landmarks']
    zone_landmark_map = _lm['zone_landmark_map']

    # One marker per pending water request (centroid of its zones + zone ids),
    # for the single-marker-per-request map display.
    from datetime import date as _date
    from .models import WaterRequest as _WR
    _center_by_id = {z['id']: z.get('center') for z in zones_list}
    _zone_id_set = set(_center_by_id.keys())
    pending_water_requests = []
    for wr in _WR.objects.filter(
        status='submitted',
        start_datetime__date__lte=today,
        end_datetime__date__gte=today,
    ).prefetch_related('zones'):
        zids = [z.id for z in wr.zones.all() if z.id in _zone_id_set]
        centers = [c for zid in zids if (c := _center_by_id.get(zid))]
        if centers:
            clat = round(sum(c['lat'] for c in centers) / len(centers), 6)
            clng = round(sum(c['lng'] for c in centers) / len(centers), 6)
            # Hover details: department (user_type) + request_type + duration.
            dept = wr.get_user_type_display()
            if wr.user_type == '其他' and wr.user_type_other:
                dept = wr.user_type_other
            rtype = wr.get_request_type_display()
            if wr.request_type == '其他需求' and wr.request_type_other:
                rtype = wr.request_type_other
            pending_water_requests.append({
                'id': wr.id, 'count': len(zids), 'zone_ids': zids,
                'center': {'lat': clat, 'lng': clng}, 'type_display': '浇水协调',
                'request_type': rtype,
                'user_type': dept,
                'start': wr.start_datetime.strftime('%m-%d %H:%M'),
                'end': wr.end_datetime.strftime('%m-%d %H:%M'),
            })

    # Pending remarks: collapse all zones with unconfirmed remarks into one group
    # (one marker at the centroid with the count + orange outline on each zone).
    _remark_zones = [z for z in zones_list if z.get('has_remarks')]
    _rm_centers = [z['center'] for z in _remark_zones if z.get('center')]
    pending_remarks_data = None
    if _remark_zones and _rm_centers:
        pending_remarks_data = {
            'count': len(_remark_zones),
            'zone_ids': [z['id'] for z in _remark_zones],
            'center': {
                'lat': round(sum(c['lat'] for c in _rm_centers) / len(_rm_centers), 6),
                'lng': round(sum(c['lng'] for c in _rm_centers) / len(_rm_centers), 6),
            },
        }

    # Patch list for map filter plugin
    from .models import Patch
    patches_list = []
    for p in Patch.objects.order_by('code'):
        patches_list.append({
            'id': p.id,
            'name': p.name,
            'code': p.code,
        })

    # Land list for map filter plugin
    lands_list = []
    for l in Land.objects.filter(active=True).order_by('order', 'name'):
        lands_list.append({
            'id': l.id,
            'name': l.name,
        })

    context = {
        # zones_json no longer inlined — fetched async from /api/zones-payload/
        'grouped_zones': grouped_zones,  # For hierarchical sidebar display
        'pending_water_requests_json': json.dumps(pending_water_requests, ensure_ascii=False),
        'pending_remarks_json': json.dumps(pending_remarks_data),
        'pipelines_json': json.dumps(pipelines_list),
        'all_plant_names': all_plant_names,
        'landmarks_json': json.dumps(landmarks_data),
        'landmark_names': [lm['name'] for lm in landmarks_data],
        'zone_landmark_map_json': json.dumps(zone_landmark_map),
        'patches_json': json.dumps(patches_list),
        'lands_json': json.dumps(lands_list),
        'is_admin': is_admin,
        'is_manager': is_manager,
        'is_dept_user': is_dept_user,
        'is_field_worker': is_field_worker,
        'pending_counts': pending_counts,
        'status_distribution': status_distribution,
        'recent_activity': recent_activity,
        'total_zones': len(zones_list),
        'total_plants': sum(z['plant_count'] for z in zones_list),
        'map_style_json': json.dumps(_cached('dashboard:map_style', 300, lambda: MapStyleSettings.get_style())),
        'announcements_json': json.dumps(_unacked_announcements_for(request.user), ensure_ascii=False),
    }

    # PM tasks for the dashboard FAB "PM安排" panel: the logged-in worker's
    # crew's dispatched/overdue tasks (managers see all). Capped for the panel.
    try:
        pm_tasks_qs = _pm_gwo_queryset(request.user, is_admin or is_manager)
        context['pm_tasks_json'] = json.dumps(
            _serialize_pm_tasks(pm_tasks_qs[:50], is_admin or is_manager), ensure_ascii=False)
        context['pm_tasks_total'] = pm_tasks_qs.count()
    except Exception:
        context['pm_tasks_json'] = '[]'
        context['pm_tasks_total'] = 0

    return render(request, 'core/dashboard.html', context)


# ─── Announcements (通知公告) ────────────────────────────────────────

def _announcement_eligible(user):
    """Whether a user is in the announcement audience: 灌溉一线 or 管理员/经理.

    Department users and other account types are excluded — announcements are
    only surfaced to field workers and managers.
    """
    from core.models import ROLE_FIELD_WORKER, ROLE_MANAGER, ROLE_SUPER_ADMIN
    from core.role_utils import get_user_role
    if not user or not getattr(user, 'is_authenticated', False):
        return False
    return get_user_role(user) in (ROLE_FIELD_WORKER, ROLE_MANAGER, ROLE_SUPER_ADMIN)


def _unacked_announcements_for(user):
    """Active announcements the user hasn't acknowledged, newest-first.

    Returns a list of plain dicts (JSON-serializable) for the dashboard popup.
    Only 灌溉一线 / 管理员 users receive announcements; others get an empty list.
    """
    from core.models import Announcement, AnnouncementAcknowledgment
    if not _announcement_eligible(user):
        return []
    acked_ids = AnnouncementAcknowledgment.objects.filter(user=user).values_list('announcement_id', flat=True)
    qs = (Announcement.objects.filter(active=True).exclude(id__in=acked_ids)
          .order_by('-created_at'))
    return [
        {
            'id': a.id,
            'title': a.title,
            'body': a.body,
            'time': a.created_at.strftime('%Y-%m-%d %H:%M'),
        }
        for a in qs
    ]


def _unacked_count_for(user):
    from core.models import Announcement, AnnouncementAcknowledgment
    if not _announcement_eligible(user):
        return 0
    acked_ids = AnnouncementAcknowledgment.objects.filter(user=user).values_list('announcement_id', flat=True)
    return Announcement.objects.filter(active=True).exclude(id__in=acked_ids).count()


@require_POST
@login_required(login_url='core:login')
def announcement_acknowledge(request, pk):
    """Acknowledge one announcement for the current user (idempotent).

    Only 灌溉一线 / 管理员 can acknowledge — the announcement audience.
    """
    if not _announcement_eligible(request.user):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)
    from core.models import Announcement, AnnouncementAcknowledgment
    ann = get_object_or_404(Announcement, pk=pk, active=True)
    AnnouncementAcknowledgment.objects.get_or_create(announcement=ann, user=request.user)
    return JsonResponse({'success': True, 'remaining': _unacked_count_for(request.user)})


@login_required(login_url='core:login')
def announcement_save(request):
    """Create or update an Announcement (manager / super-admin only)."""
    from core.models import Announcement
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        messages.error(request, '无权限')
        return redirect('core:dashboard')

    pid = request.POST.get('id', '').strip()
    title = (request.POST.get('title') or '').strip()
    body = (request.POST.get('body') or '').strip()
    active = request.POST.get('active') in ('1', 'on', 'true', 'True')
    if not title:
        messages.error(request, '标题不能为空')
        return redirect(f"{reverse('core:user_management')}?tab=announcements")

    if pid:
        ann = get_object_or_404(Announcement, pk=pid)
        ann.title = title
        ann.body = body
        ann.active = active
        ann.save()
        messages.success(request, f'通知已更新：{ann.title}')
    else:
        ann = Announcement.objects.create(
            title=title, body=body, active=active, created_by=request.user,
        )
        messages.success(request, f'通知已发布：{ann.title}')
    return redirect(f"{reverse('core:user_management')}?tab=announcements")


@require_POST
@login_required(login_url='core:login')
def announcement_delete(request, pk):
    """Delete an Announcement (manager / super-admin only)."""
    from core.models import Announcement
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        messages.error(request, '无权限')
        return redirect('core:dashboard')

    ann = get_object_or_404(Announcement, pk=pk)
    title = ann.title
    ann.delete()
    messages.success(request, f'通知已删除：{title}')
    return redirect(f"{reverse('core:user_management')}?tab=announcements")


@login_required(login_url='core:login')
def announcement_unacked_api(request, pk):
    """List users who have NOT acknowledged an announcement (manager / admin only).

    Returns the display name + role of every active user lacking an
    AnnouncementAcknowledgment row for this announcement. Used by the expandable
    "未确认" panel on the management page (lazy-loaded on expand).
    """
    from core.models import (
        Announcement, ROLE_SUPER_ADMIN, ROLE_MANAGER, ROLE_FIELD_WORKER, ROLE_DEPT_USER,
    )
    from core.role_utils import get_user_role
    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        return JsonResponse({'error': '无权限'}, status=403)

    ann = get_object_or_404(Announcement, pk=pk)
    acked_user_ids = set(ann.acknowledgments.values_list('user_id', flat=True))
    # The announcement audience is 灌溉一线 + 管理员/经理 only; department users and
    # other accounts are never tracked here. Resolve each to a display name + role.
    label_map = {
        ROLE_SUPER_ADMIN: '超管', ROLE_MANAGER: '管理员',
        ROLE_FIELD_WORKER: '灌溉一线', ROLE_DEPT_USER: '部门用户',
    }
    users = []
    from django.contrib.auth import get_user_model
    # Prefetch all three profile types in ONE query each (was 3 queries/user via
    # get_user_role's .exists() calls + 3 FK lookups in _user_display_name → ~6N
    # queries for the whole panel). select_related turns each into an in-memory
    # attribute read with zero extra queries.
    User = get_user_model()
    for u in (User.objects.filter(is_active=True)
              .order_by('username')
              .select_related('worker_profile', 'manager_profile', 'dept_profile')):
        role = _role_from_prefetched(u)
        if role not in (ROLE_FIELD_WORKER, ROLE_MANAGER, ROLE_SUPER_ADMIN):
            continue  # dept users / others are not part of the audience
        if u.id in acked_user_ids:
            continue
        users.append({
            'name': _user_display_name(u),
            'role': label_map.get(role, '—'),
        })
    return JsonResponse({'success': True, 'count': len(users), 'users': users})


def _user_display_name(user):
    """Friendly name for any user, preferring the linked profile's full_name."""
    name = (user.get_full_name() or '').strip()
    if name:
        return name
    for attr in ('worker_profile', 'manager_profile', 'dept_profile'):
        # Reverse OneToOne access raises RelatedObjectDoesNotExist (an
        # ObjectDoesNotExist subclass, not AttributeError) when no row exists,
        # so getattr's default doesn't apply — guard explicitly.
        try:
            profile = getattr(user, attr, None)
        except Exception:
            profile = None
        if profile and getattr(profile, 'full_name', None):
            return profile.full_name
    return user.username


def _role_from_prefetched(user):
    """Resolve a user's role from already-prefetched profile relations.

    Mirror of role_utils.get_user_role but with ZERO queries: it reads the
    select_related profiles instead of running .exists(). Use only when the
    queryset was built with select_related('worker_profile','manager_profile',
    'dept_profile'); otherwise the related access would lazy-load (N+1).
    """
    from core.role_utils import (
        ROLE_SUPER_ADMIN, ROLE_MANAGER, ROLE_FIELD_WORKER, ROLE_DEPT_USER,
    )
    if not user or not getattr(user, 'is_authenticated', False):
        return None
    if getattr(user, 'is_superuser', False) or getattr(user, 'is_staff', False):
        return ROLE_SUPER_ADMIN
    # For a reverse OneToOne, accessing the attribute raises
    # RelatedObjectDoesNotExist (a subclass of ObjectDoesNotExist, NOT
    # AttributeError) when no related row exists. Catch that and treat as None
    # — the profile was prefetched (or absent), so this never triggers a query.
    def _get(attr):
        try:
            return getattr(user, attr, None)
        except Exception:  # ObjectDoesNotExist / RelatedObjectDoesNotExist
            return None

    mp = _get('manager_profile')
    if mp and getattr(mp, 'active', True):
        return ROLE_SUPER_ADMIN if getattr(mp, 'is_super_admin', False) else ROLE_MANAGER
    wp = _get('worker_profile')
    if wp and getattr(wp, 'active', True):
        return ROLE_FIELD_WORKER
    dp = _get('dept_profile')
    if dp and getattr(dp, 'active', True):
        return ROLE_DEPT_USER
    return None




# Excel headers matching Zone list V0.xlsx column order
_ZONE_EXPORT_HEADERS = [
    '编号', '位置重要程度', '所属Land', '通用名称', '灌溉管理用的位置',
    '当前状态', '灌水器类型', '灌溉强度 mm/h', '区域面积',
    '灌溉分区', 'CCU编号', '电磁阀尺寸', '景观系数',
    '植物类型', '灌溉领班', '绿化分区', '绿化领班',
    '植保分区', '植保领班', '地形特点', '植物特点',
    '土壤湿度', '灌溉设备维护记录', '灌溉管理以往记录',
]

# Priority display text → model value (for import)
_PRIORITY_IMPORT_MAP = {
    '超级重点位置': Zone.PRIORITY_CRITICAL,
    '重点位置': Zone.PRIORITY_HIGH,
    '一般位置': Zone.PRIORITY_MEDIUM,
    '次要位置': Zone.PRIORITY_LOW,
    '废除': Zone.PRIORITY_ABOLISHED,
}

# Model value → display text (for export)
_PRIORITY_EXPORT_MAP = {v: k for k, v in _PRIORITY_IMPORT_MAP.items()}

# Fields that map directly to zone model attributes (col index 0-based → field name)
# Excluding: col 7 (area), col 8 (irrigation zone name), col 9 (CCU number) which are computed
_ZONE_FIELD_COLUMNS = {
    0: 'code',
    2: 'name',
    3: 'description',
    4: 'current_status',
    5: 'sprinkler_type',
    6: 'irrigation_intensity',
    10: 'solenoid_valve_size',
    11: 'landscape_coefficient',
    12: 'plant_type',
    13: 'irrigation_foreman',
    14: 'greenery_zone',
    15: 'greenery_foreman',
    16: 'pest_control_zone',
    17: 'pest_control_foreman',
    18: 'terrain_feature',
    19: 'plant_feature',
    20: 'soil_moisture',
    21: 'equipment_maintenance_notes',
    22: 'irrigation_management_notes',
}


def _check_zone_admin(request):
    """Return True if user has zone admin permission."""
    if request.user.is_superuser or request.user.is_staff:
        return True
    from .models import ManagerProfile
    try:
        ManagerProfile.objects.get(user=request.user, active=True)
        return True
    except ManagerProfile.DoesNotExist:
        return False


def _get_user_display_name(request):
    """Get the current user's display name from their profile."""
    from .models import ManagerProfile, Worker
    for Model in (ManagerProfile, Worker):
        try:
            profile = Model.objects.get(user=request.user, active=True)
            return profile.full_name or request.user.username
        except Model.DoesNotExist:
            continue
    return request.user.username


@login_required(login_url='core:login')
def zone_export_excel(request):
    """Export all zones to an Excel file matching Zone list V0.xlsx format."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Alignment
    except ImportError:
        return JsonResponse({'error': 'openpyxl not installed'}, status=500)

    if not _check_zone_admin(request):
        return JsonResponse({'error': '无权限'}, status=403)

    zones = Zone.objects.select_related('patch').order_by('code')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Zone info 仅色块'

    # Write headers with same styling as V0
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col_idx, header in enumerate(_ZONE_EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.alignment = center_wrap

    # Set column widths matching V0
    col_widths = [16, 13, 14, 13, 18, 13, 16, 13, 14, 13, 12, 8, 9, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w

    # Write data rows
    for zone in zones:
        row_num = ws.max_row + 1
        # CCU number from patch code (e.g. "CCU1" → "1")
        ccu_num = zone.patch.code.replace('CCU', '') if zone.patch else ''

        values = [
            zone.code,                                          # A: 编号
            _PRIORITY_EXPORT_MAP.get(zone.priority, ''),         # B: 位置重要程度
            zone.land.name if zone.land else None,              # C: 所属Land
            zone.name,                                          # D: 通用名称
            zone.description,                                   # E: 灌溉管理用的位置
            zone.current_status or None,                        # F: 当前状态
            zone.sprinkler_type or None,                        # G: 灌水器类型
            zone.irrigation_intensity,                          # H: 灌溉强度
            zone.area_display if zone.area_sqm else None,       # I: 区域面积
            zone.patch.name if zone.patch else None,            # J: 灌溉分区
            int(ccu_num) if ccu_num.isdigit() else ccu_num,     # K: CCU编号
            zone.solenoid_valve_size,                           # L: 电磁阀尺寸
            zone.landscape_coefficient,                         # M: 景观系数
            zone.plant_type or None,                            # N: 植物类型
            zone.irrigation_foreman or None,                    # O: 灌溉领班
            zone.greenery_zone or None,                         # P: 绿化分区
            zone.greenery_foreman or None,                      # Q: 绿化领班
            zone.pest_control_zone or None,                     # R: 植保分区
            zone.pest_control_foreman or None,                  # S: 植保领班
            zone.terrain_feature or None,                       # T: 地形特点
            zone.plant_feature or None,                         # U: 植物特点
            zone.soil_moisture or None,                         # V: 土壤湿度
            zone.equipment_maintenance_notes or None,           # W: 灌溉设备维护记录
            zone.irrigation_management_notes or None,           # X: 灌溉管理以往记录
        ]
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_num, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from django.utils import timezone
    filename = f"zones_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _parse_zone_row(row):
    """Parse an Excel row (23 columns, values_only) into a dict of zone fields."""
    code = str(row[0] or '').strip() if row[0] else ''
    if not code:
        return None

    priority_raw = str(row[1] or '').strip()
    priority = _PRIORITY_IMPORT_MAP.get(priority_raw, Zone.PRIORITY_MEDIUM)

    def _float(val):
        if val is None or val == '':
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def _str(val):
        return str(val or '').strip()

    return {
        'code': code,
        'land': _str(row[2]),                # 所属Land (resolved to Land in confirm)
        'name': _str(row[3]) or code,
        'description': _str(row[4]),
        'priority': priority,
        'current_status': _str(row[5]),
        'sprinkler_type': _str(row[6]),
        'irrigation_intensity': _float(row[7]),
        # row[8] = 区域面积 (computed) and row[9]/[10] = 灌溉分区/CCU (from patch) — skipped
        'solenoid_valve_size': _float(row[11]),
        'landscape_coefficient': _float(row[12]),
        'plant_type': _str(row[13]),
        'irrigation_foreman': _str(row[14]),
        'greenery_zone': _str(row[15]),
        'greenery_foreman': _str(row[16]),
        'pest_control_zone': _str(row[17]),
        'pest_control_foreman': _str(row[18]),
        'terrain_feature': _str(row[19]),
        'plant_feature': _str(row[20]),
        'soil_moisture': _str(row[21]),
        'equipment_maintenance_notes': _str(row[22]),
        # row[23] = 灌溉管理以往记录 (protected) — skipped on import
    }


# Field display names for the preview table
_FIELD_LABELS = {
    'code': '编号', 'land': '所属Land', 'name': '通用名称', 'description': '灌溉管理用的位置',
    'priority': '位置重要程度', 'current_status': '当前状态',
    'sprinkler_type': '灌水器类型', 'irrigation_intensity': '灌溉强度',
    'solenoid_valve_size': '电磁阀尺寸', 'landscape_coefficient': '景观系数',
    'plant_type': '植物类型', 'irrigation_foreman': '灌溉领班',
    'greenery_zone': '绿化分区', 'greenery_foreman': '绿化领班',
    'pest_control_zone': '植保分区', 'pest_control_foreman': '植保领班',
    'terrain_feature': '地形特点', 'plant_feature': '植物特点',
    'soil_moisture': '土壤湿度', 'equipment_maintenance_notes': '设备维护记录',
    'irrigation_management_notes': '灌溉管理记录',
}

# Fields that are DB-primary / auto-calculated — never overwritten on import
# (面积 is computed from the boundary; 灌溉管理以往记录 is accumulated history).
_ZONE_IMPORT_PROTECTED = {'area_sqm', 'irrigation_management_notes'}


@login_required(login_url='core:login')
def zone_import_preview(request):
    """Upload an xlsx file, compare with DB, return change list as JSON."""
    try:
        import openpyxl
    except ImportError:
        return JsonResponse({'error': 'openpyxl not installed'}, status=500)

    if not _check_zone_admin(request):
        return JsonResponse({'error': '无权限'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    uploaded = request.FILES.get('file')
    if not uploaded:
        return JsonResponse({'error': '未选择文件'}, status=400)

    try:
        wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return JsonResponse({'error': f'文件解析失败: {e}'}, status=400)

    # Preload all zones by code
    existing_zones = {z.code: z for z in Zone.objects.select_related('patch').all()}

    # Fields to compare (exclude code + DB-primary/auto-calculated fields)
    compare_fields = [k for k in _FIELD_LABELS if k != 'code' and k not in _ZONE_IMPORT_PROTECTED]

    changes = []
    counts = {'new': 0, 'modified': 0, 'unchanged': 0}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=24, values_only=True), start=2):
        parsed = _parse_zone_row(row)
        if not parsed:
            continue

        code = parsed['code']
        zone = existing_zones.get(code)

        if zone is None:
            changes.append({
                'row': row_idx, 'code': code, 'action': 'new',
                'name': parsed['name'],
                'fields': [{'field': k, 'label': _FIELD_LABELS.get(k, k), 'old': '', 'new': str(v)}
                           for k, v in parsed.items() if k != 'code' and v],
            })
            counts['new'] += 1
        else:
            field_changes = []
            for field in compare_fields:
                new_val = parsed.get(field)
                old_val = getattr(zone, field, None)
                # Normalize for comparison
                old_norm = str(old_val or '').strip()
                new_norm = str(new_val or '').strip()
                if old_norm != new_norm:
                    field_changes.append({
                        'field': field,
                        'label': _FIELD_LABELS.get(field, field),
                        'old': old_norm or '(空)',
                        'new': new_norm or '(空)',
                    })
            if field_changes:
                changes.append({
                    'row': row_idx, 'code': code, 'action': 'modified',
                    'name': zone.name,
                    'fields': field_changes,
                })
                counts['modified'] += 1
            else:
                counts['unchanged'] += 1

    wb.close()
    return JsonResponse({'success': True, 'summary': counts, 'changes': changes})


@login_required(login_url='core:login')
def zone_import_confirm(request):
    """Apply confirmed import changes. Expects multipart with file + rows JSON."""
    try:
        import openpyxl
    except ImportError:
        return JsonResponse({'error': 'openpyxl not installed'}, status=500)

    if not _check_zone_admin(request):
        return JsonResponse({'error': '无权限'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    uploaded = request.FILES.get('file')
    rows_json = request.POST.get('rows', '[]')
    try:
        import json as _json
        confirmed_rows = set(_json.loads(rows_json))
    except (_json.JSONDecodeError, TypeError):
        return JsonResponse({'error': '无效的行号数据'}, status=400)

    if not uploaded:
        return JsonResponse({'error': '未选择文件'}, status=400)

    try:
        wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return JsonResponse({'error': f'文件解析失败: {e}'}, status=400)

    from .models import Patch, Land

    # Preload patches by CCU number
    patch_map = {}
    for p in Patch.objects.filter(code__startswith='CCU'):
        patch_map[p.code.replace('CCU', '')] = p
    # Preload lands by name
    land_map = {l.name: l for l in Land.objects.all()}

    created = 0
    updated = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=24, values_only=True), start=2):
        if row_idx not in confirmed_rows:
            continue

        parsed = _parse_zone_row(row)
        if not parsed:
            continue

        code = parsed['code']
        # Resolve patch from code prefix
        parts = code.split('-')
        ccu_prefix = parts[0] if parts else ''
        patch = patch_map.get(ccu_prefix)
        # Resolve Land (create on demand); protected fields are already excluded by _parse_zone_row
        land_name = (parsed.get('land') or '').strip()
        land = None
        if land_name:
            land = land_map.get(land_name)
            if not land:
                land = Land.objects.create(name=land_name)
                land_map[land_name] = land

        zone = Zone.objects.filter(code=code).first()
        if zone:
            for k, v in parsed.items():
                if k in ('code', 'land'):
                    continue
                setattr(zone, k, v)
            zone.land = land
            if patch:
                zone.patch = patch
            zone.save()
            updated += 1
        else:
            zone = Zone(code=code, patch=patch, land=land)
            for k, v in parsed.items():
                if k in ('code', 'land'):
                    continue
                setattr(zone, k, v)
            zone.save()
            created += 1

    wb.close()
    return JsonResponse({'success': True, 'created': created, 'updated': updated})


@login_required(login_url='core:login')
def settings_page(request):
    """
    Settings page to manage zones and system configuration.
    Accessible to all authenticated users.
    """
    from .models import Pipeline, Patch, Plant, Region

    zones = Zone.objects.all().order_by('code')
    pipelines = Pipeline.objects.all().order_by('code')
    regions = Region.objects.filter(active=True).order_by('order', 'name')
    site_patches = Patch.objects.order_by('code')

    # Precompute zone counts per patch (FK + derived from code prefix)
    grouped_zones_data = _build_grouped_zones(zones)
    priority_zones_data = _build_grouped_zones(zones, group_by='priority')
    patch_zone_counts = {}
    for group in grouped_zones_data:
        pid = group.get('id')
        if pid:
            patch_zone_counts[pid] = group['zone_count']

    # Precompute linked patch counts (children via parent FK)
    child_counts = {}
    for p in site_patches:
        child_counts[p.id] = p.children.count()

    # Precompute patch counts per region
    region_patch_counts = {}
    for r in regions:
        region_patch_counts[r.id] = r.patches.count()

    all_plant_names = list(Plant.objects.values_list('name', flat=True).distinct().order_by('name'))

    from .models import Landmark
    landmarks = Landmark.objects.annotate(zone_count=Count('zone_assignments')).order_by('order', 'name')

    # SAT satellite controllers shown under the 片区管理 tab.
    from .models import Satellite
    satellites = Satellite.objects.all().order_by('code')

    context = {
        'zones': zones,
        'grouped_zones': grouped_zones_data,
        'priority_zones': priority_zones_data,
        'priority_choices': Zone.PRIORITY_CHOICES,
        'all_plant_names': all_plant_names,
        'status_choices': Zone.STATUS_CHOICES,
        'pipelines': pipelines,
        'site_patches': site_patches,
        'patch_zone_counts': patch_zone_counts,
        'child_counts': child_counts,
        'regions': regions,
        'region_patch_counts': region_patch_counts,
        'landmarks': landmarks,
        'satellites': satellites,
    }

    return render(request, 'core/settings.html', context)


@login_required(login_url='core:login')
def crew_new(request):
    """Create a new Crew (班组)."""
    from .role_utils import is_admin
    if not is_admin(request.user):
        messages.error(request, '无权限执行此操作')
        return redirect('/user-management/?tab=crews')
    from .models import Crew, Land, Patch
    ccu_patches = Patch.objects.filter(code__regex=r'^CCU[0-9]+$').order_by('code')
    if request.method == 'POST':
        crew = Crew.objects.create(
            name=request.POST.get('name', '').strip(),
            leader_id=request.POST.get('leader') or None,
        )
        crew.members.set(request.POST.getlist('members'))
        crew.lands.set(request.POST.getlist('lands'))
        crew.patches.set(request.POST.getlist('patches'))
        messages.success(request, f'班组「{crew.name}」已创建')
        return redirect('/user-management/?tab=crews')
    from .models import Worker
    context = {'crew': None, 'workers': Worker.objects.filter(active=True).order_by('full_name'),
               'lands': Land.objects.order_by('order', 'name'),
               'ccu_patches': ccu_patches}
    return render(request, 'core/crew_form.html', context)


@login_required(login_url='core:login')
def crew_edit(request, crew_id):
    """Edit an existing Crew."""
    from .role_utils import is_admin
    if not is_admin(request.user):
        messages.error(request, '无权限执行此操作')
        return redirect('/user-management/?tab=crews')
    from .models import Crew, Worker, Land, Patch
    ccu_patches = Patch.objects.filter(code__regex=r'^CCU[0-9]+$').order_by('code')
    crew = get_object_or_404(Crew, pk=crew_id)
    if request.method == 'POST':
        crew.name = request.POST.get('name', '').strip()
        crew.leader_id = request.POST.get('leader') or None
        crew.save()
        crew.members.set(request.POST.getlist('members'))
        crew.lands.set(request.POST.getlist('lands'))
        crew.patches.set(request.POST.getlist('patches'))
        messages.success(request, f'班组「{crew.name}」已更新')
        return redirect('/user-management/?tab=crews')
    context = {'crew': crew, 'workers': Worker.objects.filter(active=True).order_by('full_name'),
               'lands': Land.objects.order_by('order', 'name'),
               'ccu_patches': ccu_patches}
    return render(request, 'core/crew_form.html', context)


@login_required(login_url='core:login')
def pm_plan_new(request):
    """Create a new MaintenancePlan (PM 计划)."""
    from .role_utils import is_admin
    if not is_admin(request.user):
        messages.error(request, '无权限执行此操作')
        return redirect('core:pm_management')
    from .models import JobPlanTemplate, Crew, Zone, Satellite, MaintenancePlan, Patch
    from django.utils import timezone as _tz
    if request.method == 'POST':
        plan = MaintenancePlan.objects.create(
            pm_number=request.POST.get('pm_number', '').strip(),
            job_plan_id=request.POST.get('job_plan') or None,
            crew_id=request.POST.get('crew') or None,
            frequency_value=int(request.POST.get('frequency_value', 1) or 1),
            frequency_unit=request.POST.get('frequency_unit', 'weeks'),
            start_date=_tz.localdate(),
            lead_days=int(request.POST.get('lead_days', 28) or 28),
            active=bool(request.POST.get('active')),
            remark_template=request.POST.get('remark_template', '').strip(),
            patch_id=request.POST.get('patch') or None,
            satellite_id=request.POST.get('satellite') or None,
        )
        plan.zones.set(request.POST.getlist('zones'))
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'PM计划「{plan.pm_number}」已创建',
                                 'plan': _serialize_pm_plan(plan)})
        messages.success(request, f'PM计划「{plan.pm_number}」已创建')
        return redirect('core:pm_management')
    context = {
        'plan': None,
        'job_plans': JobPlanTemplate.objects.all().order_by('name'),
        'crews': Crew.objects.all().order_by('name'),
        'patches': Patch.objects.order_by('code'),
        'satellites': Satellite.objects.all().order_by('code'),
        'zones': Zone.objects.all().order_by('code')[:500],
    }
    return render(request, 'core/pm_plan_form.html', context)


@login_required(login_url='core:login')
def pm_plan_edit(request, plan_id):
    """Edit an existing MaintenancePlan."""
    from .role_utils import is_admin
    if not is_admin(request.user):
        messages.error(request, '无权限执行此操作')
        return redirect('core:pm_management')
    from .models import JobPlanTemplate, Crew, Zone, Satellite, MaintenancePlan, Patch
    plan = get_object_or_404(MaintenancePlan, pk=plan_id)
    if request.method == 'POST':
        plan.pm_number = request.POST.get('pm_number', '').strip()
        plan.job_plan_id = request.POST.get('job_plan') or None
        plan.crew_id = request.POST.get('crew') or None
        plan.frequency_value = int(request.POST.get('frequency_value', 1) or 1)
        plan.frequency_unit = request.POST.get('frequency_unit', 'weeks')
        plan.lead_days = int(request.POST.get('lead_days', 28) or 28)
        plan.active = bool(request.POST.get('active'))
        plan.remark_template = request.POST.get('remark_template', '').strip()
        plan.patch_id = request.POST.get('patch') or None
        plan.satellite_id = request.POST.get('satellite') or None
        # 到期基准日：经理可改。改后仅影响未来派发（rrule 锚点），已有 GWO 不动。
        from datetime import date as _dt
        start_date_str = request.POST.get('start_date', '').strip()
        if start_date_str:
            try:
                plan.start_date = _dt.fromisoformat(start_date_str)
            except ValueError:
                pass  # 格式无效时保持原值
        plan.save()
        plan.zones.set(request.POST.getlist('zones'))
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'PM计划「{plan.pm_number}」已更新',
                                 'plan': _serialize_pm_plan(plan)})
        messages.success(request, f'PM计划「{plan.pm_number}」已更新')
        return redirect('core:pm_management')
    context = {
        'plan': plan,
        'job_plans': JobPlanTemplate.objects.all().order_by('name'),
        'crews': Crew.objects.all().order_by('name'),
        'patches': Patch.objects.order_by('code'),
        'satellites': Satellite.objects.all().order_by('code'),
        'zones': Zone.objects.all().order_by('code')[:500],
    }
    return render(request, 'core/pm_plan_form.html', context)


def _pm_period_window(freq_value, freq_unit, anchor, offset, today):
    """Return (start, end, label) for a plan's own frequency-based period.

    The period is anchored to ``anchor`` (the plan's start_date). offset=0 →
    the period containing ``today``, -1 → the previous period, etc. Uses
    frequency_value × frequency_unit to determine the period span:

      每1天  → 1-day windows
      每3周  → 21-day windows
      每6月  → 6-month windows (via relativedelta for correct month math)
      每12月 → 12-month windows

    For offset < 0 that would land before the anchor, returns (None, None, label)
    so the caller can show "暂无" (the plan hasn't existed long enough).
    """
    from datetime import timedelta
    from dateutil.relativedelta import relativedelta

    if freq_unit == 'days':
        period = timedelta(days=freq_value)
    elif freq_unit == 'weeks':
        period = timedelta(weeks=freq_value)
    else:  # months
        period = relativedelta(months=freq_value)

    # Find how many full periods have elapsed since anchor → current period #.
    n = 0
    cur_start = anchor
    while cur_start + period <= today:
        n += 1
        cur_start = cur_start + period
    target_n = n + offset
    if target_n < 0:
        return None, None, f'锚点前{abs(target_n)}期'

    start = anchor
    for _ in range(target_n):
        start = start + period
    end = start + period - (timedelta(days=1) if freq_unit != 'months' else relativedelta(days=1))
    # Label: show date range, abbreviated for readability.
    if freq_unit == 'days':
        label = start.strftime('%m-%d')
    elif freq_unit == 'weeks':
        label = f'{start.strftime("%m-%d")}~{end.strftime("%m-%d")}'
    else:
        if start.year == end.year:
            label = f'{start.strftime("%Y-%m-%d")}~{end.strftime("%m-%d")}'
        else:
            label = f'{start.strftime("%y-%m")}~{end.strftime("%y-%m")}'
    return start, end, label


def _serialize_jobplan(jp):
    """Serialize a JobPlanTemplate for AJAX modal responses."""
    return {
        'id': jp.id,
        'name': jp.name,
        'asset_level': jp.asset_level,
        'asset_level_display': jp.get_asset_level_display(),
        'description': jp.description or '',
        'active': jp.active,
        'plan_count': jp.plans.count(),
    }


def _serialize_pm_plan(p):
    """Serialize a MaintenancePlan for AJAX modal responses."""
    if p.zones.exists():
        asset = f'{p.zones.count()} zones'
    elif p.satellite_id:
        asset = f'SAT {p.satellite.code}'
    elif p.patch_id:
        asset = p.patch.code
    else:
        asset = '—'
    return {
        'id': p.id,
        'pm_number': p.pm_number,
        'job_plan_id': p.job_plan_id,
        'job_plan_name': p.job_plan.name if p.job_plan_id else '',
        'frequency_value': p.frequency_value,
        'frequency_unit': p.frequency_unit,
        'freq_display': f'每{p.frequency_value}{p.get_frequency_unit_display()}',
        'asset_info': asset,
        'crew_id': p.crew_id or '',
        'crew_name': p.crew.name if p.crew_id else '',
        'lead_days': p.lead_days,
        'remark_template': p.remark_template or '',
        'patch_id': p.patch_id or '',
        'satellite_id': p.satellite_id or '',
        'start_date': p.start_date.isoformat() if p.start_date else '',
        'last_generated_date': p.last_generated_date.isoformat() if p.last_generated_date else '',
        'active': p.active,
        'order_count': p.generated_orders.count(),
    }


@login_required(login_url='core:login')
def jobplan_new(request):
    """Create a new JobPlanTemplate."""
    from .role_utils import is_admin
    if not is_admin(request.user):
        messages.error(request, '无权限执行此操作')
        return redirect('core:pm_management')
    from .models import JobPlanTemplate
    if request.method == 'POST':
        jp = JobPlanTemplate.objects.create(
            name=request.POST.get('name', '').strip(),
            description=request.POST.get('description', '').strip(),
            asset_level=request.POST.get('asset_level', 'zone_group'),
            active=bool(request.POST.get('active')),
        )
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'作业计划模板「{jp.name}」已创建',
                                 'jp': _serialize_jobplan(jp)})
        messages.success(request, f'作业计划模板「{jp.name}」已创建')
        return redirect('core:pm_management')
    from .models import JobPlanTemplate
    context = {'jp': None}
    return render(request, 'core/jobplan_form.html', context)


@login_required(login_url='core:login')
def jobplan_edit(request, jp_id):
    """Edit an existing JobPlanTemplate."""
    from .role_utils import is_admin
    if not is_admin(request.user):
        messages.error(request, '无权限执行此操作')
        return redirect('core:pm_management')
    from .models import JobPlanTemplate
    jp = get_object_or_404(JobPlanTemplate, pk=jp_id)
    if request.method == 'POST':
        jp.name = request.POST.get('name', '').strip()
        jp.description = request.POST.get('description', '').strip()
        jp.asset_level = request.POST.get('asset_level', 'zone_group')
        jp.active = bool(request.POST.get('active'))
        jp.save()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'作业计划模板「{jp.name}」已更新',
                                 'jp': _serialize_jobplan(jp)})
        messages.success(request, f'作业计划模板「{jp.name}」已更新')
        return redirect('core:pm_management')
    context = {'jp': jp}
    return render(request, 'core/jobplan_form.html', context)


@require_POST
@login_required(login_url='core:login')
def pm_generate_now(request):
    """Manually trigger PM dispatch (POST only — GET would mutate the DB)."""
    from .role_utils import is_admin
    if not is_admin(request.user):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': '无权限'}, status=403)
        messages.error(request, '无权限执行此操作')
        return redirect('core:pm_management')
    from django.core.management import call_command
    from io import StringIO
    out = StringIO()
    try:
        # Transition past-due dispatched orders to 'overdue' before generating,
        # so the completion tab reflects reality even on a manual trigger.
        call_command('mark_pm_overdue', stdout=out)
        call_command('generate_pm_workorders', stdout=out)
        summary = out.getvalue().strip().splitlines()[-1] if out.getvalue().strip() else '完成'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'PM派发完成：{summary}'})
        messages.success(request, f'PM派发完成：{summary}')
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception('pm_generate_now failed')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': f'PM派发失败：{e}'})
        messages.error(request, f'PM派发失败：{e}')
    return redirect('core:pm_management')


@require_POST
@login_required(login_url='core:login')
def pm_assign_crews(request):
    """Auto-match PM plans to crews (POST only — mutates plan.crew).

    Runs the assign_pm_crews management command: CCU/SAT-level PMs match by
    Crew.patches first, falling back to Land coverage for zone_group PMs.
    """
    from .role_utils import is_admin
    if not is_admin(request.user):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': '无权限'}, status=403)
        messages.error(request, '无权限执行此操作')
        return redirect('core:pm_management')
    from django.core.management import call_command
    from io import StringIO
    out = StringIO()
    try:
        call_command('assign_pm_crews', stdout=out)
        summary = out.getvalue().strip().splitlines()[-1] if out.getvalue().strip() else '完成'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'班组分配完成：{summary}'})
        messages.success(request, f'班组分配完成：{summary}')
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception('pm_assign_crews failed')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': f'班组分配失败：{e}'})
        messages.error(request, f'班组分配失败：{e}')
    return redirect('core:pm_management')


@require_POST
@login_required(login_url='core:login')
def pm_reset_overdue(request):
    """Re-anchor past-due PM start_dates and clear stale GWOs (POST only).

    Runs the reset_pm_overdue management command with --apply: for each active
    plan whose start_date is in the past, deletes its uncompleted work orders
    and re-anchors start_date to the next future occurrence.
    """
    from .role_utils import is_admin
    if not is_admin(request.user):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': '无权限'}, status=403)
        messages.error(request, '无权限执行此操作')
        return redirect('core:pm_management')
    from django.core.management import call_command
    from io import StringIO
    out = StringIO()
    try:
        call_command('reset_pm_overdue', '--apply', stdout=out)
        summary = out.getvalue().strip().splitlines()[-1] if out.getvalue().strip() else '完成'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'重置完成：{summary}'})
        messages.success(request, f'重置完成：{summary}')
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception('pm_reset_overdue failed')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': f'重置失败：{e}'})
        messages.error(request, f'重置失败：{e}')
    return redirect('core:pm_management')


@login_required(login_url='core:login')
def pm_plan_orders(request, plan_id):
    """AJAX: return a PM plan's generated work orders as JSON (lazy-load)."""
    from .models import MaintenancePlan
    plan = get_object_or_404(MaintenancePlan, pk=plan_id)
    orders = []
    for gwo in plan.generated_orders.select_related('pm_order__worker', 'crew', 'worker').order_by('-scheduled_date'):
        pmwo = gwo.pm_order
        # Zones: from the PMWorkOrder after completion, else from the GWO (dispatch snapshot).
        zones_src = pmwo if pmwo else gwo
        orders.append({
            'scheduled_date': gwo.scheduled_date.strftime('%Y-%m-%d'),
            'generated_at': gwo.generated_at.strftime('%Y-%m-%d %H:%M') if gwo.generated_at else '',
            'status': gwo.status,
            'pm_order_id': pmwo.id if pmwo else None,
            'report_number': pmwo.display_number if pmwo else f'PM-{gwo.id}',
            'crew_name': gwo.crew.name if gwo.crew_id else '',
            'worker_name': (pmwo.worker.full_name if pmwo and pmwo.worker_id
                            else (gwo.worker.full_name if gwo.worker_id else '')),
            'zone_count': zones_src.zones.count(),
            'entry_count': pmwo.entries.count() if pmwo else 0,
            'work_content': '',
        })
    return JsonResponse({'orders': orders})


@login_required(login_url='core:login')
def pm_extension_request(request, gwo_id):
    """Field worker submits a PM work order extension request."""
    from core.role_utils import get_worker_for_user, is_admin
    from .models import GeneratedWorkOrder, ExtensionRequest
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '仅支持POST'}, status=405)
    gwo = get_object_or_404(GeneratedWorkOrder, pk=gwo_id)
    reason = request.POST.get('reason', '').strip()
    req_date_str = request.POST.get('requested_date', '').strip()
    if not reason or not req_date_str:
        return JsonResponse({'success': False, 'message': '请填写延期理由和期望日期'}, status=400)
    from datetime import date as _date
    try:
        requested_date = _date.fromisoformat(req_date_str)
    except ValueError:
        return JsonResponse({'success': False, 'message': '日期格式无效'}, status=400)
    worker = get_worker_for_user(request.user)
    # Lock the GWO row so concurrent extension requests serialize at the DB
    # level — two crew members POSTing at the same time can't both pass the
    # pending-check and both create a request.
    from django.db import transaction
    with transaction.atomic():
        GeneratedWorkOrder.objects.select_for_update().get(pk=gwo_id)
        # Only one pending request per GWO.
        if ExtensionRequest.objects.filter(gwo=gwo, status='pending').exists():
            return JsonResponse({'success': False, 'message': '该工单已有待审批的延期申请'}, status=400)
        ext = ExtensionRequest.objects.create(
            gwo=gwo, requester=worker, reason=reason, requested_date=requested_date,
        )
    return JsonResponse({'success': True, 'message': '延期申请已提交，等待经理审批',
                         'ext_id': ext.id})


@login_required(login_url='core:login')
def pm_extension_approve(request, req_id):
    """Manager approves an extension: skip old GWO + update plan.start_date."""
    from core.role_utils import is_admin
    from .models import ExtensionRequest
    from django.utils import timezone as _tz
    if not is_admin(request.user):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '仅支持POST'}, status=405)
    ext = get_object_or_404(ExtensionRequest, pk=req_id, status='pending')
    gwo = ext.gwo
    plan = gwo.plan
    # Mark the old GWO as skipped (its work_report stays but is no longer tracked).
    gwo.status = 'skipped'
    gwo.save(update_fields=['status'])
    # Shift the plan's frequency anchor so next dispatch uses the new date.
    plan.start_date = ext.requested_date
    plan.save(update_fields=['start_date'])
    ext.status = 'approved'
    ext.reviewed_by = request.user
    ext.reviewed_at = _tz.now()
    ext.review_note = request.POST.get('note', '').strip()
    ext.save()
    return JsonResponse({'success': True, 'message': f'已批准延期到 {ext.requested_date}，下次派发按新日期执行'})


@login_required(login_url='core:login')
def pm_extension_reject(request, req_id):
    """Manager rejects an extension request."""
    from core.role_utils import is_admin
    from .models import ExtensionRequest
    from django.utils import timezone as _tz
    if not is_admin(request.user):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '仅支持POST'}, status=405)
    ext = get_object_or_404(ExtensionRequest, pk=req_id, status='pending')
    ext.status = 'rejected'
    ext.reviewed_by = request.user
    ext.reviewed_at = _tz.now()
    ext.review_note = request.POST.get('note', '').strip()
    ext.save()
    return JsonResponse({'success': True, 'message': '已拒绝延期申请'})


@login_required(login_url='core:login')
def pm_gwo_skip(request, gwo_id):
    """Manager directly marks a GWO as skipped (no extension needed)."""
    from core.role_utils import is_admin
    from .models import GeneratedWorkOrder
    if not is_admin(request.user):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '仅支持POST'}, status=405)
    gwo = get_object_or_404(GeneratedWorkOrder, pk=gwo_id)
    gwo.status = 'skipped'
    gwo.save(update_fields=['status'])
    return JsonResponse({'success': True, 'message': f'已跳过 {gwo.plan.pm_number}'})


@require_POST
@login_required(login_url='core:login')
def pm_gwo_skip_all(request):
    """Mark ALL overdue GeneratedWorkOrders as skipped (one-click bulk skip)."""
    from core.role_utils import is_admin
    from .models import GeneratedWorkOrder
    if not is_admin(request.user):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)
    count = GeneratedWorkOrder.objects.filter(status='overdue').update(status='skipped')
    return JsonResponse({'success': True, 'message': f'已跳过 {count} 张逾期工单'})


@require_POST
@login_required(login_url='core:login')
def work_report_resolve_repair(request, report_id):
    """Mark a single pending-repair / difficult work order as resolved.

    Clears is_pending_repair, and if the report is flagged 疑难 (is_difficult)
    also flips is_difficult_resolved=True so the detail page's "已处理" shows
    "是". Writes an edit-log entry recording the transition.
    """
    from core.role_utils import is_admin
    from .models import WorkReport
    from core.workorder_tree_views import _record_edit
    if not is_admin(request.user):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)
    wr = get_object_or_404(WorkReport, pk=report_id)
    if not wr.is_pending_repair and not (wr.is_difficult and not wr.is_difficult_resolved):
        return JsonResponse({'success': False, 'message': '该工单非待修/疑难未处理状态'})
    changes = []
    update_fields = []
    if wr.is_pending_repair:
        wr.is_pending_repair = False
        changes.append('待修→已解决')
        update_fields.append('is_pending_repair')
    if wr.is_difficult and not wr.is_difficult_resolved:
        wr.is_difficult_resolved = True
        changes.append('疑难未处理→已处理')
        update_fields.append('is_difficult_resolved')
    note = '；'.join(changes) + f'（由 {request.user.get_username()} 标记）'
    wr.save(update_fields=update_fields)
    _record_edit(wr, request.user, note)
    return JsonResponse({'success': True, 'message': f'已标记 #{wr.id} 为已解决（{"；".join(changes)}）'})


@require_POST
@login_required(login_url='core:login')
def work_report_resolve_all_repair(request):
    """Mark ALL unresolved pending-repair / difficult work orders as resolved (bulk).

    Clears is_pending_repair and (for 疑难 reports) sets is_difficult_resolved.
    Writes an edit-log entry per report.
    """
    from core.role_utils import is_admin
    from .models import WorkReport
    from core.workorder_tree_views import _record_edit
    if not is_admin(request.user):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)
    qs = WorkReport.objects.filter(
        is_pending_repair=True, resolved_by_pm__isnull=True,
    ) | WorkReport.objects.filter(is_difficult=True, is_difficult_resolved=False)
    count = 0
    for wr in qs.distinct():
        changes = []
        update_fields = []
        if wr.is_pending_repair:
            wr.is_pending_repair = False
            changes.append('待修→已解决')
            update_fields.append('is_pending_repair')
        if wr.is_difficult and not wr.is_difficult_resolved:
            wr.is_difficult_resolved = True
            changes.append('疑难未处理→已处理')
            update_fields.append('is_difficult_resolved')
        if not changes:
            continue
        note = '；'.join(changes) + f'（由 {request.user.get_username()} 批量标记）'
        wr.save(update_fields=update_fields)
        _record_edit(wr, request.user, note)
        count += 1
    return JsonResponse({'success': True, 'message': f'已批量解决 {count} 条工单'})


@login_required(login_url='core:login')
def pm_completion_chart(request):
    """AJAX: overdue work-order counts bucketed by scheduled_date.

    Overdue = any GWO whose status is NOT 'completed' (so dispatched/overdue/
    pending/skipped all count — a skipped ticket was overdue before it was
    skipped). Buckets are built in pure Python (no DB Trunc) for SQLite
    compatibility, with zero-filled continuous buckets across the range.

    Query params:
        range: 1m | 3m | 6m | 1y | all  (default 6m)
        granularity: day | week | month | auto (default auto)
            auto picks day (≤2m), week (2-6m), month (>6m) from the range.
    Returns: {labels, counts, granularity, total}
    """
    from core.role_utils import is_admin
    from .models import GeneratedWorkOrder
    from django.utils import timezone
    from datetime import timedelta
    from collections import defaultdict

    if not is_admin(request.user):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)

    today = timezone.localdate()
    rng = request.GET.get('range', '6m')
    granularity = request.GET.get('granularity', 'auto')

    # ── Resolve the date window ─────────────────────────────────────────
    range_days = {'1m': 30, '3m': 90, '6m': 182, '1y': 365}.get(rng, 182)
    if rng == 'all':
        earliest = (GeneratedWorkOrder.objects
                    .order_by('scheduled_date').values_list('scheduled_date', flat=True).first())
        start = earliest or today
    else:
        start = today - timedelta(days=range_days)

    # ── Resolve granularity (auto picks from range) ─────────────────────
    if granularity == 'auto':
        span = (today - start).days
        granularity = 'day' if span <= 60 else ('week' if span <= 182 else 'month')

    # ── Build the continuous bucket sequence (zero-filled) ──────────────
    def _bucket_key(d):
        if granularity == 'day':
            return d
        if granularity == 'week':
            # Monday of that week.
            return d - timedelta(days=d.weekday())
        # month: collapse to first-of-month.
        return d.replace(day=1)

    def _bucket_label(d):
        if granularity == 'day':
            return f'{d.month}/{d.day}'
        if granularity == 'week':
            return f'{d.month}/{d.day}'
        return f'{d.year}-{d.month}'

    buckets = []
    cur = _bucket_key(start)
    end_key = _bucket_key(today)
    # Cap iterations to avoid runaway loops on bad input.
    for _ in range(4000):
        buckets.append(cur)
        if cur == end_key:
            break
        if granularity == 'day':
            cur = cur + timedelta(days=1)
        elif granularity == 'week':
            cur = cur + timedelta(weeks=1)
        else:
            # Next month.
            nm = cur.month % 12 + 1
            ny = cur.year + (1 if cur.month == 12 else 0)
            cur = cur.replace(year=ny, month=nm, day=1)

    counts = defaultdict(int)
    qs = (GeneratedWorkOrder.objects
          .filter(scheduled_date__gte=start, scheduled_date__lte=today)
          .exclude(status='completed')
          .values_list('scheduled_date', flat=True))
    for d in qs:
        counts[_bucket_key(d)] += 1

    labels = [_bucket_label(b) for b in buckets]
    values = [counts.get(b, 0) for b in buckets]
    return JsonResponse({
        'labels': labels,
        'counts': values,
        'granularity': granularity,
        'total': sum(values),
    })


@login_required(login_url='core:login')
def pm_management(request):
    """作业计划管理页 — JobPlan模板 + PM维护计划。"""
    from core.role_utils import is_admin
    from .models import (JobPlanTemplate, MaintenancePlan, Crew, Patch, Satellite)

    if not is_admin(request.user):
        messages.error(request, '无权限访问此页面')
        return redirect('core:dashboard')

    active_tab = request.GET.get('tab', 'jobplans')
    period = request.GET.get('period', 'current')  # current | prev | prev2 | all
    job_plans = JobPlanTemplate.objects.all().order_by('name')
    pm_plans = (MaintenancePlan.objects.all()
                .select_related('job_plan', 'crew', 'patch', 'satellite')
                .prefetch_related(
                    # Only need the count for "N 张" display; the detail rows are
                    # lazy-loaded via AJAX on click to keep the page DOM small.
                    'generated_orders',
                )
                # Order by job_plan FIRST so {% regroup %} in the template produces
                # one group per JobPlan (regroup only merges consecutive runs).
                .order_by('job_plan__name', 'pm_number'))

    # Annotate each plan with its real next due date (first rrule occurrence on/
    # after today) so the "下次到期" column shows the true schedule, not the fixed
    # start_date anchor. Materialize to a list since we attach a runtime attr.
    _today = timezone.localdate()
    pm_plans = list(pm_plans)
    for _p in pm_plans:
        _p.next_due_date = _compute_next_due(_p, _today)

    # 资产视图：把每个 PM 的关联资产按 CCU → 资产条目 聚合，让经理看清
    # "哪些资产被哪些 PM 覆盖"。三种 asset_level 各自解析：
    #   zone_group → 每个 zone 一行（合并同组）
    #   sat        → SAT 一行
    #   ccu        → CCU 一行
    asset_rows = _build_pm_asset_rows(pm_plans)

    # JSON data for the CRUD modals' dropdown options.
    import json
    modal_options = json.dumps({
        'job_plans': [{'id': jp.id, 'name': jp.name} for jp in job_plans],
        'crews': [{'id': c.id, 'name': c.name} for c in Crew.objects.all().order_by('name')],
        'patches': [{'id': p.id, 'code': p.code, 'name': p.name} for p in Patch.objects.order_by('code')],
        'satellites': [{'id': s.id, 'code': s.code, 'name': s.name} for s in Satellite.objects.all().order_by('code')],
        'asset_levels': [{'value': v, 'label': l} for v, l in JobPlanTemplate.ASSET_LEVEL_CHOICES],
        'freq_units': [{'value': v, 'label': l} for v, l in MaintenancePlan.FREQ_UNIT_CHOICES],
    }, ensure_ascii=False)

    # ── 完成情况 tab 数据 ──
    from .models import GeneratedWorkOrder, ExtensionRequest
    from django.utils import timezone as _tz
    from django.db.models import Count, Q
    today = _tz.localdate()
    # Completion stats: each JobPlan's period window is driven by its dominant
    # plan's frequency_value × frequency_unit, anchored to that plan's start_date.
    # The period selector (current/prev/prev2/all) shifts the window back.
    jp_ids_with_plans = [jp.id for jp in job_plans if jp.plans.exists()]
    jp_name_map = {jp.id: jp.name for jp in job_plans}
    # Pre-compute each JobPlan's dominant (frequency_value, frequency_unit, start_date).
    dom_plan = {}
    for row in (MaintenancePlan.objects
                 .filter(job_plan_id__in=jp_ids_with_plans)
                 .values('job_plan_id', 'frequency_value', 'frequency_unit', 'start_date')
                 .annotate(c=Count('id'))
                 .order_by('job_plan_id', '-c')):
        dom_plan.setdefault(row['job_plan_id'], row)
    period_offset = {'current': 0, 'prev': -1, 'prev2': -2}.get(period, 0)
    freq_labels = {'days': '日', 'weeks': '周', 'months': '月'}
    # Aggregate GWOs per JobPlan, applying the per-frequency date window.
    completion_stats = []
    for jp_id in jp_ids_with_plans:
        dp = dom_plan.get(jp_id, {})
        fv = dp.get('frequency_value', 1)
        fu = dp.get('frequency_unit', 'weeks')
        anchor = dp.get('start_date') or today
        if period == 'all':
            start, end, label = None, None, '全部'
        else:
            start, end, label = _pm_period_window(fv, fu, anchor, period_offset, today)
        gwos = GeneratedWorkOrder.objects.filter(plan__job_plan_id=jp_id)
        if start:
            gwos = gwos.filter(scheduled_date__range=(start, end))
        total = gwos.count()
        completed = gwos.filter(status='completed').count()
        dispatched = gwos.filter(status__in=['dispatched', 'overdue']).count()
        overdue = gwos.filter(status__in=['dispatched', 'overdue'], scheduled_date__lt=today).count()
        rate = round(completed * 100 / total) if total else 0
        completion_stats.append({
            'name': jp_name_map.get(jp_id, ''),
            'freq_label': f'每{fv}{freq_labels.get(fu, fu)}',
            'period_label': label or '',
            'total': total, 'completed': completed,
            'dispatched': dispatched, 'overdue': overdue, 'rate': rate,
        })
    completion_stats.sort(key=lambda s: s['name'])
    # Overdue orders: dispatched/overdue GWOs past their scheduled_date.
    overdue_orders = (GeneratedWorkOrder.objects
        .filter(status__in=['dispatched', 'overdue'], scheduled_date__lt=today)
        .select_related('plan__job_plan', 'plan__crew', 'work_report')
        .order_by('scheduled_date')[:100])
    overdue_list = [{
        'gwo_id': g.id, 'pm_number': g.plan.pm_number,
        'job_plan_name': g.plan.job_plan.name if g.plan.job_plan_id else '',
        'scheduled_date': g.scheduled_date, 'days_overdue': (today - g.scheduled_date).days,
        'crew_name': g.plan.crew.name if g.plan.crew_id else '—',
    } for g in overdue_orders]
    # Pending extension requests.
    pending_exts = (ExtensionRequest.objects
        .filter(status='pending')
        .select_related('gwo__plan__job_plan', 'requester')
        .order_by('-created_at')[:50])
    ext_list = [{
        'id': e.id, 'pm_number': e.gwo.plan.pm_number,
        'job_plan_name': e.gwo.plan.job_plan.name if e.gwo.plan.job_plan_id else '',
        'requester_name': e.requester.full_name if e.requester_id else '—',
        'reason': e.reason, 'requested_date': e.requested_date,
        'created_at': e.created_at.strftime('%Y-%m-%d %H:%M') if e.created_at else '',
    } for e in pending_exts]

    context = {
        'active_tab': active_tab,
        'period': period,
        'job_plans': job_plans,
        'pm_plans': pm_plans,
        'crews': Crew.objects.all().order_by('name'),
        'asset_rows': asset_rows,
        'modal_options_json': modal_options,
        'completion_stats': completion_stats,
        'overdue_orders': overdue_list,
        'pending_extensions': ext_list,
    }
    return render(request, 'core/pm_management.html', context)


def _compute_next_due(plan, today):
    """First rrule occurrence on/after today for this plan.

    Mirrors the engine's (generate_pm_workorders) scheduling math so the UI's
    "下次到期" column shows the real next due date, not the fixed start_date
    anchor. Returns None if frequency_unit is unknown or no future occurrence
    exists within 730 days.
    """
    from datetime import timedelta
    from dateutil.rrule import rrule, DAILY, WEEKLY, MONTHLY
    freq_map = {'days': DAILY, 'weeks': WEEKLY, 'months': MONTHLY}
    freq = freq_map.get(plan.frequency_unit)
    if not freq or not plan.start_date:
        return None
    occs = list(rrule(freq, interval=plan.frequency_value,
                      dtstart=plan.start_date, until=today + timedelta(days=730)))
    for o in occs:
        d = o.date() if hasattr(o, 'date') else o
        if d >= today:
            return d
    return None


def _build_pm_asset_rows(pm_plans):
    """Aggregate PM→asset links into rows grouped by CCU for the 资产 tab.

    Returns a list of plain dicts:
        {'ccu_code', 'ccu_name', 'asset_label', 'asset_type', 'pm_list'}
    where pm_list is a list of (pm_number, job_plan_name, frequency) tuples.
    Only assets that ARE linked to a PM appear (2508 zones are not listed).
    """
    from .models import Patch
    patches_by_id = {p.id: p for p in Patch.objects.all()}
    rows = []
    for plan in pm_plans:
        level = plan.job_plan.asset_level
        pm_info = (plan.pm_number, plan.job_plan.name,
                   f'{plan.frequency_value}{plan.get_frequency_unit_display()}')
        if level == 'sat' and plan.satellite_id:
            sat = plan.satellite
            ccu = sat.patch
            rows.append({
                'ccu_code': ccu.code if ccu else '—', 'ccu_name': ccu.name if ccu else '',
                'asset_label': f'SAT {sat.code} {sat.name}'.strip(),
                'asset_type': 'SAT', 'pm_list': [pm_info],
            })
        elif level == 'ccu' and plan.patch_id:
            ccu = patches_by_id.get(plan.patch_id)
            rows.append({
                'ccu_code': ccu.code if ccu else '—', 'ccu_name': ccu.name if ccu else '',
                'asset_label': f'CCU {ccu.code} {ccu.name}'.strip() if ccu else 'CCU',
                'asset_type': 'CCU', 'pm_list': [pm_info],
            })
        else:
            # zone_group (default): one row per zone, grouped by CCU.
            for z in plan.zones.all().select_related('patch'):
                ccu = z.patch
                rows.append({
                    'ccu_code': ccu.code if ccu else '—', 'ccu_name': ccu.name if ccu else '',
                    'asset_label': f'{z.code} {z.name}'.strip(),
                    'asset_type': 'Zone', 'pm_list': [pm_info],
                })
    # Merge rows with the same (ccu_code, asset_label) so a zone covered by
    # multiple PMs shows them together rather than as duplicate rows.
    merged = {}
    for r in rows:
        key = (r['ccu_code'], r['asset_label'])
        if key in merged:
            merged[key]['pm_list'].extend(r['pm_list'])
        else:
            merged[key] = r
    return sorted(merged.values(), key=lambda x: (x['ccu_code'], x['asset_label']))


def _get_patches_by_region():
    """Return patches grouped by region for the settings page."""
    from .models import Region, Patch
    result = []
    for region in Region.objects.filter(active=True).order_by('order', 'name'):
        patches = Patch.objects.filter(region=region).order_by('code')
        if patches.exists():
            result.append({
                'region': region,
                'patches': patches,
            })
    # Include unassigned patches
    unassigned = Patch.objects.filter(region__isnull=True).order_by('code')
    if unassigned.exists():
        result.append({
            'region': None,
            'patches': unassigned,
        })
    return result


@login_required(login_url='core:login')
def zone_edit(request, zone_id):
    """
    Edit a specific zone - admin only.
    """
    from .models import ManagerProfile, Plant, ZoneEquipment, Patch

    # Check admin permission
    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        if _wants_json(request):
            return JsonResponse({'success': False, 'message': '无权限修改区域'}, status=403)
        messages.error(request, '无权限修改区域')
        return redirect('core:dashboard')

    zone = get_object_or_404(Zone, pk=zone_id)

    if request.method == 'POST':
        zone.name = request.POST.get('name', zone.name)
        zone.code = request.POST.get('code', zone.code)
        zone.description = request.POST.get('description', zone.description)
        zone.boundary_color = request.POST.get('boundary_color', zone.boundary_color)
        zone.priority = request.POST.get('priority', zone.priority)

        # Irrigation attributes
        zone.current_status = request.POST.get('current_status', '')
        zone.sprinkler_type = request.POST.get('sprinkler_type', '')
        zone.irrigation_intensity = _parse_float(request.POST.get('irrigation_intensity'))
        zone.solenoid_valve_size = _parse_float(request.POST.get('solenoid_valve_size'))
        zone.landscape_coefficient = _parse_float(request.POST.get('landscape_coefficient'))
        zone.plant_type = request.POST.get('plant_type', '')
        zone.soil_moisture = request.POST.get('soil_moisture', '')
        zone.terrain_feature = request.POST.get('terrain_feature', '')
        zone.plant_feature = request.POST.get('plant_feature', '')
        zone.irrigation_foreman = request.POST.get('irrigation_foreman', '')
        zone.greenery_zone = request.POST.get('greenery_zone', '')
        zone.greenery_foreman = request.POST.get('greenery_foreman', '')
        zone.pest_control_zone = request.POST.get('pest_control_zone', '')
        zone.pest_control_foreman = request.POST.get('pest_control_foreman', '')
        zone.equipment_maintenance_notes = request.POST.get('equipment_maintenance_notes', '')
        zone.irrigation_management_notes = request.POST.get('irrigation_management_notes', '')

        # Label settings
        label_lat = request.POST.get('label_lat', '')
        label_lng = request.POST.get('label_lng', '')
        zone.label_lat = float(label_lat) if label_lat else None
        zone.label_lng = float(label_lng) if label_lng else None
        zone.label_scale = float(request.POST.get('label_scale', 1.0) or 1.0)
        zone.label_angle = int(request.POST.get('label_angle', 0) or 0)

        # Smooth override
        smooth_val = request.POST.get('smooth_override', '')
        zone.smooth_override = int(smooth_val) if smooth_val != '' else None

        # Handle patch selection
        patch_id = request.POST.get('patch')
        new_patch_name = request.POST.get('new_patch_name', '').strip()

        if new_patch_name and not patch_id:
            # Create new patch
            patch_code = new_patch_name[:10].replace(' ', '-')
            patch, created = Patch.objects.get_or_create(
                name=new_patch_name,
                defaults={'code': patch_code}
            )
            zone.patch = patch
        elif patch_id:
            zone.patch_id = patch_id
        else:
            zone.patch = None

        # Parse boundary points from JSON
        boundary_json = request.POST.get('boundary_points', '[]')
        try:
            boundary_points = json.loads(boundary_json)
            # Auto-close any incomplete polygons (points defined but not explicitly "completed")
            zone.boundary_points = auto_close_boundary_points(boundary_points)
        except json.JSONDecodeError:
            if _wants_json(request):
                return JsonResponse({'success': False, 'message': '边界点数据格式有误，请重新绘制区域边界后重试'}, status=400)
            messages.error(request, '边界点数据格式有误，请重新绘制区域边界后重试')
            return redirect('core:zone_edit', zone_id=zone.id)

        zone.drawn_by = request.user
        zone.save()

        # Handle plants - JSON array in hidden field
        plants_data = request.POST.get('plants_data', '')
        if plants_data:
            try:
                plants_list = json.loads(plants_data)
                zone.plants.all().delete()
                for item in plants_list:
                    Plant.objects.create(
                        zone=zone,
                        name=item.get('name', ''),
                        quantity=item.get('quantity', 1),
                        planting_date=item.get('planting_date') or None,
                        end_date=item.get('end_date') or None,
                        notes=item.get('notes', ''),
                    )
            except json.JSONDecodeError:
                pass

        # Handle equipment - JSON array in hidden field
        equipment_data = request.POST.get('equipment_data', '')
        if equipment_data:
            try:
                equipment_list = json.loads(equipment_data)
                zone.equipments.all().delete()
                for item in equipment_list:
                    equipment_id = item.get('equipment_id')
                    if equipment_id:
                        ZoneEquipment.objects.create(
                            zone=zone,
                            equipment_id=equipment_id,
                            quantity=item.get('quantity', 1),
                            installation_date=item.get('installation_date') or None,
                            status=item.get('status', 'working'),
                            location_in_zone=item.get('location_in_zone', ''),
                            notes=item.get('notes', '')
                        )
            except json.JSONDecodeError:
                pass

        return _zone_save_response(request, zone, f'区域「{zone.name}」已保存成功')

    # Get available plants (distinct names from all plants)
    available_plants = list(Plant.objects.values_list('name', flat=True).distinct().order_by('name'))

    # Get zone equipment
    zone_equipments = zone.equipments.select_related('equipment').all()

    # Get all patches for selection
    patches = Patch.objects.all()

    ref_zones_json, ref_pipelines_json = _get_reference_map_data(exclude_zone_id=zone.id)

    # Sibling zones: other zones in the same patch (for sidebar navigation)
    sibling_zones = []
    if zone.patch_id:
        sibling_zones = list(
            Zone.objects.filter(patch_id=zone.patch_id)
            .exclude(pk=zone.pk)
            .values('id', 'name', 'code', 'priority', 'current_status')
            .order_by('code')
        )

    context = {
        'zone': zone,
        'boundary_json': json.dumps(zone.boundary_points),
        'available_plants': available_plants,
        'zone_equipments': zone_equipments,
        'patches': patches,
        'ref_zones_json': ref_zones_json,
        'ref_pipelines_json': ref_pipelines_json,
        'sibling_zones': sibling_zones,
        'equip_notes_json': json.dumps(json.loads(zone.equipment_maintenance_notes)) if zone.equipment_maintenance_notes else '[]',
        'irrig_notes_json': json.dumps(json.loads(zone.irrigation_management_notes)) if zone.irrigation_management_notes else '[]',
        'boundary_count': len(zone.boundary_points) if zone.boundary_points else 0,
        'equip_notes_count': len(json.loads(zone.equipment_maintenance_notes)) if zone.equipment_maintenance_notes else 0,
        'irrig_notes_count': len(json.loads(zone.irrigation_management_notes)) if zone.irrigation_management_notes else 0,
        'plant_count': zone.plants.count(),
        'equipment_count': zone.equipments.count(),
        **_get_zone_dropdown_options(),
    }

    return render(request, 'core/zone_form.html', context)


@login_required(login_url='core:login')
def zone_new(request):
    """
    Create a new zone - admin only.
    """
    from .models import ManagerProfile, Plant, ZoneEquipment, Patch

    # Check admin permission
    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        if _wants_json(request):
            return JsonResponse({'success': False, 'message': '无权限创建区域'}, status=403)
        messages.error(request, '无权限创建区域')
        return redirect('core:dashboard')

    if request.method == 'POST':
        zone = Zone(
            name=request.POST.get('name'),
            code=request.POST.get('code'),
            description=request.POST.get('description', ''),
            boundary_color=request.POST.get('boundary_color', '#52B788'),
            priority=request.POST.get('priority', Zone.PRIORITY_MEDIUM),
            current_status=request.POST.get('current_status', ''),
            sprinkler_type=request.POST.get('sprinkler_type', ''),
            irrigation_intensity=_parse_float(request.POST.get('irrigation_intensity')),
            solenoid_valve_size=_parse_float(request.POST.get('solenoid_valve_size')),
            landscape_coefficient=_parse_float(request.POST.get('landscape_coefficient')),
            plant_type=request.POST.get('plant_type', ''),
            soil_moisture=request.POST.get('soil_moisture', ''),
            terrain_feature=request.POST.get('terrain_feature', ''),
            plant_feature=request.POST.get('plant_feature', ''),
            irrigation_foreman=request.POST.get('irrigation_foreman', ''),
            greenery_zone=request.POST.get('greenery_zone', ''),
            greenery_foreman=request.POST.get('greenery_foreman', ''),
            pest_control_zone=request.POST.get('pest_control_zone', ''),
            pest_control_foreman=request.POST.get('pest_control_foreman', ''),
            equipment_maintenance_notes=request.POST.get('equipment_maintenance_notes', ''),
            irrigation_management_notes=request.POST.get('irrigation_management_notes', ''),
        )

        # Label settings
        label_lat = request.POST.get('label_lat', '')
        label_lng = request.POST.get('label_lng', '')
        zone.label_lat = float(label_lat) if label_lat else None
        zone.label_lng = float(label_lng) if label_lng else None
        zone.label_scale = float(request.POST.get('label_scale', 1.0) or 1.0)
        zone.label_angle = int(request.POST.get('label_angle', 0) or 0)

        # Smooth override
        smooth_val = request.POST.get('smooth_override', '')
        zone.smooth_override = int(smooth_val) if smooth_val != '' else None

        # Handle patch selection
        patch_id = request.POST.get('patch')
        new_patch_name = request.POST.get('new_patch_name', '').strip()

        if new_patch_name and not patch_id:
            # Create new patch
            patch_code = new_patch_name[:10].replace(' ', '-')
            patch, created = Patch.objects.get_or_create(
                name=new_patch_name,
                defaults={'code': patch_code}
            )
            zone.patch = patch
        elif patch_id:
            zone.patch_id = patch_id

        # Parse boundary points from JSON
        boundary_json = request.POST.get('boundary_points', '[]')
        try:
            boundary_points = json.loads(boundary_json)
            # Auto-close any incomplete polygons (points defined but not explicitly "completed")
            zone.boundary_points = auto_close_boundary_points(boundary_points)
        except json.JSONDecodeError:
            if _wants_json(request):
                return JsonResponse({'success': False, 'message': '边界点数据格式有误，请重新绘制区域边界后重试'}, status=400)
            messages.error(request, '边界点数据格式有误，请重新绘制区域边界后重试')
            return render(request, 'core/zone_form.html', {
                'zone': zone,
                'boundary_json': boundary_json,
            })

        zone.drawn_by = request.user
        zone.save()

        # Handle plants - JSON array with full details
        plants_data = request.POST.get('plants_data', '')
        if plants_data:
            try:
                plants_list = json.loads(plants_data)
                for item in plants_list:
                    Plant.objects.create(
                        zone=zone,
                        name=item.get('name', ''),
                        quantity=item.get('quantity', 1),
                        planting_date=item.get('planting_date') or None,
                        end_date=item.get('end_date') or None,
                        notes=item.get('notes', ''),
                    )
            except json.JSONDecodeError:
                # Fallback to comma-separated names
                plant_names = [p.strip() for p in plants_data.split(',') if p.strip()]
                for name in plant_names:
                    Plant.objects.create(zone=zone, name=name)

        # Handle equipment - JSON array in hidden field
        equipment_data = request.POST.get('equipment_data', '')
        if equipment_data:
            try:
                equipment_list = json.loads(equipment_data)
                for item in equipment_list:
                    equipment_id = item.get('equipment_id')
                    if equipment_id:
                        ZoneEquipment.objects.create(
                            zone=zone,
                            equipment_id=equipment_id,
                            quantity=item.get('quantity', 1),
                            installation_date=item.get('installation_date') or None,
                            status=item.get('status', 'working'),
                            location_in_zone=item.get('location_in_zone', ''),
                            notes=item.get('notes', '')
                        )
            except json.JSONDecodeError:
                pass

        return _zone_save_response(request, zone, f'区域「{zone.name}」已创建成功', created=True)

    # Get available plants
    available_plants = list(Plant.objects.values_list('name', flat=True).distinct().order_by('name'))

    # Get all patches for selection
    patches = Patch.objects.all()

    ref_zones_json, ref_pipelines_json = _get_reference_map_data()

    context = {
        'zone': None,
        'boundary_json': '[]',
        'available_plants': available_plants,
        'patches': patches,
        'ref_zones_json': ref_zones_json,
        'ref_pipelines_json': ref_pipelines_json,
        'equip_notes_json': '[]',
        'irrig_notes_json': '[]',
        **_get_zone_dropdown_options(),
    }

    return render(request, 'core/zone_form.html', context)


@require_POST
@login_required(login_url='core:login')
def zone_delete(request, zone_id):
    """
    Delete a zone - admin only.
    """
    from .models import ManagerProfile

    # Check admin permission
    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        messages.error(request, '无权限删除区域')
        return redirect('core:dashboard')

    zone = get_object_or_404(Zone, pk=zone_id)
    zone_name = zone.name
    zone.delete()
    messages.success(request, f'Zone "{zone_name}" deleted successfully.')
    return redirect('core:settings')


@login_required(login_url='core:login')
def zone_batch_draw(request):
    """Batch zone boundary drawing page."""
    from .models import ManagerProfile, Patch

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        messages.error(request, '无权限访问此页面')
        return redirect('core:dashboard')

    if request.method == 'POST':
        new_zone = request.POST.get('new_zone') == 'true'
        if new_zone:
            patch_id = request.POST.get('patch_id')
            code = request.POST.get('code', '').strip()
            name = request.POST.get('name', '').strip()
            if not patch_id:
                return JsonResponse({'success': False, 'message': '缺少片区ID'}, status=400)
            if not code:
                return JsonResponse({'success': False, 'message': '请输入区域编号'}, status=400)
            if Zone.objects.filter(code=code).exists():
                return JsonResponse({'success': False, 'message': f'区域编号 {code} 已存在'}, status=400)
            patch = get_object_or_404(Patch, pk=int(patch_id))
            zone = Zone.objects.create(
                patch=patch, code=code, name=name or code,
                priority='medium', boundary_points=[], boundary_color='#52B788'
            )
        else:
            zone_id = request.POST.get('zone_id')
            if not zone_id:
                return JsonResponse({'success': False, 'message': '缺少区域ID'}, status=400)
            zone = get_object_or_404(Zone, pk=int(zone_id))

        boundary_raw = request.POST.get('boundary_points', '[]')
        try:
            boundary_data = json.loads(boundary_raw)
        except (json.JSONDecodeError, TypeError):
            boundary_data = []

        boundary_data = auto_close_boundary_points(boundary_data)
        zone.boundary_points = boundary_data

        label_lat = request.POST.get('label_lat', '')
        label_lng = request.POST.get('label_lng', '')
        zone.label_lat = float(label_lat) if label_lat else None
        zone.label_lng = float(label_lng) if label_lng else None
        zone.label_scale = float(request.POST.get('label_scale', '1.0') or '1.0')
        zone.label_angle = int(request.POST.get('label_angle', '0') or '0')
        smooth_val = request.POST.get('smooth_override', '')
        zone.smooth_override = int(smooth_val) if smooth_val != '' else None

        zone.drawn_by = request.user
        zone.save()

        return JsonResponse({
            'success': True,
            'message': f'区域 {zone.code} 边界已保存',
            'zone_id': zone.id,
            'zone_name': zone.name,
            'zone_code': zone.code,
            'area_display': zone.area_display,
            'boundary_count': len(zone.boundary_points),
            'boundary_points': zone.boundary_points,
            'boundary_color': zone.boundary_color,
            'label_lat': zone.label_lat,
            'label_lng': zone.label_lng,
            'label_scale': zone.label_scale,
            'label_angle': zone.label_angle,
            'smooth_override': zone.smooth_override,
            'patch_id': zone.patch_id,
            'is_new': new_zone if request.method == 'POST' else False,
        })

    patches = Patch.objects.all().order_by('code')

    # All zones with boundaries for reference layer on map
    all_drawn_zones = []
    for z in Zone.objects.select_related('patch').only(
        'id', 'code', 'name', 'boundary_points', 'boundary_color',
        'label_lat', 'label_lng', 'label_scale', 'label_angle', 'smooth_override', 'patch_id'
    ):
        all_drawn_zones.append({
            'id': z.id,
            'code': z.code,
            'name': z.name,
            'boundary_points': z.active_boundary_points,
            'boundary_color': z.boundary_color,
            'label_lat': z.label_lat,
            'label_lng': z.label_lng,
            'label_scale': z.label_scale,
            'label_angle': z.label_angle,
            'smooth_override': z.smooth_override,
            'patch_id': z.patch_id,
        })

    context = {
        'patches': patches,
        'all_drawn_zones_json': json.dumps(all_drawn_zones),
        'nav_settings': True,
    }
    return render(request, 'core/zone_batch_draw.html', context)


@login_required(login_url='core:login')
def zone_batch_draw_zones_api(request):
    """API: return zones for a given patch_id."""
    from .models import ManagerProfile

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        return JsonResponse({'error': '无权限'}, status=403)

    patch_id = request.GET.get('patch_id')
    if not patch_id:
        return JsonResponse({'error': '缺少patch_id参数'}, status=400)

    zones = Zone.objects.filter(patch_id=patch_id).order_by('code')
    result = []
    for z in zones:
        has_boundary = bool(z.active_boundary_points)
        result.append({
            'id': z.id,
            'code': z.code,
            'name': z.name,
            'has_boundary': has_boundary,
            'boundary_points': z.active_boundary_points if has_boundary else [],
            'boundary_color': z.boundary_color,
            'label_lat': z.label_lat,
            'label_lng': z.label_lng,
            'label_scale': z.label_scale,
            'label_angle': z.label_angle,
            'smooth_override': z.smooth_override,
            'area_display': z.area_display if has_boundary else '',
        })
    return JsonResponse({'zones': result})


@login_required(login_url='core:login')
@ensure_csrf_cookie
def zone_quick_draw(request):
    """Quick zone boundary drawing page — draw first, assign zone code after."""
    from .models import ManagerProfile, Worker, MapStyleSettings

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        try:
            Worker.objects.get(user=request.user, active=True)
        except Worker.DoesNotExist:
            messages.error(request, '无权限访问此页面')
            return redirect('core:dashboard')

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'delete_boundary':
            zone_code = request.POST.get('zone_code', '').strip()
            if not zone_code:
                return JsonResponse({'success': False, 'message': '缺少区域编号'}, status=400)
            try:
                zone = Zone.objects.get(code=zone_code)
                zone.boundary_points = []
                zone.label_lat = None
                zone.label_lng = None
                zone.drawn_by = None
                zone.save()
                return JsonResponse({'success': True, 'message': f'区域 {zone.code} 边界已删除'})
            except Zone.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'区域 {zone_code} 不存在'}, status=404)

        if action == 'update_color':
            zone_code = request.POST.get('zone_code', '').strip()
            boundary_color = request.POST.get('boundary_color', '').strip()
            if not zone_code:
                return JsonResponse({'success': False, 'message': '缺少区域编号'}, status=400)
            if not boundary_color:
                return JsonResponse({'success': False, 'message': '缺少颜色值'}, status=400)
            try:
                zone = Zone.objects.get(code=zone_code)
                zone.boundary_color = boundary_color
                zone.drawn_by = request.user
                zone.save()
                return JsonResponse({'success': True, 'message': f'区域 {zone.code} 颜色已更新', 'boundary_color': zone.boundary_color})
            except Zone.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'区域 {zone_code} 不存在'}, status=404)

        if action == 'update_ring_display_mode':
            zone_code = request.POST.get('zone_code', '').strip()
            ring_index = request.POST.get('ring_index', '')
            mode = request.POST.get('mode', 'line')
            if not zone_code:
                return JsonResponse({'success': False, 'message': '缺少区域编号'}, status=400)
            if mode not in ('line', 'sublabel'):
                return JsonResponse({'success': False, 'message': '无效的显示模式'}, status=400)
            try:
                zone = Zone.objects.get(code=zone_code)
                modes = dict(zone.ring_display_modes or {})
                if mode == 'line':
                    modes.pop(str(ring_index), None)
                else:
                    modes[str(ring_index)] = mode
                # Prune stale indices beyond ring count
                bp = zone.boundary_points or []
                ring_count = len(bp) if bp and isinstance(bp[0], (list,)) else 0
                if ring_count > 0:
                    modes = {k: v for k, v in modes.items() if int(k) < ring_count}
                zone.ring_display_modes = modes
                zone.save()
                mode_label = '引导线' if mode == 'line' else '子标签'
                return JsonResponse({'success': True, 'message': f'已切换为{mode_label}', 'ring_display_modes': modes})
            except Zone.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'区域 {zone_code} 不存在'}, status=404)

        if action == 'update_all_ring_display_modes':
            zone_code = request.POST.get('zone_code', '').strip()
            mode = request.POST.get('mode', 'line')
            if not zone_code:
                return JsonResponse({'success': False, 'message': '缺少区域编号'}, status=400)
            if mode not in ('line', 'sublabel'):
                return JsonResponse({'success': False, 'message': '无效的显示模式'}, status=400)
            try:
                zone = Zone.objects.get(code=zone_code)
                bp = zone.boundary_points or []
                ring_count = len(bp) if bp and isinstance(bp[0], (list,)) else 0
                if mode == 'line':
                    modes = {}
                else:
                    modes = {str(i): mode for i in range(ring_count)}
                zone.ring_display_modes = modes
                zone.save()
                mode_label = '引导线' if mode == 'line' else '子标签'
                return JsonResponse({'success': True, 'message': f'已全部切换为{mode_label}', 'ring_display_modes': modes})
            except Zone.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'区域 {zone_code} 不存在'}, status=404)

        # Save boundary action
        zone_code = request.POST.get('zone_code', '').strip()
        if not zone_code:
            return JsonResponse({'success': False, 'message': '缺少区域编号'}, status=400)

        try:
            zone = Zone.objects.get(code=zone_code)
        except Zone.DoesNotExist:
            return JsonResponse({'success': False, 'message': f'区域编号 {zone_code} 不存在，请检查输入'}, status=404)

        boundary_raw = request.POST.get('boundary_points', '[]')
        try:
            boundary_data = json.loads(boundary_raw)
        except (json.JSONDecodeError, TypeError):
            boundary_data = []

        boundary_data = auto_close_boundary_points(boundary_data)
        zone.boundary_points = boundary_data

        label_lat = request.POST.get('label_lat', '')
        label_lng = request.POST.get('label_lng', '')
        zone.label_lat = float(label_lat) if label_lat else None
        zone.label_lng = float(label_lng) if label_lng else None
        zone.label_scale = float(request.POST.get('label_scale', '1.0') or '1.0')
        zone.label_angle = int(request.POST.get('label_angle', '0') or '0')
        smooth_val = request.POST.get('smooth_override', '')
        zone.smooth_override = int(smooth_val) if smooth_val != '' else None
        boundary_color = request.POST.get('boundary_color')
        if boundary_color:
            zone.boundary_color = boundary_color
        zone.drawn_by = request.user

        zone.save()

        patch_name = zone.patch.name if zone.patch else ''
        return JsonResponse({
            'success': True,
            'message': f'区域 {zone.code} 边界已保存',
            'zone_id': zone.id,
            'zone_name': zone.name,
            'zone_code': zone.code,
            'area_display': zone.area_display,
            'boundary_count': len(zone.boundary_points),
            'boundary_points': zone.boundary_points,
            'boundary_color': zone.boundary_color,
            'label_lat': zone.label_lat,
            'label_lng': zone.label_lng,
            'label_scale': zone.label_scale,
            'label_angle': zone.label_angle,
            'smooth_override': zone.smooth_override,
            'ring_display_modes': zone.ring_display_modes or {},
            'patch_id': zone.patch_id,
            'patch_name': patch_name,
        })

    # GET — render page with all zones data
    all_zones = []
    for z in Zone.objects.select_related('patch').order_by('code').only(
        'id', 'code', 'name', 'patch_id', 'patch__name'
    ):
        all_zones.append({
            'id': z.id,
            'code': z.code,
            'name': z.name,
            'patch_name': z.patch.name if z.patch else '',
        })

    # All zones with boundaries for reference layer on map
    all_drawn_zones = []
    for z in Zone.objects.select_related('patch').only(
        'id', 'code', 'name', 'boundary_points', 'boundary_source', 'dxf_boundary_points', 'boundary_color',
        'label_lat', 'label_lng', 'label_scale', 'label_angle', 'smooth_override',
        'ring_display_modes', 'patch_id', 'patch__name', 'area_sqm'
    ):
        all_drawn_zones.append({
            'id': z.id,
            'code': z.code,
            'name': z.name,
            'boundary_points': z.active_boundary_points,
            'boundary_color': z.boundary_color,
            'label_lat': z.label_lat,
            'label_lng': z.label_lng,
            'label_scale': z.label_scale,
            'label_angle': z.label_angle,
            'smooth_override': z.smooth_override,
            'ring_display_modes': z.ring_display_modes or {},
            'patch_id': z.patch_id,
            'patch_name': z.patch.name if z.patch else '',
            'area_display': z.area_display,
        })

    # Drawing stats per patch
    from .models import Patch
    patch_stats = []
    for p in Patch.objects.order_by('name'):
        total = Zone.objects.filter(patch=p).count()
        drawn = Zone.objects.filter(patch=p).exclude(boundary_points__isnull=True).exclude(boundary_points=[]).count() + Zone.objects.filter(patch=p, boundary_source='dxf').exclude(dxf_boundary_points__isnull=True).exclude(dxf_boundary_points=[]).count()
        patch_stats.append({'name': p.name, 'total': total, 'drawn': drawn})
    total_all = Zone.objects.count()
    drawn_all = Zone.objects.exclude(boundary_points__isnull=True).exclude(boundary_points=[]).count() + Zone.objects.filter(boundary_source='dxf').exclude(dxf_boundary_points__isnull=True).exclude(dxf_boundary_points=[]).count()

    context = {
        'all_zones_json': json.dumps(all_zones),
        'all_drawn_zones_json': json.dumps(all_drawn_zones),
        'map_style_json': json.dumps(MapStyleSettings.get_style()),
        'patch_stats_json': json.dumps(patch_stats),
        'drawn_total': drawn_all,
        'zone_total': total_all,
        'nav_settings': True,
    }
    return render(request, 'core/zone_quick_draw.html', context)


@login_required(login_url='core:login')
@ensure_csrf_cookie
def zone_quick_draw_mobile(request):
    """Mobile-optimized quick zone boundary drawing page.
    Accessible to all authenticated users."""
    from .models import MapStyleSettings

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'delete_boundary':
            zone_code = request.POST.get('zone_code', '').strip()
            if not zone_code:
                return JsonResponse({'success': False, 'message': '缺少区域编号'}, status=400)
            try:
                zone = Zone.objects.get(code=zone_code)
                zone.boundary_points = []
                zone.label_lat = None
                zone.label_lng = None
                zone.drawn_by = None
                zone.save()
                return JsonResponse({'success': True, 'message': f'区域 {zone.code} 边界已删除'})
            except Zone.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'区域 {zone_code} 不存在'}, status=404)

        if action == 'update_color':
            zone_code = request.POST.get('zone_code', '').strip()
            boundary_color = request.POST.get('boundary_color', '').strip()
            if not zone_code:
                return JsonResponse({'success': False, 'message': '缺少区域编号'}, status=400)
            if not boundary_color:
                return JsonResponse({'success': False, 'message': '缺少颜色值'}, status=400)
            try:
                zone = Zone.objects.get(code=zone_code)
                zone.boundary_color = boundary_color
                zone.drawn_by = request.user
                zone.save()
                return JsonResponse({'success': True, 'message': f'区域 {zone.code} 颜色已更新', 'boundary_color': zone.boundary_color})
            except Zone.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'区域 {zone_code} 不存在'}, status=404)

        if action == 'update_ring_display_mode':
            zone_code = request.POST.get('zone_code', '').strip()
            ring_index = request.POST.get('ring_index', '')
            mode = request.POST.get('mode', 'line')
            if not zone_code:
                return JsonResponse({'success': False, 'message': '缺少区域编号'}, status=400)
            if mode not in ('line', 'sublabel'):
                return JsonResponse({'success': False, 'message': '无效的显示模式'}, status=400)
            try:
                zone = Zone.objects.get(code=zone_code)
                modes = dict(zone.ring_display_modes or {})
                if mode == 'line':
                    modes.pop(str(ring_index), None)
                else:
                    modes[str(ring_index)] = mode
                bp = zone.boundary_points or []
                ring_count = len(bp) if bp and isinstance(bp[0], (list,)) else 0
                if ring_count > 0:
                    modes = {k: v for k, v in modes.items() if int(k) < ring_count}
                zone.ring_display_modes = modes
                zone.save()
                mode_label = '引导线' if mode == 'line' else '子标签'
                return JsonResponse({'success': True, 'message': f'已切换为{mode_label}', 'ring_display_modes': modes})
            except Zone.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'区域 {zone_code} 不存在'}, status=404)

        if action == 'update_all_ring_display_modes':
            zone_code = request.POST.get('zone_code', '').strip()
            mode = request.POST.get('mode', 'line')
            if not zone_code:
                return JsonResponse({'success': False, 'message': '缺少区域编号'}, status=400)
            if mode not in ('line', 'sublabel'):
                return JsonResponse({'success': False, 'message': '无效的显示模式'}, status=400)
            try:
                zone = Zone.objects.get(code=zone_code)
                bp = zone.boundary_points or []
                ring_count = len(bp) if bp and isinstance(bp[0], (list,)) else 0
                if mode == 'line':
                    modes = {}
                else:
                    modes = {str(i): mode for i in range(ring_count)}
                zone.ring_display_modes = modes
                zone.save()
                mode_label = '引导线' if mode == 'line' else '子标签'
                return JsonResponse({'success': True, 'message': f'已全部切换为{mode_label}', 'ring_display_modes': modes})
            except Zone.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'区域 {zone_code} 不存在'}, status=404)

        zone_code = request.POST.get('zone_code', '').strip()
        if not zone_code:
            return JsonResponse({'success': False, 'message': '缺少区域编号'}, status=400)

        try:
            zone = Zone.objects.get(code=zone_code)
        except Zone.DoesNotExist:
            return JsonResponse({'success': False, 'message': f'区域编号 {zone_code} 不存在，请检查输入'}, status=404)

        boundary_raw = request.POST.get('boundary_points', '[]')
        try:
            boundary_data = json.loads(boundary_raw)
        except (json.JSONDecodeError, TypeError):
            boundary_data = []

        boundary_data = auto_close_boundary_points(boundary_data)
        zone.boundary_points = boundary_data

        label_lat = request.POST.get('label_lat', '')
        label_lng = request.POST.get('label_lng', '')
        zone.label_lat = float(label_lat) if label_lat else None
        zone.label_lng = float(label_lng) if label_lng else None
        zone.label_scale = float(request.POST.get('label_scale', '1.0') or '1.0')
        zone.label_angle = int(request.POST.get('label_angle', '0') or '0')
        smooth_val = request.POST.get('smooth_override', '')
        zone.smooth_override = int(smooth_val) if smooth_val != '' else None
        boundary_color = request.POST.get('boundary_color')
        if boundary_color:
            zone.boundary_color = boundary_color
        zone.drawn_by = request.user

        zone.save()

        patch_name = zone.patch.name if zone.patch else ''
        return JsonResponse({
            'success': True,
            'message': f'区域 {zone.code} 边界已保存',
            'zone_id': zone.id,
            'zone_name': zone.name,
            'zone_code': zone.code,
            'area_display': zone.area_display,
            'boundary_count': len(zone.boundary_points),
            'boundary_points': zone.boundary_points,
            'boundary_color': zone.boundary_color,
            'label_lat': zone.label_lat,
            'label_lng': zone.label_lng,
            'label_scale': zone.label_scale,
            'label_angle': zone.label_angle,
            'smooth_override': zone.smooth_override,
            'ring_display_modes': zone.ring_display_modes or {},
            'patch_id': zone.patch_id,
            'patch_name': patch_name,
        })

    all_zones = []
    for z in Zone.objects.select_related('patch').order_by('code').only(
        'id', 'code', 'name', 'patch_id', 'patch__name'
    ):
        all_zones.append({
            'id': z.id,
            'code': z.code,
            'name': z.name,
            'patch_name': z.patch.name if z.patch else '',
        })

    all_drawn_zones = []
    for z in Zone.objects.select_related('patch').only(
        'id', 'code', 'name', 'boundary_points', 'boundary_source', 'dxf_boundary_points', 'boundary_color',
        'label_lat', 'label_lng', 'label_scale', 'label_angle', 'smooth_override',
        'ring_display_modes', 'patch_id', 'patch__name', 'area_sqm'
    ):
        all_drawn_zones.append({
            'id': z.id,
            'code': z.code,
            'name': z.name,
            'boundary_points': z.active_boundary_points,
            'boundary_color': z.boundary_color,
            'label_lat': z.label_lat,
            'label_lng': z.label_lng,
            'label_scale': z.label_scale,
            'label_angle': z.label_angle,
            'smooth_override': z.smooth_override,
            'ring_display_modes': z.ring_display_modes or {},
            'patch_id': z.patch_id,
            'patch_name': z.patch.name if z.patch else '',
            'area_display': z.area_display,
        })

    # Drawing stats per patch
    from .models import Patch
    patch_stats = []
    for p in Patch.objects.order_by('name'):
        total = Zone.objects.filter(patch=p).count()
        drawn = Zone.objects.filter(patch=p).exclude(boundary_points__isnull=True).exclude(boundary_points=[]).count() + Zone.objects.filter(patch=p, boundary_source='dxf').exclude(dxf_boundary_points__isnull=True).exclude(dxf_boundary_points=[]).count()
        patch_stats.append({'name': p.name, 'total': total, 'drawn': drawn})
    total_all = Zone.objects.count()
    drawn_all = Zone.objects.exclude(boundary_points__isnull=True).exclude(boundary_points=[]).count() + Zone.objects.filter(boundary_source='dxf').exclude(dxf_boundary_points__isnull=True).exclude(dxf_boundary_points=[]).count()

    context = {
        'all_zones_json': json.dumps(all_zones),
        'all_drawn_zones_json': json.dumps(all_drawn_zones),
        'map_style_json': json.dumps(MapStyleSettings.get_style()),
        'patch_stats_json': json.dumps(patch_stats),
        'drawn_total': drawn_all,
        'zone_total': total_all,
    }
    return render(request, 'core/zone_quick_draw_mobile.html', context)


@login_required(login_url='core:login')
@ensure_csrf_cookie
def map_style_editor(request):
    """Map style customization page with live preview."""
    from .models import ManagerProfile, MapStyleSettings, Worker

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        try:
            Worker.objects.get(user=request.user, active=True)
        except Worker.DoesNotExist:
            messages.error(request, '无权限访问此页面')
            return redirect('core:dashboard')

    if request.method == 'POST':
        try:
            data = json.loads(request.body) if request.body else {}
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': '无效数据'}, status=400)
        MapStyleSettings.save_style(data, user=request.user)
        return JsonResponse({'success': True, 'message': '样式已保存'})

    # Load all drawn zones for map preview
    all_drawn_zones = []
    for z in Zone.objects.select_related('patch').only(
        'id', 'code', 'name', 'boundary_points', 'boundary_color',
        'label_lat', 'label_lng', 'label_scale', 'label_angle', 'smooth_override', 'patch_id', 'patch__name', 'current_status'
    ):
        center = None
        bp = z.active_boundary_points
        if bp:
            try:
                if isinstance(bp[0], list):
                    all_pts = [p for ring in bp for p in ring]
                else:
                    all_pts = bp
                if all_pts:
                    lats = [p[0] if isinstance(p, list) else p.get('lat', 0) for p in all_pts]
                    lngs = [p[1] if isinstance(p, list) else p.get('lng', 0) for p in all_pts]
                    center = {'lat': sum(lats) / len(lats), 'lng': sum(lngs) / len(lngs)}
            except Exception:
                pass

        all_drawn_zones.append({
            'id': z.id,
            'code': z.code,
            'name': z.name,
            'boundary_points': z.active_boundary_points,
            'boundary_color': z.boundary_color,
            'label_lat': z.label_lat,
            'label_lng': z.label_lng,
            'label_scale': z.label_scale,
            'label_angle': z.label_angle,
            'smooth_override': z.smooth_override,
            'status': z.current_status,
            'center': center,
        })

    current_style = MapStyleSettings.get_style()

    context = {
        'zones_json': json.dumps(all_drawn_zones),
        'map_style_json': json.dumps(current_style),
    }
    return render(request, 'core/map_style_editor.html', context)


@login_required(login_url='core:login')
@ensure_csrf_cookie
def zone_detail_page(request, zone_id):
    """Zone detail page showing all zone parameters, plants, equipment, notes, and stats."""
    import json
    from django.db.models import Sum
    from datetime import date, timedelta
    from .models import Plant, ZoneEquipment, WorkReport

    zone = get_object_or_404(Zone, pk=zone_id)
    today = date.today()
    thirty_days_ago = today - timedelta(days=30)

    # Current plants
    plants = Plant.objects.filter(zone=zone).filter(
        Q(planting_date__lte=today) | Q(planting_date__isnull=True)
    ).filter(
        Q(end_date__gte=today) | Q(end_date__isnull=True)
    ).order_by('name')

    # Active equipment
    equipment = ZoneEquipment.objects.filter(zone=zone).select_related('equipment').filter(
        Q(installation_date__lte=today) | Q(installation_date__isnull=True)
    ).filter(status__in=['working', 'needs_repair'])

    plant_count = plants.count()
    equipment_count = equipment.count()
    work_report_count = WorkReport.objects.filter(zone_location=zone).count()

    for eq in equipment:
        eq.status_display = eq.get_status_display()
        eq.equipment_details = {
            'equipment_type_display': eq.equipment.get_equipment_type_display(),
            'manufacturer': eq.equipment.manufacturer,
            'model_name': eq.equipment.model_name,
        }

    # Recent work reports (by 工单号 descending — newest #id first).
    recent_reports = WorkReport.objects.filter(zone_location=zone).select_related(
        'worker', 'location'
    ).order_by('-id')[:10]

    # Sibling zones (same patch)
    sibling_zones = []
    if zone.patch_id:
        sibling_zones = list(Zone.objects.filter(patch_id=zone.patch_id)
                             .exclude(pk=zone.pk).order_by('code')[:20])

    # Parse maintenance notes
    def _sort_notes(notes):
        def _sort_key(n):
            d = n.get('date', '')
            if not d or not isinstance(d, str) or d == '日期格式错误':
                return (1, '')
            return (0, d)
        return sorted(notes, key=_sort_key, reverse=True)

    try:
        equip_notes = _sort_notes(json.loads(zone.equipment_maintenance_notes)) if zone.equipment_maintenance_notes else []
    except (json.JSONDecodeError, TypeError):
        equip_notes = []
    try:
        irrig_notes = _sort_notes(json.loads(zone.irrigation_management_notes)) if zone.irrigation_management_notes else []
    except (json.JSONDecodeError, TypeError):
        irrig_notes = []
    try:
        remarks = _sort_notes(json.loads(zone.remarks)) if zone.remarks else []
    except (json.JSONDecodeError, TypeError):
        remarks = []
    try:
        confirmed_remarks = _sort_notes(json.loads(zone.confirmed_remarks)) if zone.confirmed_remarks else []
    except (json.JSONDecodeError, TypeError):
        confirmed_remarks = []
    is_manager = _check_zone_admin(request)

    # Priority display
    priority_display = dict(Zone.PRIORITY_CHOICES).get(zone.priority, '一般')
    status_display = zone.current_status if zone.current_status else '正常'

    context = {
        'zone': zone,
        'plants': plants,
        'equipment': equipment,
        'plant_count': plant_count,
        'equipment_count': equipment_count,
        'work_report_count': work_report_count,
        'recent_reports': recent_reports,
        'sibling_zones': sibling_zones,
        'equip_notes': equip_notes,
        'irrig_notes': irrig_notes,
        'remarks': remarks,
        'confirmed_remarks': confirmed_remarks,
        'is_manager': is_manager,
        'today': today.isoformat(),
        'priority_display': priority_display,
        'status_display': status_display,
    }

    return render(request, 'core/zone_detail_page.html', context)


@login_required(login_url='core:login')
def zone_smooth_update(request, zone_id):
    """API: update per-zone smooth override. POST with {smooth_override: null|int}."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    if not _check_zone_admin(request):
        return JsonResponse({'error': '无权限'}, status=403)
    zone = get_object_or_404(Zone, pk=zone_id)
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': '无效数据'}, status=400)
    val = data.get('smooth_override')
    zone.smooth_override = int(val) if val is not None else None
    zone.save(update_fields=['smooth_override', 'updated_at'])
    return JsonResponse({'success': True, 'smooth_override': zone.smooth_override})


@login_required(login_url='core:login')
def zone_remark_add(request, zone_id):
    import json
    from datetime import date as date_cls
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    zone = get_object_or_404(Zone, pk=zone_id)
    date = request.POST.get('date', '').strip()
    content = request.POST.get('content', '').strip()
    if not content:
        return JsonResponse({'error': '内容不能为空'}, status=400)
    if not date:
        date = date_cls.today().isoformat()
    author = _get_user_display_name(request)
    remarks = json.loads(zone.remarks) if zone.remarks else []
    remarks.insert(0, {'date': date, 'content': content, 'author': author})
    zone.remarks = json.dumps(remarks, ensure_ascii=False)
    zone.save(update_fields=['remarks'])
    return JsonResponse({'success': True, 'remark': {'date': date, 'content': content, 'author': author}})


@login_required(login_url='core:login')
def zone_remark_confirm(request, zone_id, index):
    import json
    from datetime import date as date_cls
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    if not _check_zone_admin(request):
        return JsonResponse({'error': '无权限'}, status=403)
    zone = get_object_or_404(Zone, pk=zone_id)
    remarks = json.loads(zone.remarks) if zone.remarks else []
    if index < 0 or index >= len(remarks):
        return JsonResponse({'error': '索引无效'}, status=400)
    remark = remarks.pop(index)
    confirm_reply = request.POST.get('confirm_reply', '').strip()
    confirm_author = _get_user_display_name(request)
    confirmed = {
        **remark,
        'confirm_date': date_cls.today().isoformat(),
        'confirm_reply': confirm_reply,
        'confirm_author': confirm_author,
    }
    confirmed_list = json.loads(zone.confirmed_remarks) if zone.confirmed_remarks else []
    confirmed_list.insert(0, confirmed)
    zone.remarks = json.dumps(remarks, ensure_ascii=False)
    zone.confirmed_remarks = json.dumps(confirmed_list, ensure_ascii=False)
    zone.save(update_fields=['remarks', 'confirmed_remarks'])
    return JsonResponse({'success': True, 'confirmed': confirmed})


@login_required(login_url='core:login')
def zone_remark_move(request, zone_id, index):
    import json
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    if not _check_zone_admin(request):
        return JsonResponse({'error': '无权限'}, status=403)
    zone = get_object_or_404(Zone, pk=zone_id)
    target = request.POST.get('target', '').strip()
    if target not in ('irrigation', 'equipment'):
        return JsonResponse({'error': '目标无效'}, status=400)
    confirmed_list = json.loads(zone.confirmed_remarks) if zone.confirmed_remarks else []
    if index < 0 or index >= len(confirmed_list):
        return JsonResponse({'error': '索引无效'}, status=400)
    entry = confirmed_list.pop(index)
    note = {'date': entry.get('date', ''), 'content': entry.get('content', '')}
    if target == 'irrigation':
        notes = json.loads(zone.irrigation_management_notes) if zone.irrigation_management_notes else []
        notes.insert(0, note)
        zone.irrigation_management_notes = json.dumps(notes, ensure_ascii=False)
        zone.save(update_fields=['irrigation_management_notes'])
    else:
        notes = json.loads(zone.equipment_maintenance_notes) if zone.equipment_maintenance_notes else []
        notes.insert(0, note)
        zone.equipment_maintenance_notes = json.dumps(notes, ensure_ascii=False)
        zone.save(update_fields=['equipment_maintenance_notes'])
    zone.confirmed_remarks = json.dumps(confirmed_list, ensure_ascii=False)
    zone.save(update_fields=['confirmed_remarks'])
    return JsonResponse({'success': True, 'target': target, 'note': note})


@login_required(login_url='core:login')
def zone_remark_archive(request, zone_id, index):
    """归档一条已确认备注 - 不写入灌溉/设备记录，仅就地打 archived 标记。

    保留在 confirmed_remarks 里（不从列表删除），工单详情页可据此反查展示；
    _group_zone_remarks 过滤掉 archived 条目，使其不再出现在「已确认」列表。
    """
    import json
    from datetime import date as date_cls
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    if not _check_zone_admin(request):
        return JsonResponse({'error': '无权限'}, status=403)
    zone = get_object_or_404(Zone, pk=zone_id)
    confirmed_list = json.loads(zone.confirmed_remarks) if zone.confirmed_remarks else []
    if index < 0 or index >= len(confirmed_list):
        return JsonResponse({'error': '索引无效'}, status=400)
    # 就地打标记，不删除：保留在 confirmed_remarks 里供工单详情页反查。
    entry = confirmed_list[index]
    entry['archived'] = True
    entry['archived_date'] = date_cls.today().isoformat()
    entry['archived_author'] = _get_user_display_name(request)
    zone.confirmed_remarks = json.dumps(confirmed_list, ensure_ascii=False)
    zone.save(update_fields=['confirmed_remarks'])
    return JsonResponse({'success': True})


@login_required(login_url='core:login')
def pipeline_new(request):
    from .models import ManagerProfile, Pipeline

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        messages.error(request, '无权限创建水管')
        return redirect('core:dashboard')

    if request.method == 'POST':
        pipeline_type = request.POST.get('pipeline_type', Pipeline.TYPE_IRRIGATION)
        zone_ids = request.POST.getlist('zones')
        name, code = _auto_pipeline_name_code(zone_ids, pipeline_type)

        pipeline = Pipeline(
            name=name,
            code=code,
            description=request.POST.get('description', ''),
            pipeline_type=pipeline_type,
            line_weight=int(request.POST.get('line_weight', 3)),
        )

        line_json = request.POST.get('line_points', '[]')
        try:
            pipeline.line_points = json.loads(line_json)
        except json.JSONDecodeError:
            messages.error(request, '坐标数据格式无效')
            rz, rp = _get_reference_map_data()
            return render(request, 'core/pipeline_form.html', {
                'pipeline': pipeline,
                'line_json': line_json,
                'zones': Zone.objects.all().order_by('code'),
                'ref_zones_json': rz,
                'ref_pipelines_json': rp,
            })

        pipeline.save()

        if zone_ids:
            pipeline.zones.set(Zone.objects.filter(id__in=zone_ids))

        messages.success(request, f'水管 "{pipeline.name}" 创建成功')
        return redirect('core:settings')

    ref_zones_json, ref_pipelines_json = _get_reference_map_data()

    zones = Zone.objects.all().order_by('code')
    context = {
        'pipeline': None,
        'line_json': '[]',
        'zones': zones,
        'grouped_zones': _build_grouped_zones(zones),
        'ref_zones_json': ref_zones_json,
        'ref_pipelines_json': ref_pipelines_json,
    }
    return render(request, 'core/pipeline_form.html', context)


@login_required(login_url='core:login')
def pipeline_edit(request, pipeline_id):
    from .models import ManagerProfile, Pipeline

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        messages.error(request, '无权限修改水管')
        return redirect('core:dashboard')

    pipeline = get_object_or_404(Pipeline, pk=pipeline_id)

    if request.method == 'POST':
        pipeline_type = request.POST.get('pipeline_type', pipeline.pipeline_type)
        zone_ids = request.POST.getlist('zones')
        name, code = _auto_pipeline_name_code(zone_ids, pipeline_type)

        pipeline.name = name
        pipeline.code = code
        pipeline.pipeline_type = pipeline_type
        pipeline.line_weight = int(request.POST.get('line_weight', pipeline.line_weight))

        line_json = request.POST.get('line_points', '[]')
        try:
            pipeline.line_points = json.loads(line_json)
        except json.JSONDecodeError:
            messages.error(request, '坐标数据格式无效')
            return redirect('core:pipeline_edit', pipeline_id=pipeline.id)

        pipeline.save()

        pipeline.zones.set(Zone.objects.filter(id__in=zone_ids))
        pipeline.zones.set(Zone.objects.filter(id__in=zone_ids))

        messages.success(request, f'水管 "{pipeline.name}" 更新成功')
        return redirect('core:settings')

    ref_zones_json, ref_pipelines_json = _get_reference_map_data(exclude_pipeline_id=pipeline.id)

    zones = Zone.objects.all().order_by('code')
    context = {
        'pipeline': pipeline,
        'line_json': json.dumps(pipeline.line_points),
        'zones': zones,
        'grouped_zones': _build_grouped_zones(zones),
        'selected_zone_ids': list(pipeline.zones.values_list('id', flat=True)),
        'ref_zones_json': ref_zones_json,
        'ref_pipelines_json': ref_pipelines_json,
    }
    return render(request, 'core/pipeline_form.html', context)


@require_POST
@login_required(login_url='core:login')
def pipeline_delete(request, pipeline_id):
    from .models import ManagerProfile, Pipeline

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        messages.error(request, '无权限删除水管')
        return redirect('core:dashboard')

    pipeline = get_object_or_404(Pipeline, pk=pipeline_id)
    pipeline_name = pipeline.name
    pipeline.delete()
    messages.success(request, f'水管 "{pipeline_name}" 删除成功')
    return redirect('core:settings')


# ─── Region CRUD ───

@login_required(login_url='core:login')
def region_new(request):
    """Create a new region — admin only."""
    from .models import ManagerProfile, Region

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass
    if not is_admin:
        messages.error(request, '无权限')
        return redirect('core:dashboard')

    if request.method == 'POST':
        region = Region(
            name=request.POST['name'],
            description=request.POST.get('description', ''),
            order=int(request.POST.get('order', 0)),
            active=request.POST.get('active') is not None,
        )
        region.full_clean()
        region.save()
        messages.success(request, f'大区 "{region.name}" 创建成功')
        return redirect('core:settings')

    return render(request, 'core/region_form.html', {'mode': 'new'})


@login_required(login_url='core:login')
def region_edit(request, region_id):
    """Edit a region — admin only."""
    from .models import ManagerProfile, Region

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass
    if not is_admin:
        messages.error(request, '无权限')
        return redirect('core:dashboard')

    region = get_object_or_404(Region, pk=region_id)

    if request.method == 'POST':
        region.name = request.POST['name']
        region.description = request.POST.get('description', '')
        region.order = int(request.POST.get('order', 0))
        region.active = request.POST.get('active') is not None
        region.full_clean()
        region.save()
        messages.success(request, f'大区 "{region.name}" 更新成功')
        return redirect('core:settings')

    return render(request, 'core/region_form.html', {'mode': 'edit', 'region': region})


@login_required(login_url='core:login')
@require_POST
def region_delete(request, region_id):
    """Delete a region — admin only. Patches are SET_NULL."""
    from .models import ManagerProfile, Region

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass
    if not is_admin:
        messages.error(request, '无权限')
        return redirect('core:dashboard')

    region = get_object_or_404(Region, pk=region_id)
    region_name = region.name
    region.delete()
    messages.success(request, f'大区 "{region_name}" 已删除')
    return redirect('core:settings')


@login_required(login_url='core:login')
@require_POST
def batch_delete_region(request):
    """Batch delete regions — admin only."""
    from .models import ManagerProfile, Region
    try:
        body = json.loads(request.body)
        ids = body.get('ids', [])
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass
    if not is_admin:
        return JsonResponse({'error': '无权限'}, status=403)

    count, _ = Region.objects.filter(pk__in=ids).delete()
    messages.success(request, f'已删除 {count} 个大区')
    return JsonResponse({'success': True, 'deleted': count})


def _boundary_points_to_shapely(boundary_points):
    """Convert JSON boundary_points to a Shapely geometry."""
    from shapely.geometry import Polygon, MultiPolygon

    if not boundary_points or len(boundary_points) == 0:
        return None

    def to_coord(p):
        if isinstance(p, dict):
            return (p.get('lng', p.get('longitude', 0)), p.get('lat', p.get('latitude', 0)))
        if isinstance(p, (list, tuple)) and len(p) >= 2:
            return (p[1], p[0])
        return None

    first = boundary_points[0]
    if isinstance(first, list) and len(first) > 0 and (isinstance(first[0], (list, dict))):
        rings = boundary_points
    elif isinstance(first, (dict, list)):
        rings = [boundary_points]
    else:
        return None

    polygons = []
    for ring in rings:
        coords = [to_coord(p) for p in ring]
        coords = [c for c in coords if c is not None]
        if len(coords) >= 3:
            if coords[0] != coords[-1]:
                coords.append(coords[0])
            polygons.append(Polygon(coords))

    if len(polygons) == 0:
        return None
    if len(polygons) == 1:
        return polygons[0]
    return MultiPolygon(polygons)


def _recalculate_landmark_assignments():
    """Recalculate all zone↔landmark assignments based on boundary overlap."""
    try:
        from shapely.geometry import MultiPolygon as _MP  # noqa: F401 — verify import
    except ImportError:
        return -1

    from .models import Landmark, ZoneLandmarkAssignment, Zone

    ZoneLandmarkAssignment.objects.all().delete()

    landmarks = Landmark.objects.exclude(boundary_points=[])
    zones = Zone.objects.exclude(boundary_points=[])
    count = 0

    for landmark in landmarks:
        landmark_geom = _boundary_points_to_shapely(landmark.boundary_points)
        if landmark_geom is None:
            continue
        for zone in zones:
            zone_geom = _boundary_points_to_shapely(zone.boundary_points)
            if zone_geom is None:
                continue
            if landmark_geom.intersects(zone_geom):
                ZoneLandmarkAssignment.objects.create(zone=zone, landmark=landmark)
                count += 1

    return count


def _landmark_draw_context(landmark):
    """Build context for the landmark draw page with reference layers."""
    from .models import Landmark
    all_landmarks = []
    for lm in Landmark.objects.order_by('order', 'name'):
        all_landmarks.append({
            'id': lm.id,
            'name': lm.name,
            'boundary_points': lm.boundary_points,
            'boundary_color': lm.boundary_color,
            'area_sqm': lm.area_sqm,
        })
    all_zones = []
    for z in Zone.objects.select_related('patch').only(
        'id', 'code', 'name', 'boundary_points', 'boundary_color',
        'label_lat', 'label_lng', 'label_scale', 'label_angle', 'patch_id'
    ):
        all_zones.append({
            'id': z.id,
            'code': z.code,
            'name': z.name,
            'boundary_points': z.active_boundary_points,
            'boundary_color': z.boundary_color,
        })
    return {
        'landmark': landmark,
        'all_landmarks_json': json.dumps(all_landmarks),
        'all_zones_json': json.dumps(all_zones),
        'nav_settings': True,
    }


@login_required(login_url='core:login')
def landmark_new(request):
    from .models import Landmark
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            return JsonResponse({'error': '名称不能为空'}, status=400)
        boundary_raw = request.POST.get('boundary_points', '[]')
        boundary_color = request.POST.get('boundary_color', '#E8590C')
        try:
            boundary_data = json.loads(boundary_raw)
        except json.JSONDecodeError:
            return JsonResponse({'error': '边界数据格式错误'}, status=400)
        boundary_data = auto_close_boundary_points(boundary_data)
        landmark = Landmark(name=name, boundary_points=boundary_data, boundary_color=boundary_color)
        landmark.center = get_zone_center(boundary_data)
        landmark.save()
        return JsonResponse({'success': True, 'id': landmark.id, 'name': landmark.name})
    return render(request, 'core/landmark_draw.html', _landmark_draw_context(None))


@login_required(login_url='core:login')
def landmark_edit(request, landmark_id):
    from .models import Landmark
    landmark = get_object_or_404(Landmark, pk=landmark_id)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            landmark.name = name
        boundary_raw = request.POST.get('boundary_points', '')
        if boundary_raw:
            try:
                boundary_data = json.loads(boundary_raw)
                boundary_data = auto_close_boundary_points(boundary_data)
                landmark.boundary_points = boundary_data
                landmark.center = get_zone_center(boundary_data)
            except json.JSONDecodeError:
                return JsonResponse({'error': '边界数据格式错误'}, status=400)
        boundary_color = request.POST.get('boundary_color')
        if boundary_color:
            landmark.boundary_color = boundary_color
        landmark.save()
        return JsonResponse({'success': True, 'id': landmark.id, 'name': landmark.name})
    return render(request, 'core/landmark_draw.html', _landmark_draw_context(landmark))


@login_required(login_url='core:login')
@require_POST
def landmark_delete(request, landmark_id):
    from .models import Landmark
    landmark = get_object_or_404(Landmark, pk=landmark_id)
    landmark.delete()
    return JsonResponse({'success': True})


@login_required(login_url='core:login')
def landmarks_api(request):
    from .models import Landmark, ZoneLandmarkAssignment
    landmarks = Landmark.objects.order_by('order', 'name')
    data = []
    for lm in landmarks:
        zone_count = ZoneLandmarkAssignment.objects.filter(landmark=lm).count()
        data.append({
            'id': lm.id,
            'name': lm.name,
            'boundary_points': lm.boundary_points,
            'boundary_color': lm.boundary_color,
            'center': lm.center,
            'area_sqm': lm.area_sqm,
            'zone_count': zone_count,
        })
    return JsonResponse(data, safe=False)


@login_required(login_url='core:login')
@require_POST
def landmarks_recalculate(request):
    from .models import Landmark
    count = _recalculate_landmark_assignments()
    if count < 0:
        return JsonResponse({'error': 'Shapely library not installed. Run: pip install shapely'}, status=500)
    total_landmarks = Landmark.objects.count()
    return JsonResponse({'success': True, 'assignments': count, 'landmarks': total_landmarks})


@login_required(login_url='core:login')
@require_POST
def batch_delete_landmark(request):
    from .models import Landmark
    try:
        body = json.loads(request.body)
        ids = body.get('ids', [])
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    count, _ = Landmark.objects.filter(pk__in=ids).delete()
    return JsonResponse({'success': True, 'deleted': count})


@login_required(login_url='core:login')
def patch_new(request):
    """Create a new patch — admin only."""
    from .models import ManagerProfile, Patch

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        messages.error(request, '无权限创建片区')
        return redirect('core:dashboard')

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()
        description = request.POST.get('description', '').strip()
        region_id = request.POST.get('region') or None

        if not name:
            messages.error(request, '片区名称不能为空')
        elif not code:
            messages.error(request, '片区编号不能为空')
        elif Patch.objects.filter(name=name).exists():
            messages.error(request, f'片区名称 "{name}" 已存在')
        elif Patch.objects.filter(code=code).exists():
            messages.error(request, f'片区编号 "{code}" 已存在')
        else:
            from .models import Region
            region = None
            if region_id:
                region = Region.objects.filter(pk=region_id).first()
            Patch.objects.create(name=name, code=code, description=description, region=region)
            messages.success(request, f'片区 "{name}" 创建成功')
            return redirect('core:settings')

    from .models import Region
    return render(request, 'core/patch_form.html', {
        'mode': 'new',
        'regions': Region.objects.order_by('order', 'name'),
    })


@login_required(login_url='core:login')
def patch_edit(request, patch_id):
    """Edit a patch with zone assignment — admin only."""
    from .models import ManagerProfile, Patch

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        messages.error(request, '无权限编辑片区')
        return redirect('core:dashboard')

    patch = get_object_or_404(Patch, pk=patch_id)

    # Fields that propagate from 片区 to linked child patches
    PROPAGATE_FIELDS = [
        'description', 'active', 'time_zone', 'water_pricing',
        'et_current', 'et_default', 'et_minimum', 'et_maximum',
        'crop_coefficient', 'rain_shutdown', 'date_open', 'date_close',
    ]

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()
        description = request.POST.get('description', '').strip()
        region_id = request.POST.get('region') or None

        # Validate uniqueness excluding self
        errors = []
        if not name:
            errors.append('片区名称不能为空')
        elif not code:
            errors.append('片区编号不能为空')
        elif Patch.objects.filter(name=name).exclude(pk=patch.pk).exists():
            errors.append(f'片区名称 "{name}" 已存在')
        elif Patch.objects.filter(code=code).exclude(pk=patch.pk).exists():
            errors.append(f'片区编号 "{code}" 已存在')

        if errors:
            for err in errors:
                messages.error(request, err)
        else:
            from .models import Region
            patch.name = name
            patch.code = code
            patch.description = description
            patch.region_id = region_id
            patch.active = request.POST.get('active') is not None
            patch.time_zone = request.POST.get('time_zone', 'China')
            patch.water_pricing = request.POST.get('water_pricing') or None
            patch.et_current = request.POST.get('et_current') or None
            patch.et_default = request.POST.get('et_default') or None
            patch.et_minimum = request.POST.get('et_minimum') or None
            patch.et_maximum = request.POST.get('et_maximum') or None
            patch.crop_coefficient = request.POST.get('crop_coefficient') or None
            patch.rain_shutdown = request.POST.get('rain_shutdown') is not None
            patch.date_open = request.POST.get('date_open', '')
            patch.date_close = request.POST.get('date_close', '')
            patch.save()

            # Handle zone assignment
            all_zone_ids = request.POST.getlist('zones')
            Zone.objects.filter(patch=patch).exclude(id__in=all_zone_ids).update(patch=None)
            Zone.objects.filter(id__in=all_zone_ids, patch__isnull=True).update(patch=patch)
            Zone.objects.filter(id__in=all_zone_ids).exclude(patch=patch).update(patch=patch)

            # Handle linked patch assignment (location/zone_text only, stations managed by sync)
            linked_ids = request.POST.getlist('linked_patches')
            # Clear parent for children that were linked but are now unchecked
            Patch.objects.filter(parent=patch).exclude(
                id__in=linked_ids
            ).update(parent=None)
            # Set parent for newly checked patches
            Patch.objects.filter(id__in=linked_ids).update(parent=patch)

            # Propagate shared fields to all children
            children = patch.children.all()
            if children:
                update_kwargs = {}
                for f in PROPAGATE_FIELDS:
                    update_kwargs[f] = getattr(patch, f)
                children.update(**update_kwargs)

            messages.success(request, f'片区 "{name}" 已更新')
            return redirect('core:settings')

    # Build grouped zones for the zone picker
    all_zones = Zone.objects.all().order_by('code')
    grouped_zones = _build_grouped_zones(all_zones)

    # IDs of zones assigned to this patch (FK + derived from code prefix)
    selected_zone_ids = set()
    for group in grouped_zones:
        for z in group['zones']:
            if z['patch_id'] == patch.id:
                selected_zone_ids.add(z['id'])

    # Linked patches (children via parent FK)
    linked_patch_ids = set(patch.children.values_list('id', flat=True))
    linkable = Patch.objects.exclude(id=patch.id).order_by('code')
    linked_groups = [{
        'label': '关联片区',
        'patches': linkable,
    }] if linkable.exists() else []

    from .models import Region
    return render(request, 'core/patch_form.html', {
        'mode': 'edit',
        'patch': patch,
        'grouped_zones': grouped_zones,
        'selected_zone_ids': selected_zone_ids,
        'linked_patch_ids': linked_patch_ids,
        'linked_groups': linked_groups,
        'regions': Region.objects.order_by('order', 'name'),
    })


@login_required(login_url='core:login')
@require_POST
def patch_delete(request, patch_id):
    """Delete a patch — admin only. Zones are SET_NULL, not deleted."""
    from .models import ManagerProfile, Patch

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        messages.error(request, '无权限删除片区')
        return redirect('core:dashboard')

    patch = get_object_or_404(Patch, pk=patch_id)
    patch_name = patch.name
    patch.delete()
    messages.success(request, f'片区 "{patch_name}" 已删除')
    return redirect('core:settings')


@login_required(login_url='core:login')
@require_POST
def batch_delete_patch(request):
    """Batch delete patches — admin only."""
    from .models import ManagerProfile, Patch
    try:
        body = json.loads(request.body)
        ids = body.get('ids', [])
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass
    if not is_admin:
        return JsonResponse({'error': '无权限'}, status=403)

    count, _ = Patch.objects.filter(pk__in=ids).delete()
    messages.success(request, f'已删除 {count} 个片区')
    return JsonResponse({'success': True, 'deleted': count})


@login_required(login_url='core:login')
@require_POST
def batch_delete_zone(request):
    """Batch delete zones — admin only."""
    from .models import ManagerProfile, Zone
    try:
        body = json.loads(request.body)
        ids = body.get('ids', [])
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass
    if not is_admin:
        return JsonResponse({'error': '无权限'}, status=403)

    count, _ = Zone.objects.filter(pk__in=ids).delete()
    messages.success(request, f'已删除 {count} 个区域')
    return JsonResponse({'success': True, 'deleted': count})


@login_required(login_url='core:login')
@require_POST
def batch_delete_pipeline(request):
    """Batch delete pipelines — admin only."""
    from .models import ManagerProfile, Pipeline
    try:
        body = json.loads(request.body)
        ids = body.get('ids', [])
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass
    if not is_admin:
        return JsonResponse({'error': '无权限'}, status=403)

    count, _ = Pipeline.objects.filter(pk__in=ids).delete()
    messages.success(request, f'已删除 {count} 条水管')
    return JsonResponse({'success': True, 'deleted': count})


@login_required(login_url='core:login')
def equipment_catalog_autocomplete(request):
    """AJAX autocomplete endpoint for equipment catalog."""
    from .models import EquipmentCatalog

    equipment_type = request.GET.get('equipment_type', '')
    search = request.GET.get('search', '')

    queryset = EquipmentCatalog.objects.all()

    if equipment_type:
        queryset = queryset.filter(equipment_type=equipment_type)

    if search:
        queryset = queryset.filter(
            Q(model_name__icontains=search) |
            Q(manufacturer__icontains=search)
        )

    queryset = queryset.order_by('manufacturer', 'model_name')[:20]

    seen = set()
    results = []
    for item in queryset:
        key = (item.manufacturer, item.model_name)
        if key in seen:
            continue
        seen.add(key)
        label = f"{item.manufacturer} {item.model_name}" if item.manufacturer else item.model_name
        results.append({
            'id': item.id,
            'label': label,
            'model_name': item.model_name,
            'manufacturer': item.manufacturer,
            'equipment_type': item.equipment_type,
        })

    return JsonResponse({'results': results})


# (removed: requests_page, request_detail, update_request_status — legacy request-approval views)
def register(request):
    """
    User registration page - submit request for admin approval.

    Department type determines role:
    - 园艺一线 (field_worker) → Field Worker
    - 园艺经理 (manager) → Manager
    - 其他部门 (dept_user) → Department User (requires sub-department selection)
    """
    from core.models import RegistrationRequest, ROLE_FIELD_WORKER, ROLE_DEPT_USER, ROLE_MANAGER
    from django.contrib.auth.hashers import make_password
    from django.contrib.auth.models import User

    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').strip()
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        phone = request.POST.get('phone', '').strip()
        employee_id = request.POST.get('employee_id', '').strip()
        department_type = request.POST.get('department_type', '').strip()
        requested_role = request.POST.get('requested_role', ROLE_FIELD_WORKER).strip()
        department = request.POST.get('department', '').strip()
        department_other = request.POST.get('department_other', '').strip()

        # Map department_type to role if needed
        if department_type == 'field_worker':
            requested_role = ROLE_FIELD_WORKER
        elif department_type == 'manager':
            requested_role = ROLE_MANAGER
        elif department_type == 'dept_user':
            requested_role = ROLE_DEPT_USER

        # Valid role choices
        valid_roles = [ROLE_FIELD_WORKER, ROLE_DEPT_USER, ROLE_MANAGER]

        # Validation
        if not full_name:
            messages.error(request, '请输入姓名')
        elif not username:
            messages.error(request, '请输入用户名')
        elif len(username) < 3:
            messages.error(request, '用户名至少需要3个字符')
        elif User.objects.filter(username=username).exists():
            messages.error(request, '该用户名已存在')
        elif not password:
            messages.error(request, '请输入密码')
        elif len(password) < 6:
            messages.error(request, '密码至少需要6个字符')
        elif not phone:
            messages.error(request, '请输入手机号')
        elif not department_type:
            messages.error(request, '请选择部门')
        elif requested_role not in valid_roles:
            messages.error(request, '请选择有效的角色')
        elif department_type == 'dept_user' and not department:
            messages.error(request, '请选择所属部门')
        elif department == '其他' and not department_other:
            messages.error(request, '请输入其他部门名称')
        elif RegistrationRequest.objects.filter(username=username).exists():
            messages.error(request, '该用户名已有待审批的注册申请')
        elif RegistrationRequest.objects.filter(phone=phone, status='pending').exists():
            messages.error(request, '该手机号已有待审批的注册申请')
        else:
            # For non-dept users, clear department info
            if department_type != 'dept_user':
                department = ''
                department_other = ''

            RegistrationRequest.objects.create(
                full_name=full_name,
                username=username,
                password=make_password(password),  # Hash the password
                phone=phone,
                employee_id=employee_id,
                department=department,
                department_other=department_other if department == '其他' else '',
                requested_role=requested_role,
            )
            messages.success(request, '注册申请已提交，请等待管理员审批')
            return redirect('core:register')

    return render(request, 'core/register.html')


@login_required(login_url='core:login')
def registration_approval(request):
    """
    Registration approval page for admins/managers.
    Shows pending registration requests with approve/reject actions.
    Uses applicant-submitted username and password.
    """
    from django.contrib.auth.models import User
    from core.models import (
        RegistrationRequest, ManagerProfile, DepartmentUserProfile, Worker,
        ROLE_MANAGER, ROLE_FIELD_WORKER, ROLE_DEPT_USER
)
    from core.role_utils import is_admin

    # Check admin permission
    if not is_admin(request.user):
        messages.error(request, '无权限访问此页面')
        return redirect('core:dashboard')

    # Handle POST actions
    if request.method == 'POST':
        request_id = request.POST.get('request_id')
        action = request.POST.get('action')
        reason = request.POST.get('reason', '')

        try:
            reg = RegistrationRequest.objects.get(pk=request_id, status='pending')

            if action == 'approve':
                # Use submitted username and password
                username = reg.username
                password = reg.password  # Already hashed during submission

                # Generate employee_id
                if reg.requested_role == ROLE_MANAGER:
                    prefix = 'ADM'
                    model_class = ManagerProfile
                elif reg.requested_role == ROLE_DEPT_USER:
                    prefix = 'DEPT'
                    model_class = DepartmentUserProfile
                else:
                    prefix = 'EMP'
                    model_class = Worker

                # Use submitted employee_id if provided, otherwise auto-generate
                if reg.employee_id:
                    employee_id = reg.employee_id
                else:
                    # Find max employee_id number by scanning all records
                    max_num = 0
                    for profile in model_class.objects.all():
                        eid = profile.employee_id or ''
                        if eid.startswith(prefix):
                            try:
                                num = int(eid[len(prefix):])
                                if num > max_num:
                                    max_num = num
                            except ValueError:
                                continue

                    next_num = max_num + 1
                    employee_id = f'{prefix}{next_num:03d}'

                # Create Django User with submitted credentials
                user = User.objects.create(
                    username=username,
                    password=password,  # Already hashed
                    first_name=reg.full_name,
                )

                # Create profile based on role
                if reg.requested_role == ROLE_MANAGER:
                    ManagerProfile.objects.create(
                        user=user,
                        employee_id=employee_id,
                        full_name=reg.full_name,
                        phone=reg.phone,
                        is_super_admin=False,
                        can_approve_registrations=True,
                        can_approve_work_orders=True,
                    )
                elif reg.requested_role == ROLE_DEPT_USER:
                    DepartmentUserProfile.objects.create(
                        user=user,
                        employee_id=employee_id,
                        full_name=reg.full_name,
                        phone=reg.phone,
                        department=reg.department or 'ENT',
                        department_other=reg.department_other if reg.department == '其他' else '',
                    )
                else:  # field_worker
                    Worker.objects.create(
                        user=user,
                        employee_id=employee_id,
                        full_name=reg.full_name,
                        phone=reg.phone,
                        department=reg.department or '',
                        department_other=reg.department_other if reg.department == '其他' else '',
                    )

                # Update registration
                reg.status = 'approved'
                reg.employee_id = employee_id
                reg.processed_at = timezone.now()
                reg.created_user = user
                reg.save()

                messages.success(request, f'已批准 {reg.full_name} 的注册申请，用户名：{username}，工号：{employee_id}')

            elif action == 'reject':
                reg.status = 'rejected'
                reg.status_notes = reason
                reg.processed_at = timezone.now()
                reg.save()
                messages.success(request, f'已拒绝 {reg.full_name} 的注册申请')

        except RegistrationRequest.DoesNotExist:
            messages.error(request, '注册申请不存在')

        return redirect('core:registration_approval')

    # GET request - show list
    filter_status = request.GET.get('filter', 'pending')

    if filter_status == 'all':
        requests_qs = RegistrationRequest.objects.all()
    else:
        requests_qs = RegistrationRequest.objects.filter(status=filter_status)

    requests_qs = requests_qs.order_by('-created_at')

    # Stats
    stats = {
        'pending': RegistrationRequest.objects.filter(status='pending').count(),
        'approved': RegistrationRequest.objects.filter(status='approved').count(),
        'rejected': RegistrationRequest.objects.filter(status='rejected').count(),
    }

    context = {
        'requests': requests_qs,
        'stats': stats,
        'filter': filter_status,
    }

    return render(request, 'core/registration_approval.html', context)


@login_required(login_url='core:login')
def user_management(request):
    """
    User management page with two tabs: User List and Registration Approval.
    """
    from django.contrib.auth.models import User
    from core.models import (
        RegistrationRequest, ManagerProfile, DepartmentUserProfile, Worker,
        ROLE_MANAGER, ROLE_FIELD_WORKER, ROLE_DEPT_USER,
        ROLE_SUPER_ADMIN
)
    from core.role_utils import is_admin

    if not is_admin(request.user):
        messages.error(request, '无权限访问此页面')
        return redirect('core:dashboard')

    # Handle POST actions (registration approval/rejection)
    if request.method == 'POST':
        request_id = request.POST.get('request_id')
        action = request.POST.get('action')
        reason = request.POST.get('reason', '')

        try:
            reg = RegistrationRequest.objects.get(pk=request_id, status='pending')

            if action == 'approve':
                username = reg.username
                password = reg.password

                if reg.requested_role == ROLE_MANAGER:
                    prefix = 'ADM'
                    model_class = ManagerProfile
                elif reg.requested_role == ROLE_DEPT_USER:
                    prefix = 'DEPT'
                    model_class = DepartmentUserProfile
                else:
                    prefix = 'EMP'
                    model_class = Worker

                if reg.employee_id:
                    employee_id = reg.employee_id
                else:
                    max_num = 0
                    for profile in model_class.objects.all():
                        eid = profile.employee_id or ''
                        if eid.startswith(prefix):
                            try:
                                num = int(eid[len(prefix):])
                                if num > max_num:
                                    max_num = num
                            except ValueError:
                                continue
                    next_num = max_num + 1
                    employee_id = f'{prefix}{next_num:03d}'

                user = User.objects.create(
                    username=username,
                    password=password,
                    first_name=reg.full_name,
                )

                if reg.requested_role == ROLE_MANAGER:
                    ManagerProfile.objects.create(
                        user=user, employee_id=employee_id,
                        full_name=reg.full_name, phone=reg.phone,
                        is_super_admin=False,
                        can_approve_registrations=True,
                        can_approve_work_orders=True,
                    )
                elif reg.requested_role == ROLE_DEPT_USER:
                    DepartmentUserProfile.objects.create(
                        user=user, employee_id=employee_id,
                        full_name=reg.full_name, phone=reg.phone,
                        department=reg.department or 'ENT',
                        department_other=reg.department_other if reg.department == '其他' else '',
                    )
                else:
                    Worker.objects.create(
                        user=user, employee_id=employee_id,
                        full_name=reg.full_name, phone=reg.phone,
                        department=reg.department or '',
                        department_other=reg.department_other if reg.department == '其他' else '',
                    )

                reg.status = 'approved'
                reg.employee_id = employee_id
                reg.processed_at = timezone.now()
                reg.created_user = user
                reg.save()
                messages.success(request, f'已批准 {reg.full_name} 的注册申请，用户名：{username}，工号：{employee_id}')

            elif action == 'reject':
                reg.status = 'rejected'
                reg.status_notes = reason
                reg.processed_at = timezone.now()
                reg.save()
                messages.success(request, f'已拒绝 {reg.full_name} 的注册申请')

        except RegistrationRequest.DoesNotExist:
            messages.error(request, '注册申请不存在')

        return redirect('/user-management/?tab=approval')

    # GET request
    active_tab = request.GET.get('tab', 'users')
    filter_status = request.GET.get('filter', 'pending')

    # Zone draw counts per user
    draw_counts = {}
    for row in Zone.objects.exclude(drawn_by__isnull=True).values('drawn_by').annotate(cnt=Count('id')):
        draw_counts[row['drawn_by']] = row['cnt']
    drawn_zone_map = {}
    for z in Zone.objects.exclude(drawn_by__isnull=True).exclude(boundary_points=[]).select_related('patch').order_by('code'):
        uid = z.drawn_by_id
        drawn_zone_map.setdefault(uid, []).append({'code': z.code, 'name': z.name or z.code, 'patch': z.patch.name if z.patch else ''})

    # Build user list
    users_list = []

    for mp in ManagerProfile.objects.select_related('user').all():
        users_list.append({
            'profile_type': 'manager',
            'profile_id': mp.id,
            'user_id': mp.user.id if mp.user else None,
            'username': mp.user.username if mp.user else '-',
            'full_name': mp.full_name,
            'employee_id': mp.employee_id,
            'phone': mp.phone,
            'department': '',
            'role': ROLE_SUPER_ADMIN if mp.is_super_admin else ROLE_MANAGER,
            'role_display': '超级管理员' if mp.is_super_admin else '管理员',
            'active': mp.active,
            'drawn_zones': draw_counts.get(mp.user.id, 0) if mp.user else 0,
        })

    for dup in DepartmentUserProfile.objects.select_related('user').all():
        users_list.append({
            'profile_type': 'dept_user',
            'profile_id': dup.id,
            'user_id': dup.user.id if dup.user else None,
            'username': dup.user.username if dup.user else '-',
            'full_name': dup.full_name,
            'employee_id': dup.employee_id,
            'phone': dup.phone,
            'department': dup.get_department_display_name() if hasattr(dup, 'get_department_display_name') else (dup.department_other or dup.department),
            'role': ROLE_DEPT_USER,
            'role_display': '部门用户',
            'active': dup.active,
            'drawn_zones': draw_counts.get(dup.user.id, 0) if dup.user else 0,
        })

    for w in Worker.objects.select_related('user').all():
        users_list.append({
            'profile_type': 'worker',
            'profile_id': w.id,
            'user_id': w.user.id if w.user else None,
            'username': w.user.username if w.user else '-',
            'full_name': w.full_name,
            'employee_id': w.employee_id,
            'phone': w.phone,
            'department': w.get_department_display_name() if hasattr(w, 'get_department_display_name') else (w.department_other or w.department),
            'role': ROLE_FIELD_WORKER,
            'role_display': '灌溉一线',
            'active': w.active,
            'drawn_zones': draw_counts.get(w.user.id, 0) if w.user else 0,
        })

    users_list.sort(key=lambda x: x['full_name'])

    # Registration approval data
    if filter_status == 'all':
        requests_qs = RegistrationRequest.objects.all()
    else:
        requests_qs = RegistrationRequest.objects.filter(status=filter_status)
    requests_qs = requests_qs.order_by('-created_at')

    stats = {
        'pending': RegistrationRequest.objects.filter(status='pending').count(),
        'approved': RegistrationRequest.objects.filter(status='approved').count(),
        'rejected': RegistrationRequest.objects.filter(status='rejected').count(),
    }

    # Announcement data for the 通知管理 tab (manager-visible under 用户管理).
    from core.models import Announcement
    announcements = (Announcement.objects.all()
                     .annotate(ack_count=Count('acknowledgments'))
                     .order_by('-created_at'))
    edit_obj = None
    edit_id = request.GET.get('edit')
    if edit_id:
        edit_obj = get_object_or_404(Announcement, pk=edit_id)

    # Crew data for the 班组管理 tab (manager-visible under 用户管理).
    from core.models import Crew
    crews = Crew.objects.all().prefetch_related('leader', 'members', 'lands', 'patches').order_by('name')

    context = {
        'active_tab': active_tab,
        'filter': filter_status,
        'users_list': users_list,
        'total_users': len(users_list),
        'active_users': sum(1 for u in users_list if u['active']),
        'inactive_users': sum(1 for u in users_list if not u['active']),
        'requests': requests_qs,
        'stats': stats,
        'drawn_zone_map_json': json.dumps(drawn_zone_map),
        'announcements': announcements,
        'edit_obj': edit_obj,
        'crews': crews,
    }

    return render(request, 'core/user_management.html', context)


@login_required(login_url='core:login')
def user_edit(request, profile_type, profile_id):
    """Edit user profile via AJAX."""
    from core.role_utils import is_admin
    from core.models import ManagerProfile, DepartmentUserProfile, Worker

    if not is_admin(request.user):
        return JsonResponse({'error': '无权限'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    model_map = {
        'manager': ManagerProfile,
        'dept_user': DepartmentUserProfile,
        'worker': Worker,
    }
    model_class = model_map.get(profile_type)
    if not model_class:
        return JsonResponse({'error': 'Invalid profile type'}, status=400)

    profile = get_object_or_404(model_class, pk=profile_id)

    if 'full_name' in request.POST:
        profile.full_name = request.POST['full_name']
    if 'phone' in request.POST:
        profile.phone = request.POST['phone']
    if 'active' in request.POST:
        profile.active = request.POST['active'] in ('true', '1', 'True')
    if profile_type in ('worker', 'dept_user') and 'department' in request.POST:
        dept = request.POST['department']
        profile.department = dept
        if dept == '其他' and 'department_other' in request.POST:
            profile.department_other = request.POST['department_other']
        else:
            profile.department_other = ''

    profile.save()
    return JsonResponse({'success': True, 'message': f'已更新 {profile.full_name}'})


@login_required(login_url='core:login')
def user_preferences_api(request):
    """GET/PUT user preferences (e.g. zone card field visibility)."""
    from .models import ManagerProfile

    if not request.user.is_authenticated:
        return JsonResponse({'preferences': {}})

    profile = None
    try:
        profile = ManagerProfile.objects.get(user=request.user, active=True)
    except ManagerProfile.DoesNotExist:
        pass

    # For users without a ManagerProfile, use a simple request.session fallback
    if profile is None:
        if request.method == 'GET':
            return JsonResponse({'preferences': request.session.get('preferences', {})})
        if request.method in ('PUT', 'POST'):
            import json as _json
            try:
                data = _json.loads(request.body) if request.body else {}
            except _json.JSONDecodeError:
                data = {}
            request.session['preferences'] = data.get('preferences', {})
            request.session.save()
            return JsonResponse({'success': True, 'preferences': request.session['preferences']})
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    if request.method == 'GET':
        return JsonResponse({'preferences': profile.preferences or {}})

    if request.method in ('PUT', 'POST'):
        import json as _json
        try:
            data = _json.loads(request.body) if request.body else {}
        except _json.JSONDecodeError:
            data = {}
        profile.preferences = data.get('preferences', {})
        profile.save(update_fields=['preferences', 'updated_at'])
        return JsonResponse({'success': True, 'preferences': profile.preferences})

    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required(login_url='core:login')
def maxicom_dashboard_api(request):
    """API endpoint providing Maxicom2 irrigation data for the dashboard."""
    from core.models import (
        MaxicomController, MaxicomSchedule,
        MaxicomFlowZone, MaxicomWeatherStation, MaxicomWeatherLog,
        MaxicomEvent, MaxicomETCheckbook, MaxicomRuntime,
        Patch,
)

    # System overview stats
    stats = {
        'sites': Patch.objects.count(),
        'controllers': MaxicomController.objects.count(),
        'stations': Patch.objects.filter(parent__isnull=False).count(),
        'schedules': MaxicomSchedule.objects.count(),
        'flow_zones': MaxicomFlowZone.objects.count(),
        'weather_stations': MaxicomWeatherStation.objects.count(),
        'weather_logs': MaxicomWeatherLog.objects.count(),
        'events': MaxicomEvent.objects.count(),
        'locked_stations': Patch.objects.filter(parent__isnull=False, lockout=True).count(),
    }

    # Site hierarchy: sites with controller/station counts and station details.
    # Annotate the four reverse counts in ONE query (was 4 queries/site → N+1)
    # and prefetch children so the per-site station loop adds no extra queries.
    sites = []
    root_sites = (Patch.objects.filter(parent__isnull=True)
                  .annotate(ctrl_count=Count('controllers', distinct=True),
                            stn_count=Count('children', distinct=True),
                            sched_count=Count('schedules', distinct=True),
                            fz_count=Count('flow_zones', distinct=True))
                  .prefetch_related('children'))
    for site in root_sites:
        # Station details for hierarchy table (uses prefetched children — no query)
        station_list = []
        for stn in site.children.all():
            station_list.append({
                'id': stn.id,
                'name': stn.name or f'点位 {stn.controller_channel}',
                'controller_channel': stn.controller_channel,
                'controller_name': '-',
                'precip_rate': stn.precip_rate,
                'flow_rate': stn.flow_rate,
                'cycle_time': stn.cycle_time,
                'soak_time': stn.soak_time,
                'lockout': stn.lockout,
                'memo': stn.description,
            })

        sites.append({
            'id': site.id,
            'mdb_index': site.mdb_index,
            'name': site.name,
            'site_number': site.site_number,
            'et_current': site.et_current,
            'et_default': site.et_default,
            'water_pricing': site.water_pricing,
            'rain_shutdown': site.rain_shutdown,
            'controller_count': site.ctrl_count,
            'station_count': site.stn_count,
            'schedule_count': site.sched_count,
            'flow_zone_count': site.fz_count,
            'stations': station_list,
        })

    # Recent events (last 50)
    recent_events = []
    for ev in MaxicomEvent.objects.all()[:50]:
        recent_events.append({
            'timestamp': ev.timestamp,
            'source': ev.source,
            'flag': ev.flag,
            'text': ev.text,
        })

    # Weather summary: latest reading per station. Previously this ran one
    # `readings.order_by('-timestamp').first()` query PER station (N+1). Now a
    # single query fetches the latest-timestamp log per station via a Max()
    # subquery, then we read those rows. (weather_station,timestamp) is indexed.
    weather_summary = []
    stations = list(MaxicomWeatherStation.objects.all())
    if stations:
        from django.db.models import Max as _Max
        latest_per_station = {}
        for log in (MaxicomWeatherLog.objects
                    .filter(weather_station__in=stations,
                            timestamp__in=MaxicomWeatherLog.objects
                                .filter(weather_station__in=stations)
                                .values('weather_station')
                                .annotate(lt=_Max('timestamp'))
                                .values('lt'))
                    .select_related('weather_station')):
            # Keep only the first (latest) per station in case of timestamp ties.
            ws_id = log.weather_station_id
            if ws_id not in latest_per_station:
                latest_per_station[ws_id] = log
        for ws in stations:
            latest = latest_per_station.get(ws.id)
            if latest:
                weather_summary.append({
                    'station': ws.name,
                    'timestamp': latest.timestamp,
                    'temperature': latest.temperature,
                    'humidity': latest.humidity,
                    'rainfall': latest.rainfall,
                    'et': latest.et,
                    'wind_run': latest.wind_run,
                    'solar_radiation': latest.solar_radiation,
                })

    # ET trend: last 30 days of ET readings aggregated by day
    et_trend = []
    et_readings = MaxicomWeatherLog.objects.order_by('-timestamp')[:720]  # ~30 days * 24hrs
    et_by_date = {}
    for r in et_readings:
        ts = r.timestamp or ''
        if len(ts) >= 8:
            date_str = ts[:8]  # YYYYMMDD
            if date_str not in et_by_date:
                et_by_date[date_str] = []
            if r.et is not None:
                et_by_date[date_str].append(r.et)
    for date_str in sorted(et_by_date.keys()):
        vals = et_by_date[date_str]
        et_trend.append({
            'date': f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}",
            'avg_et': round(sum(vals) / len(vals), 3) if vals else 0,
            'count': len(vals),
        })

    # Station status breakdown
    station_status = {
        'total': Patch.objects.filter(parent__isnull=False).count(),
        'locked': Patch.objects.filter(parent__isnull=False, lockout=True).count(),
        'active': Patch.objects.filter(parent__isnull=False, lockout=False).count(),
    }

    # Top sites by station count
    top_sites = sorted(sites, key=lambda x: x['station_count'], reverse=True)[:10]

    # ET Checkbook latest
    et_checkbook_latest = []
    for ec in MaxicomETCheckbook.objects.select_related('site').order_by('-timestamp')[:20]:
        et_checkbook_latest.append({
            'site': ec.site.name,
            'timestamp': ec.timestamp,
            'soil_moisture': ec.soil_moisture,
            'rainfall': ec.rainfall,
            'et': ec.et,
            'irrigation': ec.irrigation,
        })

    return JsonResponse({
        'stats': stats,
        'sites': sites,
        'recent_events': recent_events,
        'weather_summary': weather_summary,
        'et_trend': et_trend,
        'station_status': station_status,
        'top_sites': top_sites,
        'et_checkbook': et_checkbook_latest,
    })


# ==========================================================================
# 维修工单系统 Views
# ==========================================================================


def _resolve_stats_window(request):
    """Resolve the stats date window from ?from=/&to= or ?week= (default: current week).

    Shared by stats_dashboard and the Excel export endpoint so they stay in sync.
    Returns (start_date, end_date).
    """
    from django.utils import timezone
    from datetime import datetime, timedelta
    today = timezone.now().date()
    from_param = request.GET.get('from')
    to_param = request.GET.get('to')
    week_param = request.GET.get('week')
    is_custom_range = bool(from_param or to_param)

    def _parse(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except Exception:
            return None

    if is_custom_range:
        start = _parse(from_param) or (today - timedelta(days=6))
        end = _parse(to_param) or today
        if end < start:
            start, end = end, start
    elif week_param:
        parsed = _parse(week_param)
        start = (parsed - timedelta(days=parsed.weekday())) if parsed else (today - timedelta(days=today.weekday()))
        end = start + timedelta(days=6)
    else:
        start = today - timedelta(days=today.weekday())  # current week, Monday
        end = start + timedelta(days=6)
    return start, end


def _work_report_count_columns():
    """Fixed fault-matrix columns for the Excel export: every count leaf node in
    the routine_maint section, returned as [(work_item_id, path_segments), ...].

    ``path_segments`` is the group hierarchy from the top level down to the leaf
    name (the section-root name itself is dropped), e.g.
    ``['计划性维修', '喷头', '喷嘴丢/坏']``. This lets the caller build a proper
    multi-row merged header instead of cramming the whole path into one cell.

    Sorted by path then id so columns are stable across exports regardless of
    which date range is selected.
    """
    from core.models import WorkItem
    items = {it.id: it for it in WorkItem.objects.filter(section='routine_maint')}

    def segs_of(it):
        parts, cur, seen = [], it, set()
        while cur and cur.id not in seen:
            seen.add(cur.id)
            parts.append(cur.name_zh)
            cur = items.get(cur.parent_id) if cur.parent_id else None
        parts.reverse()
        return parts[1:] or [it.name_zh]   # drop the section-root name itself

    cols = [(it.id, segs_of(it)) for it in items.values() if it.value_type == 'count']
    # Stable ordering: by path then by id (deterministic, same every export).
    cols.sort(key=lambda c: (c[1], c[0]))
    return cols


def _dedup_zone_names(raw):
    """Collapse repeated zone names in a stored zone_names string, preserving
    first-seen order. Older reports were filled before name-dedup existed, so a
    single value could list the same name dozens of times (e.g. 'BOH, BOH, ...').
    Returns '' for empty/None input.
    """
    if not raw:
        return ''
    parts = [p.strip() for p in raw.split(',') if p.strip()]
    out, seen = [], set()
    for p in parts:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return ', '.join(out)


def _report_section_label(report):
    """Work category label for a report (template's F column): derived from the
    sections of its entries, preferring non-routine ones; falls back to 常规维护."""
    from core.models import WorkReportEntry, WorkItem
    secs = WorkReportEntry.objects.filter(work_report=report).values_list('work_item__section', flat=True)
    label_map = dict(WorkItem.SECTION_CHOICES)
    # prefer the first non-routine section present; else 常规维护.
    for s in secs:
        if s and s != 'routine_maint':
            return label_map.get(s, s)
    return label_map.get('routine_maint', '常规维护')


def _report_materials_summary(report):
    """One-line summary of a report's material consumption, e.g. "喷头1804 ×2, 4寸管 ×3m".

    Reads the report's linked outbound transaction lines. Returns '' when no
    materials were consumed. Used to fill the previously-empty 消耗材料 column
    in both the Excel export and the on-screen preview."""
    parts = []
    for txn in getattr(report, '_materials_txn', []):
        for ln in getattr(txn, '_lines', []):
            unit = ln.unit or ''
            parts.append(f"{ln.category.name_zh} ×{int(ln.quantity)}{unit}")
    return '、'.join(parts)


@login_required(login_url='core:login')
def work_reports_excel(request):
    """Export work orders in the current stats date window as a reporttemplate-style
    Excel: one row per work order, fixed columns = the routine_maint count leaf
    nodes (fault matrix), blanks where a node wasn't filled. Admins see all
    reports; others see only their own — same scoping as stats_dashboard.
    """
    import io
    from core.models import WorkReport, WorkReportEntry, WorkItem, InventoryTransaction, InventoryTransactionLine
    from core.role_utils import is_admin
    from django.db.models import Prefetch
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    user = request.user
    admin = is_admin(user)
    start, end = _resolve_stats_window(request)

    # 1) Fixed fault-matrix columns (deterministic across exports). Each entry
    # is (work_item_id, [seg, seg, leaf]) — the group hierarchy with the section
    # root stripped, so we can render a proper multi-row merged header instead
    # of cramming the whole breadcrumb into one cell.
    count_nodes = _work_report_count_columns()

    # 2) Work orders in window (role-scoped). _scoped_work_reports_qs already
    # prefetches 'entries', so fetch the count-only subset into a distinct attr
    # via to_attr to avoid a duplicate-prefetch conflict.
    qs = _scoped_work_reports_qs(user, admin).filter(date__gte=start, date__lte=end)
    reports = list(qs.prefetch_related(
        Prefetch('entries', queryset=WorkReportEntry.objects.select_related('work_item').filter(work_item__value_type='count'), to_attr='_count_entries'),
        Prefetch('material_consumptions', queryset=InventoryTransaction.objects.prefetch_related(
            Prefetch('lines', queryset=InventoryTransactionLine.objects.select_related('category'), to_attr='_lines')
        ), to_attr='_materials_txn'),
    ).order_by('date', 'id'))

    # 3) Build workbook with a grouped (merged) multi-row header.
    wb = Workbook()
    ws = wb.active
    ws.title = '维修记录'
    base_header = ['序号', '日期', '工单号', '处理人', '位置', '工作分类',
                   '故障/事件位置', '区域', '灌溉组人数', '灌溉组工时',
                   '第三方人数', '第三方工时', '消耗材料', '备注',
                   '信息来源', '疑难问题', '疑难已处理']
    n_base = len(base_header)
    n_cols = n_base + len(count_nodes)
    hdr_rows = max((len(segs) for _, segs in count_nodes), default=1)

    # Header styling. Style every header cell up front so merged ranges keep
    # their borders (openpyxl only draws borders per underlying cell).
    hfill = PatternFill('solid', fgColor='1B4332')
    hfont = Font(color='FFFFFF', bold=True, size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='FFFFFF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for rr in range(1, hdr_rows + 1):
        for cc in range(1, n_cols + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = hfill; cell.font = hfont; cell.alignment = center; cell.border = border

    # Base columns span every header row (vertical merge).
    for ci, label in enumerate(base_header, 1):
        ws.cell(row=1, column=ci, value=label)
        if hdr_rows > 1:
            ws.merge_cells(start_row=1, start_column=ci, end_row=hdr_rows, end_column=ci)

    # Count columns: walk each depth level and merge consecutive siblings that
    # share the same path prefix. Columns are sorted by path, so siblings are
    # always contiguous. Shorter paths get merged down to fill remaining rows.
    for depth in range(hdr_rows):
        row = depth + 1
        ci = 0
        while ci < len(count_nodes):
            segs = count_nodes[ci][1]
            if depth >= len(segs):
                ci += 1          # already merged down at a shallower terminal depth
                continue
            prefix = segs[:depth + 1]
            cj = ci + 1
            while cj < len(count_nodes) and count_nodes[cj][1][:depth + 1] == prefix:
                cj += 1
            col_start = n_base + ci + 1
            col_end = n_base + cj              # inclusive
            ws.cell(row=row, column=col_start, value=segs[depth])
            need_h = col_end > col_start                 # siblings → merge across
            need_v = depth == len(segs) - 1 and row < hdr_rows   # terminal → merge down
            if need_h and need_v:              # one rectangular block merge
                ws.merge_cells(start_row=row, start_column=col_start,
                               end_row=hdr_rows, end_column=col_end)
            elif need_h:
                ws.merge_cells(start_row=row, start_column=col_start,
                               end_row=row, end_column=col_end)
            elif need_v:
                ws.merge_cells(start_row=row, start_column=col_start,
                               end_row=hdr_rows, end_column=col_start)
            # else: single cell, value already set — no merge needed
            ci = cj

    # work_item_id → matrix column index (1-based, offset after the base cols).
    id_to_col = {wid: n_base + i + 1 for i, (wid, _) in enumerate(count_nodes)}

    for idx, r in enumerate(reports, 1):
        # Zone codes: show only when the workorder covers ≤5 zones (sum of all
        # linked zone rows, regardless of duplicate names). Over 5 → blank, since
        # a sprawling multi-zone ticket's codes would be unreadable in one cell.
        zone_codes = ''
        zones = list(r.zones.all())
        if len(zones) <= 5:
            zone_codes = ', '.join(z.code for z in zones if z.code)
        row = [idx, r.date.isoformat() if r.date else '',
               r.display_number if r.id else '',
               r.worker.full_name if r.worker_id and r.worker else '',
               r.location.code if r.location_id and r.location else '',
               _report_section_label(r),
               _dedup_zone_names(r.zone_names),
               zone_codes,
               r.team_size or '',
               r.team_hours if r.team_hours else '',
               r.third_party_count or '',
               r.third_party_hours if r.third_party_hours else '',
               _report_materials_summary(r),
               (r.work_content or r.remark or ''),
               '',  # 信息来源 (no dedicated field on WorkReport)
               '是' if r.is_difficult else '',
               '是' if r.is_difficult_resolved else '']
        # pad matrix columns so indices line up
        row += [None] * len(count_nodes)
        for e in getattr(r, '_count_entries', []):
            ci = id_to_col.get(e.work_item_id)
            if ci is not None and e.count:
                row[ci - 1] = (row[ci - 1] or 0) + e.count
        ws.append(row)

    # Data cell styling: thin grey borders everywhere, centre numerics, wrap 备注.
    gside = Side(style='thin', color='D0D0D0')
    gborder = Border(left=gside, right=gside, top=gside, bottom=gside)
    dcenter = Alignment(horizontal='center', vertical='center')
    dwrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    last_row = hdr_rows + len(reports)
    # Wrap the long text columns: 区域 (codes list), 消耗材料, 备注 (free text).
    wrap_cols = {8, 13, 14}
    for rr in range(hdr_rows + 1, last_row + 1):
        for cc in range(1, n_cols + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.border = gborder
            cell.alignment = dwrap if cc in wrap_cols else dcenter

    # Column widths + freeze the header rows and the first five ID columns
    # (序号/日期/工单号/处理人/位置) so they stay visible while scrolling the matrix.
    from openpyxl.utils import get_column_letter
    # Per-column width overrides for the base columns (1-based):
    # 1序号 2日期 3工单号 4处理人 5位置 6工作分类 7故障位置 8区域
    # 9灌溉组人数 10灌溉组工时 11第三方人数 12第三方工时 13消耗材料 14备注
    # 15信息来源 16疑难问题 17疑难已处理
    base_widths = {3: 10, 4: 11, 5: 12, 6: 14, 7: 18, 8: 20,
                   13: 18, 14: 28}
    for ci in range(1, n_cols + 1):
        if ci <= n_base:
            width = base_widths.get(ci, 11)
        else:
            leaf = count_nodes[ci - n_base - 1][1][-1]
            width = max(8, min(20, len(leaf) * 1.7 + 2))
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = ws.cell(row=hdr_rows + 1, column=6)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f'workreports_{start}_{end}.xlsx'
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@never_cache
@login_required(login_url='core:login')
def stats_dashboard(request):
    """维修日志 数据报表 — hours, counts, distribution across section / project /
    land→name→zone, reflecting the new workorder tree (WorkItem + WorkReportEntry
    + Project) and the multi-zone WorkReport.zones M2M.

    Scope: a custom date range (?from= / ?to=) OR a calendar week (?week=). The
    default is the current week. Admins see all reports; other users see their own.
    """
    from core.models import WorkReport, WorkReportEntry, WorkItem, Patch, Zone, Project
    from core.role_utils import is_admin
    from django.db.models import Count, Q, Sum
    from django.utils import timezone
    from datetime import datetime, timedelta, date
    from collections import defaultdict

    user = request.user
    admin = is_admin(user)

    # === Resolve the date window: custom range takes precedence, else a week. ===
    today = timezone.now().date()
    from_param = request.GET.get('from')
    to_param = request.GET.get('to')
    week_param = request.GET.get('week')
    is_custom_range = bool(from_param or to_param)

    def _parse(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except Exception:
            return None

    if is_custom_range:
        start = _parse(from_param) or (today - timedelta(days=6))
        end = _parse(to_param) or today
        if end < start:
            start, end = end, start
    elif week_param:
        parsed = _parse(week_param)
        if parsed:
            start = parsed - timedelta(days=parsed.weekday())  # snap to Monday
        else:
            start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
    else:
        start = today - timedelta(days=today.weekday())  # current week, Monday
        end = start + timedelta(days=6)

    week_number = start.isocalendar()[1]

    # Week options for the <select> nav (3-year window).
    years = []
    year_weeks = {}
    current_year = today.year
    for year in [current_year - 1, current_year, current_year + 1]:
        weeks = []
        first_monday = date(year, 1, 1)
        while first_monday.weekday() != 0:
            first_monday = first_monday + timedelta(days=1)
        ws = first_monday
        wn = 1
        while ws.year == year or (ws.year == year + 1 and ws.month == 1 and wn <= 53):
            weeks.append({'week': wn, 'start': ws, 'end': ws + timedelta(days=6)})
            ws = ws + timedelta(days=7)
            wn += 1
            if wn > 53:
                break
        if weeks:
            years.append(year)
            year_weeks[year] = weeks

    # === Base queryset (role-scoped + windowed) ===
    base_qs = WorkReport.objects.select_related('worker', 'location')
    if not admin:
        try:
            base_qs = base_qs.filter(worker=user.worker_profile)
        except Exception:
            base_qs = base_qs.none()
    base_qs = base_qs.filter(date__gte=start, date__lte=end)

    # Materialize once — most aggregations below reuse this list. Prefetch the
    # material-consumption outbound lines too, so the Excel preview's 消耗材料
    # column can be filled without N+1 queries.
    from django.db.models import Prefetch
    from core.models import InventoryTransaction, InventoryTransactionLine
    reports = list(base_qs.prefetch_related(
        'zones__land', 'entries__work_item', 'entries__project',
        Prefetch('material_consumptions', queryset=InventoryTransaction.objects.prefetch_related(
            Prefetch('lines', queryset=InventoryTransactionLine.objects.select_related('category'), to_attr='_lines')
        ), to_attr='_materials_txn'),
    ))
    report_ids = [r.id for r in reports]

    section_labels = dict(WorkItem.SECTION_CHOICES)

    # ── Optional 通用名称 (zone.name) filter for the 工作内容·按章节 tree ────
    # ?zone_name=<name> scopes the tree to reports touching any zone whose 通用名称
    # matches. Empty = no filter (all reports).
    zone_filter = (request.GET.get('zone_name') or '').strip()
    zone_filter_label = '全部'
    # Distinct 通用名称 touched by the windowed reports, for the dropdown.
    _names = set()
    for r in reports:
        for z in r.zones.all():
            if z.name:
                _names.add(z.name)
    zone_filter_names = sorted(_names)

    if zone_filter:
        zone_filter_label = zone_filter
        def _name_match(report):
            return any(z.name == zone_filter for z in report.zones.all())
        tree_report_ids = {r.id for r in reports if _name_match(r)}
    else:
        tree_report_ids = set(report_ids)

    # ── 1. Overview tiles ───────────────────────────────────────────────────
    total_reports = len(reports)
    team_hours = sum(r.team_hours or 0 for r in reports)
    third_hours = sum(r.third_party_hours or 0 for r in reports)
    difficult_count = sum(1 for r in reports if r.is_difficult)
    pending_count = sum(1 for r in reports if r.is_pending_repair)

    # ── 2. Shift × hours breakdown ──────────────────────────────────────────
    shift_agg = defaultdict(lambda: {'count': 0, 'team_hours': 0.0, 'third_hours': 0.0})
    for r in reports:
        key = r.shift or '未指定'
        shift_agg[key]['count'] += 1
        shift_agg[key]['team_hours'] += r.team_hours or 0
        shift_agg[key]['third_hours'] += r.third_party_hours or 0
    shift_order = ['早班', '白班', '夜班', '未指定']
    shift_stats = [dict(name=k, **shift_agg[k]) for k in shift_order if k in shift_agg]

    # ── 3. By section → full WorkItem tree (every level shows its totals) ───
    # Build the section's WorkItem subtree from the entries' leaf nodes up their
    # parent chain, aggregating entries/counts/reports/hours at EVERY node so each
    # hierarchy level (section → group → subgroup → … → leaf) has its own totals.
    # Hours live on the WorkReport header; a report's hours credit a node once
    # (per distinct report) so a multi-node report isn't double-counted at one node.
    # Scoped to tree_report_ids when a Land/Zone filter is active.
    entries_qs = WorkReportEntry.objects.filter(
        work_report_id__in=tree_report_ids, work_item__active=True
    ).select_related('work_item', 'project')

    entries_total = entries_qs.count()
    entries_count_sum = entries_qs.filter(
        work_item__value_type='count'
    ).aggregate(s=Sum('count'))['s'] or 0

    # report_id → team_hours lookup (avoids re-fetching per entry).
    rep_hours = {r.id: (r.team_hours or 0) for r in reports}

    # Load the WorkItem tree skeleton (id → {name, parent_id, section, order,
    # value_type, unit}) once, so ancestor walks don't hit the DB per entry.
    wi_rows = WorkItem.objects.filter(active=True).values(
        'id', 'name_zh', 'parent_id', 'section', 'order', 'value_type', 'unit')
    wi_by_id = {r['id']: r for r in wi_rows}

    def _blank_node():
        return {'entries': 0, 'counts': 0, 'report_ids': set(), 'hours': 0.0, 'children': {}}

    # tree: section_code -> {node_agg, children:{work_item_id -> {node_agg, children:{...}}}}
    tree = {}

    def _ensure_chain(section, leaf_id):
        """Create/return the node-dicts along leaf_id's ancestor chain within
        this section, top (section root) → leaf, creating missing levels."""
        root = tree.setdefault(section, _blank_node())
        # Build the ancestor chain (root → ... → leaf), top-first.
        chain = []
        cur = leaf_id
        seen = set()
        while cur and cur in wi_by_id and cur not in seen:
            seen.add(cur)
            chain.append(cur)
            cur = wi_by_id[cur]['parent_id']
        chain.reverse()  # now top → leaf
        node = root
        for wid in chain:
            node = node['children'].setdefault(wid, _blank_node())
        return node

    for e in entries_qs:
        wi = wi_by_id.get(e.work_item_id)
        if not wi:
            continue
        section = wi['section'] or 'other'
        # Credit the leaf, then walk UP crediting every ancestor (so each level
        # shows its subtree total). Hours credited once per distinct report/node.
        credit_ids = [e.work_item_id]
        cur = wi['parent_id']
        seen = set()
        while cur and cur in wi_by_id and cur not in seen:
            seen.add(cur)
            credit_ids.append(cur)
            cur = wi_by_id[cur]['parent_id']
        # Credit the section root aggregate first (it carries the section total).
        root = tree.setdefault(section, _blank_node())
        root['entries'] += 1
        root['counts'] += e.count or 0
        if e.work_report_id not in root['report_ids']:
            root['report_ids'].add(e.work_report_id)
            root['hours'] += rep_hours.get(e.work_report_id, 0)
        for wid in credit_ids:
            node = _ensure_chain(section, wid)
            node['entries'] += 1
            node['counts'] += e.count or 0
            if e.work_report_id not in node['report_ids']:
                node['report_ids'].add(e.work_report_id)
                node['hours'] += rep_hours.get(e.work_report_id, 0)

    def _flatten(children_map):
        """Recursively turn the nested children dict into a sorted list of
        {name, entries, counts, reports, hours, children}. Empty subtrees pruned."""
        out = []
        for wid, node in children_map.items():
            wi = wi_by_id.get(wid, {})
            kids = _flatten(node['children'])
            # Only include a node if it (or any descendant) had entries.
            if node['entries'] == 0 and not kids:
                continue
            out.append({
                'id': wid,
                'name': wi.get('name_zh', str(wid)),
                'value_type': wi.get('value_type', ''),
                'unit': wi.get('unit', ''),
                'entries': node['entries'], 'counts': node['counts'],
                'reports': len(node['report_ids']), 'hours': round(node['hours'], 1),
                'children': kids,
            })
        # Sort: branches first (have children) by hours, then leaves by entries.
        out.sort(key=lambda x: (len(x['children']) == 0, -x['hours'], -x['entries']))
        return out

    section_rows = []
    for section, root in tree.items():
        kids = _flatten(root['children'])
        # A section is included if any descendant had entries.
        if root['entries'] == 0 and not kids:
            continue
        section_rows.append({
            'label': section_labels.get(section, section),
            'entries': root['entries'], 'counts': root['counts'],
            'reports': len(root['report_ids']), 'hours': round(root['hours'], 1),
            'children': kids,
        })
    section_rows.sort(key=lambda x: x['hours'], reverse=True)

    # Flat list for CSV export: every node (section + each descendant) on its own
    # row, with the full breadcrumb path and per-node totals. Mirrors what the
    # on-screen tree shows at every level.
    section_export = []

    def _export_walk(section_label, children, breadcrumb):
        for node in children:
            path = (breadcrumb + [node['name']]) if breadcrumb else [section_label, node['name']]
            section_export.append({
                'section': section_label,
                'path': ' › '.join(path),
                'level': len(path),
                'reports': node['reports'], 'hours': node['hours'],
                'entries': node['entries'], 'counts': node['counts'],
                'value_type': node.get('value_type', ''), 'unit': node.get('unit', ''),
            })
            _export_walk(section_label, node['children'], path)

    for sec in section_rows:
        # Section-root row first (the section total), then its subtree.
        section_export.append({
            'section': sec['label'], 'path': sec['label'], 'level': 1,
            'reports': sec['reports'], 'hours': sec['hours'],
            'entries': sec['entries'], 'counts': sec['counts'],
            'value_type': '', 'unit': '',
        })
        _export_walk(sec['label'], sec['children'], [])

    # ── 4. By project (hours credited per distinct report touching it) ──────
    project_rows = []
    proj_agg = defaultdict(lambda: {'entries': 0, 'counts': 0, 'report_ids': set(),
                                     'category': '', 'subcategory': '', 'name': ''})
    for e in entries_qs:
        if not e.project_id:
            continue
        p = e.project
        g = proj_agg[e.project_id]
        g['name'] = p.name
        g['category'] = p.get_category_display() if p.category else ''
        g['subcategory'] = p.get_subcategory_display() if p.subcategory else ''
        g['entries'] += 1
        g['counts'] += e.count or 0
        g['report_ids'].add(e.work_report_id)
    for pid, g in proj_agg.items():
        # Credit each project the sum of team_hours over the distinct reports that
        # touched it (a report may span multiple projects; each gets its full hours).
        hours = sum((r.team_hours or 0) for r in reports if r.id in g['report_ids'])
        project_rows.append({
            'name': g['name'], 'category': g['category'], 'subcategory': g['subcategory'],
            'entries': g['entries'], 'counts': g['counts'],
            'reports': len(g['report_ids']), 'hours': round(hours, 1),
        })
    project_rows.sort(key=lambda x: x['hours'], reverse=True)

    # ── 5. Land → 通用名称(Zone.name) → zone code (multi-zone tree) ─────────
    # Aggregate from WorkReport.zones (a report now carries many zones). Count a
    # report once per zone it touches, and credit that report's team_hours to each
    # of its zones (mirrors the project crediting rule).
    land_agg = defaultdict(lambda: {  # land_id -> {...}
        'name': '', 'report_ids': set(), 'hours': 0.0,
        'names': defaultdict(lambda: {  # 通用名称 -> {...}
            'report_ids': set(), 'hours': 0.0,
            'codes': defaultdict(lambda: {'report_ids': set(), 'hours': 0.0, 'name': ''}),
        }),
    })
    for r in reports:
        for z in r.zones.all():
            land_id = z.land_id
            land_name = z.land.name if (z.land_id and z.land) else '未分类'
            zname = z.name or z.code
            zcode = z.code
            # land level
            L = land_agg[land_id if land_id else '__none__']
            L['name'] = land_name
            L['report_ids'].add(r.id)
            L['hours'] += r.team_hours or 0
            # name level
            N = L['names'][zname]
            N['report_ids'].add(r.id)
            N['hours'] += r.team_hours or 0
            # code level
            C = N['codes'][zcode]
            C['name'] = zname
            C['report_ids'].add(r.id)
            C['hours'] += r.team_hours or 0
    # Flatten to sorted lists.
    land_rows = []
    for _lid, L in land_agg.items():
        land_rows.append({
            'name': L['name'],
            'reports': len(L['report_ids']),
            'hours': round(L['hours'], 1),
            'names': [{
                'name': nname,
                'reports': len(N['report_ids']),
                'hours': round(N['hours'], 1),
                'codes': [{
                    'code': ccode, 'reports': len(C['report_ids']),
                    'hours': round(C['hours'], 1),
                } for ccode, C in sorted(N['codes'].items(), key=lambda kv: -len(kv[1]['report_ids']))],
            } for nname, N in sorted(L['names'].items(), key=lambda kv: -len(kv[1]['report_ids']))],
        })
    land_rows.sort(key=lambda x: x['reports'], reverse=True)

    # ── 6. Worker hours (admin only) ────────────────────────────────────────
    worker_stats = []
    if admin:
        wmap = defaultdict(lambda: {'total': 0, 'team_hours': 0.0, 'third_hours': 0.0,
                                    'difficult': 0, 'name': '', 'employee_id': ''})
        for r in reports:
            if not r.worker_id:
                continue
            w = wmap[r.worker_id]
            w['name'] = r.worker.full_name if r.worker else ''
            w['employee_id'] = r.worker.employee_id if r.worker else ''
            w['total'] += 1
            w['team_hours'] += r.team_hours or 0
            w['third_hours'] += r.third_party_hours or 0
            if r.is_difficult:
                w['difficult'] += 1
        worker_stats = [{
            'name': w['name'], 'employee_id': w['employee_id'],
            'total': w['total'], 'team_hours': round(w['team_hours'], 1),
            'third_hours': round(w['third_hours'], 1), 'difficult': w['difficult'],
        } for w in sorted(wmap.values(), key=lambda x: x['team_hours'], reverse=True)]

    # ── 7. Excel-format preview table (admin only) ──────────────────────────
    # Same row/column layout as work_reports_excel so managers can preview the
    # export on screen: the base columns + the routine_maint count-leaf matrix
    # with a multi-row merged header. Built only for admins since the export
    # itself is admin-only. Keep this list in sync with work_reports_excel.
    excel_base_header = ['序号', '日期', '工单号', '处理人', '位置', '工作分类',
                         '故障/事件位置', '区域', '灌溉组人数', '灌溉组工时',
                         '第三方人数', '第三方工时', '消耗材料', '备注',
                         '信息来源', '疑难问题', '疑难已处理']
    excel_matrix_header, excel_preview_rows, excel_hdr_rows = [], [], 1
    excel_matrix_col_count = 0
    if admin:
        excel_matrix_cols = _work_report_count_columns()
        excel_matrix_col_count = len(excel_matrix_cols)
        excel_hdr_rows = max((len(segs) for _, segs in excel_matrix_cols), default=1)
        # Multi-row merged header for the count matrix, rendered as rows of
        # {text, colspan, rowspan}. Mirrors the openpyxl merge logic so the
        # on-screen header matches the exported spreadsheet exactly.
        for depth in range(excel_hdr_rows):
            hrow, ci = [], 0
            while ci < len(excel_matrix_cols):
                segs = excel_matrix_cols[ci][1]
                if depth >= len(segs):
                    ci += 1
                    continue
                prefix = segs[:depth + 1]
                cj = ci + 1
                while cj < len(excel_matrix_cols) and excel_matrix_cols[cj][1][:depth + 1] == prefix:
                    cj += 1
                terminal = depth == len(segs) - 1
                hrow.append({'text': segs[depth], 'colspan': cj - ci,
                             'rowspan': excel_hdr_rows - depth if terminal else 1})
                ci = cj
            if hrow:
                excel_matrix_header.append(hrow)
        # Data rows in Excel order (date, id), one cell per matrix column.
        excel_count_wids = {wid for wid, _ in excel_matrix_cols}
        for idx, r in enumerate(sorted(reports, key=lambda x: (x.date, x.id)), 1):
            zones = list(r.zones.all())
            zone_codes = ', '.join(z.code for z in zones if z.code) if len(zones) <= 5 else ''
            counts_map = {}
            for e in r.entries.all():
                if e.work_item_id in excel_count_wids and e.count:
                    counts_map[e.work_item_id] = (counts_map.get(e.work_item_id, 0) or 0) + e.count
            excel_preview_rows.append({
                'idx': idx,
                'date': r.date.isoformat() if r.date else '',
                'workorder_no': f'#{r.id}' if r.id else '',
                'worker': r.worker.full_name if (r.worker_id and r.worker) else '',
                'location': r.location.code if (r.location_id and r.location) else '',
                'section': _report_section_label(r),
                'zone_names': _dedup_zone_names(r.zone_names),
                'zone_codes': zone_codes,
                'team_size': r.team_size or '',
                'team_hours': r.team_hours if r.team_hours else '',
                'third_party_count': r.third_party_count or '',
                'third_party_hours': r.third_party_hours if r.third_party_hours else '',
                'materials': _report_materials_summary(r),
                'remark': r.work_content or r.remark or '',
                'difficult': '是' if r.is_difficult else '',
                'resolved': '是' if r.is_difficult_resolved else '',
                'counts': [counts_map.get(wid, '') for wid, _ in excel_matrix_cols],
            })

    context = {
        'is_admin': admin,
        'start': start, 'end': end,
        'is_custom_range': is_custom_range,
        'from_param': from_param or start.isoformat(),
        'to_param': to_param or end.isoformat(),
        'today_iso': today.isoformat(),
        'week_number': week_number,
        'week_iso': start.isoformat(),
        'years': years, 'year_weeks': year_weeks,
        # 1. overview
        'total_reports': total_reports,
        'team_hours': round(team_hours, 1),
        'third_hours': round(third_hours, 1),
        'difficult_count': difficult_count,
        'pending_count': pending_count,
        # 2. shift
        'shift_stats': shift_stats,
        # 3. section + node drilldown (optionally scoped to a 通用名称 filter)
        'entries_total': entries_total,
        'entries_count_sum': entries_count_sum,
        'section_rows': section_rows,
        'section_export': section_export,
        'zone_filter': zone_filter,
        'zone_filter_label': zone_filter_label,
        'zone_filter_names': zone_filter_names,
        # 4. project
        'project_rows': project_rows,
        # 5. land → name → zone
        'land_rows': land_rows,
        # 6. worker hours
        'worker_stats': worker_stats,
        # 7. Excel-format preview (admin only)
        'excel_base_header': excel_base_header,
        'excel_matrix_header': excel_matrix_header,
        'excel_matrix_col_count': excel_matrix_col_count,
        'excel_hdr_rows': excel_hdr_rows,
        'excel_preview_rows': excel_preview_rows,
    }
    return render(request, 'core/stats_dashboard.html', context)


# ==========================================================================
# Irrigation Dashboard — Maxicom runtime pivot (station × controller)
# ==========================================================================

def _sat_sort_key(name):
    """Sort satellite controller names (e.g. 'AI 7-1') by (site num, sat num)."""
    import re as _re
    m = _re.search(r'(\d+)-(\d+)', name)
    site_num = int(m.group(1)) if m else 0
    sat_num = int(m.group(2)) if m else 0
    return (site_num, sat_num, name)


def _ccu_sort_key(patch):
    """Natural numeric sort for CCU patches by the number in the code (CCU1 < CCU2 < ... < CCU10)."""
    import re as _re
    m = _re.search(r'(\d+)', patch.code or '')
    return (int(m.group(1)) if m else 0, patch.code or '')


def _ccu_queryset():
    """CCU patches in natural numeric order (CCU1, CCU2, ... CCU9, CCU10, CCU11)."""
    ccus = list(Patch.objects.filter(parent__isnull=True, code__iregex=r'^CCU\d+$'))
    ccus.sort(key=_ccu_sort_key)
    return ccus


def _irrig_window(request):
    """Resolve the irrigation dashboard's `from`/`to` GET params.

    Returns (date_from_compact, date_to_compact, ts_from_14, ts_to_14) where
    the compact form is the raw YYYYMMDDHHMM the user passed (or yesterday's
    full day by default) and the ts_* forms are padded to 14-char strings
    suitable for string-range matching against MaxicomRuntime.timestamp.

    Shared by the irrigation dashboard view and the zone-heatmap API so both
    honour the exact same date-window semantics.
    """
    yesterday = timezone.localdate() - timedelta(days=1)
    ys = yesterday.strftime('%Y%m%d')
    default_from = ys + '0000'
    default_to = ys + '2359'
    date_from = request.GET.get('from', default_from).strip()
    date_to = request.GET.get('to', default_to).strip()
    ts_from = date_from.ljust(14, '0')[:14] if date_from else ''
    ts_to = date_to.ljust(14, '9')[:14] if date_to else ''
    return date_from, date_to, ts_from, ts_to


def _irrig_data_span():
    """Return (first_ts, last_ts) — the min/max MaxicomRuntime.timestamp.

    Used by the dashboard (for the 数据范围 hint and the 重置 default) and by
    the PDF/Excel exports (for their default window when no params given).
    Implemented as a single Min/Max aggregate rather than separate
    .exists()/.first()/.last() calls (3 queries → 1).
    """
    from core.models import MaxicomRuntime
    from django.db.models import Min, Max
    row = MaxicomRuntime.objects.aggregate(min_ts=Min('timestamp'), max_ts=Max('timestamp'))
    return (row['min_ts'] or '', row['max_ts'] or '')


def _mapped_station_ids():
    """Return the set of station Patch IDs that at least one landscape Zone
    claims via its ``maxicom_runtime`` field.

    A station NOT in this set is an "orphan": its runtime is real hardware
    data, but no Zone maps to it, so it cannot be attributed to any landscape
    area. The dashboard / PDF / Excel flag orphan rows in red so reviewers
    know those valves have no Zone backing.
    """
    from core.models import Zone
    mapped = set()
    for raw in Zone.objects.exclude(maxicom_runtime=[]).values_list('maxicom_runtime', flat=True):
        if not raw:
            continue
        if isinstance(raw, list):
            for sid in raw:
                try:
                    mapped.add(int(sid))
                except (TypeError, ValueError):
                    pass
        # JSONField never returns non-list here, but be defensive.
    return mapped


def _build_ccu_matrix(ccu, rt_qs, ctrl_map, mapped_station_ids=None):
    """Build the 24 x N-satellite runtime matrix for one CCU.

    Returns dict: {controllers, rows, col_totals, grand_total, max_cell}
      - controllers: sorted list of satellite names (columns)
      - rows: 24 entries, station '01'..'24', each with values[] + total,
        plus ``orphan`` (True if no Zone maps to this row's station).
      - col_totals: list aligned to controllers
    Used by both the dashboard view and the PDF/Excel exports.
    """
    from core.models import Patch
    from collections import defaultdict

    if mapped_station_ids is None:
        mapped_station_ids = set()

    ccu_sats = sorted(
        {c.name for c in ctrl_map.values() if c.site_id == ccu.id},
        key=_sat_sort_key,
    )
    all_controllers = ccu_sats
    ctrl_index = {c: i for i, c in enumerate(all_controllers)}

    # All valves under this CCU, ensure every (satellite, channel) exists.
    # sat_chan_minutes: satellite_name -> {channel: [minutes, station_patch_id]}
    # Tracking the station_id per (sat, channel) lets us flag orphan rows.
    sat_chan_minutes = defaultdict(dict)
    sat_chan_station = defaultdict(dict)
    for st in Patch.objects.filter(parent=ccu, controller_channel__isnull=False):
        ctrl = ctrl_map.get(st.controller_number)
        sat_name = ctrl.name if ctrl else f'SAT {st.controller_number}'
        sat_chan_minutes[sat_name][st.controller_channel] = 0
        sat_chan_station[sat_name][st.controller_channel] = st.id

    # Fill in actual runtime minutes within the date range.
    for rt in rt_qs:
        st = rt.station
        if st is None:
            continue
        ctrl = ctrl_map.get(st.controller_number)
        sat_name = ctrl.name if ctrl else f'SAT {st.controller_number}'
        if sat_name in ctrl_index:
            sat_chan_minutes[sat_name][st.controller_channel] = \
                sat_chan_minutes[sat_name].get(st.controller_channel, 0) + (rt.run_time or 0)
            sat_chan_station[sat_name][st.controller_channel] = st.id

    rows = []
    col_totals = defaultdict(int)
    for ch in range(1, 25):
        vals = [0] * len(all_controllers)
        orphan_vals = [False] * len(all_controllers)   # per-cell orphan flags
        for sat_name, idx in ctrl_index.items():
            v = sat_chan_minutes.get(sat_name, {}).get(ch, 0)
            sid = sat_chan_station.get(sat_name, {}).get(ch)
            if v:
                vals[idx] = v
                col_totals[sat_name] += v
            # A cell is orphan if its station Patch exists but no Zone claims
            # it. Per-cell because one row spans many satellites, each with its
            # own valve; flagging the whole row would smear one valve's status
            # across siblings (e.g. CCU1 ch23 has SAT 1-1 orphan but SAT 1-4
            # mapped via zone 1-4-23).
            if sid is not None and sid not in mapped_station_ids:
                orphan_vals[idx] = True
        non_zero = [v for v in vals if v > 0]
        non_zero_orphan = [orphan_vals[i] for i, v in enumerate(vals) if v > 0]
        rows.append({
            'station': f"{ch:02d}",
            'channel': ch,
            'site': ccu.code,
            'values': vals,
            'orphan_values': orphan_vals,
            # cell_views (value, orphan, alpha) is filled in after max_cell is
            # known — see the second pass below.
            'total': sum(vals),
            # Row is "fully orphan" only when every runtime cell is orphan —
            # used for the side label hint. Individual cells are flagged
            # separately via orphan_values.
            'orphan': bool(non_zero) and all(non_zero_orphan),
        })
    grand_total = sum(col_totals.values())
    # max_cell drives the heatmap color scale. Exclude orphan valves so their
    # (often large) un-attributable runtime doesn't wash out the scale for the
    # valid cells — orphans get a flat light-red fill instead of heat shading.
    max_cell = 0
    for r in rows:
        ovals = r.get('orphan_values') or []
        for i, v in enumerate(r['values']):
            if v > 0 and not (i < len(ovals) and ovals[i]):
                if v > max_cell:
                    max_cell = v
    # Second pass: build cell_views tuples with a pre-computed alpha (v/max_cell
    # rounded to 2 decimals). Templates can't do float division safely — Django's
    # {% widthratio %} emits an integer, which would produce invalid CSS like
    # rgba(82,183,136,50/100). Doing the math here keeps the template a clean
    # interpolation. Tuple shape: (value, orphan, alpha_string).
    for r in rows:
        ovals = r.get('orphan_values') or []
        cv = []
        for i, v in enumerate(r['values']):
            is_orphan = i < len(ovals) and ovals[i]
            alpha = '0.00'
            if max_cell and v > 0:
                alpha = f'{min(v / max_cell, 1.0):.2f}'
            cv.append((v, is_orphan, alpha))
        r['cell_views'] = cv
    return {
        'controllers': all_controllers,
        'rows': rows,
        'col_totals': [col_totals[c] for c in all_controllers],
        'grand_total': grand_total,
        'max_cell': max_cell,
    }


@never_cache
@login_required(login_url='core:login')
def irrigation_dashboard(request):
    """Maxicom irrigation runtime dashboard.

    Pivot: rows = station, columns = controller (satellite), cell = sum of
    runtime minutes within the selected CCU and date range. A runtime row
    carries run_time=1 per active minute, so SUM(run_time) = minutes run.
    """
    from core.models import Patch, MaxicomController, MaxicomRuntime
    from django.db.models import Sum
    from collections import defaultdict
    from datetime import timedelta
    import json as _json

    # --- filter inputs ---
    ccus = _ccu_queryset()   # natural numeric order: CCU1, CCU2, ... CCU10

    # CCU scope (default: all)
    ccu_param = request.GET.get('ccu', '')
    ccu_obj = None
    if ccu_param:
        ccu_obj = next((c for c in ccus if str(c.id) == ccu_param), None)

    # full available data span — surfaced as the "数据范围" hint and restored by
    # the 重置 button (zoom back out to everything). Single Min/Max aggregate
    # (P2 fix) instead of three .exists()/.first()/.last() round-trips.
    first_ts, last_ts = _irrig_data_span()
    span_from = first_ts[:12] if len(first_ts) >= 12 else (first_ts[:8] if first_ts else '')
    span_to = last_ts[:12] if len(last_ts) >= 12 else (last_ts[:8] if last_ts else '')

    # default range = the last complete irrigation day (yesterday 00:00→23:59).
    # Today is partial: overnight runs are rolled to the next date by the
    # importer, so yesterday is the most recent full day. Shared with the
    # zone-heatmap endpoint via _irrig_window so both honour identical windows.
    date_from, date_to, ts_from, ts_to = _irrig_window(request)

    # --- build the pivot ---
    rt_qs = MaxicomRuntime.objects.select_related('station', 'station__parent', 'site')
    if ccu_obj is not None:
        rt_qs = rt_qs.filter(site=ccu_obj)
    if ts_from:
        rt_qs = rt_qs.filter(timestamp__gte=ts_from)
    if ts_to:
        rt_qs = rt_qs.filter(timestamp__lte=ts_to)

    # satellite lookup: MaxicomController.mdb_index -> controller.
    ctrl_map = {c.mdb_index: c for c in MaxicomController.objects.exclude(name__icontains='CCU')}

    # Station Patch IDs claimed by at least one landscape Zone. A station not
    # in this set is an "orphan" (no Zone backing) and its row gets flagged.
    mapped_station_ids = _mapped_station_ids()

    # For a specific CCU, build a true 24 x N-satellites matrix: rows are the
    # Maxicom station numbers 01-24 (always shown, even with no runtime), columns
    # are the CCU's satellites, and each cell is the runtime of the valve at that
    # (channel, satellite). For "all CCUs" the valves span many satellites so we
    # keep the per-valve-row layout (one row per valve that ran).
    if ccu_obj is not None:
        m = _build_ccu_matrix(ccu_obj, rt_qs, ctrl_map, mapped_station_ids)
        all_controllers = m['controllers']
        rows = m['rows']
        col_totals = defaultdict(int, dict(zip(all_controllers, m['col_totals'])))
        grand_total = m['grand_total']
        max_cell = m['max_cell']
        total_stations_with_runtime = sum(1 for r in rows if r['total'] > 0)

    else:
        # "All CCUs" — one row per valve that actually ran (across many satellites).
        # rt_qs is select_related('station','station__parent','site'), so we
        # collect everything we need (minutes + station metadata) in a single
        # pass — no second Patch.objects.filter round-trip (P3 fix).
        station_minutes = defaultdict(int)   # station_id -> total minutes
        station_meta = {}                    # station_id -> (channel, site_code, controller_number)
        for rt in rt_qs:
            st = rt.station
            if st is None:
                continue
            station_minutes[st.id] += (rt.run_time or 0)
            station_meta[st.id] = (
                st.controller_channel,
                rt.site.code if rt.site else '',
                st.controller_number,
            )

        station_rows = []
        sat_names_seen = set()
        for sid, total in station_minutes.items():
            channel, site_code, ctrl_num = station_meta.get(sid, (None, '', None))
            ctrl = ctrl_map.get(ctrl_num) if ctrl_num is not None else None
            sat_name = ctrl.name if ctrl else f'SAT {ctrl_num}'
            sat_names_seen.add(sat_name)
            station_rows.append({
                'station': f"{channel:02d}" if channel is not None else "—",
                'channel': channel,
                'site': site_code,
                'satellite': sat_name,
                'total': total,
                'orphan': sid not in mapped_station_ids,
            })

        all_controllers = sorted(sat_names_seen, key=_sat_sort_key)
        ctrl_index = {c: i for i, c in enumerate(all_controllers)}
        rows = []
        col_totals = defaultdict(int)
        station_rows.sort(key=lambda r: (r['site'], _sat_sort_key(r['satellite']), r['channel'] or 0))
        # All-CCU rows are one-valve-per-row, so orphan is per-row here. We
        # also build orphan_values (a single-element list) so the JS renderer
        # — which reads orphan_values uniformly — flags these rows too (B2 fix).
        for meta in station_rows:
            vals = [0] * len(all_controllers)
            vals[ctrl_index[meta['satellite']]] = meta['total']
            meta['values'] = vals
            meta['orphan_values'] = [meta['orphan']]
            rows.append(meta)
            col_totals[meta['satellite']] += meta['total']
        grand_total = sum(col_totals.values())
        # All-CCU rows are one-valve-per-row, so orphan is per-row here.
        # Excluded so un-attributable runtime doesn't wash out the heat scale.
        max_cell = max((max(r['values'], default=0) for r in rows if not r.get('orphan')), default=0)
        # Build cell_views with pre-computed alpha (mirrors _build_ccu_matrix's
        # second pass) for the server-rendered pivot partial.
        for r in rows:
            cv = []
            for i, v in enumerate(r['values']):
                is_orphan = i < len(r['orphan_values']) and r['orphan_values'][i]
                alpha = '0.00'
                if max_cell and v > 0:
                    alpha = f'{min(v / max_cell, 1.0):.2f}'
                cv.append((v, is_orphan, alpha))
            r['cell_views'] = cv
        total_stations_with_runtime = len(rows)

    # Note: single-CCU path inherits max_cell from _build_ccu_matrix, which
    # already excludes orphan cells (per-cell, since one row spans multiple
    # satellites). The all-CCU path computes max_cell just above.

    # user role (mirrors other dashboard views)
    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            from core.models import ManagerProfile
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except Exception:
            pass

    # The CCU column is only useful when multiple CCUs are in view (no CCU picked)
    show_ccu_col = ccu_obj is None

    # ISO datetime forms for the native <input type="datetime-local"> picker
    # (YYYY-MM-DDTHH:MM), which the template binds directly. The compact
    # YYYYMMDDHHMM form is kept for export URLs / backend filtering.
    def _iso_dt(s):
        s = (s or '').strip()
        if len(s) < 8:
            return s
        hh = s[8:10] if len(s) >= 10 else '00'
        mm = s[10:12] if len(s) >= 12 else '00'
        return s[:4] + '-' + s[4:6] + '-' + s[6:8] + 'T' + hh + ':' + mm
    datetime_from_iso = _iso_dt(date_from)
    datetime_to_iso = _iso_dt(date_to)
    today_iso = timezone.localdate().isoformat()
    data_span_dt = (_iso_dt(span_from), _iso_dt(span_to))

    context = {
        'is_admin': is_admin,
        'ccus': ccus,
        'selected_ccu': ccu_obj,
        'date_from': date_from,
        'date_to': date_to,
        'data_span': (span_from, span_to),
        'datetime_from_iso': datetime_from_iso,
        'datetime_to_iso': datetime_to_iso,
        'today_iso': today_iso,
        'data_span_dt': data_span_dt,
        'controllers': all_controllers,
        'rows': rows,
        'col_totals': [col_totals[c] for c in all_controllers],
        'grand_total': grand_total,
        'max_cell': max_cell,
        'total_stations': total_stations_with_runtime,
        'show_ccu_col': show_ccu_col,
    }

    # JSON response for AJAX refresh (filters changed without full reload).
    # Triggered by either the AJAX header (sent by the page's fetch) or an
    # explicit ?format=json for testing.
    if request.GET.get('format') == 'json' or \
            request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'controllers': all_controllers,
            'rows': rows,
            'col_totals': [col_totals[c] for c in all_controllers],
            'grand_total': grand_total,
            'max_cell': max_cell,
            'total_stations': total_stations_with_runtime,
            'date_from': date_from,
            'date_to': date_to,
            'show_ccu_col': show_ccu_col,
        })

    return render(request, 'core/irrigation_dashboard.html', context)


@login_required(login_url='core:login')
def irrigation_zone_heatmap(request):
    """API: per-zone irrigation runtime minutes for the /irrigation/ heatmap.

    Same date-window semantics as ``irrigation_dashboard`` (via
    ``_irrig_window``). Each zone carries its boundary points + center plus the
    sum of ``MaxicomRuntime.run_time`` over the station Patch ids stored in
    ``Zone.maxicom_runtime`` (populated by populate_zone_maxicom_runtime).

    Returns ``{zones: [...], max: <minutes>}`` shaped to match stats_zone_heatmap
    so the client can reuse the same Leaflet polygon renderer.
    """
    from core.models import Zone, MaxicomRuntime
    from django.db.models import Sum
    from collections import defaultdict

    _date_from, _date_to, ts_from, ts_to = _irrig_window(request)

    # active_boundary_points is a Python @property (resolves dxf vs manual), so
    # we can't filter on it in the ORM — load all zones and filter in Python.
    # Only zones with both a boundary AND a maxicom_runtime mapping are useful
    # for the map. Zones WITH a mapping but NO boundary are surfaced separately
    # as "unmapped" so reviewers know which zones need polygons drawn.
    all_zones = list(Zone.objects.select_related('land'))
    zones_out = []
    unmapped_entries = []      # zones with runtime mapping but no boundary
    station_ids = set()
    for z in all_zones:
        try:
            sids = [int(sid) for sid in (z.maxicom_runtime or []) if sid is not None]
        except (TypeError, ValueError):
            continue
        if not sids:
            continue
        station_ids.update(sids)
        bp = z.active_boundary_points
        if bp:
            zones_out.append({'zone': z, 'bp': bp, 'station_ids': sids})
        else:
            unmapped_entries.append({'zone': z, 'station_ids': sids})

    # Single aggregated query: station_id -> sum of runtime minutes.
    rt_qs = MaxicomRuntime.objects.filter(station_id__in=station_ids)
    if ts_from:
        rt_qs = rt_qs.filter(timestamp__gte=ts_from)
    if ts_to:
        rt_qs = rt_qs.filter(timestamp__lte=ts_to)
    minutes_by_station = defaultdict(int)
    for row in rt_qs.values('station_id').annotate(minutes=Sum('run_time')):
        minutes_by_station[row['station_id']] = row['minutes'] or 0

    out = []
    max_minutes = 0
    for entry in zones_out:
        z = entry['zone']
        minutes = sum(minutes_by_station.get(sid, 0) for sid in entry['station_ids'])
        if minutes > max_minutes:
            max_minutes = minutes
        out.append({
            'id': z.id,
            'code': z.code,
            'name': z.name or '',
            'land_name': z.land.name if z.land else '',
            'boundary_points': entry['bp'],
            'center': get_zone_center(entry['bp']),
            'runtime_minutes': minutes,
        })

    # Unmapped zones (runtime mapping exists, but no polygon to draw). Sorted
    # by runtime desc so the heaviest-irrigated missing zones surface first —
    # those are the most valuable to draw next.
    unmapped_out = []
    for entry in unmapped_entries:
        z = entry['zone']
        minutes = sum(minutes_by_station.get(sid, 0) for sid in entry['station_ids'])
        unmapped_out.append({
            'id': z.id,
            'code': z.code,
            'name': z.name or '',
            'land_name': z.land.name if z.land else '',
            'runtime_minutes': minutes,
        })
    unmapped_out.sort(key=lambda r: r['runtime_minutes'], reverse=True)

    return JsonResponse({
        'zones': out,
        'max': max_minutes,
        'unmapped': unmapped_out,
        'date_from': _date_from,
        'date_to': _date_to,
    }, json_dumps_params={'ensure_ascii': False})


@login_required(login_url='core:login')
def stats_zone_heatmap(request):
    """API: per-zone work-order stats for the 数据报表 heatmap tab.

    Returns ``{zones: [...], max: {reports, hours, difficult}}`` scoped to the
    same date window the stats page uses (via _resolve_stats_window). Each zone
    carries its boundary points + center (so the JS can draw polygons without a
    second fetch) and four metric values: report count, total hours, entry count
    (填报数量), and difficult/pending count.
    """
    from core.models import Zone, WorkReport, WorkReportEntry
    from collections import defaultdict

    # --- resolve date range (same helper as stats_dashboard) ---
    start, end = _resolve_stats_window(request)
    wr_qs = WorkReport.objects.filter(date__gte=start, date__lte=end)

    # --- optional worker filter (?workers=1,3,5) ---
    workers_param = request.GET.get('workers', '').strip()
    if workers_param:
        wids = [int(x) for x in workers_param.split(',') if x.strip().isdigit()]
        if wids:
            wr_qs = wr_qs.filter(worker_id__in=wids)

    # active_boundary_points is a Python @property (resolves dxf vs manual), so
    # we can't filter on it in the ORM — load all zones and filter in Python.
    all_zones = list(Zone.objects.select_related('land'))
    zones_with_boundary = [z for z in all_zones if z.active_boundary_points]
    zone_id_set = set(z.id for z in zones_with_boundary)
    # zone_id -> zone object (for boundary lookup during output).
    zone_map = {z.id: z for z in zones_with_boundary}

    # Pre-aggregate entry counts per report (sum of WorkReportEntry.count).
    report_ids = list(wr_qs.values_list('id', flat=True))
    entry_count_map = defaultdict(int)
    if report_ids:
        from django.db.models import Sum
        for row in (WorkReportEntry.objects
                    .filter(work_report_id__in=report_ids)
                    .values('work_report_id')
                    .annotate(total=Sum('count'))):
            entry_count_map[row['work_report_id']] = row['total'] or 0

    # ── Deduplicated aggregation at all 3 hierarchy levels ────────────────
    # At zone level, a report touching N zones credits each zone once (full
    # hours). But at land/name level, the same report must count only ONCE per
    # land (or name) — otherwise a multi-zone report is double-counted when its
    # zones are aggregated upward.
    #
    # We do this by iterating reports and, for each, crediting each DISTINCT
    # land / name / zone it touches exactly once.
    SCOPE_KEYS = ['all', 'difficult', 'pending', 'risk']

    def _blank():
        return {k: {'reports': 0, 'team_hours': 0.0, 'third_hours': 0.0}
                for k in SCOPE_KEYS}

    # agg[entity_key][scope] = counters. entity_key is prefixed by level so the
    # same land appearing in different names doesn't collide.
    agg = defaultdict(lambda: {'all': _blank(), 'difficult': _blank(),
                               'pending': _blank(), 'risk': _blank()})
    # Actually simpler: 4 separate dicts, one per scope.
    zone_agg = {k: defaultdict(lambda: {'reports': 0, 'team_hours': 0.0, 'third_hours': 0.0})
                for k in SCOPE_KEYS}
    name_agg = {k: defaultdict(lambda: {'reports': 0, 'team_hours': 0.0, 'third_hours': 0.0})
                for k in SCOPE_KEYS}
    land_agg = {k: defaultdict(lambda: {'reports': 0, 'team_hours': 0.0, 'third_hours': 0.0})
                for k in SCOPE_KEYS}

    for r in wr_qs.prefetch_related('zones').iterator(chunk_size=200):
        # Collect the distinct zones / names / lands this report touches.
        r_zone_ids = []
        seen_names = set()
        seen_lands = set()
        for z in r.zones.all():
            if z.id not in zone_id_set:
                continue
            land = z.land.name if z.land else '未分类'
            nm = z.name or z.code
            name_key = land + '||' + nm
            r_zone_ids.append(z.id)
            seen_names.add(name_key)
            seen_lands.add(land)
        if not r_zone_ids:
            continue

        team_h = r.team_hours or 0
        third_h = r.third_party_hours or 0
        scopes = ['all']
        if r.is_difficult:
            scopes.append('difficult')
        if r.is_pending_repair:
            scopes.append('pending')
        if r.is_difficult or r.is_pending_repair:
            scopes.append('risk')

        # Credit zone level: each touched zone gets full credit.
        for zid in r_zone_ids:
            for k in scopes:
                d = zone_agg[k][zid]
                d['reports'] += 1
                d['team_hours'] += team_h
                d['third_hours'] += third_h
        # Credit name level: each distinct name gets credit ONCE.
        for nk in seen_names:
            for k in scopes:
                d = name_agg[k][nk]
                d['reports'] += 1
                d['team_hours'] += team_h
                d['third_hours'] += third_h
        # Credit land level: each distinct land gets credit ONCE.
        for lk in seen_lands:
            for k in scopes:
                d = land_agg[k][lk]
                d['reports'] += 1
                d['team_hours'] += team_h
                d['third_hours'] += third_h

    def _scope_dict(agg_dict, key):
        """Build the 4-scope breakdown (all/difficult/pending/risk) for one key."""
        def _round_sc(d):
            th = round(d['team_hours'], 1)
            tph = round(d['third_hours'], 1)
            return {
                'reports': d['reports'],
                'team_hours': th,
                'third_hours': tph,
                'hours': round(th + tph, 1),
            }
        return {
            'all': _round_sc(agg_dict['all'][key]),
            'difficult': _round_sc(agg_dict['difficult'][key]),
            'pending': _round_sc(agg_dict['pending'][key]),
            'risk': _round_sc(agg_dict['risk'][key]),
        }

    # ── Build the zone-level output (same shape as before, for the map) ────
    zones_out = []
    for z in zones_with_boundary:
        bp = z.active_boundary_points
        center = get_zone_center(bp)
        scopes = _scope_dict(zone_agg, z.id)
        a = scopes['all']
        zones_out.append({
            'id': z.id,
            'code': z.code,
            'name': z.name or '',
            'land_name': z.land.name if z.land else '',
            'boundary_points': bp,
            'center': center,
            'reports': a['reports'],
            'team_hours': a['team_hours'],
            'third_hours': a['third_hours'],
            'hours': a['hours'],
            'f_difficult': scopes['difficult'],
            'f_pending': scopes['pending'],
            'f_risk': scopes['risk'],
        })

    # ── Build name-level + land-level aggregates (for the higher granularities) ─
    # Collect all distinct name_keys and land_keys that appeared.
    name_keys = set()
    land_keys = set()
    for k in SCOPE_KEYS:
        name_keys.update(name_agg[k].keys())
        land_keys.update(land_agg[k].keys())

    names_out = []
    for nk in name_keys:
        parts = nk.split('||', 1)
        land = parts[0] if len(parts) > 0 else '未分类'
        nm = parts[1] if len(parts) > 1 else ''
        scopes = _scope_dict(name_agg, nk)
        a = scopes['all']
        names_out.append({
            'key': nk,
            'land_name': land,
            'name': nm,
            'reports': a['reports'],
            'team_hours': a['team_hours'],
            'third_hours': a['third_hours'],
            'hours': a['hours'],
            'f_difficult': scopes['difficult'],
            'f_pending': scopes['pending'],
            'f_risk': scopes['risk'],
        })

    lands_out = []
    for lk in land_keys:
        scopes = _scope_dict(land_agg, lk)
        a = scopes['all']
        lands_out.append({
            'key': lk,
            'land_name': lk,
            'reports': a['reports'],
            'team_hours': a['team_hours'],
            'third_hours': a['third_hours'],
            'hours': a['hours'],
            'f_difficult': scopes['difficult'],
            'f_pending': scopes['pending'],
            'f_risk': scopes['risk'],
        })

    def _max_from(rows, metric):
        return max((r[metric] for r in rows), default=0)

    # Distinct workers in the (unfiltered) date window — for the dropdown. We
    # query the base window (ignoring the workers param) so the list is stable
    # regardless of the current filter selection.
    worker_rows = (WorkReport.objects
                   .filter(date__gte=start, date__lte=end, worker__isnull=False)
                   .values('worker__id', 'worker__full_name')
                   .distinct()
                   .order_by('worker__full_name'))
    workers_out = [{'id': r['worker__id'], 'name': r['worker__full_name'] or ''}
                   for r in worker_rows]

    # Max is computed at zone level (the finest granularity) — the heatmap
    # recalculates its own max client-side when the granularity changes.
    return JsonResponse({
        'zones': zones_out,
        'names': names_out,
        'lands': lands_out,
        'workers': workers_out,
        'max': {
            'reports': _max_from(zones_out, 'reports'),
            'team_hours': _max_from(zones_out, 'team_hours'),
            'third_hours': _max_from(zones_out, 'third_hours'),
            'hours': _max_from(zones_out, 'hours'),
        },
    }, json_dumps_params={'ensure_ascii': False})


@login_required(login_url='core:login')
def irrigation_report_pdf(request):
    """One-click PDF export: one page per CCU (sorted by CCU code), each page a
    24 x N-satellite runtime matrix fitting the page. Same format as the
    dashboard. Honors the date-range filter from the query string.
    """
    from core.models import Patch, MaxicomController, MaxicomRuntime
    from collections import defaultdict
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    # Register a CJK font so Chinese renders (STSong-Light ships with reportlab).
    CJK = 'STSong-Light'
    try:
        pdfmetrics.registerFont(UnicodeCIDFont(CJK))
    except Exception:
        CJK = 'Helvetica'  # fallback (Chinese will be missing)

    # --- date range (same logic as the dashboard, defaulting to full span) ---
    # Single Min/Max aggregate (P2 fix) instead of three queries.
    first_ts, last_ts = _irrig_data_span()
    default_from = first_ts[:12] if len(first_ts) >= 12 else (first_ts[:8] if first_ts else '')
    default_to = last_ts[:12] if len(last_ts) >= 12 else (last_ts[:8] if last_ts else '')
    date_from = request.GET.get('from', default_from).strip()
    date_to = request.GET.get('to', default_to).strip()
    ts_from = date_from.ljust(14, '0')[:14] if date_from else ''
    ts_to = date_to.ljust(14, '9')[:14] if date_to else ''

    # --- precompute satellite lookup and runtime queryset once ---
    ctrl_map = {c.mdb_index: c for c in MaxicomController.objects.exclude(name__icontains='CCU')}
    rt_qs = MaxicomRuntime.objects.select_related('station', 'station__parent', 'site')
    if ts_from:
        rt_qs = rt_qs.filter(timestamp__gte=ts_from)
    if ts_to:
        rt_qs = rt_qs.filter(timestamp__lte=ts_to)
    rt_by_site = defaultdict(list)
    for rt in rt_qs:
        rt_by_site[rt.site_id].append(rt)

    # Station Patch IDs claimed by at least one Zone — rows whose station is
    # NOT in this set are flagged red as orphan (no Zone backing = un-attributable).
    mapped_station_ids = _mapped_station_ids()

    # --- PDF setup (landscape A4) ---
    buf = BytesIO()
    page_size = landscape(A4)
    # The title (Paragraph, leading 14) + subtitle (leading 9) + spacer (2mm)
    # consume real vertical space reportlab won't compress; budget 28mm for it.
    frame_h = page_size[1] - 16 * mm              # page minus top/bottom margins
    title_block_h = 28 * mm
    avail_for_table = frame_h - title_block_h
    ROW_H = avail_for_table / 26.5                # 1.4 + 24 + 1.1 = 26.5 row-units
    FS = 6.0                                       # cell font size

    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=8 * mm, bottomMargin=8 * mm,
        title='灌溉运行报告',
    )
    title_style = ParagraphStyle('T', fontName=CJK, fontSize=12,
                                 textColor=colors.HexColor('#1B4332'), spaceAfter=1, leading=14)
    sub_style = ParagraphStyle('S', fontName=CJK, fontSize=7.5, leading=9,
                               textColor=colors.HexColor('#6B7B6E'))

    avail_w = page_size[0] - 20 * mm

    ccus = _ccu_queryset()   # natural numeric order
    elements = []

    for idx, ccu in enumerate(ccus):
        m = _build_ccu_matrix(ccu, rt_by_site.get(ccu.id, []), ctrl_map, mapped_station_ids)
        controllers = m['controllers']
        rows = m['rows']
        col_totals = m['col_totals']
        grand_total = m['grand_total']
        max_cell = m['max_cell'] or 1
        ran = sum(1 for r in rows if r['total'] > 0)
        orphan_count = sum(1 for r in rows if r.get('orphan') and r['total'] > 0)

        elements.append(Paragraph(
            f'{ccu.code} ({ccu.name}) — 站点 × 卫星控制器 运行时间（分钟）', title_style))
        subtitle = f'卫星控制器 {len(controllers)} 个，运行站点 {ran} 个，总运行 {grand_total} 分钟'
        if orphan_count:
            subtitle += f'  ·  其中 {orphan_count} 个运行站点无 Zone 对应（红色标记）'
        elements.append(Paragraph(subtitle, sub_style))
        elements.append(Spacer(1, 2 * mm))

        # table data
        header = ['站#'] + list(controllers) + ['合计']
        data = [header]
        for r in rows:
            data.append([r['station']] + [str(v) if v else '·' for v in r['values']]
                        + [str(r['total']) if r['total'] else '·'])
        data.append(['合计'] + [str(t) if t else '·' for t in col_totals] + [str(grand_total)])

        # column widths
        n_sat = len(controllers)
        side_w = 11 * mm
        sat_col_w = max(7 * mm, (avail_w - 2 * side_w) / n_sat) if n_sat else (avail_w - 2 * side_w)
        col_widths = [side_w] + [sat_col_w] * n_sat + [side_w]
        # row heights: header slightly taller
        row_heights = [ROW_H * 1.4] + [ROW_H] * 24 + [ROW_H * 1.1]

        tbl = Table(data, colWidths=col_widths, rowHeights=row_heights, repeatRows=1)
        ts = [
            ('FONTNAME', (0, 0), (-1, -1), CJK),
            ('FONTSIZE', (0, 0), (-1, 0), FS),
            ('FONTSIZE', (0, 1), (-1, -1), FS),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('LEFTPADDING', (0, 0), (-1, -1), 1),
            ('RIGHTPADDING', (0, 0), (-1, -1), 1),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4332')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D9D0C0')),
            ('BACKGROUND', (0, 1), (0, -2), colors.HexColor('#F5F0E8')),
            ('BACKGROUND', (-1, 1), (-1, -2), colors.HexColor('#EDE8DC')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EDE8DC')),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#1B4332')),
        ]
        # heatmap shading on non-zero VALID cells (interpolate light->dark green)
        gl = (0.929, 0.969, 0.847)   # #EDF7F1
        gd = (0.106, 0.263, 0.196)   # #1B4332
        orphan_fill = colors.HexColor('#FDECEA')     # light red cell wash
        orphan_text = colors.HexColor('#C62828')     # red text on orphan cells
        for ri, r in enumerate(rows, start=1):
            orphan_vals = r.get('orphan_values') or []
            for ci, v in enumerate(r['values'], start=1):
                is_orphan = (ci - 1) < len(orphan_vals) and orphan_vals[ci - 1]
                if is_orphan:
                    # Orphan valve: flat light-red fill, no heat shading (its
                    # runtime is excluded from max_cell so valid cells keep
                    # their full color range).
                    ts.append(('BACKGROUND', (ci, ri), (ci, ri), orphan_fill))
                    if v > 0:
                        ts.append(('TEXTCOLOR', (ci, ri), (ci, ri), orphan_text))
                elif v > 0:
                    t = min(v / max_cell, 1.0)
                    c = colors.Color(gl[0] + (gd[0] - gl[0]) * t,
                                     gl[1] + (gd[1] - gl[1]) * t,
                                     gl[2] + (gd[2] - gl[2]) * t)
                    ts.append(('BACKGROUND', (ci, ri), (ci, ri), c))
                    if t > 0.6:
                        ts.append(('TEXTCOLOR', (ci, ri), (ci, ri), colors.white))
        # Side label hint: when every runtime cell in a row is orphan, tint the
        # 站# cell red as a row-level cue.
        for ri, r in enumerate(rows, start=1):
            if r.get('orphan'):
                ts.append(('TEXTCOLOR', (0, ri), (0, ri), orphan_text))
        tbl.setStyle(TableStyle(ts))
        elements.append(tbl)

        if idx < len(ccus) - 1:
            elements.append(PageBreak())

    doc.build(elements)
    buf.seek(0)
    fname = f'irrigation_report_{date_from}_{date_to}.pdf'
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required(login_url='core:login')
def irrigation_report_excel(request):
    """One-click Excel export: one sheet per CCU (sorted naturally), each sheet
    a 24 x N-satellite runtime matrix. Same data as the dashboard / PDF.
    Honors the date-range filter from the query string.
    """
    from core.models import MaxicomController, MaxicomRuntime
    from collections import defaultdict
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # --- date range (same logic as the dashboard) ---
    # Single Min/Max aggregate (P2 fix) instead of three queries.
    first_ts, last_ts = _irrig_data_span()
    default_from = first_ts[:12] if len(first_ts) >= 12 else (first_ts[:8] if first_ts else '')
    default_to = last_ts[:12] if len(last_ts) >= 12 else (last_ts[:8] if last_ts else '')
    date_from = request.GET.get('from', default_from).strip()
    date_to = request.GET.get('to', default_to).strip()
    ts_from = date_from.ljust(14, '0')[:14] if date_from else ''
    ts_to = date_to.ljust(14, '9')[:14] if date_to else ''

    ctrl_map = {c.mdb_index: c for c in MaxicomController.objects.exclude(name__icontains='CCU')}
    rt_qs = MaxicomRuntime.objects.select_related('station', 'station__parent', 'site')
    if ts_from:
        rt_qs = rt_qs.filter(timestamp__gte=ts_from)
    if ts_to:
        rt_qs = rt_qs.filter(timestamp__lte=ts_to)
    rt_by_site = defaultdict(list)
    for rt in rt_qs:
        rt_by_site[rt.site_id].append(rt)

    # Station Patch IDs claimed by at least one Zone — rows whose station is
    # NOT in this set are flagged red as orphan (no Zone backing = un-attributable).
    mapped_station_ids = _mapped_station_ids()

    # --- styles ---
    header_fill = PatternFill('solid', fgColor='1B4332')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    side_fill = PatternFill('solid', fgColor='F5F0E8')
    total_fill = PatternFill('solid', fgColor='EDE8DC')
    total_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='D9D0C0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # Orphan row (station has no Zone backing): light-red wash + red label font.
    orphan_fill = PatternFill('solid', fgColor='FDECEA')
    orphan_font = Font(color='C62828', bold=True)

    wb = Workbook()
    # remove the default sheet (we add per-CCU sheets)
    default_ws = wb.active

    ccus = _ccu_queryset()

    for idx, ccu in enumerate(ccus):
        m = _build_ccu_matrix(ccu, rt_by_site.get(ccu.id, []), ctrl_map, mapped_station_ids)
        controllers = m['controllers']
        rows = m['rows']
        col_totals = m['col_totals']
        grand_total = m['grand_total']
        ran = sum(1 for r in rows if r['total'] > 0)
        orphan_count = sum(1 for r in rows if r.get('orphan') and r['total'] > 0)

        # Excel sheet names: max 31 chars, no special chars
        sheet_name = f"{ccu.code}_{ccu.name}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        # title row
        ws.cell(row=1, column=1,
                value=f"{ccu.code} ({ccu.name}) — 站点 × 卫星控制器 运行时间（分钟）  {date_from}~{date_to}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12, color='1B4332')
        subtitle = f"卫星控制器 {len(controllers)} 个，运行站点 {ran} 个，总运行 {grand_total} 分钟"
        if orphan_count:
            subtitle += f"  ·  其中 {orphan_count} 个运行站点无 Zone 对应（红色标记）"
        ws.cell(row=2, column=1, value=subtitle)
        ws.cell(row=2, column=1).font = Font(size=9, color='6B7B6E')

        # matrix starts at row 4
        header_row = 4
        n_sat = len(controllers)
        # header
        ws.cell(row=header_row, column=1, value='站#')
        for ci, cname in enumerate(controllers):
            ws.cell(row=header_row, column=2 + ci, value=cname)
        ws.cell(row=header_row, column=2 + n_sat, value='合计')
        for col in range(1, 2 + n_sat + 1):
            c = ws.cell(row=header_row, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
            c.border = border

        # 24 station rows
        for ri, r in enumerate(rows):
            rownum = header_row + 1 + ri
            sc = ws.cell(row=rownum, column=1, value=r['station'])
            sc.fill = side_fill
            sc.alignment = center
            sc.border = border
            orphan_vals = r.get('orphan_values') or []
            for ci, v in enumerate(r['values']):
                is_orphan = ci < len(orphan_vals) and orphan_vals[ci]
                vc = ws.cell(row=rownum, column=2 + ci, value=v if v else None)
                vc.alignment = center
                vc.border = border
                if is_orphan:
                    # Orphan valve: flat light-red fill regardless of value
                    # (no border emphasis — fill alone signals "no Zone backing").
                    vc.fill = orphan_fill
                    if v:
                        vc.font = orphan_font
                elif v:
                    vc.font = Font(bold=True)
            tc = ws.cell(row=rownum, column=2 + n_sat, value=r['total'] if r['total'] else None)
            tc.fill = total_fill
            tc.font = total_font
            tc.alignment = center
            tc.border = border

        # totals row
        trow = header_row + 1 + len(rows)
        ws.cell(row=trow, column=1, value='合计')
        for ci, t in enumerate(col_totals):
            ws.cell(row=trow, column=2 + ci, value=t if t else None)
        ws.cell(row=trow, column=2 + n_sat, value=grand_total)
        for col in range(1, 2 + n_sat + 1):
            c = ws.cell(row=trow, column=col)
            c.fill = total_fill
            c.font = total_font
            c.alignment = center
            c.border = border

        # column widths + freeze
        ws.column_dimensions['A'].width = 6
        for ci in range(n_sat):
            ws.column_dimensions[get_column_letter(2 + ci)].width = 11
        ws.column_dimensions[get_column_letter(2 + n_sat)].width = 8
        ws.freeze_panes = 'B5'

    # remove default empty sheet
    if default_ws is not None and default_ws.title in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(default_ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'irrigation_report_{date_from}_{date_to}.xlsx'
    resp = HttpResponse(
        buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


# ==========================================================================
# Custom Report API
# ==========================================================================

from django.http import JsonResponse
from django.views.decorators.http import require_POST
import json


@require_POST
@login_required(login_url='core:login')
def custom_report_api(request):
    """Return aggregated data for custom chart configurations."""
    from core.models import WorkReport
    from django.db.models import Count, Q
    from django.utils import timezone
    from collections import defaultdict

    try:
        body = json.loads(request.body)
        charts = body.get('charts', [])
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    user = request.user
    is_worker_user = False
    worker = None
    try:
        worker = user.worker_profile
        is_worker_user = True
    except Exception:
        pass

    from core.role_utils import is_admin as check_is_admin

    is_admin_user = check_is_admin(user)

    results = []

    for chart_cfg in charts:
        try:
            chart_id = chart_cfg.get('id', '')
            data_source = chart_cfg.get('dataSource', '')
            metric = chart_cfg.get('metric', '')
            date_from = chart_cfg.get('dateFrom', '')
            date_to = chart_cfg.get('dateTo', '')
            chart_type = chart_cfg.get('chartType', 'bar')
            bar_mode = chart_cfg.get('barMode', 'stacked')
            stack_by = chart_cfg.get('stackBy', '')
            title = chart_cfg.get('title', '')

            chart_data = _build_chart_data(data_source, metric, date_from, date_to,
                                            is_admin_user, is_worker_user, worker,
                                            bar_mode, stack_by)
            if chart_data is None:
                results.append({'id': chart_id, 'error': 'No data'})
                continue

            chart_data['id'] = chart_id
            chart_data['chartType'] = chart_type
            chart_data['barMode'] = bar_mode
            chart_data['title'] = title
            results.append(chart_data)
        except Exception as e:
            import traceback
            results.append({'id': chart_cfg.get('id', ''), 'error': str(e), 'traceback': traceback.format_exc()})

    return JsonResponse({'charts': results})


def _build_chart_data(data_source, metric, date_from, date_to,
                      is_admin, is_worker, worker, bar_mode='stacked', stack_by=''):
    """Build chart data dict for a given metric. Returns None if no data."""
    from core.models import WorkReport, WorkReportEntry, WorkItem
    from django.db.models import Count, Q
    from django.utils import timezone
    from datetime import timedelta, datetime

    # Apply date filter helper
    def date_filter(qs, date_field='date'):
        if date_from:
            qs = qs.filter(**{f'{date_field}__gte': date_from})
        if date_to:
            qs = qs.filter(**{f'{date_field}__lte': date_to})
        return qs

    # === WORK REPORTS ===
    if data_source == 'work_reports':
        qs = WorkReport.objects.select_related('worker', 'location')
        if not is_admin:
            if is_worker:
                qs = qs.filter(worker=worker)
            else:
                qs = qs.none()

        qs = date_filter(qs)

        # Stacked bar chart: two-level group-by
        if stack_by:
            primary, stack_field = _get_metric_groupings(data_source, metric, stack_by)
            if primary is None:
                return None
            entries = list(qs.values(primary, stack_field)
                           .annotate(count=Count('id'))
                           .order_by(primary))
            if not entries:
                return None
            labels_seen = []
            label_set = set()
            stack_values = []
            data_map = {}
            for e in entries:
                label = str(e[primary] or '未指定')
                sv = e.get(stack_field) or '未指定'
                if label not in label_set:
                    labels_seen.append(label)
                    label_set.add(label)
                if sv not in stack_values:
                    stack_values.append(sv)
                data_map.setdefault(label, {})[sv] = e['count']

            datasets = []
            for i, sv in enumerate(stack_values[:15]):
                ds_data = []
                for lbl in labels_seen:
                    ds_data.append(data_map.get(lbl, {}).get(sv, 0))
                color_idx = i % len(_chart_colors_palette())
                datasets.append({
                    'label': str(sv),
                    'data': ds_data,
                    'backgroundColor': _chart_colors_palette()[color_idx],
                    'borderColor': 'rgba(255,255,255,1)',
                    'borderWidth': 1,
                })

            return {
                'labels': labels_seen,
                'datasets': datasets,
            }

        if metric == 'daily_trend':
            if not date_from or not date_to:
                today = timezone.now().date()
                date_from = (today - timedelta(days=29)).isoformat()
                date_to = today.isoformat()
                qs = date_filter(qs)
            entries = list(qs.values('date').annotate(count=Count('id')).order_by('date'))
            if not entries:
                return None
            return {
                'labels': [str(e['date']) for e in entries],
                'datasets': [{
                    'label': '维修日志数',
                    'data': [e['count'] for e in entries],
                    'backgroundColor': _chart_colors(len(entries)),
                }]
            }

        elif metric == 'by_location':
            entries = list(qs.values('location__name').annotate(count=Count('id')).order_by('-count')[:15])
            if not entries:
                return None
            return {
                'labels': [e['location__name'] or '未指定' for e in entries],
                'datasets': [{
                    'label': '日志数',
                    'data': [e['count'] for e in entries],
                    'backgroundColor': _chart_colors(len(entries)),
                }]
            }

        elif metric == 'by_zone':
            entries = list(qs.values('zone_location__name').annotate(count=Count('id')).exclude(zone_location__isnull=True).order_by('-count')[:15])
            if not entries:
                return None
            return {
                'labels': [e['zone_location__name'] or '未指定' for e in entries],
                'datasets': [{
                    'label': '日志数',
                    'data': [e['count'] for e in entries],
                    'backgroundColor': _chart_colors(len(entries)),
                }]
            }

        # Removed metrics (fields/relations deleted): by_category, by_fault_type,
        # by_fault_category, by_info_source, by_equipment (via fault_entries).
        elif metric in ('by_category', 'by_fault_type', 'by_fault_category',
                        'by_info_source', 'by_equipment'):
            return None


        elif metric == 'by_worker_department':
            entries = list(qs.values('worker__department').annotate(count=Count('id')).order_by('-count'))
            if not entries:
                return None
            return {
                'labels': [e['worker__department'] or '未指定' for e in entries],
                'datasets': [{
                    'label': '日志数',
                    'data': [e['count'] for e in entries],
                    'backgroundColor': _chart_colors(len(entries)),
                }]
            }

        elif metric in ('by_entry_section', 'by_entry_node', 'by_entry_project'):
            # 工作内容明细 (WorkReportEntry) — additive, reads the new tree-form data
            eqs = WorkReportEntry.objects.filter(work_item__active=True)
            eqs = date_filter(eqs, 'work_report__date')
            if not is_admin:
                if is_worker:
                    eqs = eqs.filter(work_report__worker=worker)
                else:
                    eqs = eqs.none()
            if metric == 'by_entry_section':
                labels_map = dict(WorkItem.SECTION_CHOICES)
                rows = list(eqs.values('work_item__section')
                            .annotate(c=Count('id')).order_by('-c')[:15])
                labels = [labels_map.get(r['work_item__section'], r['work_item__section']) for r in rows]
            elif metric == 'by_entry_node':
                rows = list(eqs.values('work_item__name_zh')
                            .annotate(c=Count('id')).order_by('-c')[:15])
                labels = [r['work_item__name_zh'] or '未指定' for r in rows]
            else:  # by_entry_project
                rows = list(eqs.exclude(project__isnull=True)
                            .values('project__name').annotate(c=Count('id')).order_by('-c')[:15])
                labels = [r['project__name'] or '未指定' for r in rows]
            if not rows:
                return None
            return {
                'labels': labels,
                'datasets': [{
                    'label': '填报次数',
                    'data': [r['c'] for r in rows],
                    'backgroundColor': _chart_colors(len(rows)),
                }],
            }

        elif metric == 'difficult_rate_by_category':
            return None  # work_category field removed

        elif metric == 'by_worker' and is_admin:
            entries = list(qs.values('worker__full_name').annotate(
                total=Count('id'),
                difficult=Count('id', filter=Q(is_difficult=True))
            ).order_by('-total')[:15])
            if not entries:
                return None
            return {
                'labels': [e['worker__full_name'] or '未指定' for e in entries],
                'datasets': [
                    {
                        'label': '总数',
                        'data': [e['total'] for e in entries],
                        'backgroundColor': 'rgba(27, 67, 50, 0.7)',
                        'borderColor': 'rgba(27, 67, 50, 1)',
                    },
                    {
                        'label': '疑难',
                        'data': [e['difficult'] for e in entries],
                        'backgroundColor': 'rgba(204, 119, 34, 0.7)',
                        'borderColor': 'rgba(204, 119, 34, 1)',
                    }
                ]
            }

        elif metric == 'difficult_rate_by_worker' and is_admin:
            entries = list(qs.values('worker__full_name').annotate(
                total=Count('id'),
                difficult=Count('id', filter=Q(is_difficult=True)),
                resolved=Count('id', filter=Q(is_difficult=True, is_difficult_resolved=True))
            ).order_by('-total')[:15])
            if not entries:
                return None
            return {
                'labels': [e['worker__full_name'] or '未指定' for e in entries],
                'datasets': [
                    {
                        'label': '总数',
                        'data': [e['total'] for e in entries],
                        'backgroundColor': 'rgba(27, 67, 50, 0.7)',
                        'borderColor': 'rgba(27, 67, 50, 1)',
                    },
                    {
                        'label': '疑难',
                        'data': [e['difficult'] for e in entries],
                        'backgroundColor': 'rgba(204, 119, 34, 0.7)',
                        'borderColor': 'rgba(204, 119, 34, 1)',
                    },
                    {
                        'label': '已处理',
                        'data': [e['resolved'] for e in entries],
                        'backgroundColor': 'rgba(64, 145, 108, 0.7)',
                        'borderColor': 'rgba(64, 145, 108, 1)',
                    }
                ]
            }

    # === DEMAND RECORDS (removed — model deleted) ===
    elif data_source == 'demand_records':
        return None

        # Stacked bar chart: two-level group-by
        if stack_by:
            primary, stack_field = _get_metric_groupings(data_source, metric, stack_by)
            if primary is None:
                return None
            entries = list(qs.values(primary, stack_field)
                           .annotate(count=Count('id'))
                           .order_by(primary))
            if not entries:
                return None
            labels_seen = []
            label_set = set()
            stack_values = []
            data_map = {}
            for e in entries:
                label = str(e[primary] or '未指定')
                sv = e.get(stack_field) or '未指定'
                if label not in label_set:
                    labels_seen.append(label)
                    label_set.add(label)
                if sv not in stack_values:
                    stack_values.append(sv)
                data_map.setdefault(label, {})[sv] = e['count']

            datasets = []
            for i, sv in enumerate(stack_values[:15]):
                ds_data = []
                for lbl in labels_seen:
                    ds_data.append(data_map.get(lbl, {}).get(sv, 0))
                color_idx = i % len(_chart_colors_palette())
                datasets.append({
                    'label': str(sv),
                    'data': ds_data,
                    'backgroundColor': _chart_colors_palette()[color_idx],
                    'borderColor': 'rgba(255,255,255,1)',
                    'borderWidth': 1,
                })

            return {
                'labels': labels_seen,
                'datasets': datasets,
            }

        if metric == 'daily_trend':
            if not date_from or not date_to:
                today = timezone.now().date()
                date_from = (today - timedelta(days=29)).isoformat()
                date_to = today.isoformat()
                qs = date_filter(qs)
            entries = list(qs.values('date').annotate(count=Count('id')).order_by('date'))
            if not entries:
                return None
            return {
                'labels': [str(e['date']) for e in entries],
                'datasets': [{
                    'label': '需求日志数',
                    'data': [e['count'] for e in entries],
                    'backgroundColor': _chart_colors(len(entries)),
                }]
            }

        elif metric == 'by_category':
            return None  # demand_records data source removed

        elif metric == 'by_department':
            entries = list(qs.values('demand_department__name').annotate(count=Count('id')).exclude(demand_department__isnull=True).order_by('-count')[:15])
            if not entries:
                return None
            return {
                'labels': [e['demand_department__name'] or '未指定' for e in entries],
                'datasets': [{
                    'label': '需求数',
                    'data': [e['count'] for e in entries],
                    'backgroundColor': _chart_colors(len(entries)),
                }]
            }

        elif metric == 'by_status':
            status_map = {'submitted': '已提交', 'approved': '已批准', 'rejected': '已拒绝', 'in_progress': '进行中', 'completed': '已完成'}
            entries = list(qs.values('status').annotate(count=Count('id')).order_by('-count'))
            if not entries:
                return None
            return {
                'labels': [status_map.get(e['status'], e['status']) for e in entries],
                'datasets': [{
                    'label': '需求数',
                    'data': [e['count'] for e in entries],
                    'backgroundColor': _chart_colors(len(entries)),
                }]
            }

        elif metric == 'global_events':
            entries = list(qs.filter(is_global_event=True).values('category__name').annotate(count=Count('id')).order_by('-count')[:15])
            if not entries:
                return None
            return {
                'labels': [e['category__name'] or '未指定' for e in entries],
                'datasets': [{
                    'label': '全局事件数',
                    'data': [e['count'] for e in entries],
                    'backgroundColor': _chart_colors(len(entries)),
                }]
            }

        elif metric == 'by_zone':
            entries = list(qs.filter(zone__isnull=False)
                           .values('zone__name').annotate(count=Count('id'))
                           .order_by('-count')[:15])
            if not entries:
                return None
            return {
                'labels': [e['zone__name'] or '未指定' for e in entries],
                'datasets': [{
                    'label': '需求数',
                    'data': [e['count'] for e in entries],
                    'backgroundColor': _chart_colors(len(entries)),
                }]
            }

        elif metric == 'by_contact':
            entries = list(qs.exclude(demand_contact='')
                           .values('demand_contact').annotate(count=Count('id'))
                           .order_by('-count')[:15])
            if not entries:
                return None
            return {
                'labels': [e['demand_contact'] or '未指定' for e in entries],
                'datasets': [{
                    'label': '需求数',
                    'data': [e['count'] for e in entries],
                    'backgroundColor': _chart_colors(len(entries)),
                }]
            }

        elif metric == 'by_submitter':
            # Merge two attribution paths: irrigation workers (submitter) and
            # department users (submitter_user). The WaterRequest model now allows
            # either; grouping only on submitter__full_name silently drops dept users.
            agg = {}
            for e in qs.filter(submitter__isnull=False).values('submitter__full_name').annotate(c=Count('id')):
                nm = e['submitter__full_name'] or '未指定'
                agg[nm] = agg.get(nm, 0) + e['c']
            from django.contrib.auth import get_user_model as _gum
            _User = _gum()
            for e in qs.filter(submitter_user__isnull=False).values('submitter_user_id').annotate(c=Count('id')):
                u = _User.objects.filter(pk=e['submitter_user_id']).first()
                nm = (u.get_full_name() or u.username) if u else '未指定'
                agg[nm] = agg.get(nm, 0) + e['c']
            entries = sorted(agg.items(), key=lambda kv: kv[1], reverse=True)[:15]
            if not entries:
                return None
            return {
                'labels': [nm for nm, _c in entries],
                'datasets': [{
                    'label': '需求数',
                    'data': [c for _nm, c in entries],
                    'backgroundColor': _chart_colors(len(entries)),
                }]
            }

        elif metric == 'global_event_volume':
            if not date_from or not date_to:
                today = timezone.now().date()
                date_from = (today - timedelta(days=29)).isoformat()
                date_to = today.isoformat()
                qs = date_filter(qs)
            entries = list(qs.filter(is_global_event=True)
                           .values('date').annotate(count=Count('id')).order_by('date'))
            if not entries:
                return None
            return {
                'labels': [str(e['date']) for e in entries],
                'datasets': [{
                    'label': '全局事件数',
                    'data': [e['count'] for e in entries],
                    'backgroundColor': _chart_colors(len(entries)),
                }]
            }

    return None


def _get_metric_groupings(data_source, metric, stack_by):
    """Return (primary_field, stack_field) for stacked bar chart. Returns (None, None) if invalid."""
    mapping = {
        'work_reports': {
            'worker': ('worker__full_name', 'worker__full_name'),
            'location': ('location__name', 'location__name'),
            'zone': ('zone_location__name', 'zone_location__name'),
            'date': ('date', 'date'),
        },
        'demand_records': {},
    }

    table = mapping.get(data_source, {})
    stack_field = table.get(stack_by)
    if not stack_field:
        return None, None

    # Determine primary (label) field based on metric
    primary_map = {
        'work_reports': {
            'daily_trend': 'date',
            'by_location': ('location__name', 'location__name'),
            'by_zone': ('zone_location__name', 'zone_location__name'),
            'by_worker_department': ('worker__department', 'worker__department'),
            'by_worker': ('worker__full_name', 'worker__full_name'),
            'difficult_rate_by_worker': ('worker__full_name', 'worker__full_name'),
        },
        'demand_records': {},
    }

    primary_entry = primary_map.get(data_source, {}).get(metric)
    if primary_entry is None:
        return None, None

    if isinstance(primary_entry, str):
        # Single field metric (e.g. daily_trend) — use that as primary, stack_by as stack
        return primary_entry, stack_field[0]

    # Two-element tuple: (group_field, display_field)
    return primary_entry[0], stack_field[0]


def _chart_colors_palette():
    """Return the full color palette for stacked charts."""
    return [
        'rgba(27, 67, 50, 0.7)',
        'rgba(45, 106, 79, 0.7)',
        'rgba(64, 145, 108, 0.7)',
        'rgba(82, 183, 136, 0.7)',
        'rgba(204, 119, 34, 0.7)',
        'rgba(155, 34, 38, 0.7)',
        'rgba(45, 106, 150, 0.7)',
        'rgba(108, 117, 125, 0.7)',
        'rgba(27, 67, 100, 0.7)',
        'rgba(120, 50, 120, 0.7)',
        'rgba(180, 80, 40, 0.7)',
        'rgba(60, 140, 140, 0.7)',
        'rgba(140, 100, 40, 0.7)',
        'rgba(80, 60, 140, 0.7)',
        'rgba(160, 60, 80, 0.7)',
    ]


def _chart_colors(n):
    """Return a list of n distinct chart background colors."""
    palette = [
        'rgba(27, 67, 50, 0.7)',
        'rgba(45, 106, 79, 0.7)',
        'rgba(64, 145, 108, 0.7)',
        'rgba(82, 183, 136, 0.7)',
        'rgba(204, 119, 34, 0.7)',
        'rgba(155, 34, 38, 0.7)',
        'rgba(45, 106, 150, 0.7)',
        'rgba(108, 117, 125, 0.7)',
        'rgba(27, 67, 100, 0.7)',
        'rgba(120, 50, 120, 0.7)',
        'rgba(180, 80, 40, 0.7)',
        'rgba(60, 140, 140, 0.7)',
        'rgba(140, 100, 40, 0.7)',
        'rgba(80, 60, 140, 0.7)',
        'rgba(160, 60, 80, 0.7)',
    ]
    return palette[:n]


@login_required(login_url='core:login')
def custom_report(request):
    """Custom report page - user selects data and generates charts."""
    from core.role_utils import is_admin
    admin = is_admin(request.user)
    return render(request, 'core/custom_report.html', {'is_admin': admin})


WORK_REPORTS_PAGE_SIZE = 30
# Default window shown on first visit / when no explicit date range is chosen.
WORK_REPORTS_DEFAULT_DAYS = 7
# Sort options for the work-report list. Key = URL value, value = ORM order_by.
# created  — newest by created_at (the model default; id is monotonic with it,
#            kept as a deterministic tiebreaker)
# id       — newest by workorder number (id DESC), independent of creation time
# worker   — alphabetical by creator full_name; same id tiebreaker for stability
WORK_REPORTS_SORT_OPTIONS = {
    'created': ('-created_at', '-id'),
    'id': ('-id',),
    'worker': ('worker__full_name', '-id'),
}
WORK_REPORTS_DEFAULT_SORT = 'created'


def _work_report_filters(request):
    """Collect + normalize the work-report list query filters.

    Shared by the server-rendered list and the AJAX "load more" endpoint so both
    apply exactly the same scoping. Returns ``(date_from, date_to, land_id,
    worker_id, difficult, pending, before_id, sort, q)`` where ``date_from``/
    ``date_to`` are strings (or '') ready to echo back into the filter form,
    ``land_id`` filters by the report's zones' Land (top-level area), ``sort``
    is one of ``WORK_REPORTS_SORT_OPTIONS`` keys, and the rest are already-coerced
    values. ``before_id`` is the cursor for "load more".
    """
    date_from = (request.GET.get('date_from') or '').strip()
    date_to = (request.GET.get('date_to') or '').strip()
    land_id = request.GET.get('land') or ''
    worker_id = request.GET.get('worker') or ''
    difficult = request.GET.get('is_difficult')
    pending = request.GET.get('is_pending_repair')
    before_id = request.GET.get('before_id')
    # Free-text search by work-order number (id). Accepts "42", "#42", or a
    # fragment; strips the leading # and matches as a prefix on the id.
    q = (request.GET.get('q') or '').strip().lstrip('#')
    sort = (request.GET.get('sort') or WORK_REPORTS_DEFAULT_SORT).strip()
    if sort not in WORK_REPORTS_SORT_OPTIONS:
        sort = WORK_REPORTS_DEFAULT_SORT
    # Workorder type = the 工作类别 (WorkItem section) derived from the report's
    # entries. Validated against SECTION_CHOICES so bogus values yield '' (all).
    section = (request.GET.get('section') or '').strip()
    if section not in dict(WorkItem.SECTION_CHOICES):
        section = ''

    # Default to the most recent N days only when the user hasn't picked any
    # explicit start (neither a preset nor a manual date). This keeps the first
    # page light; older records come in on demand via "加载更多".
    if not date_from:
        today = timezone.localdate()
        date_from = (today - timedelta(days=WORK_REPORTS_DEFAULT_DAYS)).isoformat()

    return date_from, date_to, land_id, worker_id, difficult, pending, before_id, sort, q, section


def _scoped_work_reports_qs(user, admin, sort=WORK_REPORTS_DEFAULT_SORT):
    """Base queryset with role-based visibility applied (no filters yet).

    Both 灌溉一线 (field workers) and managers/admins see ALL workorders — the
    visibility is shared so the team can review and comment on each other's
    records. Only dept users (no field-worker and no admin role) are scoped to
    their own submissions, if any.
    """
    from core.models import WorkReport
    from core.role_utils import get_worker_for_user, is_field_worker

    order = WORK_REPORTS_SORT_OPTIONS.get(sort, WORK_REPORTS_SORT_OPTIONS[WORK_REPORTS_DEFAULT_SORT])
    qs = WorkReport.objects.select_related(
        'worker', 'location'
    ).prefetch_related(
        'entries__work_item', 'entries__project', 'zones__land', 'edit_logs__editor'
    ).filter(is_pm=False).order_by(*order)
                        # 游标分页用 id__lt(before_id) — 任何排序下都按 "更老的工单"
                        # 取下一页,再以所选排序渲染。对 created/id 完全等价;
                        # 对 worker 排序,每页内字母序连续,翻页时仍是按 id 取更老批次。

    if not admin and not is_field_worker(user):
        worker = get_worker_for_user(user)
        qs = qs.filter(worker=worker) if worker else qs.none()
    return qs


def _serialize_work_reports(reports):
    """Turn enriched WorkReport instances into JSON-serializable dicts.

    Mirrors what the template renders per card. Expects ``enrich_reports`` and
    ``attach_zone_hierarchy`` to have already attached ``entry_groups``,
    ``section_labels``, ``entry_count``, ``zone_hierarchy`` and ``zone_summary``.
    """
    from django.urls import reverse

    out = []
    for r in reports:
        out.append({
            'id': r.id,
            'display_number': r.display_number,
            'date': r.date.isoformat() if r.date else '',
            'shift': r.shift or '',
            'shift_display': r.get_shift_display() if r.shift else '',
            'worker_name': r.worker.full_name if r.worker_id and r.worker else '',
            'worker_employee_id': r.worker.employee_id if r.worker_id and r.worker else '',
            'work_start_time': r.work_start_time.strftime('%H:%M') if r.work_start_time else '',
            'work_end_time': r.work_end_time.strftime('%H:%M') if r.work_end_time else '',
            'team_hours': float(r.team_hours) if r.team_hours else 0,
            'third_party_hours': float(r.third_party_hours) if r.third_party_hours else 0,
            'team_size': r.team_size or 0,
            'third_party_count': r.third_party_count or 0,
            'zone_summary': r.zone_summary or '',
            'zone_hierarchy': getattr(r, 'zone_hierarchy', []),
            'section_labels': getattr(r, 'section_labels', []),
            'entry_count': getattr(r, 'entry_count', 0),
            'entry_groups': getattr(r, 'entry_groups', []),
            'comment_count': getattr(r, 'comment_count', 0),
            'photos': r.photos or [],
            'remark': r.remark or '',
            # Material consumption (材料消耗) outbound lines, flattened to
            # [{name, quantity, unit}]. Reads the prefetched _materials_txn attr
            # so the load-more AJAX path stays N+1-free.
            'materials': [
                {'name': ln.category.name_zh, 'quantity': ln.quantity, 'unit': ln.unit or ''}
                for txn in getattr(r, '_materials_txn', [])
                for ln in getattr(txn, '_lines', [])
            ],
            'is_pending_repair': bool(r.is_pending_repair),
            'is_difficult': bool(r.is_difficult),
            'is_difficult_resolved': bool(r.is_difficult_resolved),
            'detail_url': reverse('core:work_report_detail', args=[r.id]),
            'edit_url': reverse('core:dashboard') + '?edit_workorder=' + str(r.id),
            # Edit history (oldest→newest). Empty when never edited.
            'edit_logs': [
                {
                    'editor': log.editor.full_name if log.editor_id and log.editor else '(未知)',
                    'time': log.created_at.strftime('%Y-%m-%d %H:%M') if log.created_at else '',
                }
                for log in r.edit_logs.all()
            ],
        })
    return out


@login_required(login_url='core:login')
def work_report_reassign(request, report_id):
    """Admin-only: re-assign a work order's creator (worker).

    Used to fix reports that were attributed to the wrong worker (e.g. after a
    DB mishap). POST {worker_id}. Returns JSON.
    """
    from core.models import WorkReport, Worker
    from core.role_utils import is_admin

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '仅支持 POST'}, status=405)
    if not is_admin(request.user):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)

    report = get_object_or_404(WorkReport, pk=report_id)
    worker_id = request.POST.get('worker_id')
    if not worker_id:
        return JsonResponse({'success': False, 'message': '缺少 worker_id'}, status=400)
    worker = Worker.objects.filter(pk=worker_id).first()
    if not worker:
        return JsonResponse({'success': False, 'message': '处理人不存在'}, status=400)

    report.worker = worker
    report.save(update_fields=['worker'])
    return JsonResponse({
        'success': True,
        'message': f'已改为 {worker.full_name}',
        'worker_name': worker.full_name,
        'worker_employee_id': worker.employee_id or '',
    })


def _pm_gwo_queryset(user, is_mgr, include_done=False):
    """Return the GWO queryset visible to this user for the PM tab.

    By default only dispatched/overdue tasks (today's to-do). With
    ``include_done=True`` also includes completed/skipped GWOs so the PM tab can
    show history. Field workers: their crew's tasks. Managers: all.
    """
    from core.models import GeneratedWorkOrder, Worker
    from django.utils import timezone as _tz
    today = _tz.localdate()
    if include_done:
        qs = GeneratedWorkOrder.objects.all()
    else:
        qs = GeneratedWorkOrder.objects.filter(
            status__in=['dispatched', 'overdue'], scheduled_date__lte=today,
        )
    if not is_mgr:
        try:
            worker = Worker.objects.get(user=user, active=True)
        except Worker.DoesNotExist:
            return GeneratedWorkOrder.objects.none()
        crew_ids = set(worker.crews.values_list('id', flat=True)) | \
            set(worker.led_crews.values_list('id', flat=True))
        if not crew_ids:
            return GeneratedWorkOrder.objects.none()
        qs = qs.filter(crew_id__in=crew_ids)
    return qs.select_related('pm_order', 'plan__job_plan', 'crew', 'worker').prefetch_related('zones').order_by('-scheduled_date', '-id')


def _serialize_pm_tasks(gwos, is_mgr):
    """Serialize a GWO iterable into PM-task dicts for the work-reports PM tab."""
    from django.utils import timezone as _tz
    today = _tz.localdate()
    result = []
    for gwo in gwos:
        pmwo = gwo.pm_order
        plan = gwo.plan
        level = plan.job_plan.asset_level if plan.job_plan_id else 'zone_group'
        # Smart area description: zone_group → zone codes; ccu/sat → the CCU/SAT
        # label only (don't list every included zone for device-level tasks).
        # patch.code already starts with "CCU" (e.g. CCU2), so no extra prefix.
        if level == 'ccu' and plan.patch_id:
            area_desc = f'{plan.patch.code} {plan.patch.name}'.strip()
            area_count = 1
        elif level == 'sat' and plan.satellite_id:
            area_desc = f'{plan.satellite.code} {plan.satellite.name}'.strip()
            area_count = 1
        else:
            zones_qs = list(gwo.zones.all())
            area_desc = '、'.join(z.code for z in zones_qs[:3])
            area_count = len(zones_qs)
        result.append({
            'gwo_id': gwo.id,
            'ticket': f'PM-{gwo.id}',
            'pm_order_id': pmwo.id if pmwo else None,
            'pm_number': plan.pm_number,
            'job_plan_name': plan.job_plan.name if plan.job_plan_id else '',
            'scheduled_date': gwo.scheduled_date.strftime('%Y-%m-%d'),
            'area_desc': area_desc,
            'area_count': area_count,
            'freq_label': f'每{plan.frequency_value}{plan.get_frequency_unit_display()}',
            # Legacy fields (kept for existing PM-tab template compatibility).
            'zone_count': area_count,
            'zone_preview': area_desc,
            'overdue': gwo.scheduled_date < today,
            'crew_name': gwo.crew.name if gwo.crew_id else '—',
            'status': gwo.status,
            'completed_at': gwo.completed_at.strftime('%Y-%m-%d') if gwo.completed_at else '',
            'worker_name': gwo.worker.full_name if gwo.worker_id and gwo.worker else '',
        })
    return result


@login_required(login_url='core:login')
def pm_gwo_detail(request, gwo_id):
    """AJAX: return a single GeneratedWorkOrder's seed data for the completion form.

    The PM-tab "去完成" button opens a CREATE workorder form seeded from the GWO
    (dispatch stores worker/zones/remark on the GWO). This endpoint feeds those
    fields so the form can pre-select zones, date and remark without the worker
    re-entering them.
    """
    from core.models import GeneratedWorkOrder
    gwo = get_object_or_404(GeneratedWorkOrder, pk=gwo_id)
    return JsonResponse({
        'gwo_id': gwo.id,
        'ticket': f'PM-{gwo.id}',
        'pm_number': gwo.plan.pm_number,
        'scheduled_date': gwo.scheduled_date.strftime('%Y-%m-%d'),
        'remark': gwo.remark or '',
        'zone_codes': list(gwo.zones.values_list('code', flat=True)),
        'worker_id': gwo.worker_id,
    })


@login_required(login_url='core:login')
def work_reports_pm_tasks(request):
    """AJAX endpoint: return the next batch of PM tasks (cursor-paginated by GWO id).

    GET params: ``before_id`` — only return GWOs with id < before_id.
    """
    from core.role_utils import is_admin
    admin = is_admin(request.user)
    include_done = request.GET.get('done') in ('1', 'true', 'True')
    qs = _pm_gwo_queryset(request.user, admin, include_done=include_done)
    before_id = request.GET.get('before_id', '').strip()
    if before_id and before_id.isdigit():
        qs = qs.filter(id__lt=int(before_id))
    batch = list(qs[:20])
    return JsonResponse({
        'tasks': _serialize_pm_tasks(batch, admin),
        'total': _pm_gwo_queryset(request.user, admin, include_done=include_done).count(),
    })


@never_cache
@login_required(login_url='core:login')
def work_reports_list(request):
    """Unified, responsive work-order list (维修日志).

    Server-renders the initial page (latest 7 days by default) and also serves
    cursor-paginated "load more" batches as JSON when requested via AJAX.

    Filters (GET): ``date_from`` / ``date_to`` (YYYY-MM-DD), ``location`` (Patch
    id), ``worker`` (Worker id, admin only), ``is_pending_repair`` / ``is_difficult``
    (any truthy value), and ``before_id`` (report id cursor for load-more).
    """
    from core.models import Patch, Worker, Land, WorkReport
    from core.role_utils import is_admin
    from core.workorder_tree_views import workitem_path_map, enrich_reports, attach_zone_hierarchy

    user = request.user
    admin = is_admin(user)

    date_from, date_to, land_id, worker_id, difficult, pending, before_id, sort, q, section = _work_report_filters(request)

    qs = _scoped_work_reports_qs(user, admin, sort=sort)
    qs = qs.annotate(comment_count=Count('comments'))

    # Apply filters (shared with the load-more endpoint).
    qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if land_id:
        # Subquery (not a JOIN) so the comment_count annotation above isn't
        # multiplied by the number of matching zones.
        qs = qs.filter(zones__in=Zone.objects.filter(land_id=land_id))
    if worker_id:
        qs = qs.filter(worker_id=worker_id)
    if difficult:
        qs = qs.filter(is_difficult=True)
    if pending:
        qs = qs.filter(is_pending_repair=True)
    if section:
        # Subquery (not a JOIN) so the comment_count annotation above isn't
        # multiplied, and reports with several matching entries aren't duped.
        qs = qs.filter(entries__in=WorkReportEntry.objects.filter(work_item__section=section))
    # Work-order number search: match the id (or #id) as a prefix so "4" finds
    # #4, #40-49, #400-499, etc. An exact integer narrows to that one report.
    if q:
        if q.isdigit():
            qs = qs.filter(id__startswith=q)
    # Cursor pagination — older records have smaller ids under (-date, -id).
    if before_id:
        try:
            qs = qs.filter(id__lt=int(before_id))
        except (TypeError, ValueError):
            pass

    page = list(qs[:WORK_REPORTS_PAGE_SIZE + 1])  # +1 to detect a next page.
    has_more = len(page) > WORK_REPORTS_PAGE_SIZE
    reports = page[:WORK_REPORTS_PAGE_SIZE]
    # Material consumption (材料消耗) outbound lines for the expanded card.
    # Prefetched here (not in _scoped_work_reports_qs) so other callers of that
    # helper — the stats/Excel path — can attach their own material_consumptions
    # prefetch without a duplicate-to_attr conflict.
    from django.db.models import prefetch_related_objects, Prefetch
    from core.models import InventoryTransaction, InventoryTransactionLine
    prefetch_related_objects(reports, Prefetch(
        'material_consumptions',
        queryset=InventoryTransaction.objects.prefetch_related(
            Prefetch('lines', queryset=InventoryTransactionLine.objects.select_related('category'), to_attr='_lines')
        ), to_attr='_materials_txn',
    ))
    enrich_reports(reports, workitem_path_map())
    attach_zone_hierarchy(reports)

    # AJAX "load more": hand the next batch to the client as JSON.
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('ajax'):
        return JsonResponse({
            'reports': _serialize_work_reports(reports),
            'has_more': has_more,
            'last_id': reports[-1].id if reports else None,
            'is_admin': admin,
            # Worker list so the admin-only reassign control can render on
            # AJAX-loaded cards (mirrors the server-rendered workers context).
            'workers': [{'id': w.id, 'name': w.full_name, 'employee_id': w.employee_id or ''}
                        for w in Worker.objects.all().order_by('full_name')] if admin else [],
        })

    lands = Land.objects.filter(active=True).order_by('order', 'name')
    workers = Worker.objects.all().order_by('full_name') if admin else []

    # Remarks tab (managers only): pending remarks (Step 1: confirm) + confirmed
    # remarks (Step 2: transfer to 灌溉/设备 records). Grouped by workorder across
    # the zones each touches. Non-managers get empty lists (tab hidden).
    is_manager = admin
    pending_remark_groups = _group_zone_remarks('remarks') if is_manager else []
    confirmed_remark_groups = _group_zone_remarks('confirmed_remarks') if is_manager else []
    # Per-group data for the JS: a flat [[zone_id, index]] list (for confirm, which
    # still fans out to all zones) AND a Land → 通用名称 → zone tree (for transfer,
    # where the user picks which names/zones to write the record to — avoids duplicate
    # records across zones). Keys are unique across both lists.
    remark_groups_map = {}
    for g in list(pending_remark_groups) + list(confirmed_remark_groups):
        tree = []
        for h in g['hierarchy']:
            names = []
            for nm in h['names']:
                zones = [{'code': z.code, 'zid': z.id, 'idx': i} for z, i in nm['pairs']]
                names.append({'name': nm['name'], 'zones': zones})
            tree.append({'land': h['land'], 'names': names})
        remark_groups_map[g['key']] = {
            'pairs': [[z.id, i] for z, i in g['zones']],
            'tree': tree,
        }

    filter_json = {
        'filters': {
            'date_from': date_from,
            'date_to': date_to,
            'land': int(land_id) if land_id else '',
            'worker': int(worker_id) if worker_id else '',
            'is_difficult': bool(difficult),
            'is_pending_repair': bool(pending),
            'section': section,
            'sort': sort,
            'q': q,
        },
        'last_id': reports[-1].id if reports else None,
        'is_admin': admin,
        # Worker list for the admin-only reassign control on AJAX-loaded cards.
        'workers': [{'id': w.id, 'name': w.full_name, 'employee_id': w.employee_id or ''}
                    for w in workers] if admin else [],
    }

    # ── PM工单 tab: field workers see their crew's tasks, managers see all ──
    # Initial batch is capped (20); the full count drives the badge. A dedicated
    # AJAX endpoint (/work-reports/pm-tasks/) handles "load more".
    from core.role_utils import is_field_worker
    pm_include_done = request.GET.get('pm_done') in ('1', 'true', 'True')
    pm_qs = _pm_gwo_queryset(request.user, admin, include_done=pm_include_done)
    pm_tasks_total = pm_qs.count()
    pm_tasks = _serialize_pm_tasks(pm_qs[:20], admin)

    # ── 待修工单 tab (managers only): unresolved pending-repair reports ──
    # These are what paint zones orange on the map (needs_attention). Listed
    # here so a manager can resolve them without filing a 计划性维修 work order.
    pending_repairs = []
    if is_manager:
        # Unresolved 待修 OR 疑难未处理 — both block a report from being "done"
        # and both are cleared by the resolve-repair action.
        pr_qs = ((WorkReport.objects.filter(is_pending_repair=True, resolved_by_pm__isnull=True)
                  | WorkReport.objects.filter(is_difficult=True, is_difficult_resolved=False))
                 .distinct().select_related('worker', 'location').order_by('-date', '-id'))
        for wr in pr_qs:
            zone_codes = ', '.join(z.code for z in wr.zones.all()[:4])
            flags = []
            if wr.is_pending_repair:
                flags.append('待修')
            if wr.is_difficult and not wr.is_difficult_resolved:
                flags.append('疑难未处理')
            pending_repairs.append({
                'id': wr.id,
                'date': wr.date,
                'worker_name': wr.worker.full_name if wr.worker_id and wr.worker else '—',
                'location_code': wr.location.code if wr.location_id and wr.location else '',
                'remark': (wr.remark or '')[:120],
                'zone_preview': zone_codes,
                'zone_count': wr.zones.count(),
                'flags': '、'.join(flags),
            })

    return render(request, 'core/work_reports.html', {
        'reports': reports,
        'lands': lands,
        'workers': workers,
        'sections': WorkItem.SECTION_CHOICES,
        'is_admin': admin,
        'has_more': has_more,
        'last_id': reports[-1].id if reports else '',
        'default_days': WORK_REPORTS_DEFAULT_DAYS,
        'filter_json': filter_json,
        'filters': filter_json['filters'],
        'is_manager': is_manager,
        'pending_remark_groups': pending_remark_groups,
        'confirmed_remark_groups': confirmed_remark_groups,
        'remark_groups_map': remark_groups_map,
        'pm_tasks': pm_tasks,
        'pm_tasks_total': pm_tasks_total,
        'pm_include_done': pm_include_done,
        'pending_repairs': pending_repairs,
        'active_tab': request.GET.get('tab', 'workorders'),
    })


# ── 媒体下载 tab ─────────────────────────────────────────────────────────
# 经理/管理员专用的「媒体下载」tab：列出带照片/视频的工单，勾选后打 zip 下载。
VIDEO_EXTS = {'.mp4', '.mov', '.m4v', '.webm', '.ogg', '.ogv', '.avi', '.mkv'}


def _is_media_video(path):
    p = (path or '').lower()
    return any(p.endswith(ext) for ext in VIDEO_EXTS)


@login_required(login_url='core:login')
def work_report_photos(request):
    """「媒体下载」tab 的工单-媒体列表 (AJAX JSON)。

    沿用 work_reports_list 的日期/位置/人员筛选；只返回 photos 非空的工单。
    每条工单带工作类别(section)、所属Land、处理人等简要信息 + 媒体路径列表。
    """
    from core.models import Patch, Worker, Land
    from core.role_utils import is_admin
    from core.workorder_tree_views import workitem_path_map, enrich_reports, attach_zone_hierarchy

    user = request.user
    admin = is_admin(user)
    if not admin:
        return JsonResponse({'reports': [], 'error': '无权限'}, status=403)

    date_from, date_to, land_id, worker_id, difficult, pending, before_id, sort, q, section = _work_report_filters(request)
    qs = _scoped_work_reports_qs(user, admin, sort=sort).filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if land_id:
        qs = qs.filter(zones__in=Zone.objects.filter(land_id=land_id))
    if worker_id:
        qs = qs.filter(worker_id=worker_id)
    if difficult:
        qs = qs.filter(is_difficult=True)
    if pending:
        qs = qs.filter(is_pending_repair=True)
    if section:
        qs = qs.filter(entries__in=WorkReportEntry.objects.filter(work_item__section=section))
    if q and q.isdigit():
        qs = qs.filter(id__startswith=q)

    # 只取带媒体的工单（排除空 photos / 空 list）。限制条数避免一次拉太多。
    page = list(qs.exclude(photos=[]).exclude(photos__isnull=True)[:200])
    enrich_reports(page, workitem_path_map())
    attach_zone_hierarchy(page)

    out = []
    for r in page:
        zh = getattr(r, 'zone_hierarchy', None) or []
        section_labels = getattr(r, 'section_labels', None) or []
        out.append({
            'id': r.id,
            'display_number': r.display_number,
            'date': r.date.isoformat() if r.date else '',
            'category': section_labels[0] if section_labels else '',   # 工作类别
            'land': zh[0]['land'] if zh and zh[0].get('land') else '',
            'worker_name': r.worker.full_name if r.worker_id and r.worker else '',
            'shift_display': r.get_shift_display() if hasattr(r, 'get_shift_display') else '',
            'photos': [
                {'path': p, 'is_video': _is_media_video(p)}
                for p in (r.photos or [])
            ],
        })
    return JsonResponse({'reports': out})


@login_required(login_url='core:login')
def work_report_photos_download(request):
    """打包选中媒体为 zip 流式下载。

    前端收集勾选的原始媒体路径（相对 MEDIA_ROOT）POST 到这里。安全：白名单
    前缀(workorder_photos/ work_reports/) + 拒绝 '..', 防路径穿越。

    quality 参数：
      - 'thumb'（默认）：图片打包缩略图(_thumb.jpg，体积小、下载快)；视频无
        小尺寸版，始终打包原文件。
      - 'original'：图片和视频都打包原文件。
    """
    import io
    import zipfile
    from django.http import FileResponse, HttpResponseForbidden
    from django.conf import settings
    from django.utils import timezone
    from core.role_utils import is_admin
    from core.workorder_tree_views import thumb_path

    if not is_admin(request.user):
        return HttpResponseForbidden('无权限')
    if request.method != 'POST':
        return HttpResponseForbidden('POST only')

    items = request.POST.getlist('items')
    if not items:
        try:
            import json as _json
            items = _json.loads(request.POST.get('items', '[]') or '[]')
        except (ValueError, TypeError):
            items = []

    quality = (request.POST.get('quality') or 'thumb').strip()
    use_thumb = (quality == 'thumb')

    ALLOWED_PREFIXES = ('workorder_photos/', 'work_reports/')
    media_root = settings.MEDIA_ROOT
    VIDEO_EXTS = ('.mp4', '.mov', '.avi', '.m4v', '.webm', '.mkv')

    def _resolve(rel):
        """Resolve a relative media path to an absolute path inside MEDIA_ROOT,
        or None if it's invalid / escapes / missing. Guards against traversal."""
        rel = (rel or '').lstrip('/')
        if not rel or '..' in rel or not rel.startswith(ALLOWED_PREFIXES):
            return None
        abs_path = (media_root / rel).resolve()
        try:
            abs_path.relative_to(media_root.resolve())
        except ValueError:
            return None
        return abs_path if (abs_path.exists() and abs_path.is_file()) else None

    buf = io.BytesIO()
    added = 0
    import os
    # Track archive entry names to avoid collisions (e.g. two originals mapping to
    # the same thumb name is unlikely but possible after manual file moves).
    seen_names = set()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for p in items:
            abs_path = _resolve(p)
            if abs_path is None:
                continue
            is_video = p.lower().endswith(VIDEO_EXTS)
            # 缩略图模式仅对图片生效；视频始终用原文件。
            if use_thumb and not is_video:
                thumb_rel = thumb_path(p)
                thumb_abs = _resolve(thumb_rel)
                if thumb_abs:
                    # 缩略图存在 → 用它（小、快）。文件名沿用 _thumb 形式。
                    arc_name = thumb_rel
                    abs_path = thumb_abs
                else:
                    # 缩略图缺失（旧数据未生成）→ 回退原图，但文件名保持原始名，
                    # 避免和其它原图/缩略图在 zip 内重名冲突。
                    arc_name = p
            else:
                arc_name = p
            # Flatten to a single folder: strip the on-disk directory structure so
            # the zip holds all files at its root (no per-workorder subfolders).
            arc_name = os.path.basename(arc_name)
            # Dedupe archive entry names (append a counter suffix on collision).
            final_name = arc_name
            if final_name in seen_names:
                base, dot, ext = arc_name.rpartition('.')
                i = 2
                while f'{base}_{i}{dot}{ext}' in seen_names:
                    i += 1
                final_name = f'{base}_{i}{dot}{ext}'
            seen_names.add(final_name)
            zf.write(abs_path, final_name)
            added += 1

    if added == 0:
        return JsonResponse({'error': '没有可下载的文件（路径无效或不存在）'}, status=400)

    buf.seek(0)
    resp = FileResponse(buf, content_type='application/zip')
    tag = '_thumb' if use_thumb else '_original'
    resp['Content-Disposition'] = 'attachment; filename="workreport_media{0}_{1:%Y%m%d_%H%M}.zip"'.format(tag, timezone.now())
    return resp


@login_required(login_url='core:login')
def water_requests_list(request):
    """需求列表 — 浇水协调 (WaterRequest) records.

    Managers/admins see all requests and can approve/reject inline; other users see
    only their own submissions. Mirrors the work_reports_list filter+card structure.
    """
    from core.models import WaterRequest, Zone
    from core.role_utils import is_admin, get_worker_for_user

    user = request.user
    admin = is_admin(user)
    qs = WaterRequest.objects.select_related('zone', 'submitter', 'submitter_user', 'approver').prefetch_related('zones').order_by('-created_at', '-id')
    if not admin:
        # Show this user's own submissions. Department users are stored on
        # submitter_user (they have no Worker row); irrigation workers on submitter.
        # Match either attribution so both paths see their own requests.
        from django.db.models import Q
        worker = get_worker_for_user(user)
        q = Q(submitter_user=user)
        if worker:
            q |= Q(submitter=worker)
        qs = qs.filter(q)

    status_filter = request.GET.get('status', '')
    zone_id = request.GET.get('zone', '')
    request_type = request.GET.get('request_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if status_filter:
        qs = qs.filter(status=status_filter)
    if zone_id:
        qs = qs.filter(zones__id=zone_id).distinct()
    if request_type:
        qs = qs.filter(request_type=request_type)
    if date_from:
        qs = qs.filter(start_datetime__date__gte=date_from)
    if date_to:
        qs = qs.filter(end_datetime__date__lte=date_to)

    requests = list(qs[:200])
    # Attach a Land → 通用名称 → [codes] hierarchy to each request so the list
    # collapses repetitive zones instead of printing every zone name.
    from core.workorder_tree_views import attach_zone_hierarchy
    attach_zone_hierarchy(requests)
    zones = Zone.objects.all().order_by('code')
    status_choices = WaterRequest.STATUS_CHOICES
    request_type_choices = WaterRequest.REQUEST_TYPE_CHOICES

    context = {
        'requests': requests,
        'zones': zones,
        'is_admin': admin,
        'status_choices': status_choices,
        'request_type_choices': request_type_choices,
        'filters': {
            'status': status_filter,
            'zone': int(zone_id) if zone_id else '',
            'request_type': request_type,
            'date_from': date_from,
            'date_to': date_to,
        },
    }
    return render(request, 'core/water_requests.html', context)


def _group_zone_remarks(notes_field):
    """Group a Zone notes JSON-list field into remark groups.

    Works for any of the four Zone note fields (remarks / confirmed_remarks /
    irrigation_management_notes / equipment_maintenance_notes). Remarks sourced
    from one workorder are duplicated across every zone it touches; group those
    by ``workorder_id`` (one card, all its zones). Hand-added remarks (no
    workorder_id) stay per-zone.

    Returns a list of dicts, each: ``{remark, key, is_grouped, zones:[(zone,idx)],
    hierarchy:[{land, names:[{name,count,pairs}], zone_count}]}``, sorted grouped-first
    then by date desc. ``zone`` items carry ``.id``/``.name``/``.land`` so callers
    can build transfer endpoints. Shared by the /remarks/ page and the work-reports
    待确认备注 tab.
    """
    import json as _json
    from collections import OrderedDict
    from core.models import Zone, WorkReportEntry, WorkItem

    grouped = OrderedDict()
    for z in (Zone.objects.select_related('land', 'patch')
              .exclude(**{notes_field: ''}).exclude(**{notes_field + '__isnull': True})
              .order_by('code')):
        raw = getattr(z, notes_field)
        try:
            items = _json.loads(raw)
        except (ValueError, TypeError):
            continue
        for idx, it in enumerate(items):
            if it.get('archived'):          # 归档备注不进列表（仍保留在 confirmed_remarks 供工单详情反查）
                continue
            woid = it.get('workorder_id')
            key = 'wo:' + str(woid) if woid else 'z:%d:%d' % (z.id, idx)
            entry = grouped.setdefault(key, {
                'remark': it, 'zones': [], 'key': key, 'is_grouped': bool(woid),
            })
            entry['zones'].append((z, idx))

    # Enrich workorder-sourced groups with the source workorder's section labels
    # (工单类别, e.g. 常规维护/灌溉项目) so the UI can show the category alongside
    # the (often-generic) remark content. One batched query, not per-group.
    section_labels = dict(WorkItem.SECTION_CHOICES)
    woids = [g['remark'].get('workorder_id') for g in grouped.values()
             if g['remark'].get('workorder_id')]
    wo_sections = {}
    if woids:
        for row in (WorkReportEntry.objects
                    .filter(work_report_id__in=woids, work_item__active=True)
                    .values_list('work_report_id', 'work_item__section')):
            wo_sections.setdefault(row[0], set()).add(row[1])
    for g in grouped.values():
        woid = g['remark'].get('workorder_id')
        secs = sorted(wo_sections.get(woid, [])) if woid else []
        g['section_labels'] = [section_labels.get(s, s) for s in secs]

    for g in grouped.values():
        lands = OrderedDict()
        for z, idx in g['zones']:
            ln = (z.land.name if z.land_id and z.land else '其它') or '其它'
            nm = z.name or z.code
            lands.setdefault(ln, {}).setdefault(nm, []).append((z, idx))
        g['hierarchy'] = [
            {'land': ln, 'names': [
                {'name': nm, 'count': len(pairs), 'pairs': pairs}
                for nm, pairs in sorted(names.items())
            ], 'zone_count': sum(len(p) for p in names.values())}
            for ln, names in lands.items()
        ]

    groups = list(grouped.values())
    groups.sort(key=lambda g: (0 if g['is_grouped'] else 1,
                               g['remark'].get('date', '')), reverse=True)
    return groups


@login_required(login_url='core:login')
def remarks_list(request):
    """待确认备注 — zones with unconfirmed remarks, with inline confirm for managers.

    Reads the raw Zone.remarks JSON (not the lossy dashboard payload) so each remark's
    date/content/author is shown. Confirming a remark POSTs to the existing
    zone_remark_confirm endpoint (manager-only).
    """
    from core.role_utils import is_admin

    admin = is_admin(request.user)
    groups = _group_zone_remarks('remarks')

    context = {
        'groups': groups,
        'is_admin': admin,
        'total': len(groups),
    }
    return render(request, 'core/remarks.html', context)


@require_POST
@login_required(login_url='core:login')
def water_request_update(request, pk):
    """Approve / reject / request-info on a WaterRequest (admin only, AJAX)."""
    from core.models import WaterRequest
    from core.role_utils import is_admin, get_worker_for_user
    if not is_admin(request.user):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)
    wr = get_object_or_404(WaterRequest, pk=pk)
    new_status = request.POST.get('status', '')
    valid = {c for c, _ in WaterRequest.STATUS_CHOICES}
    if new_status not in valid:
        return JsonResponse({'success': False, 'message': '无效状态'}, status=400)
    wr.status = new_status
    wr.status_notes = request.POST.get('status_notes', '').strip()
    wr.approver = get_worker_for_user(request.user)
    wr.processed_at = timezone.now()
    wr.save()

    # Notify the submitter on approve/reject/info_needed (not on bare re-submit).
    # Recipient is the dept User if set (the common case), else the Worker's User.
    from core.notifications import notify
    recipient = wr.submitter_user or (getattr(wr.submitter, 'user', None) if wr.submitter_id else None)
    zone_label = wr.zone.code if (wr.zone_id and wr.zone) else '区域'
    _STATUS_NOTIFY = {
        'approved':     ('request_approved',     '需求已批准',     '你的浇水协调需求已批准'),
        'rejected':     ('request_rejected',     '需求已拒绝',     '你的浇水协调需求已被拒绝'),
        'info_needed':  ('request_info_needed',  '需补充信息',     '你的浇水协调需求需补充信息'),
    }
    if recipient is not None and new_status in _STATUS_NOTIFY:
        verb, title, prefix = _STATUS_NOTIFY[new_status]
        note_suffix = f'：{wr.status_notes}' if wr.status_notes else ''
        notify(recipient, verb, f'【{zone_label}】{title}',
               prefix + note_suffix, '/requests/')

    return JsonResponse({
        'success': True,
        'message': '已更新',
        'status': wr.status,
        'status_display': wr.get_status_display(),
        'status_notes': wr.status_notes,
        'approver': wr.approver.full_name if wr.approver else '',
    })


@login_required(login_url='core:login')
@require_POST
def water_request_resubmit(request, pk):
    """Re-open an info_needed WaterRequest after the submitter补充信息.

    Submitter-only (matched by submitter_user or submitter). Sets status back to
    'submitted', appends the note to status_notes (prefixed 【补充】), and notifies
    the admin who requested the info (wr.approver.user). Idempotent: if an unread
    resubmitted notification already exists since the last processed_at, don't
    create a duplicate (submitter re-editing without the admin having seen it).
    """
    from core.models import WaterRequest
    from core.role_utils import get_worker_for_user
    from core.notifications import notify, resubmit_already_notified

    wr = get_object_or_404(WaterRequest, pk=pk)
    # Only the request's owner may resubmit (dept user via submitter_user, or
    # irrigation worker via submitter).
    worker = get_worker_for_user(request.user)
    is_owner = (wr.submitter_user_id == request.user.id
                or (worker is not None and wr.submitter_id == worker.id))
    if not is_owner:
        return JsonResponse({'success': False, 'message': '无权操作此需求'}, status=403)
    if wr.status != 'info_needed':
        return JsonResponse({'success': False, 'message': '该需求当前无需补充信息'}, status=400)

    note = (request.POST.get('note') or '').strip()
    if not note:
        return JsonResponse({'success': False, 'message': '请填写补充信息'}, status=400)

    append = f'\n【补充 {timezone.now():%Y-%m-%d %H:%M}】{note}'
    wr.status_notes = (wr.status_notes + append) if wr.status_notes else append.strip()
    wr.status = 'submitted'
    wr.save()

    # Notify the admin who requested the info (the recorded approver). The approver
    # is a Worker; reach their User via .user (managers created via
    # resolve_or_create_worker carry a linked User). Idempotency guard avoids
    # duplicate unread notifications if the submitter re-edits before the admin
    # opens the popup.
    approver_user = getattr(wr.approver, 'user', None) if wr.approver_id else None
    if approver_user and not resubmit_already_notified(wr):
        zone_label = wr.zone.code if (wr.zone_id and wr.zone) else '区域'
        notify(approver_user, 'request_resubmitted',
               f'【{zone_label}】需求已补充信息重新提交',
               note[:200], '/requests/')

    return JsonResponse({
        'success': True,
        'message': '已补充信息并重新提交',
        'status': wr.status,
        'status_display': wr.get_status_display(),
        'status_notes': wr.status_notes,
    })


@login_required(login_url='core:login')
@require_POST
def notification_read(request, nid):
    """Mark a single notification as read (我已知晓 in the popup)."""
    from core.notifications import mark_read
    ok = mark_read(nid, request.user)
    return JsonResponse({'success': ok})


@login_required(login_url='core:login')
def work_report_detail(request, report_id):
    from core.models import WorkReport, WorkItem
    from core.role_utils import is_admin, is_field_worker
    from collections import OrderedDict

    report = get_object_or_404(
        WorkReport.objects.select_related(
            'worker', 'location'
        ).prefetch_related('entries__work_item', 'entries__project', 'zones__land', 'edit_logs__editor'),
        pk=report_id
    )

    # Both 灌溉一线 (field workers) and managers/admins can view any workorder
    # (so they can read/post comments on it). Other account types are denied.
    if not is_admin(request.user) and not is_field_worker(request.user):
        messages.error(request, '无权查看此记录')
        return redirect('core:work_reports')

    # Group tree-form entries (WorkReportEntry) by section for display.
    section_labels = dict(WorkItem.SECTION_CHOICES)
    grouped = OrderedDict()
    for e in report.entries.select_related('work_item', 'project'):
        sec = e.work_item.section
        grouped.setdefault(sec, {'label': section_labels.get(sec, sec), 'items': []})
        grouped[sec]['items'].append(e)

    # Deduplicated Land → name hierarchy (shared helper).
    from core.workorder_tree_views import attach_zone_hierarchy
    attach_zone_hierarchy([report])

    # 备注：与该工单关联的 remark（含已确认 / 已归档），去重后展示。一条工单
    # 的 remark 在其触及的每个 zone 的 confirmed_remarks 里重复，按 (content+date)
    # 去重取一份。注意：以 workorder_id 为权威链接，扫描全部 zone，不限于
    # report.zones m2m（后者可能因编辑/补录而与备注实际所在 zone 不一致）。
    import json as _json
    from core.models import Zone as _Zone
    seen_keys, related_remarks = set(), []
    for z in _Zone.objects.exclude(confirmed_remarks='').exclude(confirmed_remarks__isnull=True):
        try:
            clist = _json.loads(z.confirmed_remarks)
        except (ValueError, TypeError):
            continue
        for it in clist:
            if not isinstance(it, dict) or it.get('workorder_id') != report.id:
                continue
            dedup = (it.get('content', ''), it.get('date', ''), it.get('archived'))
            if dedup in seen_keys:
                continue
            seen_keys.add(dedup)
            related_remarks.append(it)

    return render(request, 'core/work_report_detail.html', {
        'report': report,
        'tree_entry_groups': list(grouped.values()),
        'zone_hierarchy': report.zone_hierarchy,
        'related_remarks': related_remarks,
        # Material consumption (材料消耗): the report's linked outbound transactions,
        # flattened to [{name, quantity, unit}] rows for display.
        'materials': [
            {'name': ln.category.name_zh, 'quantity': ln.quantity, 'unit': ln.unit}
            for txn in report.material_consumptions.all()
            for ln in txn.lines.select_related('category')
        ],
    })


@login_required(login_url='core:login')
def work_report_comments(request, report_id):
    """List (GET) and add (POST) comments on a posted workorder.

    Mirrors the "everyone sees all workorders" rule: any logged-in user —
    灌溉一线 (field worker) or manager/admin — can read the thread and post a
    comment. Author is resolved via role_utils.resolve_or_create_worker so the
    comment is attributable to a real person regardless of account type.
    """
    from core.models import WorkReport, WorkReportComment
    from core.role_utils import resolve_or_create_worker, is_admin
    from core.notifications import notify

    report = get_object_or_404(WorkReport.objects.select_related('worker'), pk=report_id)

    if request.method == 'POST':
        body = (request.POST.get('body') or '').strip()
        if not body:
            return JsonResponse({'success': False, 'message': '评论内容不能为空'}, status=400)
        author, _created = resolve_or_create_worker(request.user)
        comment = WorkReportComment.objects.create(
            work_report=report, author=author, body=body[:2000],
        )

        # Fire notifications: the report owner + prior commenters in the thread,
        # each excluding the current author (don't notify yourself) and de-duped
        # within this call. Recipient is a User (the popup key); a Worker with no
        # linked User is silently skipped by notify().
        link = f'/work-reports/{report.id}/'
        author_id = author.id if author else None
        recipients = []
        owner_user = getattr(report.worker, 'user', None) if report.worker_id else None
        if owner_user:
            recipients.append(owner_user)
        for c in report.comments.exclude(author_id=author_id).select_related('author__user'):
            u = getattr(c.author, 'user', None) if c.author_id else None
            if u and u not in recipients:
                recipients.append(u)
        author_name = author.full_name if author else '(未知)'
        for u in recipients:
            notify(u, 'comment',
                   f'{author_name} 评论了工单 #{report.id}',
                   body[:200], link)

        return JsonResponse({
            'success': True,
            'message': '评论已发布',
            'comment': {
                'id': comment.id,
                'author': author.full_name if author else '(未知)',
                'body': comment.body,
                'time': comment.created_at.strftime('%Y-%m-%d %H:%M'),
            },
        })

    # GET: list the thread, newest-first for display.
    comments = (report.comments.select_related('author')
                .order_by('-created_at').values('id', 'body', 'created_at',
                                                author_name=F('author__full_name')))
    data = [
        {
            'id': c['id'],
            'author': c['author_name'] or '(未知)',
            'body': c['body'],
            'time': c['created_at'].strftime('%Y-%m-%d %H:%M') if c['created_at'] else '',
        }
        for c in comments
    ]
    return JsonResponse({'success': True, 'comments': data, 'count': len(data)})


@login_required(login_url='core:login')
def work_report_create(request):
    from core.models import WorkReport, Patch, Worker, Zone, ZoneEquipment
    from core.role_utils import is_admin

    if request.method == 'POST':
        try:
            worker = request.user.worker_profile
        except Exception:
            messages.error(request, '当前用户未关联处理人账号')
            return redirect('core:work_reports')

        zone_code = request.POST.get('zone_location', '').strip()
        zone = Zone.objects.filter(code=zone_code).first() if zone_code else None

        report = WorkReport.objects.create(
            date=request.POST.get('date'),
            weather=request.POST.get('weather', ''),
            worker=worker,
            location_id=request.POST.get('location'),
            zone_location=zone,
            remark=request.POST.get('remark', ''),
            is_difficult=bool(request.POST.get('is_difficult')),
            is_difficult_resolved=bool(request.POST.get('is_difficult_resolved')),
        )

        # Parse fault entries
        messages.success(request, f'工作日报已创建 (ID: {report.id})')

        if request.POST.get('save_and_new'):
            return redirect('core:work_report_create')
        return redirect('core:work_reports')

    locations = Patch.objects.filter(active=True).order_by('order')
    zones = Zone.objects.order_by('code')

    # Build zone equipment map for frontend lookup
    zone_equipment_map = {}
    for ze in ZoneEquipment.objects.select_related('zone', 'equipment').all():
        zone_code = ze.zone.code
        if zone_code not in zone_equipment_map:
            zone_equipment_map[zone_code] = []
        zone_equipment_map[zone_code].append({
            'id': ze.id,
            'equipment_details': {
                'equipment_type': ze.equipment.equipment_type,
                'equipment_type_display': ze.equipment.get_equipment_type_display(),
                'model_name': ze.equipment.model_name,
            },
            'location_in_zone': ze.location_in_zone or '',
        })

    from datetime import date
    return render(request, 'core/work_report_form.html', {
        'locations': locations,
        'zones': zones,
        'grouped_zones': _build_grouped_zones(zones),
        'zone_equipment_json': json.dumps(zone_equipment_map),
        'today': date.today().isoformat(),
    })


@login_required(login_url='core:login')
def work_report_edit(request, report_id):
    # The legacy column-based edit form referenced deleted models (WorkCategory,
    # InfoSource). Redirect to the v2 tree-form editor, which is the single source
    # of truth for editing a WorkReport.
    from core.models import WorkReport
    get_object_or_404(WorkReport, pk=report_id)
    return redirect('core:workorder_tree_form_edit', report_id=report_id)


@require_POST
@login_required(login_url='core:login')
def work_report_upload_photo(request, report_id):
    import os
    from django.conf import settings
    from datetime import datetime
    from core.models import WorkReport
    from core.role_utils import is_admin

    if not is_admin(request.user):
        return JsonResponse({'error': '仅管理员可上传照片'}, status=403)

    report = get_object_or_404(WorkReport, pk=report_id)
    files = request.FILES.getlist('files')
    if not files:
        return JsonResponse({'error': '未选择文件'}, status=400)

    # Validate every file BEFORE writing any to disk: extension allow-list +
    # size cap + Pillow content check (rejects renamed non-image payloads).
    from core.upload_security import validate_upload
    for f in files:
        ok, err = validate_upload(f)
        if not ok:
            return JsonResponse({'error': f'{f.name}: {err}'}, status=400)

    photo_paths = list(report.photos or [])
    upload_dir = os.path.join(settings.MEDIA_ROOT, 'work_reports', str(report.id))
    os.makedirs(upload_dir, exist_ok=True)

    for f in files:
        ext = os.path.splitext(f.name)[1]
        filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{len(photo_paths)}{ext}"
        filepath = os.path.join(upload_dir, filename)
        with open(filepath, 'wb') as dest:
            for chunk in f.chunks():
                dest.write(chunk)
        photo_paths.append(f"work_reports/{report.id}/{filename}")

    report.photos = photo_paths
    report.save(update_fields=['photos'])

    return JsonResponse({'photos': photo_paths})


@require_POST
@login_required(login_url='core:login')
def work_report_remove_photo(request, report_id):
    import os
    from django.conf import settings
    from core.models import WorkReport
    from core.role_utils import is_admin

    if not is_admin(request.user):
        return JsonResponse({'error': '仅管理员可删除照片'}, status=403)

    report = get_object_or_404(WorkReport, pk=report_id)
    photo = request.POST.get('photo')
    if not photo:
        return JsonResponse({'error': '缺少照片参数'}, status=400)

    photo_paths = list(report.photos or [])
    if photo in photo_paths:
        photo_paths.remove(photo)
        report.photos = photo_paths
        report.save(update_fields=['photos'])
        filepath = os.path.join(settings.MEDIA_ROOT, photo)
        if os.path.exists(filepath):
            os.remove(filepath)

    return JsonResponse({'photos': photo_paths})


@require_POST
@login_required(login_url='core:login')
def work_report_delete(request, report_id):
    from core.models import WorkReport
    from core.role_utils import is_admin

    if not is_admin(request.user):
        messages.error(request, '仅管理员可删除记录')
        return redirect('core:work_reports')

    report = get_object_or_404(WorkReport, pk=report_id)
    report.delete()
    messages.success(request, f'工作日报已删除')
    return redirect('core:work_reports')


@login_required(login_url='core:login')
def zone_geo_api(request):
    from core.models import Zone
    import json as _json
    zones = Zone.objects.filter(Q(boundary_points__isnull=False) | Q(dxf_boundary_points__isnull=False)).exclude(boundary_points=[]).exclude(dxf_boundary_points=[]).only('code', 'name', 'boundary_points', 'dxf_boundary_points', 'boundary_source')
    data = []
    for z in zones:
        bp = z.active_boundary_points
        if not bp:
            continue
        all_pts = bp if isinstance(bp[0], dict) else []
        if isinstance(bp[0], list):
            all_pts = [p for ring in bp for p in ring]
        center = None
        if all_pts:
            lat = sum(p.get('lat', 0) for p in all_pts) / len(all_pts)
            lng = sum(p.get('lng', 0) for p in all_pts) / len(all_pts)
            center = [round(lat, 6), round(lng, 6)]
        data.append({'c': z.code, 'n': z.name or z.code, 'b': bp, 't': center})
    return JsonResponse(data, safe=False)


def _build_zone_geo_data():
    from core.models import Zone
    zones = Zone.objects.filter(Q(boundary_points__isnull=False) | Q(dxf_boundary_points__isnull=False)).exclude(boundary_points=[]).exclude(dxf_boundary_points=[]).only('code', 'name', 'boundary_points', 'dxf_boundary_points', 'boundary_source')
    data = []
    for z in zones:
        bp = z.active_boundary_points
        if not bp:
            continue
        all_pts = bp if isinstance(bp[0], dict) else []
        if isinstance(bp[0], list):
            all_pts = [p for ring in bp for p in ring]
        center = None
        if all_pts:
            lat = sum(p.get('lat', 0) for p in all_pts) / len(all_pts)
            lng = sum(p.get('lng', 0) for p in all_pts) / len(all_pts)
            center = [round(lat, 6), round(lng, 6)]
        data.append({'c': z.code, 'n': z.name or z.code, 'b': bp, 't': center})
    return data


@login_required(login_url='core:login')
def workorder_mobile_v2(request):
    from core.models import WorkReport, Patch, Zone, Worker
    from core.role_utils import get_worker_for_user, is_admin, get_user_role, ROLE_FIELD_WORKER
    from core.role_utils import ROLE_SUPER_ADMIN, ROLE_MANAGER, resolve_or_create_worker
    from core.workorder_tree_views import (
        _calc_hours, _save_entries, _collect_entry_photos, _save_photo,
        _resolve_pending_repairs, _record_edit, _save_workorder_materials,
        _resolve_material_dest,
    )
    from datetime import date, datetime, time

    role = get_user_role(request.user)
    if role not in (ROLE_FIELD_WORKER, ROLE_SUPER_ADMIN, ROLE_MANAGER):
        messages.error(request, '无权限访问此页面')
        return redirect('core:login')

    worker = get_worker_for_user(request.user)
    if not worker and not is_admin(request.user):
        messages.error(request, '未关联处理人账号')
        return redirect('core:login')

    if request.method == 'POST':
        try:
            if not worker and not is_admin(request.user):
                return JsonResponse({'success': False, 'message': '未关联处理人账号'}, status=400)

            # Resolve the real submitter. Managers/admins have no Worker row,
            # so provision one from their ManagerProfile (idempotent via employee_id).
            # Previously this fell back to Worker.objects.first(), which attributed
            # every manager submission to the same arbitrary worker.
            post_worker, _created = resolve_or_create_worker(request.user)
            if not post_worker:
                return JsonResponse({'success': False, 'message': '未关联处理人账号'}, status=400)

            shift = request.POST.get('shift', '')
            start_str = request.POST.get('work_start_time', '')
            end_str = request.POST.get('work_end_time', '')

            work_start = None
            work_end = None
            if start_str:
                h, m = start_str.split(':')
                work_start = time(int(h), int(m))
            if end_str:
                h, m = end_str.split(':')
                work_end = time(int(h), int(m))

            team_size = int(request.POST.get('team_size', 1) or 1)
            third_party_count = int(request.POST.get('third_party_count', 0) or 0)

            team_hours = _calc_hours(work_start, work_end, team_size)
            third_party_hours = _calc_hours(work_start, work_end, third_party_count)

            zone_codes = request.POST.getlist('zones')
            selected_zones = Zone.objects.filter(code__in=zone_codes).select_related('patch')
            # Deduplicate display names while preserving first-seen order. The same
            # 通用名称 can map to multiple zone rows (different codes), and without
            # dedup a selection of N same-named zones repeats the name N times
            # (e.g. "BOH, BOH, BOH"). Show each distinct name once.
            names = []
            seen = set()
            for z in selected_zones:
                n = z.name or z.code
                if n not in seen:
                    seen.add(n)
                    names.append(n)
            zone_names = ', '.join(names)

            # Resolve the report's location (CCU/Patch). Prefer the patch of the
            # first selected zone; some zones belong to a Land but have no Patch
            # and no boundary (e.g. 酒店3/酒店4), so scan the whole selection for a
            # non-null patch before giving up. Only null (now DB-legal) if none of
            # the selected zones carry a patch and there are no patches at all.
            first_zone = selected_zones.first()
            location = None
            if first_zone and first_zone.patch_id:
                location = first_zone.patch
            elif selected_zones.exists():
                z_with_patch = next((z for z in selected_zones if z.patch_id), None)
                location = z_with_patch.patch if z_with_patch else Patch.objects.first()
            else:
                location = Patch.objects.first()

            # Edit vs create: a posted report_id means we update an existing
            # report in place; otherwise we create a new one. A posted gwo_id
            # (PM task completion) seeds a new is_pm=True report and links it to
            # the GeneratedWorkOrder so dispatch needs no WorkReport shell.
            report_id = request.POST.get('report_id')
            pm_order_id = request.POST.get('pm_order_id')
            is_edit = (report_id and report_id.isdigit()) or (pm_order_id and pm_order_id.isdigit())
            pm_gwo_id = request.POST.get('gwo_id')
            is_pm_completion = bool(pm_gwo_id and pm_gwo_id.isdigit())
            # Wrap all DB writes in a transaction so a failure mid-save doesn't
            # leave a half-written report (the desktop path already does this).
            from django.db import transaction
            from core.models import PMWorkOrder, GeneratedWorkOrder
            with transaction.atomic():
                # PM completion uses a separate PMWorkOrder table (own id sequence),
                # so PM work never occupies a WorkReport #id slot.
                if is_pm_completion:
                    if pm_order_id and pm_order_id.isdigit():
                        report = get_object_or_404(PMWorkOrder, pk=pm_order_id)
                        report.date = request.POST.get('date') or report.date
                        report.location = location
                        report.zone_location = first_zone
                        report.remark = request.POST.get('remark', '')
                        report.shift = shift
                        report.work_start_time = work_start
                        report.work_end_time = work_end
                        report.team_size = team_size
                        report.third_party_count = third_party_count
                        report.team_hours = team_hours
                        report.third_party_hours = third_party_hours
                        report.zone_names = zone_names
                        report.save()
                        _record_edit(report, request.user)
                    else:
                        report = PMWorkOrder.objects.create(
                            gwo_id=int(pm_gwo_id),
                            date=request.POST.get('date') or date.today().isoformat(),
                            worker=post_worker,
                            location=location,
                            zone_location=first_zone,
                            remark=request.POST.get('remark', ''),
                            shift=shift,
                            work_start_time=work_start,
                            work_end_time=work_end,
                            team_size=team_size,
                            third_party_count=third_party_count,
                            team_hours=team_hours,
                            third_party_hours=third_party_hours,
                            zone_names=zone_names,
                        )
                elif is_edit:
                    report = get_object_or_404(WorkReport, pk=report_id)
                    report.date = request.POST.get('date') or report.date
                    # 编辑不改处理人：保留原 worker，仅更新内容字段。改派走管理员
                    # 专用的 work_report_reassign 入口（已校验 is_admin），不经此路径。
                    report.location = location
                    report.zone_location = first_zone
                    report.remark = request.POST.get('remark', '')
                    report.is_pending_repair = bool(request.POST.get('is_pending_repair'))
                    report.is_difficult = bool(request.POST.get('is_difficult'))
                    report.is_difficult_resolved = bool(request.POST.get('is_difficult_resolved'))
                    report.shift = shift
                    report.work_start_time = work_start
                    report.work_end_time = work_end
                    report.team_size = team_size
                    report.third_party_count = third_party_count
                    report.team_hours = team_hours
                    report.third_party_hours = third_party_hours
                    report.zone_names = zone_names
                    report.work_content = request.POST.get('work_content', '')
                    report.save()
                    # 编辑历史：edit 分支记录一次，新建工单不记。
                    _record_edit(report, request.user)
                else:
                    report = WorkReport.objects.create(
                        date=request.POST.get('date') or date.today().isoformat(),
                        weather='',
                        worker=post_worker,
                        location=location,
                        zone_location=first_zone,
                        remark=request.POST.get('remark', ''),
                        is_pending_repair=bool(request.POST.get('is_pending_repair')),
                        is_difficult=bool(request.POST.get('is_difficult')),
                        is_difficult_resolved=bool(request.POST.get('is_difficult_resolved')),
                        shift=shift,
                        work_start_time=work_start,
                        work_end_time=work_end,
                        team_size=team_size,
                        third_party_count=third_party_count,
                        team_hours=team_hours,
                        third_party_hours=third_party_hours,
                        zone_names=zone_names,
                        work_content=request.POST.get('work_content', ''),
                    )

                if selected_zones:
                    report.zones.set(selected_zones)

                # Work-content tree entries (replaces the old two-level fault model).
                # _save_entries deletes-then-recreates, so it's idempotent on edit.
                entries = json.loads(request.POST.get('entries', '[]') or '[]')
                _save_entries(report, entries, _collect_entry_photos(request))
                # Material consumption (材料消耗): rebuild the report's outbound
                # transaction from the cart (rolls back prior, then applies). The
                # destination is auto-derived from the work category or user-picked.
                materials = json.loads(request.POST.get('materials', '[]') or '[]')
                m_dest, m_proj, m_cp = _resolve_material_dest(request, entries)
                _save_workorder_materials(report, materials, entry_subtype=m_dest,
                                          related_project_id=m_proj, counterparty=m_cp)
                # 计划性维修: resolve the checked past 待修 workorders (create only —
                # re-resolving on edit would double-link).
                if not is_edit:
                    pm_ids = [x for x in (request.POST.get('pm_resolved') or '').split(',') if x.strip().isdigit()]
                    if pm_ids:
                        _resolve_pending_repairs(report, pm_ids)
                entry_count = report.entries.count()

                if not is_pm_completion and getattr(report, 'is_difficult', False) and not is_edit:
                    note = request.POST.get('remark', '').strip()
                    remark_content = note or (f'疑难工单 · {entry_count} 项' if entry_count else '疑难工单')
                    remark_entry = {
                        'date': report.date if isinstance(report.date, str) else report.date.isoformat(),
                        'content': remark_content,
                        'author': worker.full_name if worker else str(request.user),
                        'workorder_id': report.id,
                    }
                    for z in selected_zones:
                        remarks = json.loads(z.remarks) if z.remarks else []
                        remarks.insert(0, remark_entry)
                        z.remarks = json.dumps(remarks, ensure_ascii=False)
                        z.save(update_fields=['remarks'])

                # Report-level photos. On edit, MERGE: keep existing photos except
                # those the user explicitly removed, then append any new uploads.
                # (On create, just use the new uploads.)
                if is_edit:
                    removed = set(x for x in (request.POST.get('report_photos_remove') or '').split(',') if x.strip())
                    kept = [p for p in (report.photos or []) if p not in removed]
                    # Physically delete the removed media + thumbnails.
                    from core.workorder_tree_views import _delete_media
                    for p in removed:
                        _delete_media(p)
                    new_paths = [_save_photo(report, f) for f in request.FILES.getlist('report_photos')]
                    report.photos = kept + new_paths
                    report.save(update_fields=['photos'])
                else:
                    photo_paths = [_save_photo(report, f)
                                   for f in request.FILES.getlist('report_photos')]
                    if photo_paths:
                        report.photos = photo_paths
                        report.save(update_fields=['photos'])

                # PM 任务完成：如果这条工单关联了一条派发的 GeneratedWorkOrder，
                # 提交即视为完成（班组全体可见的任务从待办列表消失）。
                # 逻辑抽取到 workorder_tree_views.mark_pm_completed，两条提交路径共用。
                # pm_gwo_id 非空时是 PM 完成路径（新建 is_pm 工单）——传给 mark_pm_completed
                # 让它建立 report↔GWO 链接并标完成。
                from core.workorder_tree_views import mark_pm_completed
                mark_pm_completed(report, gwo_id=pm_gwo_id if is_pm_completion else None)

            ticket = report.display_number if hasattr(report, 'display_number') else f'#{report.id}'
            success_msg = f'工作记录已更新 ({ticket})' if is_edit else f'工作记录已提交 ({ticket})'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': success_msg})
            messages.success(request, success_msg)
            return redirect('core:dashboard')

        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': str(e)}, status=400)
            messages.error(request, f'提交失败: {e}')
            return redirect('core:dashboard')

    # GET: redirect to dashboard (modal handles form)
    return redirect('core:dashboard')


@login_required(login_url='core:login')
def water_request_mobile_v2(request):
    from core.models import WaterRequest, Zone, Patch, Worker, DepartmentUserProfile
    from core.role_utils import (
        get_worker_for_user, get_user_role, is_admin,
        ROLE_DEPT_USER, ROLE_MANAGER, ROLE_SUPER_ADMIN,
)
    from datetime import date, datetime

    role = get_user_role(request.user)
    if role not in (ROLE_DEPT_USER, ROLE_MANAGER, ROLE_SUPER_ADMIN):
        messages.error(request, '无权限访问此页面')
        return redirect('core:login')

    worker = get_worker_for_user(request.user)
    user_name = worker.full_name if worker else request.user.get_full_name() or request.user.username
    today = date.today()
    now_time = datetime.now().strftime('%H:%M')

    dept_profile = DepartmentUserProfile.objects.filter(user=request.user).first()
    user_type = dept_profile.department if dept_profile else 'ENT'

    if request.method == 'POST':
        try:
            zone_codes = json.loads(request.POST.get('zone_codes', '[]'))
            if not zone_codes:
                return JsonResponse({'success': False, 'message': '请选择至少一个区域'}, status=400)

            request_type = request.POST.get('request_type', '停水需求')
            start_dt = request.POST.get('start_datetime')
            end_dt = request.POST.get('end_datetime')
            remark = request.POST.get('remark', '')

            if not start_dt or not end_dt:
                return JsonResponse({'success': False, 'message': '请填写需求时间段'}, status=400)

            start_datetime = datetime.fromisoformat(start_dt)
            end_datetime = datetime.fromisoformat(end_dt)

            zones = list(Zone.objects.filter(code__in=zone_codes))
            if not zones:
                return JsonResponse({'success': False, 'message': '未找到选择的区域'}, status=400)

            # ONE request covering all selected zones (multi-zone M2M), so it needs only
            # one approval. Previously each zone got its own request → N rows + N labels.
            # Submitter attribution: dept users (the common case) are stored on
            # submitter_user (a Django User); irrigation workers go on submitter(Worker).
            # Most submitters are department users with no Worker row at all.
            create_kwargs = dict(
                zone=zones[0],  # legacy single-zone FK kept for backward-compat
                user_type=user_type,
                request_type=request_type,
                start_datetime=start_datetime,
                end_datetime=end_datetime,
                status='submitted',
                status_notes=remark,
            )
            if worker:
                create_kwargs['submitter'] = worker
            create_kwargs['submitter_user'] = request.user
            wr = WaterRequest.objects.create(**create_kwargs)
            wr.zones.set(zones)

            return JsonResponse({
                'success': True,
                'message': f'已提交浇水协调需求（{len(zones)} 个区域，1 个审批）',
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    # GET: redirect to dashboard (modal handles form)
    return redirect('core:dashboard')


@login_required(login_url='core:login')
def workorder_modal_data(request):
    """API: return workorder form metadata as JSON for dashboard modal."""
    from core.models import WorkReport, Worker
    from core.role_utils import get_worker_for_user, get_user_role, is_admin, ROLE_FIELD_WORKER, ROLE_SUPER_ADMIN, ROLE_MANAGER
    from core.workorder_tree_views import (
        serialize_workitem_tree, serialize_projects, IRRIGATION_SUBCATEGORIES,
        serialize_existing_entries, _report_header_dict, _serialize_workorder_materials,
        _serialize_workorder_material_dest,
    )
    from datetime import date, datetime
    from django.shortcuts import get_object_or_404

    role = get_user_role(request.user)
    if role not in (ROLE_FIELD_WORKER, ROLE_SUPER_ADMIN, ROLE_MANAGER):
        return JsonResponse({'error': '无权限'}, status=403)

    worker = get_worker_for_user(request.user)

    now = datetime.now()
    rounded_min = (now.minute // 15) * 15
    default_time = now.replace(minute=rounded_min, second=0, microsecond=0).strftime('%H:%M')

    shift_freq = {'早班': 0, '白班': 0, '夜班': 0}
    if worker:
        recent = (WorkReport.objects
                  .filter(worker=worker, shift__in=shift_freq.keys(), is_pm=False)
                  .values_list('shift', flat=True))
        for s in recent:
            if s in shift_freq:
                shift_freq[s] += 1
    sorted_shifts = sorted(shift_freq.keys(), key=lambda s: -shift_freq[s])

    # Edit mode: when a report_id (WorkReport) or pm_order_id (PMWorkOrder) is
    # supplied, also return the existing record's header fields, filled tree
    # entries, and photos so the modal can pre-fill.
    report = None
    report_id = request.GET.get('report_id')
    pm_order_id = request.GET.get('pm_order_id')
    if report_id:
        report = get_object_or_404(WorkReport, pk=report_id)
    elif pm_order_id:
        from core.models import PMWorkOrder
        report = get_object_or_404(PMWorkOrder, pk=pm_order_id)

    payload = {
        'work_tree': serialize_workitem_tree(),
        'projects': serialize_projects(),
        'irrigation_subcategories': IRRIGATION_SUBCATEGORIES(),
        'can_create_project': role in (ROLE_MANAGER, ROLE_SUPER_ADMIN),
        'sorted_shifts': sorted_shifts,
        'today': date.today().isoformat(),
        'now_time': now.strftime('%H:%M'),
        'default_time': default_time,
        'worker_name': worker.full_name if worker else request.user.get_full_name() or request.user.username,
        # Inventory catalog so the workorder modal can render the material-
        # consumption picker without a second round trip.
        'inventory_tree': serialize_inventory_tree(),
    }
    if report:
        from core.models import PMWorkOrder
        if isinstance(report, PMWorkOrder):
            payload['pm_order_id'] = report.id
        else:
            payload['report_id'] = report.id
        payload['header'] = _report_header_dict(report)
        payload['existing'] = serialize_existing_entries(report)
        payload['report_photos'] = report.photos or []
        payload['existing_materials'] = _serialize_workorder_materials(report)
        payload['existing_material_dest'] = _serialize_workorder_material_dest(report)
    return JsonResponse(payload)


@login_required(login_url='core:login')
def water_request_modal_data(request):
    """API: return water request form metadata as JSON for dashboard modal."""
    from core.models import WaterRequest, Worker
    from core.role_utils import get_worker_for_user, get_user_role, ROLE_DEPT_USER, ROLE_MANAGER, ROLE_SUPER_ADMIN
    from datetime import date, datetime

    role = get_user_role(request.user)
    if role not in (ROLE_DEPT_USER, ROLE_MANAGER, ROLE_SUPER_ADMIN):
        return JsonResponse({'error': '无权限'}, status=403)

    worker = get_worker_for_user(request.user)
    now = datetime.now()

    return JsonResponse({
        'request_type_choices': list(WaterRequest.REQUEST_TYPE_CHOICES),
        'user_name': worker.full_name if worker else request.user.get_full_name() or request.user.username,
        'today': date.today().isoformat(),
        'now_time': now.strftime('%H:%M'),
    })


# ─── Inventory Management (库存管理) ──────────────────────────────────

def serialize_inventory_tree():
    """Emit the InventoryCategory tree as nested JSON (depth-first).

    Each leaf carries ``current_stock``, ``min_stock`` and ``is_main_material``
    so the mobile form can display them read-only. Mirrors serialize_workitem_tree's algorithm.
    """
    from core.models import InventoryCategory
    qs = (InventoryCategory.objects.filter(active=True)
          .order_by('order', 'code')
          .values('id', 'code', 'name_zh', 'parent_id', 'current_stock',
                  'min_stock', 'is_main_material', 'node_type', 'unit'))
    nodes = {n['id']: {**n, 'name': n['name_zh'], 'children': []} for n in qs}
    roots = []
    for n in qs:
        node = nodes[n['id']]
        pid = n['parent_id']
        if pid in nodes:
            nodes[pid]['children'].append(node)
        else:
            roots.append(node)
    # Skip a single wrapper root (e.g. "库存材料和工具") — it just adds an extra
    # expand click. Promote its children to top-level so the picker opens
    # straight onto the real sections (喷头 / PVC管 / ...).
    if len(roots) == 1 and roots[0]['children']:
        roots = roots[0]['children']
    return roots


def inventory_category_paths():
    """Map every InventoryCategory id → full display path "大类别 › … › 品名".

    Single query + in-memory parent walk. Replaces the previous per-call
    ``c.parent`` chain walk that fired one DB query per ancestor hop and was
    copy-pasted across inventory_management, _po_received_parts,
    _po_planned_lines and _project_budget_data.
    """
    from core.models import InventoryCategory
    rows = (InventoryCategory.objects
            .values('id', 'name_zh', 'parent_id'))
    name = {r['id']: r['name_zh'] for r in rows}
    parent = {r['id']: r['parent_id'] for r in rows}
    out = {}
    for cid in name:
        chain = []
        cur = parent.get(cid)
        seen = set()
        while cur and cur not in seen:        # guard against accidental cycles
            seen.add(cur)
            chain.append(name.get(cur, '?'))
            cur = parent.get(cur)
        chain.reverse()
        out[cid] = ' › '.join(chain + [name[cid]]) if chain else name[cid]
    return out


@login_required(login_url='core:login')
def inventory_modal_data(request):
    """API: return inventory form metadata as JSON for the dashboard modal."""
    from core.models import InventoryTransaction, Project, PurchaseOrder
    from core.role_utils import get_worker_for_user, get_user_role, ROLE_FIELD_WORKER, ROLE_MANAGER, ROLE_SUPER_ADMIN
    from core.workorder_tree_views import serialize_projects
    from datetime import date, datetime

    role = get_user_role(request.user)
    if role not in (ROLE_FIELD_WORKER, ROLE_MANAGER, ROLE_SUPER_ADMIN):
        return JsonResponse({'error': '无权限'}, status=403)

    worker = get_worker_for_user(request.user)
    now = datetime.now()

    # Borrower list = distinct counterparties from past 出库-借用 transactions,
    # so the field self-populates over time. Users can still type a new one.
    borrowers = list(InventoryTransaction.objects
                     .filter(operation='出库', entry_subtype='借用')
                     .exclude(counterparty='')
                     .values_list('counterparty', flat=True)
                     .distinct().order_by('counterparty'))

    # Edit mode: when a txn_id is supplied, also return the existing txn's
    # fields + lines + edit history so the modal can pre-fill (mirrors
    # workorder_modal_data's report_id block).
    txn_id = request.GET.get('txn_id')
    edit_payload = {}
    if txn_id and txn_id.isdigit():
        from django.shortcuts import get_object_or_404
        txn = get_object_or_404(
            InventoryTransaction.objects
            .select_related('related_project', 'zone')
            .prefetch_related('lines__category', 'edit_logs__editor'),
            pk=txn_id,
        )
        edit_payload = {
            'txn_id': txn.id,
            'header': {
                'operation': txn.operation,
                'entry_subtype': txn.entry_subtype,
                'date': txn.date.isoformat(),
                'order_no': txn.order_no,
                'counterparty': txn.counterparty,
                'project_id': txn.related_project_id,
                'project_category': txn.related_project.category if txn.related_project else None,
                'zone_id': txn.zone_id,
                'remark': txn.remark,
            },
            'existing_lines': [
                {'category': ln.category_id, 'quantity': ln.quantity, 'unit': ln.unit,
                 'name': ln.category.name_zh, 'stock': ln.category.current_stock}
                for ln in txn.lines.all()
            ],
            'edit_logs': [
                {'editor': (log.editor.full_name if log.editor else '(未知)'),
                 'time': log.created_at.strftime('%Y-%m-%d %H:%M')}
                for log in txn.edit_logs.all()
            ],
        }

    return JsonResponse({
        'inventory_tree': serialize_inventory_tree(),
        'operations': [
            {'op': '入库', 'label': '入库', 'subtypes': [s for s, _ in InventoryTransaction.INBOUND_SUBTYPES]},
            {'op': '出库', 'label': '出库', 'subtypes': [s for s, _ in InventoryTransaction.OUTBOUND_DESTINATIONS]},
        ],
        'projects': serialize_projects(),
        'project_categories': [{'code': c, 'label': label} for c, label in Project.CATEGORY_CHOICES],
        'borrowers': borrowers,
        # 灌溉订单编号列表 → 入库-采购表单的订单号 datalist 选项（仍允许自由输入）。
        'purchase_orders': list(PurchaseOrder.objects
                                .values_list('order_number', flat=True)
                                .order_by('-created_at')),
        'today': date.today().isoformat(),
        'now_time': now.strftime('%H:%M'),
        'worker_name': worker.full_name if worker else request.user.get_full_name() or request.user.username,
        **edit_payload,
    })


@login_required(login_url='core:login')
def inventory_mobile_v2(request):
    """Submit an inventory transaction (a multi-item cart of stock movements).

    Mirrors workorder_mobile_v2's structure: role-gated POST that reads form
    fields + a JSON ``lines`` array, creates the transaction + its lines, then
    atomically adjusts each category's current_stock via F().
    """
    from core.models import (
        InventoryTransaction, InventoryTransactionLine, InventoryCategory, Project, Zone,
        PurchaseOrder,
    )
    from core.role_utils import get_user_role, resolve_or_create_worker
    from core.role_utils import ROLE_FIELD_WORKER, ROLE_MANAGER, ROLE_SUPER_ADMIN
    from datetime import date

    role = get_user_role(request.user)
    if role not in (ROLE_FIELD_WORKER, ROLE_MANAGER, ROLE_SUPER_ADMIN):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)

    if request.method != 'POST':
        return redirect('core:dashboard')

    try:
        worker, _created = resolve_or_create_worker(request.user)
        if not worker:
            return JsonResponse({'success': False, 'message': '未关联处理人账号'}, status=400)

        operation = (request.POST.get('operation') or '').strip()
        if operation not in dict(InventoryTransaction.OPERATION_CHOICES):
            return JsonResponse({'success': False, 'message': '操作类型无效'}, status=400)
        entry_subtype = (request.POST.get('entry_subtype') or '').strip()

        lines_raw = request.POST.get('lines', '[]')
        try:
            lines = json.loads(lines_raw)
        except (json.JSONDecodeError, TypeError):
            lines = []
        if not lines:
            return JsonResponse({'success': False, 'message': '请至少添加一个物料'}, status=400)

        # Optional zone association.
        zone_id = request.POST.get('zone_id')
        zone = None
        if zone_id and zone_id.isdigit():
            zone = Zone.objects.filter(pk=zone_id).first()

        # Optional project (出库-项目).
        project_id = request.POST.get('project_id')
        project = None
        if project_id and project_id.isdigit():
            project = Project.objects.filter(pk=project_id).first()

        # Consumption mode (出库-项目 only): 'actual' (default) or 'estimated'.
        # Estimated consumption records the txn + lines but does NOT deduct stock —
        # it's confirmed (and stock applied) later from the inventory manage page.
        consumption_mode = 'estimated' if request.POST.get('consumption_mode') == 'estimated' else 'actual'

        # Edit mode: load the existing txn and reverse its old stock delta before
        # re-applying. Both paths then share the same line-application + PO-linking
        # tail. Wrapped in a transaction so a partial failure can't corrupt stock.
        from django.db import transaction as _db_txn
        from core.models import InventoryTransactionEditLog
        edit_txn_id = request.POST.get('txn_id')
        is_edit = bool(edit_txn_id and edit_txn_id.isdigit())
        with _db_txn.atomic():
            if is_edit:
                txn = get_object_or_404(InventoryTransaction, pk=edit_txn_id)
                # Reverse the OLD delta so stock reflects only what the new lines
                # will apply below. Old direction may differ from the new one.
                # Skip reversal for the OLD txn's estimated lines — they never
                # deducted stock, so reversing them would corrupt the balance.
                if txn.consumption_mode != 'estimated':
                    old_sign = 1 if txn.operation == '入库' else -1
                    for old_ln in txn.lines.all():
                        InventoryCategory.objects.filter(pk=old_ln.category_id).update(
                            current_stock=F('current_stock') - old_sign * old_ln.quantity,
                        )
                txn.lines.all().delete()
                # Apply the edited fields.
                txn.date = request.POST.get('date') or date.today().isoformat()
                txn.operation = operation
                txn.entry_subtype = entry_subtype
                txn.order_no = (request.POST.get('order_no') or '').strip()
                txn.counterparty = (request.POST.get('counterparty') or '').strip()
                txn.related_project = project
                txn.consumption_mode = consumption_mode
                txn.remark = (request.POST.get('remark') or '').strip()
                txn.zone = zone
                txn.purchase_order = None   # re-linked below if 入库-采购 hits a PO
                txn.save()
            else:
                txn = InventoryTransaction.objects.create(
                    date=request.POST.get('date') or date.today().isoformat(),
                    worker=worker,
                    operation=operation,
                    entry_subtype=entry_subtype,
                    order_no=(request.POST.get('order_no') or '').strip(),
                    counterparty=(request.POST.get('counterparty') or '').strip(),
                    related_project=project,
                    consumption_mode=consumption_mode,
                    remark=(request.POST.get('remark') or '').strip(),
                    zone=zone,
                )

            # 入库-采购时，若订单号命中某张采购订单，则自动关联该流水到订单。
            # （自由输入且未命中的订单号保持 purchase_order=NULL，仍以文本形式保留。）
            if operation == '入库' and entry_subtype == '采购' and txn.order_no:
                po = PurchaseOrder.objects.filter(order_number=txn.order_no).first()
                if po is not None:
                    txn.purchase_order = po
                    txn.save(update_fields=['purchase_order'])

            # Direction: 入库 adds stock; 出库 subtracts. Estimated consumption
            # (预估消耗) records the line but defers the stock deduction until the
            # transaction is confirmed from the inventory manage page.
            sign = 1 if operation == '入库' else -1
            deduct_stock = consumption_mode != 'estimated'
            created_lines = 0
            for ln in lines:
                cat_id = ln.get('category')
                qty = ln.get('quantity') or 0
                try:
                    qty = abs(float(qty))
                except (ValueError, TypeError):
                    continue
                cat = InventoryCategory.objects.filter(pk=cat_id).first()
                if not cat or qty <= 0:
                    continue
                InventoryTransactionLine.objects.create(
                    transaction=txn, category=cat,
                    quantity=qty, unit=(ln.get('unit') or '').strip(),
                )
                if deduct_stock:
                    # Atomically adjust current_stock (F() avoids race on concurrent submits).
                    InventoryCategory.objects.filter(pk=cat_id).update(
                        current_stock=F('current_stock') + sign * qty,
                    )
                created_lines += 1

            if created_lines == 0:
                if is_edit:
                    raise ValueError('没有有效的物料行')   # rolls back the whole atomic block
                txn.delete()
                return JsonResponse({'success': False, 'message': '没有有效的物料行'}, status=400)

            # Edit history: record one log row per edit save (new txns get none).
            if is_edit:
                InventoryTransactionEditLog.objects.create(
                    transaction=txn, editor=worker, note='',
                )

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            sub = f'-{entry_subtype}' if entry_subtype else ''
            verb = '已更新' if is_edit else '已提交'
            return JsonResponse({
                'success': True,
                'message': f'库存{operation}{sub}{verb} ({created_lines} 项)',
            })
        return redirect('core:dashboard')

    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
        return redirect('core:dashboard')


@login_required(login_url='core:login')
def inventory_management(request):
    """Manager-only page: collapsible inventory tree + per-item stock editing.

    The catalog renders as a nested, all-collapsed-by-default tree. Leaf nodes
    show an editable current-stock input and a "查看出入库记录" trigger that
    lazy-loads that item's transactions within the selected date range.
    """
    from core.models import InventoryCategory
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    from datetime import date, timedelta

    role = get_user_role(request.user)
    # 管理員/超管可访问整个页面。灌溉一线(field_worker)可见「出入库记录」+
    # 「预估消耗确认」两个标签——出入库记录只显示本人提交的流水，其余管理功能
    # （库存警告、库存目录、采购订单）仍限管理員。
    is_full_access = role in (ROLE_MANAGER, ROLE_SUPER_ADMIN)
    if not is_full_access and role != 'field_worker':
        messages.error(request, '无权限访问库存管理')
        return redirect('core:dashboard')

    # Flat query → nested tree (same algorithm as serialize_inventory_tree,
    # but on real model objects so the template can render them directly).
    cats = list(InventoryCategory.objects.filter(active=True).order_by('order', 'code'))
    nodes = {c.id: {'obj': c, 'children': []} for c in cats}
    roots = []
    for c in cats:
        node = nodes[c.id]
        if c.parent_id and c.parent_id in nodes:
            nodes[c.parent_id]['children'].append(node)
        else:
            roots.append(node)

    # Skip a single wrapper root — promote its children to top-level (see
    # serialize_inventory_tree for the rationale).
    if len(roots) == 1 and roots[0]['children']:
        roots = roots[0]['children']

    # Default date range: last 30 days.
    today = date.today()
    date_from = request.GET.get('from') or (today - timedelta(days=30)).isoformat()
    date_to = request.GET.get('to') or today.isoformat()

    # Transactions ledger (出入库记录 tab): every transaction in the date range
    # with its lines expanded, newest first. Prefetch lines + related objects so
    # the template renders without N+1 queries.
    from core.models import InventoryTransaction
    # 出入库记录只显示已确认的实际流水——预估消耗(estimated)尚未扣库存，
    # 只出现在「预估消耗确认」标签页，确认后才进入出入库记录。
    txn_filter = {'date__gte': date_from, 'date__lte': date_to,
                  'consumption_mode': 'actual'}
    # 灌溉一线只看到本人提交的出入库记录；管理員看到全部。
    if not is_full_access:
        from core.role_utils import get_worker_for_user
        worker = get_worker_for_user(request.user)
        if worker:
            txn_filter['worker_id'] = worker.id
        else:
            txn_filter['worker_id'] = -1   # no linked worker → show nothing
    txns = (InventoryTransaction.objects
            .filter(**txn_filter)
            .select_related('worker', 'related_project', 'zone', 'work_report')
            .prefetch_related('lines__category', 'edit_logs__editor')
            .order_by('-date', '-id'))

    # Stockable leaves for the alert tab — {id, name, path, stock, min}. The
    # threshold filtering is done client-side so the slider updates instantly.
    # Also builds a cat_id → full hierarchy path map reused by the ledger tab.
    # Path computation uses one query + in-memory parent walk (was N×D queries).
    full_paths = inventory_category_paths()   # id → "大类别 › … › 品名" (leaf incl.)
    cat_paths = dict(full_paths)              # alias used by the ledger tab
    leaves = []
    for c in cats:
        if c.node_type != 'part':
            continue   # only 'part' nodes carry stock; empty categories are skipped
        # leaves[].path is ancestors only (no leaf name); strip the trailing leaf.
        full = full_paths.get(c.id, c.name_zh)
        ancestor_path = full.rsplit(' › ', 1)[0] if ' › ' in full else ''
        leaves.append({
            'id': c.id, 'name': c.name_zh, 'path': ancestor_path,
            'stock': c.current_stock, 'min': c.min_stock,
            'main': c.is_main_material, 'unit': c.unit,
        })

    # Purchase orders for the 采购订单 tab — batched (2 queries total, was 2 per PO).
    from core.models import PurchaseOrder
    pos = list(PurchaseOrder.objects.all().order_by('-created_at'))
    pos_json = json.dumps(_serialize_purchase_orders(pos), ensure_ascii=False)

    # Pending estimated-consumption transactions for the 预估消耗确认 tab.
    pending_estimates = (InventoryTransaction.objects
                         .filter(consumption_mode='estimated')
                         .select_related('worker', 'related_project')
                         .prefetch_related('lines__category')
                         .order_by('-date', '-id'))

    return render(request, 'core/inventory_management.html', {
        'roots': roots,
        'date_from': date_from,
        'date_to': date_to,
        'txns': txns,
        'pending_estimates': pending_estimates,
        'is_full_access': is_full_access,
        'cat_paths_json': json.dumps(cat_paths, ensure_ascii=False),
        'leaves_json': json.dumps(leaves, ensure_ascii=False),
        'pos_json': pos_json,
        'inventory_tree_json': json.dumps(serialize_inventory_tree(), ensure_ascii=False),
        'projects_json': json.dumps(_po_project_options(), ensure_ascii=False),
    })


@login_required(login_url='core:login')
def inventory_estimate_confirm(request, txn_id):
    """Confirm a pending estimated-consumption transaction as actual.

    Manager + field-worker. Field workers can submit estimated consumption and
    also need to confirm it. Updates each line to the (possibly revised) actual
    quantity, flips consumption_mode 'estimated' → 'actual', and only NOW deducts
    stock — this is the deferred deduction the original submission skipped.
    Writes an InventoryTransactionEditLog entry as an audit trail.

    POST params: ``quantities`` = JSON ``{line_id: qty}``.
    """
    from core.models import InventoryTransaction, InventoryTransactionEditLog, InventoryCategory
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN, ROLE_FIELD_WORKER
    from django.db import transaction as _db_txn
    import json as _json

    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN, ROLE_FIELD_WORKER):
        return JsonResponse({'success': False, 'error': '无权限'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST only'}, status=405)

    txn = get_object_or_404(InventoryTransaction, pk=txn_id)
    if txn.consumption_mode != 'estimated':
        return JsonResponse({'success': False, 'error': '该流水不是预估消耗'}, status=400)

    # Revised actual quantities: {line_id: qty}. Missing lines keep their estimate.
    try:
        quantities = _json.loads(request.POST.get('quantities') or '{}')
    except (ValueError, TypeError):
        quantities = {}

    try:
        with _db_txn.atomic():
            for ln in txn.lines.select_related('category'):
                new_qty = quantities.get(str(ln.id))
                if new_qty is not None:
                    try:
                        new_qty = abs(float(new_qty))
                    except (ValueError, TypeError):
                        continue
                    ln.quantity = new_qty
                    ln.save(update_fields=['quantity'])
                # Apply the (possibly revised) stock deduction NOW — the original
                # estimated submission recorded the line but skipped this step.
                InventoryCategory.objects.filter(pk=ln.category_id).update(
                    current_stock=F('current_stock') - ln.quantity,
                )
            txn.consumption_mode = 'actual'
            txn.save(update_fields=['consumption_mode'])
            worker = getattr(request.user, 'worker_profile', None)
            InventoryTransactionEditLog.objects.create(
                transaction=txn, editor=worker,
                note='预估消耗确认（确认前未扣库存，确认后按实际量扣减）',
            )
    except Exception as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=500)

    return JsonResponse({'success': True})


@login_required(login_url='core:login')
def inventory_estimate_void(request, txn_id):
    """Void (delete) a pending estimated-consumption transaction.

    Manager + field-worker. Only estimated txns can be voided — they never
    deducted stock, so deletion is clean (no reversal needed). Actual txns must
    be edited instead.
    """
    from core.models import InventoryTransaction
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN, ROLE_FIELD_WORKER

    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN, ROLE_FIELD_WORKER):
        return JsonResponse({'success': False, 'error': '无权限'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST only'}, status=405)

    txn = get_object_or_404(InventoryTransaction, pk=txn_id)
    if txn.consumption_mode != 'estimated':
        return JsonResponse({'success': False, 'error': '实际消耗流水不能作废，请改用编辑'}, status=400)

    txn.delete()   # CASCADE removes lines; estimated never touched stock.
    return JsonResponse({'success': True})


@login_required(login_url='core:login')
def inventory_category_transactions(request, cat_id):
    """JSON: a leaf category's in/out transactions within a date range.

    Manager-only. Used by the lazy-loaded "查看出入库记录" panel on each leaf.
    """
    from core.models import InventoryCategory, InventoryTransactionLine
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN

    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        return JsonResponse({'error': '无权限'}, status=403)

    cat = get_object_or_404(InventoryCategory, pk=cat_id)
    date_from = request.GET.get('from') or ''
    date_to = request.GET.get('to') or ''

    qs = (InventoryTransactionLine.objects
          .filter(category=cat)
          .select_related('transaction', 'transaction__worker', 'transaction__related_project'))
    if date_from:
        qs = qs.filter(transaction__date__gte=date_from)
    if date_to:
        qs = qs.filter(transaction__date__lte=date_to)
    qs = qs.order_by('-transaction__date', '-transaction__id')

    records = []
    for ln in qs:
        t = ln.transaction
        records.append({
            'date': t.date.isoformat() if t.date else '',
            'operation': t.operation,
            'subtype': t.entry_subtype,
            'quantity': ln.quantity,
            'unit': ln.unit,
            'counterparty': t.counterparty,
            'order_no': t.order_no,
            'project': t.related_project.name if t.related_project else '',
            'worker': t.worker.full_name if t.worker else '',
            'remark': t.remark,
        })
    return JsonResponse({'success': True, 'name': cat.name_zh, 'count': len(records), 'records': records})


@require_POST
@login_required(login_url='core:login')
def inventory_save_stock(request):
    """Update inventory category attributes for one or more leaves (manager only).

    Reads three kinds of POST fields:
      • ``stock_<id>`` — current_stock (int)
      • ``min_<id>``   — min_stock (int)
      • ``main_<id>``  — is_main_material (checkbox: present=on, absent=off)

    Responds with JSON when the request is AJAX (so the tree's expand state is
    preserved); falls back to a redirect for non-AJAX form posts.
    """
    from core.models import InventoryCategory
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': '无权限'}, status=403)
        messages.error(request, '无权限')
        return redirect('core:dashboard')

    # Collect the three field types into per-id update dicts so we issue one
    # .update() per touched category instead of three.
    changes = {}   # cat_id -> {field: value}
    for key, val in request.POST.items():
        if key.startswith('stock_') and key[6:].isdigit():
            cid = key[6:]
            try:
                changes.setdefault(cid, {})['current_stock'] = int(val)
            except (ValueError, TypeError):
                pass
        elif key.startswith('min_') and key[4:].isdigit():
            cid = key[4:]
            try:
                changes.setdefault(cid, {})['min_stock'] = int(val)
            except (ValueError, TypeError):
                pass
        elif key.startswith('main_') and key[5:].isdigit():
            # Checkbox: any presence means "on".
            changes.setdefault(key[5:], {})['is_main_material'] = True

    # Any category that has a main_ checkbox absent from the POST was unchecked.
    # We still need to flip it to False — but only for categories shown on the
    # page (which are exactly the ones whose stock_/min_ fields were submitted).
    for cid in changes:
        changes[cid].setdefault('is_main_material', False)

    updated = 0
    for cid, fields in changes.items():
        InventoryCategory.objects.filter(pk=cid).update(**fields)
        updated += 1

    msg = f'已更新 {updated} 个物料的库存设置'
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': msg, 'updated': updated})
    messages.success(request, msg)
    return redirect('core:inventory_management')


@require_POST
@login_required(login_url='core:login')
def inventory_category_create(request):
    """Create a new inventory category (sub-directory or leaf part) under a parent.

    Manager-only. ``parent_id`` may be empty (creates a top-level root). ``node_type``
    is 'category' (a branch that can hold children) or 'part' (a leaf — also gets
    initial stock/min/main values). The code is derived from the parent's code path
    + a slugified name + a numeric suffix to stay unique.
    """
    import re
    from core.models import InventoryCategory
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': '无权限'}, status=403)
        messages.error(request, '无权限')
        return redirect('core:inventory_management')

    name = (request.POST.get('name') or '').strip()
    if not name:
        return _iv_create_err(request, '请填写名称')
    node_type = request.POST.get('node_type') or 'category'
    parent_id = request.POST.get('parent_id') or ''

    parent = None
    if parent_id:
        parent = InventoryCategory.objects.filter(pk=parent_id).first()
        if not parent:
            return _iv_create_err(request, '父节点不存在')

    # Derive a stable, unique code: parent.code + slug(name) + .N suffix if needed.
    base = _iv_slug(name)
    if parent:
        prefix = parent.code + '.' + base
        level = parent.level + 1
    else:
        prefix = base
        level = 0
    code = prefix
    n = 1
    while InventoryCategory.objects.filter(code=code).exists():
        code = f'{prefix}.{n}'
        n += 1

    # order: append after the last sibling so new nodes show up at the end.
    last_order = (InventoryCategory.objects.filter(parent=parent)
                  .order_by('-order').values_list('order', flat=True).first()) or 0

    cat = InventoryCategory.objects.create(
        code=code, parent=parent, name_zh=name, level=level,
        order=last_order + 1, node_type=node_type,
    )
    # Leaf-only attributes (a 'part' starts as a leaf; a 'category' has no stock).
    if node_type == 'part':
        try:
            cat.current_stock = int(request.POST.get('current_stock') or 0)
        except (ValueError, TypeError):
            cat.current_stock = 0
        try:
            cat.min_stock = int(request.POST.get('min_stock') or 0)
        except (ValueError, TypeError):
            cat.min_stock = 0
        cat.is_main_material = bool(request.POST.get('is_main_material'))
        cat.save()

    msg = f'已创建「{name}」' + ('（部件）' if node_type == 'part' else '（目录）')
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Return full node data so the frontend can insert it into the live DOM
        # without a page reload (preserving tab + expand state).
        return JsonResponse({
            'success': True, 'message': msg, 'id': cat.id,
            'node': {
                'id': cat.id, 'name': cat.name_zh, 'node_type': cat.node_type,
                'current_stock': cat.current_stock, 'min_stock': cat.min_stock,
                'is_main_material': cat.is_main_material, 'level': cat.level,
                'parent_id': parent.id if parent else None,
            },
        })
    messages.success(request, msg)
    return redirect('core:inventory_management')


def _iv_create_err(request, message):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'message': message}, status=400)
    messages.error(request, message)
    return redirect('core:inventory_management')


def _iv_slug(name):
    """Best-effort ASCII slug for code derivation. Falls back to pinyin-less
    stripping: non-alphanumerics removed; Chinese chars become 'cat' so the code
    stays ASCII-stable (the real uniqueness comes from the .N suffix anyway)."""
    import re
    s = re.sub(r'[^A-Za-z0-9]+', '', name).lower()
    return s[:30] if s else 'cat'


@require_POST
@login_required(login_url='core:login')
def inventory_category_delete(request, cat_id):
    """Delete an inventory category — but only if it has no children and no
    transaction history (PROTECT on InventoryTransactionLine.category would raise
    otherwise). Soft-block with a clear message when blocked.
    """
    from core.models import InventoryCategory, InventoryTransactionLine
    from django.db.models import ProtectedError
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': '无权限'}, status=403)
        messages.error(request, '无权限')
        return redirect('core:inventory_management')

    cat = get_object_or_404(InventoryCategory, pk=cat_id)
    # Block 1: has children → must delete children first.
    if cat.children.exists():
        msg = f'「{cat.name_zh}」下还有 {cat.children.count()} 个子节点，请先删除子节点'
        return _iv_create_err(request, msg)
    # Block 2: has transaction history → PROTECT prevents deletion; refuse up front
    # with a helpful message instead of letting it raise a 500.
    if InventoryTransactionLine.objects.filter(category=cat).exists():
        n = InventoryTransactionLine.objects.filter(category=cat).count()
        msg = f'「{cat.name_zh}」有 {n} 条出入库记录，无法删除（请保留历史或改为停用）'
        return _iv_create_err(request, msg)

    label = cat.name_zh
    try:
        cat.delete()
    except ProtectedError:
        return _iv_create_err(request, f'「{label}」被引用，无法删除')
    msg = f'已删除「{label}」'
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': msg})
    messages.success(request, msg)
    return redirect('core:inventory_management')


@require_POST
@login_required(login_url='core:login')
def inventory_category_edit(request, cat_id):
    """Rename / update an inventory category or part (manager only).

    Updates ``name_zh`` (always), and for parts also ``min_stock`` /
    ``is_main_material``. ``current_stock`` is left untouched (it's adjusted by
    transactions, not by editing). ``node_type`` is preserved.
    """
    from core.models import InventoryCategory
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': '无权限'}, status=403)
        messages.error(request, '无权限')
        return redirect('core:inventory_management')

    cat = get_object_or_404(InventoryCategory, pk=cat_id)
    name = (request.POST.get('name') or '').strip()
    if not name:
        return _iv_create_err(request, '请填写名称')

    cat.name_zh = name
    # Part-only fields (categories have no stock/min/main).
    if cat.node_type == 'part':
        try:
            cat.min_stock = int(request.POST.get('min_stock') or 0)
        except (ValueError, TypeError):
            cat.min_stock = 0
        cat.is_main_material = bool(request.POST.get('is_main_material'))
        cat.unit = (request.POST.get('unit') or '').strip()
    cat.save()

    msg = f'已更新「{name}」'
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True, 'message': msg,
            'node': {
                'id': cat.id, 'name': cat.name_zh, 'node_type': cat.node_type,
                'current_stock': cat.current_stock, 'min_stock': cat.min_stock,
                'is_main_material': cat.is_main_material, 'unit': cat.unit,
            },
        })
    messages.success(request, msg)
    return redirect('core:inventory_management')


@login_required(login_url='core:login')
def inventory_export_excel(request):
    """Export the inventory catalog (leaves only) as an Excel spreadsheet.

    Columns: 大类别 / 小类别 / 系列 / 品名 / 是否主材 / 最小库存 / 现有库存.
    The hierarchy depth varies (1–5 levels); the last node is always the 品名
    (part name), the first is always 大类别. Intermediate levels fill 小类别 /
    系列 in order; when a path is shorter than 4 levels, the gaps are filled
    bottom-up by repeating the nearest available ancestor (per the user's spec:
    "如果只有2级，则系列、品名都可以重复小类别的内容").
    """
    from core.models import InventoryCategory
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        messages.error(request, '无权限')
        return redirect('core:dashboard')

    # Build the ancestor chain (root → ... → leaf) for every part leaf.
    cats = {c.id: c for c in InventoryCategory.objects.filter(active=True)}
    rows = []
    for c in cats.values():
        if c.node_type != 'part':
            continue
        chain = []
        node = c
        while node:
            chain.append(node.name_zh)
            node = cats.get(node.parent_id) if node.parent_id else None
        chain.reverse()   # root → leaf
        rows.append((chain, c))

    wb = Workbook()
    ws = wb.active
    ws.title = '库存目录'
    headers = ['大类别', '小类别', '系列', '品名', '是否主材', '单位', '最小库存', '现有库存']
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill(start_color='1B4332', end_color='1B4332', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    for ri, (chain, cat) in enumerate(rows, 2):
        # Map the variable-length chain onto 4 fixed columns.
        # 品名 = last element; 大类别 = first element.
        # 小类别 / 系列 = the middle levels; if fewer than 4 levels, fill
        # bottom-up by repeating the nearest ancestor.
        name = chain[-1]                      # 品名 (always the leaf)
        big = chain[0] if chain else ''       # 大类别 (always the root)
        mid = chain[1:-1] if len(chain) > 2 else []   # levels between root and leaf
        # Target 2 middle slots: [小类别, 系列]
        if len(mid) >= 2:
            small, series = mid[0], mid[1]
        elif len(mid) == 1:
            small = series = mid[0]
        else:
            # Only 2 levels (root + leaf): repeat root as small + series.
            small = series = big
        vals = [big, small, series, name,
                '是' if cat.is_main_material else '',
                cat.unit or '',
                cat.min_stock or 0, cat.current_stock or 0]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = border
            if ci >= 6:
                cell.alignment = Alignment(horizontal='right')
            if cat.is_main_material and ci == 5:
                cell.font = Font(bold=True, color='2D6A4F')

    # Column widths
    for ci, w in enumerate([18, 22, 22, 28, 10, 8, 10, 10], 1):
        ws.column_dimensions[chr(64 + ci)].width = w
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="inventory_catalog.xlsx"'
    return resp


@login_required(login_url='core:login')
def inventory_transactions_export(request):
    """Export the 出入库记录 ledger as an Excel spreadsheet.

    One row per transaction line (a multi-item cart expands to multiple rows).
    Columns: 日期 / 类型 / 物料(full hierarchy path) / 数量 / 单位 / 经办人 /
    去向来源 / 关联 / 备注. Date range from ?from=&to= (defaults to last 30 days).
    """
    from core.models import InventoryCategory, InventoryTransaction
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    from datetime import date, timedelta
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        messages.error(request, '无权限')
        return redirect('core:dashboard')

    today = date.today()
    date_from = request.GET.get('from') or (today - timedelta(days=30)).isoformat()
    date_to = request.GET.get('to') or today.isoformat()

    # Build cat_id → full hierarchy path (same as the ledger tab).
    cats = {c.id: c for c in InventoryCategory.objects.filter(active=True)}
    cat_paths = {}
    for c in cats.values():
        chain = []
        p = c.parent
        while p:
            chain.append(p.name_zh)
            p = cats.get(p.parent_id) if p.parent_id else None
        chain.reverse()
        cat_paths[c.id] = ' › '.join(chain + [c.name_zh]) if chain else c.name_zh

    txns = (InventoryTransaction.objects
            .filter(date__gte=date_from, date__lte=date_to)
            .select_related('worker', 'related_project', 'zone', 'work_report')
            .prefetch_related('lines__category')
            .order_by('-date', '-id'))

    wb = Workbook()
    ws = wb.active
    ws.title = '出入库记录'
    headers = ['日期', '类型', '物料', '数量', '单位', '经办人', '去向/来源', '关联', '备注']
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill(start_color='1B4332', end_color='1B4332', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = border

    ri = 2
    for t in txns:
        op = f"{t.operation}-{t.entry_subtype}" if t.entry_subtype else t.operation
        ctx = ''
        if t.operation == '出库' and t.entry_subtype == '项目' and t.related_project:
            ctx = '项目 · ' + t.related_project.name
        elif t.operation == '出库' and t.entry_subtype == '借用' and t.counterparty:
            ctx = '借用 · ' + t.counterparty
        elif t.operation == '入库' and t.entry_subtype == '采购' and t.order_no:
            ctx = '采购 · 订单 ' + t.order_no
        else:
            ctx = t.entry_subtype or ''
        link = ''
        if t.work_report_id:
            link = f'工单 #{t.work_report_id}'
        elif t.zone:
            link = t.zone.code or t.zone.name_zh
        for ln in t.lines.all():
            vals = [
                t.date.isoformat() if t.date else '',
                op,
                cat_paths.get(ln.category_id, ln.category.name_zh if ln.category_id else ''),
                ('+' if t.operation == '入库' else '-') + str(ln.quantity),
                ln.unit or '',
                t.worker.full_name if t.worker_id and t.worker else '',
                ctx, link, t.remark or '',
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = border
                if ci == 4:
                    cell.font = Font(bold=True, color='2D6A4F' if t.operation == '入库' else 'c0392b')
            ri += 1

    for ci, w in enumerate([12, 14, 36, 10, 8, 12, 22, 14, 24], 1):
        ws.column_dimensions[chr(64 + ci)].width = w
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="inventory_transactions.xlsx"'
    return resp


# ==========================================================================
# 采购订单 (Purchase Order) 管理
# ==========================================================================

def _po_received_parts_batch(po_ids, cat_paths):
    """Batched replacement for the old per-PO _po_received_parts.

    One query over 入库-采购 transaction lines for all POs in ``po_ids``,
    grouped by purchase_order_id. Returns ``{po_id: (parts_list, txn_count)}``,
    where each part is ``{category_id, name, path, qty, unit}`` (path looked up
    from the precomputed ``cat_paths`` map — no per-line parent walk).
    """
    from core.models import InventoryTransaction
    po_ids = list(po_ids)
    if not po_ids:
        return {}
    agg = {}        # (po_id, cat_id) → entry
    txn_count = {}  # po_id → int
    txns = (InventoryTransaction.objects
            .filter(purchase_order_id__in=po_ids, operation='入库', entry_subtype='采购')
            .prefetch_related('lines__category'))
    for t in txns:
        pid = t.purchase_order_id
        txn_count[pid] = txn_count.get(pid, 0) + 1
        for ln in t.lines.all():
            cat = ln.category
            key = (pid, cat.id)
            entry = agg.get(key)
            if entry is None:
                entry = {'category_id': cat.id, 'name': cat.name_zh,
                         'path': cat_paths.get(cat.id, cat.name_zh),
                         'unit': ln.unit or cat.unit, 'qty': 0}
                agg[key] = entry
            entry['qty'] += ln.quantity
    out = {}
    for pid in po_ids:
        out[pid] = ([v for (p, _c), v in agg.items() if p == pid],
                    txn_count.get(pid, 0))
    return out


def _po_planned_lines_batch(po_ids, cat_paths):
    """Batched replacement for the old per-PO _po_planned_lines.

    One query over PurchaseOrderLine for all POs in ``po_ids``, grouped by
    purchase_order_id. Returns ``{po_id: [{category_id, name, path, qty, unit}]}``.
    """
    from core.models import PurchaseOrderLine
    po_ids = list(po_ids)
    if not po_ids:
        return {}
    out = {pid: [] for pid in po_ids}
    for ln in (PurchaseOrderLine.objects
               .filter(purchase_order_id__in=po_ids)
               .select_related('category')):
        cat = ln.category
        out[ln.purchase_order_id].append({
            'category_id': cat.id, 'name': cat.name_zh,
            'path': cat_paths.get(cat.id, cat.name_zh),
            'qty': ln.quantity, 'unit': ln.unit or cat.unit,
        })
    return out


def _po_upsert_planned_lines(po, lines_raw):
    """Apply planned purchase lines (from the JSON ``lines`` POST field) to a PO.

    ``lines_raw`` is the raw JSON string (same contract as the 入库 cart):
    a list of ``{category, quantity, unit}``. Replace semantics: lines for
    categories not in the payload are removed; the rest are created/updated.
    Invalid entries (bad category id / non-positive qty) are skipped silently.
    """
    from core.models import InventoryCategory, PurchaseOrderLine
    try:
        lines = json.loads(lines_raw or '[]')
    except (json.JSONDecodeError, TypeError):
        lines = []
    if not isinstance(lines, list):
        return
    seen = {}
    for ln in lines:
        if not isinstance(ln, dict):
            continue
        cat_id = ln.get('category')
        try:
            qty = int(ln.get('quantity') or 0)
        except (ValueError, TypeError):
            continue
        if not cat_id or qty <= 0:
            continue
        # Last write wins if a category appears more than once.
        seen[cat_id] = {'qty': qty, 'unit': (str(ln.get('unit') or '').strip())}

    # Remove lines whose category isn't in the payload.
    po.planned_lines.exclude(category_id__in=seen.keys()).delete()
    # Upsert the rest.
    valid_cat_ids = set(InventoryCategory.objects.filter(pk__in=seen.keys())
                        .values_list('id', flat=True))
    for cat_id, data in seen.items():
        if int(cat_id) not in valid_cat_ids:
            continue
        PurchaseOrderLine.objects.update_or_create(
            purchase_order=po, category_id=cat_id,
            defaults={'quantity': data['qty'], 'unit': data['unit']},
        )


def _po_sync_project_link(po, project_id_raw):
    """Sync the PO's ProjectPurchaseOrder link to match the selected project.

    Called from both create and edit paths. If project_id is provided and
    valid, ensures a link row exists (replacing any prior link so a PO tracks
    at most one project via the dropdown). If project_id is empty/invalid,
    removes all existing links for this PO.
    """
    from core.models import Project, ProjectPurchaseOrder
    pid = (project_id_raw or '').strip()
    if pid and pid.isdigit() and Project.objects.filter(pk=int(pid)).exists():
        pid = int(pid)
        # Replace any existing links for this PO with the selected project.
        ProjectPurchaseOrder.objects.filter(purchase_order=po).exclude(project_id=pid).delete()
        ProjectPurchaseOrder.objects.get_or_create(project_id=pid, purchase_order=po)
    else:
        # No project selected — clear links that were auto-created via dropdown.
        # (Links created manually from the project panel are preserved by the
        # user's explicit save there; this only clears what the dropdown manages.)
        ProjectPurchaseOrder.objects.filter(purchase_order=po).delete()


def _po_project_options():
    """Build a category → subcategory grouped list of active, unfinished Projects
    for the PO modal's project <select> with <optgroup> hierarchy.
    """
    from core.models import Project
    cat_labels = dict(Project.CATEGORY_CHOICES)
    sub_labels = dict(Project.SUBCATEGORY_CHOICES)
    projects = (Project.objects.filter(active=True, is_completed=False)
                .order_by('category', 'subcategory', 'name'))
    groups = {}   # (cat_key, cat_label) → {sub_key → [items]}
    for p in projects:
        ck = p.category
        cl = cat_labels.get(ck, ck)
        sk = p.subcategory or ''
        sl = sub_labels.get(sk, sk) if sk else ''
        g = groups.setdefault((ck, cl), {})
        g.setdefault((sk, sl), []).append({
            'id': p.id, 'name': p.name, 'code': p.code or '',
        })
    return [
        {
            'category': cl,
            'subgroups': [
                {'label': sl or '—', 'items': items}
                for (_, sl), items in sorted(subs.items(), key=lambda x: x[0][0])
            ],
        }
        for (_, cl), subs in sorted(groups.items(), key=lambda x: x[0][0])
    ]


def _serialize_purchase_orders(pos):
    """Serialize a list of PurchaseOrders with 2 queries total (received + planned)
    instead of 2 queries per PO. Builds the category-path map once.
    """
    pos = list(pos)
    if not pos:
        return []
    cat_paths = inventory_category_paths()
    po_ids = [po.id for po in pos]
    received = _po_received_parts_batch(po_ids, cat_paths)
    planned = _po_planned_lines_batch(po_ids, cat_paths)
    out = []
    for po in pos:
        parts, txn_count = received.get(po.id, ([], 0))
        out.append({
            'id': po.id,
            'order_number': po.order_number,
            'po_number': po.po_number,
            'po_amount_untaxed': ('' if po.po_amount_untaxed is None
                                  else f'{po.po_amount_untaxed:.2f}'),
            'project_name': po.project_name,
            'project_code': po.project_code,
            'received_date': po.received_date.isoformat() if po.received_date else '',
            'is_completed': po.is_completed,
            'created_at': po.created_at.strftime('%Y-%m-%d') if po.created_at else '',
            'txn_count': txn_count,
            'planned_lines': planned.get(po.id, []),
            'parts': parts,
        })
    return out


def _serialize_purchase_order(po):
    """Serialize a single PurchaseOrder. Thin wrapper over the batched serializer
    so the create/edit response paths share one code path."""
    return _serialize_purchase_orders([po])[0]


@login_required(login_url='core:login')
@require_POST
def purchase_order_create(request):
    """Create a PurchaseOrder. 灌溉订单编号 is mandatory and must be unique."""
    from core.models import PurchaseOrder
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    from decimal import InvalidOperation

    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)

    order_number = (request.POST.get('order_number') or '').strip()
    if not order_number:
        return JsonResponse({'success': False, 'message': '灌溉订单编号不能为空'}, status=400)
    if PurchaseOrder.objects.filter(order_number=order_number).exists():
        return JsonResponse({'success': False, 'message': '灌溉订单编号已存在'}, status=400)

    amount_raw = (request.POST.get('po_amount_untaxed') or '').strip()
    amount = None
    if amount_raw:
        try:
            from decimal import Decimal
            amount = Decimal(amount_raw)
        except InvalidOperation:
            return JsonResponse({'success': False, 'message': 'PO未税金额格式无效'}, status=400)
    from datetime import date as _date
    received_date = None
    rd_raw = (request.POST.get('received_date') or '').strip()
    if rd_raw:
        try:
            received_date = _date.fromisoformat(rd_raw)
        except ValueError:
            return JsonResponse({'success': False, 'message': '收货日期格式无效'}, status=400)

    po = PurchaseOrder.objects.create(
        order_number=order_number,
        po_number=(request.POST.get('po_number') or '').strip(),
        po_amount_untaxed=amount,
        project_name=(request.POST.get('project_name') or '').strip(),
        project_code=(request.POST.get('project_code') or '').strip(),
        received_date=received_date,
    )
    _po_upsert_planned_lines(po, request.POST.get('lines'))
    # Auto-link the PO to the selected Project via the M2M-through table so it
    # appears in the project's budget panel without manual linking.
    _po_sync_project_link(po, request.POST.get('project_id'))
    return JsonResponse({'success': True, 'message': '采购订单已创建',
                         'node': _serialize_purchase_order(po)})


@login_required(login_url='core:login')
@require_POST
def purchase_order_edit(request, po_id):
    """Edit a PurchaseOrder's fields. 灌溉订单编号 stays unique."""
    from core.models import PurchaseOrder
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    from decimal import Decimal, InvalidOperation

    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)

    po = get_object_or_404(PurchaseOrder, pk=po_id)

    order_number = (request.POST.get('order_number') or '').strip()
    if not order_number:
        return JsonResponse({'success': False, 'message': '灌溉订单编号不能为空'}, status=400)
    if (PurchaseOrder.objects.filter(order_number=order_number)
            .exclude(pk=po_id).exists()):
        return JsonResponse({'success': False, 'message': '灌溉订单编号已存在'}, status=400)

    amount_raw = (request.POST.get('po_amount_untaxed') or '').strip()
    amount = None
    if amount_raw:
        try:
            amount = Decimal(amount_raw)
        except InvalidOperation:
            return JsonResponse({'success': False, 'message': 'PO未税金额格式无效'}, status=400)
    from datetime import date as _date
    received_date = None
    rd_raw = (request.POST.get('received_date') or '').strip()
    if rd_raw:
        try:
            received_date = _date.fromisoformat(rd_raw)
        except ValueError:
            return JsonResponse({'success': False, 'message': '收货日期格式无效'}, status=400)

    po.order_number = order_number
    po.po_number = (request.POST.get('po_number') or '').strip()
    po.po_amount_untaxed = amount
    po.project_name = (request.POST.get('project_name') or '').strip()
    po.project_code = (request.POST.get('project_code') or '').strip()
    po.received_date = received_date
    po.save()
    _po_upsert_planned_lines(po, request.POST.get('lines'))
    # Re-sync the project link (add new, remove old if project changed).
    _po_sync_project_link(po, request.POST.get('project_id'))
    return JsonResponse({'success': True, 'message': '采购订单已更新',
                         'node': _serialize_purchase_order(po)})


@login_required(login_url='core:login')
@require_POST
def purchase_order_delete(request, po_id):
    """Delete a PurchaseOrder. SET_NULL keeps linked 入库流水 intact."""
    from core.models import PurchaseOrder
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN

    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)

    po = get_object_or_404(PurchaseOrder, pk=po_id)
    order_number = po.order_number
    po.delete()
    return JsonResponse({'success': True,
                         'message': f'已删除采购订单「{order_number}」'})


@login_required(login_url='core:login')
@require_POST
def purchase_order_complete(request, po_id):
    """Mark a PurchaseOrder completed: requires the planned materials and the
    received (入库-采购) materials to match exactly — same categories AND same
    per-category quantity. On success, fills received_date with today (if empty)
    and sets is_completed=True. Completed POs remain referenceable by projects.

    The match check reuses the same aggregation the page uses: planned lines
    come from PurchaseOrderLine, received parts from 入库-采购 transaction lines.
    """
    from core.models import PurchaseOrder
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN
    from datetime import date

    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        return JsonResponse({'success': False, 'message': '无权限'}, status=403)

    po = get_object_or_404(PurchaseOrder, pk=po_id)
    if po.is_completed:
        return JsonResponse({'success': False, 'message': '该订单已完成，无需重复操作'})

    # Build {category_id: qty} for planned vs received, then compare exactly.
    cat_paths = inventory_category_paths()
    received, _ = _po_received_parts_batch([po.id], cat_paths).get(po.id, ([], 0))
    planned_map = {}
    for ln in po.planned_lines.all():
        planned_map[ln.category_id] = planned_map.get(ln.category_id, 0) + ln.quantity
    received_map = {}
    for p in received:
        received_map[p['category_id']] = received_map.get(p['category_id'], 0) + p['qty']

    if planned_map != received_map:
        # Build a human-readable diff so the user knows what's still missing/extra.
        all_cids = set(planned_map) | set(received_map)
        diffs = []
        for cid in sorted(all_cids):
            pq = planned_map.get(cid, 0)
            rq = received_map.get(cid, 0)
            if pq != rq:
                nm = next((p['name'] for p in received if p['category_id'] == cid), None) \
                     or next((ln.category.name_zh for ln in po.planned_lines.all()
                              if ln.category_id == cid), f'类别{cid}')
                diffs.append(f'{nm}（计划 {pq} / 已入库 {rq}）')
        detail = '；'.join(diffs[:5]) + ('…' if len(diffs) > 5 else '')
        return JsonResponse({'success': False,
                             'message': '计划采购与已入库物料不一致，无法完成：' + detail})

    if not po.received_date:
        po.received_date = date.today()
    po.is_completed = True
    po.save()
    return JsonResponse({'success': True,
                         'message': f'已确认完成采购订单「{po.order_number}」',
                         'received_date': po.received_date.isoformat() if po.received_date else ''})


@login_required(login_url='core:login')
def purchase_order_export_excel(request):
    """Export all purchase orders to Excel.

    Columns mirror the management table plus two material columns: 计划采购 and
    已入库. Within each cell multiple materials are newline-separated; the two
    columns are aligned by material name so rows correspond.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from core.models import PurchaseOrder
    from core.role_utils import get_user_role, ROLE_MANAGER, ROLE_SUPER_ADMIN

    role = get_user_role(request.user)
    if role not in (ROLE_MANAGER, ROLE_SUPER_ADMIN):
        return redirect('core:dashboard')

    pos = list(PurchaseOrder.objects.all().order_by('-created_at'))
    # Batched serialization (2 queries for all POs, not 2 × len(pos)).
    po_data_list = _serialize_purchase_orders(pos)
    po_data_by_id = {d['id']: d for d in po_data_list}

    wb = Workbook()
    ws = wb.active
    ws.title = '采购订单'
    headers = ['灌溉订单编号', 'PO号', '项目名称', '项目code', '收货日期',
               'PO未税金额', '计划采购', '已入库', '入库次数', '状态', '创建日期']
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='1B4332')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    thin = Side(style='thin', color='D9D0C0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical='top')
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = header_fill; c.font = header_font; c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')

    for po in pos:
        data = po_data_by_id.get(po.id)
        if data is None:
            continue
        # Align planned vs received by category_id, newline-separated.
        planned = {ln['category_id']: ln for ln in data.get('planned_lines', [])}
        received = {ln['category_id']: ln for ln in data.get('parts', [])}
        all_ids = list(planned.keys()) + [k for k in received if k not in planned]
        planned_lines, received_lines = [], []
        for cid in all_ids:
            pl = planned.get(cid); rc = received.get(cid)
            ref = pl or rc
            nm = ref['name']; unit = ref.get('unit', '') or ''
            planned_lines.append('%s ×%d %s' % (nm, pl['qty'] if pl else 0, unit))
            received_lines.append('%s ×%d %s' % (nm, rc['qty'] if rc else 0, unit))

        row = [
            po.order_number,
            po.po_number,
            po.project_name,
            po.project_code,
            po.received_date.isoformat() if po.received_date else '',
            float(po.po_amount_untaxed) if po.po_amount_untaxed is not None else '',
            '\n'.join(planned_lines),
            '\n'.join(received_lines),
            data.get('txn_count', 0),
            '已完成' if po.is_completed else '进行中',
            po.created_at.strftime('%Y-%m-%d') if po.created_at else '',
        ]
        ws.append(row)
        r = ws.max_row
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).border = border
            ws.cell(row=r, column=ci).alignment = wrap_top

    widths = [18, 12, 18, 12, 12, 12, 40, 40, 8, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="purchase_orders.xlsx"'
    return resp


@login_required(login_url='core:login')
def zones_in_area_api(request):
    """API: given a drawn polygon/rect/circle, return zone codes whose centroid falls inside."""
    from core.models import Zone

    shape_type = request.GET.get('type', 'polygon')
    points_json = request.GET.get('points', '[]')
    center_lat = request.GET.get('center_lat')
    center_lng = request.GET.get('center_lng')
    radius = request.GET.get('radius')  # in meters

    try:
        points = json.loads(points_json)
    except (json.JSONDecodeError, TypeError):
        points = []

    def point_in_polygon(lat, lng, polygon):
        """Ray casting algorithm."""
        n = len(polygon)
        inside = False
        j = n - 1
        for i in range(n):
            yi, xi = polygon[i]['lat'], polygon[i]['lng']
            yj, xj = polygon[j]['lat'], polygon[j]['lng']
            if ((yi > lat) != (yj > lat)) and (lng < (xj - xi) * (lat - yi) / (yj - yi) + xi):
                inside = not inside
            j = i
        return inside

    # Get all zones with at least one boundary source (boundary_points OR dxf_boundary_points).
    zones = Zone.objects.exclude(Q(boundary_points=[]) & Q(dxf_boundary_points=[]))
    result_codes = []

    for z in zones:
        bp = z.active_boundary_points
        if not bp:
            continue
        # Calculate centroid
        all_pts = bp if isinstance(bp[0], dict) else []
        if isinstance(bp[0], list):
            all_pts = [p for ring in bp for p in ring]
        if not all_pts:
            continue
        clat = sum(p.get('lat', 0) for p in all_pts) / len(all_pts)
        clng = sum(p.get('lng', 0) for p in all_pts) / len(all_pts)

        if shape_type == 'polygon' and len(points) >= 3:
            if point_in_polygon(clat, clng, points):
                result_codes.append(z.code)
        elif shape_type == 'circle' and center_lat and center_lng and radius:
            from math import radians, cos, sin, asin, sqrt
            dlat = radians(clat - float(center_lat))
            dlng = radians(clng - float(center_lng))
            a = sin(dlat/2)**2 + cos(radians(float(center_lat))) * cos(radians(clat)) * sin(dlng/2)**2
            d = 2 * 6371000 * asin(sqrt(a))  # distance in meters
            if d <= float(radius):
                result_codes.append(z.code)
        elif shape_type == 'rect' and len(points) >= 2:
            lats = [p['lat'] for p in points]
            lngs = [p['lng'] for p in points]
            if min(lats) <= clat <= max(lats) and min(lngs) <= clng <= max(lngs):
                result_codes.append(z.code)

    return JsonResponse({'zone_codes': result_codes, 'count': len(result_codes)})


@login_required(login_url='core:login')
@ensure_csrf_cookie
def zone_dxf_import(request):
    """DXF boundary import page — upload DXF, georeference, assign shapes to zones."""
    import json
    from .models import ManagerProfile, Worker, MapStyleSettings
    from .dxf_utils import (
        parse_dxf_shapes, detect_nesting, detect_coord_system,
        shapes_to_latlng_auto, compute_affine_transform,
        transform_shape, transform_group_to_boundary,
        shapes_to_geojson_preview, auto_calibrate,
)

    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            ManagerProfile.objects.get(user=request.user, active=True)
            is_admin = True
        except ManagerProfile.DoesNotExist:
            pass

    if not is_admin:
        try:
            Worker.objects.get(user=request.user, active=True)
        except Worker.DoesNotExist:
            messages.error(request, '无权限访问此页面')
            return redirect('core:dashboard')

    if request.method == 'POST':
        action = request.POST.get('action', '')

        # ── Upload DXF ──
        if action == 'upload_dxf':
            uploaded = request.FILES.get('file')
            if not uploaded:
                return JsonResponse({'success': False, 'message': '未选择文件'}, status=400)
            if not uploaded.name.lower().endswith('.dxf'):
                return JsonResponse({'success': False, 'message': '请上传 .dxf 格式文件'}, status=400)

            shapes = parse_dxf_shapes(uploaded)
            if isinstance(shapes, dict) and 'error' in shapes:
                return JsonResponse({'success': False, 'message': shapes['error']}, status=400)

            groups = detect_nesting(shapes)
            coord_info = detect_coord_system(shapes)

            # Store in session
            request.session['dxf_shapes'] = shapes
            request.session['dxf_groups'] = groups
            request.session['dxf_filename'] = uploaded.name
            request.session['dxf_coord_info'] = coord_info
            # Clear previous anchors/transformed data
            request.session.pop('dxf_anchors', None)
            request.session.pop('dxf_transformed', None)

            preview = shapes_to_geojson_preview(shapes, groups)

            # If WGS84 auto-detected, pre-transform
            if coord_info.get('type') == 'wgs84' and not coord_info.get('need_anchors'):
                axis_map = coord_info.get('axis_map', 'xy_to_lnglat')
                transformed = shapes_to_latlng_auto(shapes, axis_map)
                request.session['dxf_transformed'] = transformed
                preview['transformed'] = transformed
            elif coord_info.get('auto_calibrated'):
                # Local coords with site calibration — auto-transform
                transformed = auto_calibrate(shapes)
                if transformed:
                    request.session['dxf_transformed'] = transformed
                    preview['transformed'] = transformed
                    coord_info['need_anchors'] = False
                    coord_info['info'] += ' — 已自动配准'

            return JsonResponse({
                'success': True,
                'message': f'识别到 {len(shapes)} 个封闭图形，{len(groups)} 个图形组',
                'filename': uploaded.name,
                'coord_info': coord_info,
                **preview,
            })

        # ── Set anchors (georeference) ──
        if action == 'set_anchors':
            shapes = request.session.get('dxf_shapes')
            groups = request.session.get('dxf_groups')
            if not shapes:
                return JsonResponse({'success': False, 'message': '请先上传DXF文件'}, status=400)

            anchors_raw = request.POST.get('anchors', '[]')
            try:
                anchors = json.loads(anchors_raw)
            except (json.JSONDecodeError, TypeError):
                return JsonResponse({'success': False, 'message': '锚点数据格式错误'}, status=400)

            if len(anchors) < 2:
                return JsonResponse({'success': False, 'message': '需要至少2个锚点'}, status=400)

            transform_fn = compute_affine_transform(anchors)
            if isinstance(transform_fn, str):
                return JsonResponse({'success': False, 'message': transform_fn}, status=400)

            # Transform all shapes
            transformed = [transform_shape(s, transform_fn) for s in shapes]

            request.session['dxf_anchors'] = anchors
            request.session['dxf_transformed'] = transformed

            return JsonResponse({
                'success': True,
                'message': '坐标配准成功',
                'transformed': transformed,
            })

        # ── Preview transform ──
        if action == 'preview_transform':
            shapes = request.session.get('dxf_shapes')
            if not shapes:
                return JsonResponse({'success': False, 'message': '请先上传DXF文件'}, status=400)

            anchors_raw = request.POST.get('anchors', '[]')
            try:
                anchors = json.loads(anchors_raw)
            except (json.JSONDecodeError, TypeError):
                return JsonResponse({'success': False, 'message': '锚点数据格式错误'}, status=400)

            if len(anchors) < 2:
                return JsonResponse({'success': False, 'message': '需要至少2个锚点才能预览', 'status': 'need_more'})

            transform_fn = compute_affine_transform(anchors)
            if isinstance(transform_fn, str):
                return JsonResponse({'success': False, 'message': transform_fn}, status=400)

            transformed = [transform_shape(s, transform_fn) for s in shapes]
            return JsonResponse({
                'success': True,
                'transformed': transformed,
            })

        # ── Assign shape to zone ──
        if action == 'assign_shape':
            shapes = request.session.get('dxf_shapes')
            groups = request.session.get('dxf_groups')
            filename = request.session.get('dxf_filename', '')

            # Prefer client-sent transformed coords (includes drag offsets)
            transformed = None
            transformed_json = request.POST.get('transformed_json', '')
            if transformed_json:
                try:
                    transformed = json.loads(transformed_json)
                except (json.JSONDecodeError, TypeError):
                    pass
            if not transformed:
                transformed = request.session.get('dxf_transformed')
            else:
                # Sync client offset back to session for future assigns / reloads
                request.session['dxf_transformed'] = transformed

            if not transformed:
                # Check if auto WGS84
                coord_info = request.session.get('dxf_coord_info', {})
                if coord_info.get('type') == 'wgs84':
                    transformed = shapes_to_latlng_auto(shapes, coord_info.get('axis_map', 'xy_to_lnglat'))
                else:
                    return JsonResponse({'success': False, 'message': '请先完成坐标配准'}, status=400)

            zone_code = request.POST.get('zone_code', '').strip()

            # Support multi-group: accept group_ids (JSON array) or legacy group_id
            group_ids_raw = request.POST.get('group_ids', '')
            group_id_legacy = request.POST.get('group_id', '').strip()
            if group_ids_raw:
                try:
                    group_ids = json.loads(group_ids_raw)
                except (json.JSONDecodeError, TypeError):
                    group_ids = []
            elif group_id_legacy:
                group_ids = [group_id_legacy]
            else:
                group_ids = []

            if not zone_code:
                return JsonResponse({'success': False, 'message': '缺少区域编号'}, status=400)
            if not group_ids:
                return JsonResponse({'success': False, 'message': '请先选择图形组'}, status=400)

            try:
                zone = Zone.objects.get(code=zone_code)
            except Zone.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'区域 {zone_code} 不存在'}, status=404)

            # Build boundary from one or more groups
            # Multi-group format: [[outer1, hole1a, ...], [outer2, hole2a, ...]]
            # Single-group (backward compat): [outer, hole1, hole2, ...]
            target_groups = []
            for gid in group_ids:
                for g in (groups or []):
                    if g['id'] == gid:
                        target_groups.append(g)
                        break

            if not target_groups:
                return JsonResponse({'success': False, 'message': '图形组不存在'}, status=404)

            if len(target_groups) == 1:
                # Single group — flat ring list (backward compatible)
                tg = target_groups[0]
                boundary_rings = []
                outer_t = transformed[tg['outer']]
                boundary_rings.append(outer_t['vertices_latlng'])
                for hole_idx in tg['holes']:
                    hole_t = transformed[hole_idx]
                    boundary_rings.append(hole_t['vertices_latlng'])
                zone.dxf_boundary_points = boundary_rings
            else:
                # Multi group — nested format: [[group1_rings], [group2_rings], ...]
                boundary_groups = []
                for tg in target_groups:
                    group_rings = []
                    outer_t = transformed[tg['outer']]
                    group_rings.append(outer_t['vertices_latlng'])
                    for hole_idx in tg['holes']:
                        hole_t = transformed[hole_idx]
                        group_rings.append(hole_t['vertices_latlng'])
                    boundary_groups.append(group_rings)
                zone.dxf_boundary_points = boundary_groups

            zone.dxf_boundary_source = filename
            zone.boundary_source = 'dxf'
            zone.save()

            group_label = f'{len(target_groups)}个图形组' if len(target_groups) > 1 else '图形'
            return JsonResponse({
                'success': True,
                'message': f'已将{group_label}分配给区域 {zone.code}',
                'zone_id': zone.id,
                'zone_code': zone.code,
                'zone_name': zone.name,
                'boundary_source': zone.boundary_source,
                'dxf_boundary_points': zone.dxf_boundary_points,
                'area_display': zone.area_display,
            })

        # ── Switch boundary source ──
        if action == 'switch_source':
            zone_code = request.POST.get('zone_code', '').strip()
            source = request.POST.get('source', 'manual').strip()

            if source not in ('manual', 'dxf'):
                return JsonResponse({'success': False, 'message': '无效的来源类型'}, status=400)

            try:
                zone = Zone.objects.get(code=zone_code)
            except Zone.DoesNotExist:
                return JsonResponse({'success': False, 'message': f'区域 {zone_code} 不存在'}, status=404)

            if source == 'dxf' and not zone.dxf_boundary_points:
                return JsonResponse({'success': False, 'message': f'区域 {zone_code} 没有DXF边界数据'}, status=400)

            zone.boundary_source = source
            zone.save()

            label = 'DXF导入' if source == 'dxf' else '手动绘制'
            return JsonResponse({
                'success': True,
                'message': f'区域 {zone.code} 已切换为「{label}」',
                'zone_code': zone.code,
                'boundary_source': zone.boundary_source,
                'area_display': zone.area_display,
            })

        return JsonResponse({'success': False, 'message': '未知操作'}, status=400)

    # ── GET: render page ──
    all_zones = []
    for z in Zone.objects.select_related('patch').order_by('code').only(
        'id', 'code', 'name', 'patch_id', 'patch__name',
        'boundary_source', 'dxf_boundary_points',
    ):
        all_zones.append({
            'id': z.id,
            'code': z.code,
            'name': z.name,
            'patch_name': z.patch.name if z.patch else '',
            'boundary_source': z.boundary_source,
            'has_dxf_boundary': bool(z.dxf_boundary_points),
        })

    # All drawn zones for reference layer (use active boundary)
    all_drawn_zones = []
    for z in Zone.objects.select_related('patch').only(
        'id', 'code', 'name', 'boundary_points', 'dxf_boundary_points',
        'boundary_color', 'boundary_source',
        'label_lat', 'label_lng', 'label_scale', 'label_angle',
        'smooth_override', 'ring_display_modes', 'patch_id', 'patch__name'
    ):
        active_bp = z.active_boundary_points
        if not active_bp:
            continue
        all_drawn_zones.append({
            'id': z.id,
            'code': z.code,
            'name': z.name,
            'boundary_points': active_bp,
            'boundary_color': z.boundary_color,
            'label_lat': z.label_lat,
            'label_lng': z.label_lng,
            'label_scale': z.label_scale,
            'label_angle': z.label_angle,
            'smooth_override': z.smooth_override,
            'ring_display_modes': z.ring_display_modes or {},
            'patch_id': z.patch_id,
            'patch_name': z.patch.name if z.patch else '',
            'area_display': z.area_display,
        })

    # Restore session DXF data for page reload
    session_dxf = None
    if request.session.get('dxf_transformed'):
        session_dxf = {
            'shapes': request.session.get('dxf_shapes', []),
            'groups': request.session.get('dxf_groups', []),
            'transformed': request.session.get('dxf_transformed', []),
            'coord_info': request.session.get('dxf_coord_info', {}),
            'filename': request.session.get('dxf_filename', ''),
        }

    context = {
        'all_zones_json': json.dumps(all_zones),
        'all_drawn_zones_json': json.dumps(all_drawn_zones),
        'map_style_json': json.dumps(MapStyleSettings.get_style()),
        'session_dxf_json': json.dumps(session_dxf) if session_dxf else 'null',
    }
    return render(request, 'core/zone_dxf_import.html', context)
