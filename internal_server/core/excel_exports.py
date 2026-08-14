"""Reusable Excel workbook builders for the export endpoints.

Each public ``build_<report>_excel`` function returns a
``(filename_without_path, io.BytesIO)`` tuple — the Excel bytes written to an
in-memory buffer — so the same generation logic can be shared by:

  * the interactive download views (``views.py`` / ``workorder_tree_views.py``),
    which parse ``request.GET`` and wrap the result in an ``HttpResponse``; and
  * a scheduled-email job, which can call the build functions directly with
    plain Python arguments (no request object).

The builders are pure: they take explicit Python params (dates / query strings /
a user), do the data fetching + workbook styling, and return the bytes. They
must produce byte-identical output to the inline code that previously lived in
the views (same columns, same ``1B4332`` header fill, same borders, column
widths, merged headers, sheet names, and filename patterns).

Helpers that live in ``core.views`` (``_work_report_count_columns``,
``_scoped_work_reports_qs``, the irrigation matrix helpers, the zone export
header/priority maps, ``_serialize_purchase_orders``, ``inventory_category_paths``)
and in ``core.inventory_tree_views`` (``current_price_for``) are imported
lazily INSIDE each build function — the same way the views did — so that
importing ``excel_exports`` at module load never triggers ``core.views``,
which itself imports ``excel_exports`` lazily inside the views (circular import
safe).
"""

import io
from collections import defaultdict
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.utils import timezone


# ===========================================================================
# 1) Zone export
# ===========================================================================

def build_zone_export_excel():
    """Build the Zone-info Excel export.

    Mirrors the former inline body of ``views.zone_export_excel`` (from the
    permission check onward). Returns ``(filename, BytesIO)``.

    No params: exports every zone, ordered by code.
    """
    from core.models import Zone
    from core.views import _ZONE_EXPORT_HEADERS, _PRIORITY_EXPORT_MAP

    zones = Zone.objects.select_related('patch').order_by('code')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Zone info 仅色块'

    # Write headers with same styling as V0
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col_idx, header in enumerate(_ZONE_EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.alignment = center_wrap

    # Set column widths matching V0
    col_widths = [16, 13, 14, 13, 18, 13, 16, 13, 14, 13, 12, 8, 9, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Write data rows
    for zone in zones:
        row_num = ws.max_row + 1
        # CCU number from patch code (e.g. "CCU1" → "1")
        ccu_num = zone.patch.code.replace('CCU', '') if zone.patch else ''

        values = [
            zone.code,                                          # A: 编号
            _PRIORITY_EXPORT_MAP.get(zone.priority, ''),         # B: 位置重要程度
            zone.land.name if zone.land else None,              # C: 所属Land
            zone.name,                                          # D: 通用名称
            zone.description,                                   # E: 灌溉管理用的位置
            zone.current_status or None,                        # F: 当前状态
            zone.sprinkler_type or None,                        # G: 灌水器类型
            zone.irrigation_intensity,                          # H: 灌溉强度
            zone.area_display if zone.area_sqm else None,       # I: 区域面积
            zone.patch.name if zone.patch else None,            # J: 灌溉分区
            int(ccu_num) if ccu_num.isdigit() else ccu_num,     # K: CCU编号
            zone.solenoid_valve_size,                           # L: 电磁阀尺寸
            zone.landscape_coefficient,                         # M: 景观系数
            zone.plant_type or None,                            # N: 植物类型
            zone.irrigation_foreman or None,                    # O: 灌溉领班
            zone.greenery_zone or None,                         # P: 绿化分区
            zone.greenery_foreman or None,                      # Q: 绿化领班
            zone.pest_control_zone or None,                     # R: 植保分区
            zone.pest_control_foreman or None,                  # S: 植保领班
            zone.terrain_feature or None,                       # T: 地形特点
            zone.plant_feature or None,                         # U: 植物特点
            zone.soil_moisture or None,                         # V: 土壤湿度
            zone.equipment_maintenance_notes or None,           # W: 灌溉设备维护记录
            zone.irrigation_management_notes or None,           # X: 灌溉管理以往记录
        ]
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_num, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"zones_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return filename, buf


# ===========================================================================
# 2) Work reports export
# ===========================================================================

def build_work_reports_excel(start_date, end_date, user):
    """Build the work-reports Excel export for a date window.

    Mirrors the former inline body of ``views.work_reports_excel`` (after the
    ``_resolve_stats_window`` call). ``start_date`` / ``end_date`` are
    ``date`` objects; ``user`` is the request user (used for role scoping via
    ``_scoped_work_reports_qs`` + ``is_admin``). Returns
    ``(filename, BytesIO)``.
    """
    from django.db.models import Prefetch
    from core.models import (WorkReport, WorkReportEntry, WorkItem,
                             InventoryTransaction, InventoryTransactionLine)
    from core.role_utils import is_admin
    from core.views import (
        _work_report_count_columns, _scoped_work_reports_qs,
        _dedup_zone_names, _report_section_label, _report_materials_summary,
    )

    admin = is_admin(user)

    # 1) Fixed fault-matrix columns (deterministic across exports). Each entry
    # is (work_item_id, [seg, seg, leaf]) — the group hierarchy with the section
    # root stripped, so we can render a proper multi-row merged header instead
    # of cramming the whole breadcrumb into one cell.
    count_nodes = _work_report_count_columns()

    # 2) Work orders in window (role-scoped). _scoped_work_reports_qs already
    # prefetches 'entries', so fetch the count-only subset into a distinct attr
    # via to_attr to avoid a duplicate-prefetch conflict. Also prefetch the
    # FULL entry set (not just count-type) with each entry's work_item parent
    # chain + linked project, so the 工作类别细分 + 关联项目 columns can be built
    # without N+1 queries.
    qs = _scoped_work_reports_qs(user, admin).filter(date__gte=start_date, date__lte=end_date)
    reports = list(qs.prefetch_related(
        Prefetch('entries', queryset=WorkReportEntry.objects.select_related('work_item').filter(work_item__value_type='count'), to_attr='_count_entries'),
        Prefetch('entries', queryset=WorkReportEntry.objects.select_related(
            'work_item', 'work_item__parent', 'work_item__parent__parent',
            'work_item__parent__parent__parent', 'work_item__parent__parent__parent__parent',
            'work_item__parent__parent__parent__parent__parent',
            'project'), to_attr='_all_entries'),
        Prefetch('material_consumptions', queryset=InventoryTransaction.objects.prefetch_related(
            Prefetch('lines', queryset=InventoryTransactionLine.objects.select_related('category'), to_attr='_lines')
        ), to_attr='_materials_txn'),
    ).order_by('date', 'id'))

    # 3) Build workbook with a grouped (merged) multi-row header.
    wb = Workbook()
    ws = wb.active
    ws.title = '维修记录'
    base_header = ['序号', '日期', '工单号', '处理人', '位置', '工作分类',
                   '工作类别细分', '关联项目',
                   '故障/事件位置', '区域', '灌溉组人数', '灌溉组工时',
                   '第三方人数', '第三方工时', '消耗材料', '备注',
                   '信息来源', '疑难问题', '疑难已处理']
    n_base = len(base_header)
    n_cols = n_base + len(count_nodes)
    hdr_rows = max((len(segs) for _, segs in count_nodes), default=1)

    # Header styling. Style every header cell up front so merged ranges keep
    # their borders (openpyxl only draws borders per underlying cell).
    hfill = PatternFill('solid', fgColor='1B4332')
    hfont = Font(color='FFFFFF', bold=True, size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='FFFFFF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for rr in range(1, hdr_rows + 1):
        for cc in range(1, n_cols + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = hfill; cell.font = hfont; cell.alignment = center; cell.border = border

    # Base columns span every header row (vertical merge).
    for ci, label in enumerate(base_header, 1):
        ws.cell(row=1, column=ci, value=label)
        if hdr_rows > 1:
            ws.merge_cells(start_row=1, start_column=ci, end_row=hdr_rows, end_column=ci)

    # Count columns: walk each depth level and merge consecutive siblings that
    # share the same path prefix. Columns are sorted by path, so siblings are
    # always contiguous. Shorter paths get merged down to fill remaining rows.
    for depth in range(hdr_rows):
        row = depth + 1
        ci = 0
        while ci < len(count_nodes):
            segs = count_nodes[ci][1]
            if depth >= len(segs):
                ci += 1          # already merged down at a shallower terminal depth
                continue
            prefix = segs[:depth + 1]
            cj = ci + 1
            while cj < len(count_nodes) and count_nodes[cj][1][:depth + 1] == prefix:
                cj += 1
            col_start = n_base + ci + 1
            col_end = n_base + cj              # inclusive
            ws.cell(row=row, column=col_start, value=segs[depth])
            need_h = col_end > col_start                 # siblings → merge across
            need_v = depth == len(segs) - 1 and row < hdr_rows   # terminal → merge down
            if need_h and need_v:              # one rectangular block merge
                ws.merge_cells(start_row=row, start_column=col_start,
                               end_row=hdr_rows, end_column=col_end)
            elif need_h:
                ws.merge_cells(start_row=row, start_column=col_start,
                               end_row=row, end_column=col_end)
            elif need_v:
                ws.merge_cells(start_row=row, start_column=col_start,
                               end_row=hdr_rows, end_column=col_start)
            # else: single cell, value already set — no merge needed
            ci = cj

    # work_item_id → matrix column index (1-based, offset after the base cols).
    id_to_col = {wid: n_base + i + 1 for i, (wid, _) in enumerate(count_nodes)}

    for idx, r in enumerate(reports, 1):
        # Zone codes: show only when the workorder covers ≤5 zones (sum of all
        # linked zone rows, regardless of duplicate names). Over 5 → blank, since
        # a sprawling multi-zone ticket's codes would be unreadable in one cell.
        zone_codes = ''
        zones = list(r.zones.all())
        if len(zones) <= 5:
            zone_codes = ', '.join(z.code for z in zones if z.code)
        # 工作类别细分: the full multi-level breadcrumb of each entry's work_item
        # (e.g. 常规维护 → 维保定期检查 → 喷头 → 喷头盖掉/松/坏). A ticket can have
        # several entries, so collect the DISTINCT breadcrumbs and join with \n.
        # Prefetched _all_entries carries the parent chain + project already.
        seen_paths = []
        seen_set = set()
        proj_names = []
        proj_seen = set()
        for e in getattr(r, '_all_entries', []):
            wi = e.work_item
            if wi is None:
                continue
            # Walk the cached parent chain (prefetched 5 levels deep, enough for
            # the current tree; deeper nodes resolve lazily if ever needed).
            chain = []
            node = wi
            while node is not None:
                chain.append(node.name_zh)
                node = node.parent
            chain.reverse()
            # Drop the top-level section — it already has its own column
            # (工作分类), so repeating it here is redundant. Always strip the
            # first element (the section root); what's left is the level-2+
            # detail. A level-0-only entry (just the section) becomes empty,
            # which is correct — nothing more specific to show.
            chain = chain[1:]
            breadcrumb = ' → '.join(chain)
            if breadcrumb and breadcrumb not in seen_set:
                seen_set.add(breadcrumb)
                seen_paths.append(breadcrumb)
            if e.project_id and e.project:
                pname = e.project.name
                if pname not in proj_seen:
                    proj_seen.add(pname)
                    proj_names.append(pname)
        section_detail = '\n'.join(seen_paths)
        linked_project = '、'.join(proj_names)
        row = [idx, r.date.isoformat() if r.date else '',
               r.display_number if r.id else '',
               r.worker.full_name if r.worker_id and r.worker else '',
               r.location.code if r.location_id and r.location else '',
               _report_section_label(r),
               section_detail,
               linked_project,
               _dedup_zone_names(r.zone_names),
               zone_codes,
               r.team_size or '',
               r.team_hours if r.team_hours else '',
               r.third_party_count or '',
               r.third_party_hours if r.third_party_hours else '',
               _report_materials_summary(r),
               (r.work_content or r.remark or ''),
               '',  # 信息来源 (no dedicated field on WorkReport)
               '是' if r.is_difficult else '',
               '是' if r.is_difficult_resolved else '']
        # pad matrix columns so indices line up
        row += [None] * len(count_nodes)
        for e in getattr(r, '_count_entries', []):
            ci = id_to_col.get(e.work_item_id)
            if ci is not None and e.count:
                row[ci - 1] = (row[ci - 1] or 0) + e.count
        ws.append(row)

    # Data cell styling: thin grey borders everywhere, centre numerics, wrap 备注.
    gside = Side(style='thin', color='D0D0D0')
    gborder = Border(left=gside, right=gside, top=gside, bottom=gside)
    dcenter = Alignment(horizontal='center', vertical='center')
    dwrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    last_row = hdr_rows + len(reports)
    # Wrap the long text columns: 工作类别细分 (multi-line breadcrumb),
    # 区域 (codes list), 消耗材料, 备注 (free text).
    # 1序号 2日期 3工单号 4处理人 5位置 6工作分类 7工作类别细分 8关联项目
    # 9故障位置 10区域 11灌溉组人数 12灌溉组工时 13第三方人数 14第三方工时
    # 15消耗材料 16备注 17信息来源 18疑难问题 19疑难已处理
    wrap_cols = {7, 10, 15, 16}
    for rr in range(hdr_rows + 1, last_row + 1):
        for cc in range(1, n_cols + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.border = gborder
            cell.alignment = dwrap if cc in wrap_cols else dcenter

    # Column widths + freeze the header rows and the first five ID columns
    # (序号/日期/工单号/处理人/位置) so they stay visible while scrolling the matrix.
    base_widths = {3: 10, 4: 11, 5: 12, 6: 14,   # 序号/日期/工单号/处理人/位置/工作分类
                   7: 34, 8: 16,                 # 工作类别细分 (wide, multi-line) / 关联项目
                   9: 18, 10: 20,                # 故障位置 / 区域
                   15: 18, 16: 28}               # 消耗材料 / 备注
    for ci in range(1, n_cols + 1):
        if ci <= n_base:
            width = base_widths.get(ci, 11)
        else:
            leaf = count_nodes[ci - n_base - 1][1][-1]
            width = max(8, min(20, len(leaf) * 1.7 + 2))
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = ws.cell(row=hdr_rows + 1, column=6)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f'workreports_{start_date}_{end_date}.xlsx'
    return fname, buf


# ===========================================================================
# 3) Irrigation report export
# ===========================================================================

def build_irrigation_report_excel(date_from, date_to, ccu_ids=None):
    """Build the per-CCU satellite runtime Excel report.

    Mirrors the former inline body of ``views.irrigation_report_excel`` (after
    the request param parsing). ``date_from`` / ``date_to`` are the compact
    Maxicom timestamp strings (``YYYYMMDDHHMM`` / ``YYYYMMDDHH`` etc.) exactly
    as the dashboard uses — they are embedded verbatim in the filename and
    padded to 14 chars for the timestamp range filter.

    ``ccu_ids`` (optional) restricts the export to a subset of CCU Patch ids —
    used by the scheduled-email feature to distribute specific CCUs to the
    managers responsible for them. ``None``/empty = all CCUs (default).

    Returns ``(filename, BytesIO)``.
    """
    from core.models import MaxicomController, MaxicomRuntime
    from core.views import (_irrig_data_span, _mapped_station_ids,
                            _ccu_queryset, _build_ccu_matrix)

    ts_from = date_from.ljust(14, '0')[:14] if date_from else ''
    ts_to = date_to.ljust(14, '9')[:14] if date_to else ''

    ctrl_map = {c.mdb_index: c for c in MaxicomController.objects.exclude(name__icontains='CCU')}
    rt_qs = MaxicomRuntime.objects.select_related('station', 'station__parent', 'site')
    if ts_from:
        rt_qs = rt_qs.filter(timestamp__gte=ts_from)
    if ts_to:
        rt_qs = rt_qs.filter(timestamp__lte=ts_to)
    rt_by_site = defaultdict(list)
    for rt in rt_qs:
        rt_by_site[rt.site_id].append(rt)

    # Station Patch IDs claimed by at least one Zone — rows whose station is
    # NOT in this set are flagged red as orphan (no Zone backing = un-attributable).
    mapped_station_ids = _mapped_station_ids()

    # --- styles ---
    header_fill = PatternFill('solid', fgColor='1B4332')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    side_fill = PatternFill('solid', fgColor='F5F0E8')
    total_fill = PatternFill('solid', fgColor='EDE8DC')
    total_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='D9D0C0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # Orphan row (station has no Zone backing): light-red wash + red label font.
    orphan_fill = PatternFill('solid', fgColor='FDECEA')
    orphan_font = Font(color='C62828', bold=True)

    wb = Workbook()
    # remove the default sheet (we add per-CCU sheets)
    default_ws = wb.active

    ccus = _ccu_queryset()
    # Restrict to a CCU subset when an explicit id list is supplied (scheduled
    # email "distribute specific CCUs" feature). Keeps the natural numeric order.
    if ccu_ids:
        wanted = set(int(x) for x in ccu_ids if str(x).strip().lstrip('-').isdigit())
        ccus = [c for c in ccus if c.id in wanted]

    for idx, ccu in enumerate(ccus):
        m = _build_ccu_matrix(ccu, rt_by_site.get(ccu.id, []), ctrl_map, mapped_station_ids)
        controllers = m['controllers']
        rows = m['rows']
        col_totals = m['col_totals']
        grand_total = m['grand_total']
        ran = sum(1 for r in rows if r['total'] > 0)
        orphan_count = sum(1 for r in rows if r.get('orphan') and r['total'] > 0)

        # Excel sheet names: max 31 chars, no special chars
        sheet_name = f"{ccu.code}_{ccu.name}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        # title row
        ws.cell(row=1, column=1,
                value=f"{ccu.code} ({ccu.name}) — 站点 × 卫星控制器 运行时间（分钟）  {date_from}~{date_to}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12, color='1B4332')
        subtitle = f"卫星控制器 {len(controllers)} 个，运行站点 {ran} 个，总运行 {grand_total} 分钟"
        if orphan_count:
            subtitle += f"  ·  其中 {orphan_count} 个运行站点无 Zone 对应（红色标记）"
        ws.cell(row=2, column=1, value=subtitle)
        ws.cell(row=2, column=1).font = Font(size=9, color='6B7B6E')

        # matrix starts at row 4
        header_row = 4
        n_sat = len(controllers)
        # header (no 合计 column — totals add no value, see dashboard change)
        ws.cell(row=header_row, column=1, value='站#')
        for ci, cname in enumerate(controllers):
            ws.cell(row=header_row, column=2 + ci, value=cname)
        for col in range(1, 2 + n_sat):
            c = ws.cell(row=header_row, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
            c.border = border

        # 24 station rows (no per-row 合计 cell)
        for ri, r in enumerate(rows):
            rownum = header_row + 1 + ri
            sc = ws.cell(row=rownum, column=1, value=r['station'])
            sc.fill = side_fill
            sc.alignment = center
            sc.border = border
            orphan_vals = r.get('orphan_values') or []
            for ci, v in enumerate(r['values']):
                is_orphan = ci < len(orphan_vals) and orphan_vals[ci]
                vc = ws.cell(row=rownum, column=2 + ci, value=v if v else None)
                vc.alignment = center
                vc.border = border
                if is_orphan:
                    # Orphan valve: flat light-red fill regardless of value
                    # (no border emphasis — fill alone signals "no Zone backing").
                    vc.fill = orphan_fill
                    if v:
                        vc.font = orphan_font
                elif v:
                    vc.font = Font(bold=True)

        # column widths + freeze (no extra width for the removed 合计 column)
        ws.column_dimensions['A'].width = 6
        for ci in range(n_sat):
            ws.column_dimensions[get_column_letter(2 + ci)].width = 11
        ws.freeze_panes = 'B5'

    # remove default empty sheet
    if default_ws is not None and default_ws.title in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(default_ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'irrigation_report_{date_from}_{date_to}.xlsx'
    return fname, buf


# ===========================================================================
# 4) Inventory catalog export
# ===========================================================================

def build_inventory_catalog_excel():
    """Build the inventory catalog Excel export.

    Mirrors the former inline body of ``views.inventory_export_excel`` (after
    the role check). No params: exports every part leaf with its price history.
    Returns ``(filename, BytesIO)``.
    """
    from core.models import InventoryCategory, InventoryPriceRecord, InventoryTransactionLine
    from core.inventory_tree_views import current_price_for

    # Build the ancestor chain (root → ... → leaf) for every part leaf.
    cats = {c.id: c for c in InventoryCategory.objects.filter(active=True)}
    # Prefetch all price records into a {cat_id: [records]} map so we don't
    # N+1 query inside the part loop (catalog can be 300+ parts).
    price_map = {}
    for rec in (InventoryPriceRecord.objects
                .select_related('category').order_by('category_id', '-date', '-id')):
        price_map.setdefault(rec.category_id, []).append(rec)
    # Prefetch stock-in/out history per part (出入库记录) for the new read-only
    # column. One batched query across all parts (uses the core_invline_cat
    # index), grouped by category_id — same shape as price_map above. Most
    # recent first so the cell reads newest→oldest.
    txn_map = {}
    part_ids = [cid for c in cats.values() if c.node_type == 'part' for cid in [c.id]]
    if part_ids:
        # Only ledger-CONFIRMED transactions count in the per-item 出入库记录
        # column — pending (awaiting manager) and removed stay out of stats.
        for ln in (InventoryTransactionLine.objects
                   .filter(category_id__in=part_ids,
                           transaction__confirm_status='confirmed')
                   .select_related('transaction')
                   .order_by('-transaction__date', '-id')):
            txn_map.setdefault(ln.category_id, []).append(ln)

    parts = []
    for c in cats.values():
        if c.node_type != 'part':
            continue
        chain = []
        node = c
        while node:
            chain.append(node.name_zh)
            node = cats.get(node.parent_id) if node.parent_id else None
        chain.reverse()   # root → leaf
        parts.append((chain, c))

    wb = Workbook()
    ws = wb.active
    ws.title = '库存目录'
    # First 8 cols = part info; cols 9-11 = current price snapshot (part row only);
    # col 12 = record count; cols 13-16 = per-record detail (price rows only).
    # col 17 (编码) is the hidden round-trip key for import. col 18 (出入库记录)
    # is appended AFTER it so the importer's hardcoded col-17 lookup is unaffected;
    # it's a read-only history summary (date + signed qty per line, newest first).
    def _txn_cell(lines):
        """Format a part's stock-in/out history into one multi-line cell.
        e.g. "2026-07-22 出库 -10个\\n2026-07-20 入库 +50个". Newest first."""
        if not lines:
            return ''
        out = []
        for ln in lines:
            t = ln.transaction
            d = t.date.isoformat() if t.date else '?'
            sign = '+' if t.operation == '入库' else '-'
            unit = ln.unit or ''
            # 入库/出库 + optional subtype (采购/项目 etc.) for context.
            kind = t.operation + ('-' + t.entry_subtype if t.entry_subtype else '')
            out.append(f'{d} {kind} {sign}{ln.quantity}{unit}')
        return '\n'.join(out)

    headers = ['大类别', '小类别', '系列', '品名', '是否主材', '单位', '最小库存', '现有库存',
               '当前单价', '当前供应商', '当前价格日期', '记录数',
               '采购日期', '供应商', '历史单价', '单位',
               '编码',   # col 17: hidden key for import round-trip (cat.code)
               '出入库记录']   # col 18: read-only stock history (part row only)
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill(start_color='1B4332', end_color='1B4332', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # Part-row styling: light green band so it's visually distinct from the
    # price-detail rows beneath it.
    part_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    part_font_bold = Font(bold=True)
    price_font = Font(color='6B7280', size=10)
    price_indent = Alignment(horizontal='left', indent=2)   # nudge detail rows right

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    ri = 2
    for chain, cat in parts:
        # Map the variable-length chain onto 4 fixed columns (same logic as before).
        name = chain[-1]
        big = chain[0] if chain else ''
        mid = chain[1:-1] if len(chain) > 2 else []
        if len(mid) >= 2:
            small, series = mid[0], mid[1]
        elif len(mid) == 1:
            small = series = mid[0]
        else:
            small = series = big
        records = price_map.get(cat.id, [])
        # Resolve current price once per part. current_price_for honors the
        # is_current flag (else falls back to newest-by-date).
        cur = current_price_for(cat)
        cur_price = float(cur.unit_price) if cur else ''
        cur_supplier = cur.supplier if cur else ''
        cur_date = cur.date.isoformat() if (cur and cur.date) else ''
        # Part row
        vals = [big, small, series, name,
                '是' if cat.is_main_material else '',
                cat.unit or '',
                cat.min_stock or 0, cat.current_stock or 0,
                cur_price, cur_supplier, cur_date, len(records),
                '', '', '', '',
                cat.code,   # col 17: stable key for import (hidden in Excel)
                _txn_cell(txn_map.get(cat.id, []))]   # col 18: 出入库记录
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = border
            cell.fill = part_fill
            cell.font = part_font_bold
            if ci in (9, 12):
                cell.alignment = Alignment(horizontal='right')
            elif ci == 18:
                # Multi-line history cell: wrap + top-align so the stacked
                # date/qty records read cleanly within the row.
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        ri += 1
        # Price-detail rows (one per record) — col 17 intentionally blank so
        # the importer can tell part rows (code present) from detail rows.
        for rec in records:
            dvals = ['', '', '', name, '', rec.unit or cat.unit or '', '', '',
                     '', '', '', '',
                     rec.date.isoformat() if rec.date else '',
                     rec.supplier or '',
                     float(rec.unit_price),
                     rec.unit or '',
                     '',   # col 17 blank (not a part row)
                     '']   # col 18 blank (history only on the part row)
            for ci, v in enumerate(dvals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = border
                cell.font = price_font
                if ci == 4:
                    cell.alignment = price_indent   # repeat 品名 indented for context
                elif ci == 15:
                    cell.alignment = Alignment(horizontal='right')
            ri += 1

    # Column widths (18 cols). Use get_column_letter for safety beyond Z.
    widths = [18, 22, 22, 28, 10, 8, 10, 10, 12, 18, 14, 9, 14, 18, 12, 8, 10, 34]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    # Hide the code column (col 17 = Q) — it's a round-trip key for import,
    # not meant for human editing. Excel keeps the data but hides the column.
    ws.column_dimensions[get_column_letter(17)].hidden = True
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = 'inventory_catalog.xlsx'
    return fname, buf


# ===========================================================================
# 5) Inventory transactions (ledger) export
# ===========================================================================

def build_inventory_transactions_excel(date_from=None, date_to=None):
    """Build the 出入库记录 ledger Excel export.

    Mirrors the former inline body of ``views.inventory_transactions_export``
    (after the role check). ``date_from`` / ``date_to`` are optional
    ``YYYY-MM-DD`` ISO strings (the values the dashboard passes through
    ``?from=`` / ``?to=``); ``None`` falls back to the last 30 days, matching
    the original view's default. Returns ``(filename, BytesIO)``.
    """
    from core.models import InventoryCategory, InventoryTransaction

    today = date.today()
    date_from = date_from or (today - timedelta(days=30)).isoformat()
    date_to = date_to or today.isoformat()

    # Build cat_id → full hierarchy path (same as the ledger tab).
    cats = {c.id: c for c in InventoryCategory.objects.filter(active=True)}
    cat_paths = {}
    for c in cats.values():
        chain = []
        p = c.parent
        while p:
            chain.append(p.name_zh)
            p = cats.get(p.parent_id) if p.parent_id else None
        chain.reverse()
        cat_paths[c.id] = ' › '.join(chain + [c.name_zh]) if chain else c.name_zh

    txns = (InventoryTransaction.objects
            .filter(date__gte=date_from, date__lte=date_to)
            .exclude(confirm_status='removed')   # removed stays out of exports too
            .select_related('worker', 'related_project', 'zone', 'work_report')
            .prefetch_related('lines__category')
            .order_by('-date', '-id'))

    confirm_labels = dict(InventoryTransaction.CONFIRM_STATUS_CHOICES)
    wb = Workbook()
    ws = wb.active
    ws.title = '出入库记录'
    headers = ['日期', '类型', '物料', '数量', '单位', '经办人', '去向/来源', '关联', '备注', '确认状态']
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill(start_color='1B4332', end_color='1B4332', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = border

    ri = 2
    for t in txns:
        op = f"{t.operation}-{t.entry_subtype}" if t.entry_subtype else t.operation
        ctx = ''
        if t.operation == '出库' and t.entry_subtype == '项目' and t.related_project:
            ctx = '项目 · ' + t.related_project.name
        elif t.operation == '出库' and t.entry_subtype == '借用' and t.counterparty:
            ctx = '借用 · ' + t.counterparty
        elif t.operation == '入库' and t.entry_subtype == '采购' and t.order_no:
            ctx = '采购 · 订单 ' + t.order_no
        else:
            ctx = t.entry_subtype or ''
        link = ''
        if t.work_report_id:
            link = f'工单 #{t.work_report_id}'
        elif t.zone:
            link = t.zone.code or t.zone.name_zh
        for ln in t.lines.all():
            vals = [
                t.date.isoformat() if t.date else '',
                op,
                cat_paths.get(ln.category_id, ln.category.name_zh if ln.category_id else ''),
                ('+' if t.operation == '入库' else '-') + str(ln.quantity),
                ln.unit or '',
                t.worker.full_name if t.worker_id and t.worker else '',
                ctx, link, t.remark or '',
                confirm_labels.get(t.confirm_status, t.confirm_status),
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = border
                if ci == 4:
                    cell.font = Font(bold=True, color='2D6A4F' if t.operation == '入库' else 'c0392b')
            ri += 1

    for ci, w in enumerate([12, 14, 36, 10, 8, 12, 22, 14, 24, 10], 1):
        ws.column_dimensions[chr(64 + ci)].width = w
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = 'inventory_transactions.xlsx'
    return fname, buf


# ===========================================================================
# 6) Purchase order export
# ===========================================================================

def build_purchase_order_excel(q='', completed=None):
    """Build the purchase-order Excel export.

    Mirrors the former inline body of ``views.purchase_order_export_excel``
    (after the role check). ``q`` filters by order_number / po_number /
    project_name / project_code (case-insensitive contains); ``completed`` is
    ``True``/``False`` to restrict to a state, or ``None`` for all. Returns
    ``(filename, BytesIO)``.

    Note: the legacy view exported every PO unconditionally, so the default
    args (``q=''``, ``completed=None``) reproduce that exact output.
    """
    from django.db.models import Q
    from core.models import PurchaseOrder
    from core.views import _serialize_purchase_orders

    qs = PurchaseOrder.objects.all()
    if q:
        qs = qs.filter(
            Q(order_number__icontains=q) | Q(po_number__icontains=q)
            | Q(project_name__icontains=q) | Q(project_code__icontains=q)
        )
    if completed is True:
        qs = qs.filter(is_completed=True)
    elif completed is False:
        qs = qs.filter(is_completed=False)
    pos = list(qs.order_by('-created_at'))
    # Batched serialization (2 queries for all POs, not 2 × len(pos)).
    po_data_list = _serialize_purchase_orders(pos)
    po_data_by_id = {d['id']: d for d in po_data_list}

    wb = Workbook()
    ws = wb.active
    ws.title = '采购订单'
    headers = ['灌溉订单编号', 'PO号', '项目名称', '项目code', '收货日期',
               'PO未税金额', '计划采购', '已入库', '入库次数', '状态', '创建日期']
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='1B4332')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    thin = Side(style='thin', color='D9D0C0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical='top')
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = header_fill; c.font = header_font; c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')

    for po in pos:
        data = po_data_by_id.get(po.id)
        if data is None:
            continue
        # Align planned vs received by category_id, newline-separated.
        planned = {ln['category_id']: ln for ln in data.get('planned_lines', [])}
        received = {ln['category_id']: ln for ln in data.get('parts', [])}
        all_ids = list(planned.keys()) + [k for k in received if k not in planned]
        planned_lines, received_lines = [], []
        for cid in all_ids:
            pl = planned.get(cid); rc = received.get(cid)
            ref = pl or rc
            nm = ref['name']; unit = ref.get('unit', '') or ''
            planned_lines.append('%s ×%d %s' % (nm, pl['qty'] if pl else 0, unit))
            received_lines.append('%s ×%d %s' % (nm, rc['qty'] if rc else 0, unit))

        row = [
            po.order_number,
            po.po_number,
            po.project_name,
            po.project_code,
            po.received_date.isoformat() if po.received_date else '',
            float(po.po_amount_untaxed) if po.po_amount_untaxed is not None else '',
            '\n'.join(planned_lines),
            '\n'.join(received_lines),
            data.get('txn_count', 0),
            '已完成' if po.is_completed else '进行中',
            timezone.localtime(po.created_at).strftime('%Y-%m-%d') if po.created_at else '',
        ]
        ws.append(row)
        r = ws.max_row
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).border = border
            ws.cell(row=r, column=ci).alignment = wrap_top

    widths = [18, 12, 18, 12, 12, 12, 40, 40, 8, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = 'purchase_orders.xlsx'
    return fname, buf


# ===========================================================================
# 7) Project export
# ===========================================================================

def build_project_export_excel(q='', category='', completed=None,
                               date_from=None, date_to=None):
    """Build the project-management Excel export.

    Mirrors the former inline body of ``workorder_tree_views.project_export_excel``
    (after the role check and filter parsing). Params reproduce the dashboard's
    filters:

      * ``q`` — free-text search across name / symbol / code / notes.
      * ``category`` — a Project.CATEGORY_CHOICES key, or ``''`` for all.
      * ``completed`` — ``True``/``False`` to filter state, or ``None`` for all.
      * ``date_from`` / ``date_to`` — ``date`` objects bounding ``start_date``
        (projects with a null start_date are excluded when either bound is set).

    Returns ``(filename, BytesIO)``.
    """
    from django.db.models import Q
    from core.models import Project
    from core.workorder_tree_views import _project_summaries, _project_budget_data

    qs = Project.objects.all()
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(symbol__icontains=q)
                       | Q(code__icontains=q) | Q(notes__icontains=q))
    if category in {c for c, _ in Project.CATEGORY_CHOICES}:
        qs = qs.filter(category=category)
    if completed in (True, False):
        qs = qs.filter(is_completed=completed)
    if date_from or date_to:
        qs = qs.exclude(start_date__isnull=True)
        if date_from:
            qs = qs.filter(start_date__gte=date_from)
        if date_to:
            qs = qs.filter(start_date__lte=date_to)
    projects = list(qs.order_by('category', 'subcategory', 'name'))
    summ = _project_summaries(projects)
    budget_data = _project_budget_data(projects)

    cat_labels = dict(Project.CATEGORY_CHOICES)
    sub_labels = dict(Project.SUBCATEGORY_CHOICES)

    wb = Workbook()
    ws = wb.active
    ws.title = '项目管理'
    headers = ['项目', '类别', '子类别', '开工日期', '计划完工日期',
               '材料预算金额', '人工预算金额', '材料余额', '人工余额',
               '灌溉组工时', '第三方工时', '工单数', '状态', '备注']
    ws.append(headers)

    # Styles.
    header_fill = PatternFill('solid', fgColor='1B4332')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    thin = Side(style='thin', color='D9D0C0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical='top')
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = header_fill; c.font = header_font; c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')

    for p in projects:
        s = summ[p.id]
        b = budget_data.get(p.id, {})
        # 人工余额 = 人工预算 − 已用工时×率。Decimal 与 float 不能直接运算，统一转 float。
        tr = p.team_rate
        tpr = p.third_party_rate
        labor_cost = (s['hours'] * float(tr) if tr is not None else 0) + \
                     (s['third_hours'] * float(tpr) if tpr is not None else 0)
        labor_bal = (float(p.labor_budget_amount) - labor_cost) \
            if p.labor_budget_amount is not None else ''
        row = [
            p.name,
            cat_labels.get(p.category, p.category),
            sub_labels.get(p.subcategory, p.subcategory or ''),
            p.start_date.isoformat() if p.start_date else '',
            p.planned_end_date.isoformat() if p.planned_end_date else '',
            float(p.material_budget_amount) if p.material_budget_amount is not None else '',
            float(p.labor_budget_amount) if p.labor_budget_amount is not None else '',
            float(b.get('material_balance') or 0) if b.get('material_balance') != '' else '',
            labor_bal,
            round(s['hours'], 1),
            round(s['third_hours'], 1),
            len(s['reports']),
            '已完成' if p.is_completed else '进行中',
            p.notes,
        ]
        ws.append(row)
        r = ws.max_row
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).border = border
            ws.cell(row=r, column=ci).alignment = wrap_top

    # Column widths.
    widths = [20, 10, 10, 12, 12, 14, 14, 14, 14, 10, 10, 8, 8, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = 'projects.xlsx'
    return fname, buf

# ===========================================================================
# 8) Single-project detail export — three sheets: 项目信息汇总 (one header row
#    + one value row), 关联工单 (confirmed work orders with their material
#    consumption), 采购订单 (linked POs).
# ===========================================================================

def build_project_detail_excel(project):
    """Build the single-project detail Excel export (3 sheets).

    1. 项目信息汇总 — the project's header fields laid out horizontally
       (headers in row 1, the project's values in row 2).
    2. 关联工单 — confirmed work orders (hours/materials switches), one row per
       order, with its material consumption lines newline-joined in one cell.
    3. 采购订单 — POs linked to the project.

    Aggregates reuse ``_project_summaries`` / ``_project_budget_data`` so
    numbers match the on-screen panel. Returns ``(filename, BytesIO)``.
    """
    from core.models import InventoryTransaction
    from core.workorder_tree_views import _project_summaries, _project_budget_data

    summ = _project_summaries([project])[project.id]
    b = _project_budget_data([project]).get(project.id, {})
    cat_labels = dict(project.CATEGORY_CHOICES)
    sub_labels = dict(project.SUBCATEGORY_CHOICES)

    wb = Workbook()
    header_fill = PatternFill('solid', fgColor='1B4332')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    thin = Side(style='thin', color='D9D0C0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical='top')

    def styled_sheet(title, headers, rows, widths):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for ci in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=ci)
            c.fill = header_fill; c.font = header_font; c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
        for row in rows:
            ws.append(row)
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=ci).border = border
                ws.cell(row=ws.max_row, column=ci).alignment = wrap_top
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return ws

    # ── Sheet 1: 项目信息汇总 — headers in one row, values in the next ────
    ws = wb.active
    ws.title = '项目信息汇总'
    info_headers = ['项目名称', '项目代号', '项目Code', '类别', '子类别', '状态',
                    '开工日期', '计划完工日期', '材料预算金额', '人工预算金额',
                    '灌溉组费率', '第三方费率', '材料余额', '关联采购金额合计',
                    '灌溉组工时(已确认)', '第三方工时(已确认)']
    info_values = [
        project.name, project.symbol or '', project.code or '',
        cat_labels.get(project.category, project.category),
        sub_labels.get(project.subcategory, project.subcategory or ''),
        '已完成' if project.is_completed else '进行中',
        project.start_date.isoformat() if project.start_date else '',
        project.planned_end_date.isoformat() if project.planned_end_date else '',
        b.get('material_budget_amount', ''), b.get('labor_budget_amount', ''),
        b.get('team_rate', ''), b.get('third_party_rate', ''),
        b.get('material_balance', ''), b.get('po_amount_total', ''),
        round(summ['hours'], 1), round(summ['third_hours'], 1),
    ]
    ws.append(info_headers)
    ws.append(info_values)
    for ci in range(1, len(info_headers) + 1):
        h = ws.cell(row=1, column=ci)
        h.fill = header_fill; h.font = header_font; h.border = border
        h.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        v = ws.cell(row=2, column=ci)
        v.border = border; v.alignment = wrap_top
    for i, w in enumerate([24, 12, 14, 10, 10, 8, 12, 12, 13, 13, 10, 10, 12, 14, 13, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    # ── Sheet 2: 关联工单 — confirmed orders with material consumption ───
    reports = b.get('work_reports', [])
    # Per-work-order material lines for THIS project: actual, ledger-confirmed
    # or pending (removed excluded — dismissed from stats). One query, grouped
    # by work_report_id; rendered newline-joined inside one cell.
    wr_ids = [wr['id'] for wr in reports]
    mat_by_wr = {}
    sec_by_wr = {}      # 工作类别: SECTION_CHOICES labels, deduped
    subpath_by_wr = {}  # 子类别(所有层级): WorkItem path minus the section root
    remark_by_wr = {}
    if wr_ids:
        txns = (InventoryTransaction.objects
                .filter(work_report_id__in=wr_ids, related_project=project,
                        operation='出库', entry_subtype='项目',
                        consumption_mode='actual')
                .exclude(confirm_status='removed')
                .prefetch_related('lines__category'))
        for t in txns:
            cell = mat_by_wr.setdefault(t.work_report_id, [])
            for ln in t.lines.all():
                cell.append('%s ×%s%s' % (ln.category.name_zh, ln.quantity,
                                          ln.unit or (ln.category.unit or '')))

        # Work-content columns. Prefer entries bound to THIS project (that's why
        # the report appears in this sheet); fall back to all its entries when
        # the binding was recorded only at report level.
        from core.models import WorkItem, WorkReport, WorkReportEntry
        from core.workorder_tree_views import workitem_path_map
        path_lookup = workitem_path_map()
        sec_labels = dict(WorkItem.SECTION_CHOICES)
        entries = (WorkReportEntry.objects.filter(work_report_id__in=wr_ids)
                   .select_related('work_item'))
        by_wr_proj, by_wr_all = {}, {}
        for e in entries:
            by_wr_all.setdefault(e.work_report_id, []).append(e)
            if e.project_id == project.id:
                by_wr_proj.setdefault(e.work_report_id, []).append(e)
        for rid in wr_ids:
            for e in (by_wr_proj.get(rid) or by_wr_all.get(rid) or []):
                wi = e.work_item
                if not wi:
                    continue
                if wi.section:
                    lbl = sec_labels.get(wi.section, wi.section)
                    if lbl not in sec_by_wr.setdefault(rid, []):
                        sec_by_wr[rid].append(lbl)
                # Full path starts with the section-root name (duplicates
                # 工作类别) — strip that first segment for the 子类别 column.
                full = path_lookup.get(wi.id, wi.name_zh)
                segs = [x.strip() for x in full.split('›')]
                sub = '›'.join(segs[1:]) if len(segs) > 1 else full
                if sub and sub not in subpath_by_wr.setdefault(rid, []):
                    subpath_by_wr[rid].append(sub)
        for row_ in WorkReport.objects.filter(id__in=wr_ids).values('id', 'remark'):
            remark_by_wr[row_['id']] = row_['remark'] or ''
    wo_rows = [
        ['#%d' % wr['id'], wr['date'],
         {'early': '早班', 'day': '白班', 'night': '夜班'}.get(wr.get('shift'), wr.get('shift') or ''),
         (wr.get('start_time') + '-' + wr.get('end_time')) if wr.get('start_time') else '',
         wr.get('worker_name') or '',
         wr.get('team_hours') or 0, wr.get('third_hours') or 0,
         '、'.join(sec_by_wr.get(wr['id'], [])) or '—',
         '\n'.join(subpath_by_wr.get(wr['id'], [])) or '—',
         remark_by_wr.get(wr['id'], ''),
         wr.get('created_at') or '',
         '\n'.join(mat_by_wr.get(wr['id'], [])) or '—']
        for wr in reports
    ]
    styled_sheet('关联工单',
                 ['工单号', '日期', '班次', '起止时间', '处理人', '灌溉组工时',
                  '第三方工时', '工作类别', '子类别(所有层级)', '备注', '创建时间',
                  '物料消耗明细'],
                 wo_rows,
                 [10, 12, 8, 13, 10, 11, 11, 13, 26, 22, 17, 40])

    # ── Sheet 3: 采购订单 — POs linked to this project ───────────────────
    from core.models import PurchaseOrder
    po_ids = [po['id'] for po in b.get('linked_pos', [])]
    po_objs = {p.id: p for p in PurchaseOrder.objects.filter(id__in=po_ids)}
    po_rows = [[po['order_number'], po['po_number'] or '',
                po['po_amount_untaxed'],
                po_objs[po['id']].received_date.isoformat()
                    if po_objs[po['id']].received_date else '',
                po_objs[po['id']].project_name or '',
                po_objs[po['id']].project_code or '',
                '已完成' if po_objs[po['id']].is_completed else '进行中']
               for po in b.get('linked_pos', [])]
    styled_sheet('采购订单',
                 ['灌溉订单编号', 'PO号', 'PO未税金额', '收货日期', '项目名称',
                  '项目Code', '状态'],
                 po_rows,
                 [20, 16, 13, 12, 20, 14, 8])

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    import re as _re
    safe_name = _re.sub(r'[^\w\u4e00-\u9fff-]', '', project.name)[:40] or 'project'
    fname = 'project_%d_%s.xlsx' % (project.id, safe_name)
    return fname, buf
