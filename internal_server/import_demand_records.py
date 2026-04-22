"""
Import script for 002 需求周报.xlsx.
Run: python3 manage.py shell < import_demand_records.py
"""

import openpyxl
import re
from datetime import time
from core.models import DemandRecord, DemandCategory, DemandDepartment, Zone

# ==========================================================================
# Excel文件路径
# ==========================================================================

EXCEL_FILE = '/Users/chen/development/maxicom/002 需求周报.xlsx'
SHEET_NAME = '26Q1'

# ==========================================================================
# 行映射配置
# ==========================================================================

# 全局事件行（停水停电、项目施工）
GLOBAL_EVENT_ROWS = {
    3: ('停水停电', 'water_power_stop'),
    4: ('项目施工', 'project_construction'),
}

# 区域行（映射到Zone）- 行号: (显示名, 可能的Zone名称匹配)
ZONE_ROWS = {
    6: '01西门',
    7: '02东门',
    8: '03TL',
    9: '03TL水池',
    10: '05FL408',
    11: '06FL406',
    12: '07AI&ME',
    13: '08TC',
    14: '09Garden',
    15: '10H2',
    16: '11H1',
    17: '12RDE',
    18: '13PTC',
    19: '14GPL',
    20: '35-2探索路',
    21: '35-5北泵站',
    22: '36北环路',
    23: '37-3西环路',
    24: '37-6西泵站',
    25: '38-3西南环',
    26: '38-3西喷泉',
    27: '38-4西南环',
    28: '38-6高架边',
    29: '39-03南环路',
    30: '39-10星光道',
    31: '39-11灵感街',
    32: '40-3奇妙路',
    33: '43-1东大湖',
    34: '44-3南大湖',
    35: '45-2西大湖',
    36: 'Sitewalk1',
    37: 'Sitewalk2',
}

# 工作类别行（映射到DemandCategory）
CATEGORY_ROWS = {
    38: ('项目配合', 'project_support'),
    39: ('配合走场', 'site_walk'),
    40: ('Improvement', 'improvement'),
    41: ('Learning', 'learning'),
    42: ('询价1', 'inquiry_1'),
    43: ('询价2', 'inquiry_2'),
    44: ('养护维修1', 'maintenance_1'),
    45: ('养护维修2', 'maintenance_2'),
    46: ('项目维修', 'project_maintenance'),
    47: ('咨询', 'consultation'),
    48: ('设计', 'design'),
    49: ('团队', 'team'),
    51: ('安全', 'safety'),
    52: ('Meeting', 'meeting'),
    53: ('Visit', 'visit'),
    54: ('奖惩', 'reward_penalty'),
}

# 其他行（Row 5天气信息等）
OTHER_ROWS = {5}  # 天气/低温等信息行

# ==========================================================================
# 时间段解析函数
# ==========================================================================

def parse_time_segment(text):
    """
    解析时间段，格式如：
    - "2300-600" → 23:00, 06:00 (跨天)
    - "400-700" → 04:00, 07:00
    - "巡道 400-700" → 提取时间段

    Returns: (start_time, end_time, crosses_midnight, remaining_text)
    """
    if not text or not isinstance(text, str):
        return None, None, False, text

    # 匹配时间段格式：HHMM-HHMM 或 HH:MM-HH:MM
    # 也匹配 HMM-HMM（如 400-700）
    patterns = [
        r'(\d{1,2}):(\d{2})[-~至](\d{1,2}):(\d{2})',  # HH:MM-HH:MM
        r'(\d{3,4})[-~至](\d{3,4})',  # HMM-HMM or HHMM-HHMM
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            groups = match.groups()

            if len(groups) == 4:  # HH:MM format
                start_hour = int(groups[0])
                start_min = int(groups[1])
                end_hour = int(groups[2])
                end_min = int(groups[3])
            else:  # HMM or HHMM format
                start_str = groups[0]
                end_str = groups[1]

                # 解析：可能是 HHMM 或 HMM
                if len(start_str) == 4:
                    start_hour = int(start_str[:2])
                    start_min = int(start_str[2:])
                elif len(start_str) == 3:
                    start_hour = int(start_str[:1])
                    start_min = int(start_str[1:])
                else:  # len == 2 or 1
                    start_hour = int(start_str) if len(start_str) <= 2 else int(start_str[:2])
                    start_min = 0

                if len(end_str) == 4:
                    end_hour = int(end_str[:2])
                    end_min = int(end_str[2:])
                elif len(end_str) == 3:
                    end_hour = int(end_str[:1])
                    end_min = int(end_str[1:])
                else:
                    end_hour = int(end_str) if len(end_str) <= 2 else int(end_str[:2])
                    end_min = 0

            # 规范化小时（超过23视为次日）
            start_hour = start_hour % 24
            end_hour = end_hour % 24

            try:
                start_time = time(start_hour, start_min)
                end_time = time(end_hour, end_min)
            except ValueError:
                return None, None, False, text

            # 判断跨天（开始>=12且结束<12，或者开始>结束）
            crosses_midnight = (start_hour >= 12 and end_hour < 12) or (start_hour > end_hour)

            # 提取剩余内容
            remaining = text.replace(match.group(), '').strip()

            return start_time, end_time, crosses_midnight, remaining

    return None, None, False, text


# ==========================================================================
# Zone匹配函数
# ==========================================================================

def find_zone(zone_text):
    """尝试根据文本匹配Zone记录。"""
    if not zone_text:
        return None

    # 清理文本
    clean_text = zone_text.strip()

    # 尝试精确匹配
    zone = Zone.objects.filter(name__iexact=clean_text).first()
    if zone:
        return zone

    # 尝试code匹配
    zone = Zone.objects.filter(code__iexact=clean_text).first()
    if zone:
        return zone

    # 尝试部分匹配（去掉空格等）
    zone = Zone.objects.filter(name__icontains=clean_text.replace(' ', '')).first()
    if zone:
        return zone

    # 尝试数字编号匹配（如 "01西门" 匹配包含"01"的）
    code_match = re.match(r'^(\d+)', clean_text)
    if code_match:
        zone = Zone.objects.filter(code__startswith=code_match.group(1)).first()
        if zone:
            return zone

    return None


# ==========================================================================
# 主导入函数
# ==========================================================================

def import_demand_records():
    """从Excel导入需求记录。"""

    print(f"Loading Excel: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb[SHEET_NAME]

    print(f"Sheet: {SHEET_NAME}, Rows: {sheet.max_row}, Columns: {sheet.max_column}")

    # 预加载Zone和Category缓存
    zone_cache = {}
    category_cache = {}

    # 预加载Zone映射
    for row_num, zone_name in ZONE_ROWS.items():
        zone = find_zone(zone_name)
        zone_cache[row_num] = zone

    # 预加载Category映射
    for row_num, (cat_name, cat_code) in CATEGORY_ROWS.items():
        category_cache[row_num] = DemandCategory.objects.filter(code=cat_code).first()

    # 全局事件Category
    global_event_cache = {}
    for row_num, (cat_name, cat_code) in GLOBAL_EVENT_ROWS.items():
        global_event_cache[row_num] = DemandCategory.objects.filter(code=cat_code).first()

    # 统计计数
    stats = {
        'total_cells': 0,
        'imported': 0,
        'time_parsed': 0,
        'zone_matched': 0,
        'global_events': 0,
        'category_records': 0,
        'skipped': 0,
    }

    # 遍历列（日期列）
    # 每2列一组：日期列在奇数位置（3, 5, 7...），对应数据在偶数位置（4, 6, 8...）
    # 实际结构：Row 1是星期，Row 2是日期

    date_columns = []
    for col_idx in range(3, sheet.max_column + 1, 2):
        date_val = sheet.cell(row=2, column=col_idx).value
        if date_val and hasattr(date_val, 'strftime'):
            date_columns.append((col_idx, date_val))

    print(f"Found {len(date_columns)} date columns")

    # 遍历每个日期列
    for col_idx, date_val in date_columns:
        date_str = date_val.strftime('%Y-%m-%d')

        # 遍历每个行类别
        for row_idx in range(3, sheet.max_row + 1):
            # 跳过空行和天气行
            if row_idx in OTHER_ROWS:
                continue

            # 获取单元格内容
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if not cell_value:
                continue

            stats['total_cells'] += 1

            # 转换为字符串
            content = str(cell_value).strip()
            if not content:
                continue

            # 判断行类型并创建记录
            record_data = {
                'date': date_val,
                'original_text': content,
                'status': DemandRecord.STATUS_APPROVED,  # 历史数据默认已批准
            }

            # 解析时间段
            start_time, end_time, crosses_midnight, remaining_text = parse_time_segment(content)
            if start_time:
                record_data['start_time'] = start_time
                record_data['end_time'] = end_time
                record_data['crosses_midnight'] = crosses_midnight
                record_data['time_parsed'] = True
                record_data['content'] = remaining_text or content
                stats['time_parsed'] += 1
            else:
                record_data['content'] = content

            # 处理全局事件
            if row_idx in GLOBAL_EVENT_ROWS:
                cat_name, cat_code = GLOBAL_EVENT_ROWS[row_idx]
                record_data['is_global_event'] = True
                record_data['category'] = global_event_cache.get(row_idx)
                record_data['category_text'] = cat_name
                record_data['zone_text'] = None
                stats['global_events'] += 1

            # 处理区域需求
            elif row_idx in ZONE_ROWS:
                zone_name = ZONE_ROWS[row_idx]
                record_data['zone'] = zone_cache.get(row_idx)
                record_data['zone_text'] = zone_name
                record_data['category'] = DemandCategory.objects.filter(code='zone_demand').first()
                record_data['category_text'] = '区域需求'
                if record_data['zone']:
                    stats['zone_matched'] += 1

            # 处理工作类别
            elif row_idx in CATEGORY_ROWS:
                cat_name, cat_code = CATEGORY_ROWS[row_idx]
                record_data['category'] = category_cache.get(row_idx)
                record_data['category_text'] = cat_name
                record_data['zone_text'] = None
                stats['category_records'] += 1

            else:
                # 其他行跳过
                stats['skipped'] += 1
                continue

            # 创建记录
            try:
                DemandRecord.objects.create(**record_data)
                stats['imported'] += 1
            except Exception as e:
                print(f"Error creating record at Row {row_idx}, Col {col_idx}: {e}")

    # 输出统计
    print("\n=== Import Statistics ===")
    print(f"Total cells with data: {stats['total_cells']}")
    print(f"Records imported: {stats['imported']}")
    print(f"Time parsed: {stats['time_parsed']}")
    print(f"Zone matched: {stats['zone_matched']}")
    print(f"Global events: {stats['global_events']}")
    print(f"Category records: {stats['category_records']}")
    print(f"Skipped: {stats['skipped']}")

    return stats


# ==========================================================================
# 运行入口
# ==========================================================================

if __name__ == '__main__':
    print("Starting import...")
    import_demand_records()
    print("Import completed!")

# For shell execution
print("Running import_demand_records()...")
import_demand_records()