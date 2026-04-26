#!/usr/bin/env python
"""Import work report data from 001 工作日报.xlsx into Django database."""

import os
from datetime import datetime

import openpyxl

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django
django.setup()

from django.db import transaction
from core.models import (
    Worker, Location, WorkCategory, InfoSource, Zone,
    FaultCategory, FaultSubType, WorkReport, WorkReportFault,
)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', '001 工作日报.xlsx')
DATA_START_ROW = 7
DATE_COL = 3       # C
WEATHER_COL = 2    # B
WORKER_COL = 4     # D
LOCATION_COL = 5   # E
CATEGORY_COL = 6   # F
ZONE_COL = 7       # G
REMARK_COL = 8     # H
INFO_SRC_COL = 9   # I
DIFFICULT_COL = 10 # J
RESOLVED_COL = 11  # K
FAULT_START_COL = 12  # L


def build_fault_col_map(ws):
    """Build mapping: col_index -> FaultSubType from row 5 Chinese headers."""
    row5 = [cell.value for cell in ws[5]]
    subtypes = {}
    for fst in FaultSubType.objects.select_related('category').all():
        key = fst.name_zh.replace('\n', '').replace(' ', '').strip()
        subtypes[key] = fst

    col_map = {}
    for col_idx in range(FAULT_START_COL - 1, len(row5)):
        header = row5[col_idx]
        if not header:
            continue
        clean = str(header).replace('\n', '').replace(' ', '').strip()
        if clean and clean in subtypes:
            col_map[col_idx] = subtypes[clean]
    return col_map


def clean_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        val = val.strip()
        if '小计' in val or '合计' in val:
            return None
        try:
            return datetime.strptime(val[:10], '%Y-%m-%d').date()
        except ValueError:
            return None
    return None


def is_summary_row(row_vals):
    """Check if this is a weekly/yearly summary row (skip it)."""
    date_val = row_vals[DATE_COL - 1]
    if date_val is None:
        worker = row_vals[WORKER_COL - 1]
        if worker is None:
            fault_vals = row_vals[FAULT_START_COL - 1:]
            return any(v and v != 0 for v in fault_vals)
    elif isinstance(date_val, str) and ('小计' in date_val or '合计' in date_val):
        return True
    return False


def main():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['维修记录']

    # Preload dictionaries
    workers = {w.full_name: w for w in Worker.objects.all()}
    locations = {}
    for loc in Location.objects.all():
        locations[loc.name] = loc
        locations[loc.code] = loc
    categories = {c.name: c for c in WorkCategory.objects.all()}
    info_sources = {i.name: i for i in InfoSource.objects.all()}

    # Preload existing zones by code
    zone_map = {z.code: z for z in Zone.objects.all()}

    # First pass: collect all unique zone codes from Excel
    all_zone_codes = set()
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        row_vals = [cell.value for cell in ws[row_idx]]
        if is_summary_row(row_vals):
            continue
        date_val = clean_date(row_vals[DATE_COL - 1])
        if date_val is None:
            continue
        worker_name = (row_vals[WORKER_COL - 1] or '')
        if not worker_name or not str(worker_name).strip():
            continue
        zone_code = str(row_vals[ZONE_COL - 1] or '').strip()
        if zone_code:
            all_zone_codes.add(zone_code)

    # Create missing zones
    new_zones = []
    for code in sorted(all_zone_codes):
        if code not in zone_map:
            z = Zone(code=code, name=code)
            new_zones.append(z)
    if new_zones:
        Zone.objects.bulk_create(new_zones)
        print(f"Created {len(new_zones)} new zones")

    # Refresh zone_map
    zone_map = {z.code: z for z in Zone.objects.all()}
    print(f"Total zones: {len(zone_map)}")

    fault_col_map = build_fault_col_map(ws)

    print(f"Loaded: {len(workers)} workers, {len(locations)} locations, "
          f"{len(categories)} categories, {len(info_sources)} info sources")
    print(f"Mapped {len(fault_col_map)} fault subtype columns")

    # Collect all data rows
    rows_to_import = []
    skipped = 0

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        row_vals = [cell.value for cell in ws[row_idx]]

        if is_summary_row(row_vals):
            skipped += 1
            continue

        date_val = clean_date(row_vals[DATE_COL - 1])
        if date_val is None:
            all_empty = all(v is None for v in row_vals[1:11])
            if all_empty:
                continue
            skipped += 1
            continue

        worker_name = str(row_vals[WORKER_COL - 1] or '').strip()
        if not worker_name:
            skipped += 1
            continue

        rows_to_import.append((row_idx, row_vals, date_val, worker_name))

    print(f"Found {len(rows_to_import)} data rows to import ({skipped} skipped)")

    # Import in batches
    imported = 0
    errors = 0
    missing_workers = set()
    missing_locations = set()
    missing_categories = set()

    with transaction.atomic():
        for batch_start in range(0, len(rows_to_import), 200):
            batch = rows_to_import[batch_start:batch_start + 200]

            reports = []
            fault_entries = []

            for row_idx, row_vals, date_val, worker_name in batch:
                worker = workers.get(worker_name)
                if not worker:
                    missing_workers.add(worker_name)
                    errors += 1
                    continue

                loc_name = str(row_vals[LOCATION_COL - 1] or '').strip()
                location = locations.get(loc_name)
                if not location:
                    missing_locations.add(loc_name)
                    errors += 1
                    continue

                cat_name = str(row_vals[CATEGORY_COL - 1] or '').strip()
                category = categories.get(cat_name)
                if not category:
                    missing_categories.add(cat_name)
                    errors += 1
                    continue

                weather = str(row_vals[WEATHER_COL - 1] or '').strip() or ''
                zone_code = str(row_vals[ZONE_COL - 1] or '').strip()
                zone = zone_map.get(zone_code) if zone_code else None
                remark = str(row_vals[REMARK_COL - 1] or '').strip() or ''

                info_name = str(row_vals[INFO_SRC_COL - 1] or '').strip()
                info_source = info_sources.get(info_name) if info_name else None

                difficult_raw = row_vals[DIFFICULT_COL - 1]
                is_difficult = bool(difficult_raw and str(difficult_raw).strip())

                resolved_raw = row_vals[RESOLVED_COL - 1]
                is_resolved = bool(resolved_raw and str(resolved_raw).strip())

                report = WorkReport(
                    date=date_val,
                    weather=weather,
                    worker=worker,
                    location=location,
                    work_category=category,
                    zone_location=zone,
                    remark=remark,
                    info_source=info_source,
                    is_difficult=is_difficult,
                    is_difficult_resolved=is_resolved,
                )
                reports.append(report)

                # Collect fault entries for this row
                row_faults = []
                for col_idx, subtype in fault_col_map.items():
                    if col_idx < len(row_vals):
                        val = row_vals[col_idx]
                        if val and int(val) > 0:
                            row_faults.append((subtype, int(val)))

                fault_entries.append(row_faults)

            # Bulk create reports
            created_reports = WorkReport.objects.bulk_create(reports)

            # Create fault entries
            fault_objects = []
            for report, row_faults in zip(created_reports, fault_entries):
                for subtype, count in row_faults:
                    fault_objects.append(WorkReportFault(
                        work_report=report,
                        fault_subtype=subtype,
                        count=count,
                    ))

            if fault_objects:
                WorkReportFault.objects.bulk_create(fault_objects)

            imported += len(created_reports)
            print(f"  Imported {imported}/{len(rows_to_import)}...")

    print(f"\nDone! Imported: {imported}, Errors: {errors}")
    if missing_workers:
        print(f"  Missing workers: {missing_workers}")
    if missing_locations:
        print(f"  Missing locations: {missing_locations}")
    if missing_categories:
        print(f"  Missing categories: {missing_categories}")

    # Verify
    total = WorkReport.objects.count()
    total_faults = WorkReportFault.objects.count()
    print(f"\nDatabase now: {total} work reports, {total_faults} fault entries, {Zone.objects.count()} zones")


if __name__ == '__main__':
    main()
